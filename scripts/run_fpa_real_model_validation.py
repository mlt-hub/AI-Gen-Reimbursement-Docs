"""Run FPA real-model validation against selected golden cases.

This script intentionally prints only validation summaries. It does not print
API keys, prompts, or full fixture contents.
"""

from __future__ import annotations

import json
import shutil
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from ai_gen_reimbursement_docs.config_utils import load_api_key, load_base_url, load_model_name
from ai_gen_reimbursement_docs.gen_fpa import (
    _read_fpa_rows_md_for_audit,
    generate_fpa_check_xlsx_from_md,
    plan_fpa_md_from_tree,
)


FIXTURE_DIR = ROOT / "tests" / "fixtures" / "fpa_golden_cases"
OUT_ROOT = ROOT / "tmp_fpa_real_model_validation"
TEMPLATE = ROOT / "data" / "out_templates" / "FPA工作量评估-输出模板.xlsx"
CASES = [
    "mixed_internal_external_data_functions",
    "payment_gateway_refund",
    "master_data_org_reference",
]


def _write_meta_md(path: Path, meta: dict[str, str]) -> None:
    rows = "\n".join(f"| {key} | {value} |" for key, value in meta.items())
    path.write_text(f"# 文档元数据\n\n{rows}\n", encoding="utf-8")


def _write_tree_md(path: Path, rows: list[dict[str, str]]) -> None:
    lines = [
        "| 入口 | 一级模块 | 二级模块 | 三级模块 | 客户端类型 | 三级模块整体功能描述 | 功能过程 | 功能过程描述 | 变更状态 |",
        "|------|---------|---------|---------|----------|----------------------|----------|--------------|--------------|",
    ]
    for row in rows:
        lines.append(
            "| "
            + " | ".join(
                [
                    "后台",
                    row.get("一级模块", ""),
                    row.get("二级模块", ""),
                    row.get("三级模块", ""),
                    row.get("客户端类型", ""),
                    row.get("三级模块整体功能描述", ""),
                    row.get("功能过程", ""),
                    row.get("功能过程描述", ""),
                    row.get("变更状态", ""),
                ]
            )
            + " |"
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _sheet_rows(ws) -> list[dict[str, object]]:
    headers = [cell.value for cell in ws[1]]
    return [
        dict(zip(headers, values))
        for values in ws.iter_rows(min_row=2, values_only=True)
        if any(value not in (None, "") for value in values)
    ]


def _strict_expected(case: dict[str, object]) -> list[dict[str, str]]:
    expected = case.get("expected", {})
    if not isinstance(expected, dict):
        return []
    rows = expected.get("strict_fpa", [])
    if not isinstance(rows, list):
        return []
    return [
        {"name": str(item.get("name", "")), "type": str(item.get("type", ""))}
        for item in rows
        if isinstance(item, dict)
    ]


def _case_summary(case_id: str) -> dict[str, object]:
    case = json.loads((FIXTURE_DIR / f"{case_id}.json").read_text(encoding="utf-8"))
    out_dir = OUT_ROOT / case_id
    out_dir.mkdir(parents=True, exist_ok=True)

    tree_md = out_dir / "tree.md"
    meta_md = out_dir / "meta.md"
    fpa_md = out_dir / "fpa_ai.md"
    summary_md = out_dir / "summary.md"
    audit_trace = out_dir / "audit_trace.json"
    check_xlsx = out_dir / "check.xlsx"

    _write_tree_md(tree_md, case["rows"])
    _write_meta_md(meta_md, case["meta"])

    plan_fpa_md_from_tree(
        str(tree_md),
        str(meta_md),
        str(fpa_md),
        template_path=str(TEMPLATE),
        api_key=load_api_key(),
        model=load_model_name(),
        base_url=load_base_url(),
        summary_md_path=str(summary_md),
        profile_name="strict_fpa",
        strategy="ai_first",
        rule_set="strict_fpa_default",
        audit_trace_path=str(audit_trace),
    )
    generate_fpa_check_xlsx_from_md(str(fpa_md), str(tree_md), str(check_xlsx), str(audit_trace))

    execution_meta, fpa_rows = _read_fpa_rows_md_for_audit(str(fpa_md))
    actual_pairs = [
        {"name": str(row.get("新增/修改功能点", "")), "type": str(row.get("类型", ""))}
        for row in fpa_rows
    ]

    wb = openpyxl.load_workbook(check_xlsx, data_only=True)
    try:
        result_rows = _sheet_rows(wb["FPA结果"])
        coverage_rows = _sheet_rows(wb["覆盖审核"])
        warning_rows = _sheet_rows(wb["Warnings"])
        rule_hit_rows = _sheet_rows(wb["规则命中详情"])
        raw_rows = _sheet_rows(wb["AI原始返回"])
        sheet_names = wb.sheetnames
    finally:
        wb.close()

    expected_pairs = _strict_expected(case)
    return {
        "case_id": case_id,
        "execution_meta": execution_meta,
        "sheets": sheet_names,
        "sheet_row_counts": {
            "FPA结果": len(result_rows),
            "覆盖审核": len(coverage_rows),
            "Warnings": len(warning_rows),
            "规则命中详情": len(rule_hit_rows),
            "AI原始返回": len(raw_rows),
        },
        "matches_expected_strict_fpa": actual_pairs == expected_pairs,
        "expected_count": len(expected_pairs),
        "actual_count": len(actual_pairs),
        "actual_pairs": actual_pairs,
        "sources": sorted({str(row.get("生成方式") or "") for row in fpa_rows}),
        "types": sorted({str(row.get("类型") or "") for row in fpa_rows}),
        "warning_count": sum(1 for row in fpa_rows if row.get("后处理警告")),
        "warning_rule_ids": sorted(
            {
                str(row.get("rule_id") or row.get("规则ID") or "")
                for row in warning_rows
                if row.get("rule_id") or row.get("规则ID")
            }
        ),
        "rule_hit_ids": sorted(
            {
                str(row.get("rule_id") or row.get("规则ID") or "")
                for row in rule_hit_rows
                if row.get("rule_id") or row.get("规则ID")
            }
        )[:30],
        "raw_sources": sorted(
            {
                str(row.get("source") or row.get("来源") or "")
                for row in raw_rows
                if row.get("source") or row.get("来源")
            }
        ),
        "summary_one_line": " | ".join(
            line.strip()
            for line in summary_md.read_text(encoding="utf-8").splitlines()
            if line.strip()
        ),
    }


def main() -> None:
    if not load_api_key() or not load_base_url() or not load_model_name():
        raise SystemExit("missing api_key/base_url/model")
    if OUT_ROOT.exists():
        shutil.rmtree(OUT_ROOT)
    OUT_ROOT.mkdir(parents=True)

    base_url = load_base_url()
    payload = {
        "base_url_host": base_url.split("/")[2] if "://" in base_url else base_url,
        "model": load_model_name(),
        "out_root": str(OUT_ROOT),
        "results": [_case_summary(case_id) for case_id in CASES],
    }
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
