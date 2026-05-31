"""Run F3 validation on the real business sample under ``1111/md``.

The script prints only validation summaries. It does not print API keys,
prompts, or full source business content.
"""

from __future__ import annotations

import json
import shutil
import sys
from collections import Counter
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from ai_gen_reimbursement_docs.config_utils import load_api_key, load_base_url, load_model_name
from ai_gen_reimbursement_docs.gen_fpa import (
    _group_rows_by_l3,
    _read_fpa_rows_md_for_audit,
    generate_fpa_check_xlsx_from_md,
    parse_module_tree_md,
    plan_fpa_md_from_tree,
)


SOURCE_MD_DIR = ROOT / "1111" / "md"
TREE_MD = SOURCE_MD_DIR / "0.1.gen-basedata-功能清单-模块树.md"
META_MD = SOURCE_MD_DIR / "0.2.gen-basedata-录入文档元数据-模板.md"
TEMPLATE = ROOT / "data" / "out_templates" / "FPA工作量评估-输出模板.xlsx"
OUT_ROOT = ROOT / "tmp_fpa_real_business_validation"


def _sheet_rows(ws) -> list[dict[str, object]]:
    headers = [cell.value for cell in ws[1]]
    return [
        dict(zip(headers, values))
        for values in ws.iter_rows(min_row=2, values_only=True)
        if any(value not in (None, "") for value in values)
    ]


def _summary_value(summary_path: Path) -> str:
    for line in summary_path.read_text(encoding="utf-8").splitlines():
        if "FPA工作量" in line:
            return line.strip()
    return ""


def _workbook_summary(check_xlsx: Path) -> dict[str, object]:
    wb = openpyxl.load_workbook(check_xlsx, data_only=True)
    try:
        sheets = wb.sheetnames
        result_rows = _sheet_rows(wb["FPA结果"])
        coverage_rows = _sheet_rows(wb["覆盖审核"])
        warning_rows = _sheet_rows(wb["Warnings"])
        rule_hit_rows = _sheet_rows(wb["规则命中详情"])
        raw_rows = _sheet_rows(wb["AI原始返回"])
    finally:
        wb.close()

    uncovered_modules = [
        {
            "module": row.get("三级模块"),
            "missing": row.get("未覆盖功能过程"),
        }
        for row in coverage_rows
        if row.get("未覆盖数") not in (None, "", 0)
    ]
    return {
        "sheets": sheets,
        "sheet_row_counts": {
            "FPA结果": len(result_rows),
            "覆盖审核": len(coverage_rows),
            "Warnings": len(warning_rows),
            "规则命中详情": len(rule_hit_rows),
            "AI原始返回": len(raw_rows),
        },
        "coverage": {
            "modules": len(coverage_rows),
            "uncovered_modules": uncovered_modules[:10],
        },
        "warning_rule_ids": sorted(
            {
                str(row.get("来源规则ID") or row.get("rule_id") or row.get("规则ID") or "")
                for row in warning_rows
                if row.get("来源规则ID") or row.get("rule_id") or row.get("规则ID")
            }
        ),
        "rule_hit_ids": sorted(
            {
                str(row.get("规则ID") or row.get("rule_id") or "")
                for row in rule_hit_rows
                if row.get("规则ID") or row.get("rule_id")
            }
        )[:40],
        "raw_sources": sorted(
            {
                str(row.get("来源") or row.get("source") or "")
                for row in raw_rows
                if row.get("来源") or row.get("source")
            }
        ),
    }


def _run_case(name: str, *, strategy: str, use_ai: bool) -> dict[str, object]:
    out_dir = OUT_ROOT / name
    out_dir.mkdir(parents=True, exist_ok=True)
    fpa_md = out_dir / "fpa.md"
    summary_md = out_dir / "summary.md"
    audit_trace = out_dir / "audit_trace.json"
    check_xlsx = out_dir / "check.xlsx"

    plan_fpa_md_from_tree(
        str(TREE_MD),
        str(META_MD),
        str(fpa_md),
        template_path=str(TEMPLATE),
        api_key=load_api_key() if use_ai else "",
        model=load_model_name() if use_ai else "",
        base_url=load_base_url() if use_ai else "",
        summary_md_path=str(summary_md),
        profile_name="strict_fpa",
        strategy=strategy,
        rule_set="strict_fpa_default",
        audit_trace_path=str(audit_trace),
    )
    generate_fpa_check_xlsx_from_md(str(fpa_md), str(TREE_MD), str(check_xlsx), str(audit_trace))

    execution_meta, fpa_rows = _read_fpa_rows_md_for_audit(str(fpa_md))
    type_counts = Counter(str(row.get("类型") or "") for row in fpa_rows)
    source_counts = Counter(str(row.get("生成方式") or "") for row in fpa_rows)
    warning_count = sum(1 for row in fpa_rows if row.get("后处理警告"))
    return {
        "execution_meta": execution_meta,
        "row_count": len(fpa_rows),
        "type_counts": dict(sorted(type_counts.items())),
        "source_counts": dict(sorted(source_counts.items())),
        "warning_count": warning_count,
        "summary": _summary_value(summary_md),
        "workbook": _workbook_summary(check_xlsx),
        "sample_rows": [
            {
                "name": row.get("新增/修改功能点"),
                "type": row.get("类型"),
                "source": row.get("生成方式"),
            }
            for row in fpa_rows[:10]
        ],
    }


def main() -> None:
    rows = parse_module_tree_md(str(TREE_MD))
    groups = _group_rows_by_l3(rows)
    if OUT_ROOT.exists():
        shutil.rmtree(OUT_ROOT)
    OUT_ROOT.mkdir(parents=True)

    payload = {
        "input": {
            "tree_md": str(TREE_MD),
            "meta_md": str(META_MD),
            "process_rows": len(rows),
            "l3_modules": len(groups),
            "module_names": [str(group.get("l3") or "") for group in groups],
        },
        "rules_only": _run_case("rules_only", strategy="rules_only", use_ai=False),
    }
    if load_api_key() and load_base_url() and load_model_name():
        payload["ai_first"] = _run_case("ai_first", strategy="ai_first", use_ai=True)
        payload["model"] = load_model_name()
        base_url = load_base_url()
        payload["base_url_host"] = base_url.split("/")[2] if "://" in base_url else base_url
    else:
        payload["ai_first"] = "skipped: missing api_key/base_url/model"
    payload["out_root"] = str(OUT_ROOT)
    print(json.dumps(payload, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
