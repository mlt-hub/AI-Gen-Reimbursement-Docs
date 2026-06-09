from pathlib import Path

from openpyxl import Workbook

from ai_gen_reimbursement_docs.gen_fpa import _read_fpa_judgement_rules_from_template


def test_read_fpa_judgement_rules_from_template_uses_default_appendix_layout(tmp_path: Path):
    template = tmp_path / "fpa.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "附录1-FPA评估方法说明"
    ws["C1"] = "判定原则"
    ws["C2"] = "规则一"
    ws["C3"] = "规则二"
    wb.save(template)
    wb.close()

    assert _read_fpa_judgement_rules_from_template(str(template)) == ["规则一", "规则二"]


def test_read_fpa_judgement_rules_from_template_uses_manifest_sheet_column_and_range(tmp_path: Path):
    template = tmp_path / "custom-fpa.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "FPA结果"
    appendix = wb.create_sheet("客户FPA附录")
    appendix["D3"] = "忽略的说明"
    appendix["D4"] = "规则一：识别内部逻辑文件。"
    appendix["D5"] = "规则二：识别外部接口文件。"
    appendix["D6"] = "超出范围的规则"
    wb.save(template)
    wb.close()
    template.with_suffix(".manifest.yaml").write_text(
        """
template_id: custom_fpa_v1
kind: fpa
version: 1
sheets:
  judgement_rules:
    name: 客户FPA附录
    data_start_row: 4
    data_end_row: 5
    rule_column: D
""".lstrip(),
        encoding="utf-8",
    )

    assert _read_fpa_judgement_rules_from_template(str(template)) == [
        "规则一：识别内部逻辑文件。",
        "规则二：识别外部接口文件。",
    ]


def test_read_fpa_judgement_rules_from_template_uses_manifest_anchor_and_column(tmp_path: Path):
    template = tmp_path / "client-fpa.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "客户附录"
    ws["B4"] = "计算依据归类判定原则"
    ws["E5"] = "客户规则一"
    ws["E6"] = "客户规则二"
    wb.save(template)
    wb.close()
    template.with_suffix(".manifest.yaml").write_text(
        """
template_id: client_fpa_v1
kind: fpa
version: 1
sheets:
  judgement_rules:
    name: 客户附录
    anchor:
      contains: 判定原则
      offset_rows: 1
      column: E
""".lstrip(),
        encoding="utf-8",
    )

    assert _read_fpa_judgement_rules_from_template(str(template)) == ["客户规则一", "客户规则二"]
