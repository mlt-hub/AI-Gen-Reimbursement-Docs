"""gen_xlsx 核心逻辑测试 —— FPA 行构建、接收者判定、说明格式化。"""
from pathlib import Path

import pytest
import openpyxl
from ai_gen_reimbursement_docs.gen_fpa import (
    _build_fpa_rule_rows,
    _receiver_from_client_type,
    _format_fpa_explanation,
    calculate_fpa_row_workload,
    calculate_fpa_total,
    generate_fpa_xlsx_from_md,
    validate_fpa_excel_recalculation,
)


FIXTURES = Path(__file__).parent / "fixtures"


class TestReceiverFromClientType:
    def test_no_rules_returns_default(self):
        """无规则时返回默认值（Excel 模板未配置则用默认值）。"""
        assert _receiver_from_client_type("后台", "") == "操作员"
        assert _receiver_from_client_type("前台", "") == "操作员"
        assert _receiver_from_client_type("未知", "") == "操作员"

    def test_match_by_keyword(self):
        rules = "后台：后台管理员\n普通：普通用户\n渠道：渠道人员"
        result = _receiver_from_client_type("后台", rules)
        assert result == "后台管理员"

    def test_match_second_keyword(self):
        rules = "后台：后台管理员\n普通：普通用户"
        result = _receiver_from_client_type("普通", rules)
        assert result == "普通用户"

    def test_default_when_no_match(self):
        rules = "后台：后台管理员"
        result = _receiver_from_client_type("未知类型", rules)
        assert result == "操作员"


class TestBuildFpaRuleRows:
    def _make_meta(self, **kwargs):
        return {
            "子系统（模块）": kwargs.get("subsystem", "测试系统"),
            "资产标识": kwargs.get("asset", "TEST-001"),
            "新增/修改功能点前缀生成规则": kwargs.get(
                "prefix_rule",
                "【客户端类型】一级模块-二级模块-三级模块-功能过程"
            ),
            "功能用户-接收者判定": kwargs.get("receiver_rules", ""),
        }

    def _make_row(self, **kwargs):
        return {
            "入口": kwargs.get("entry", "地市后台"),
            "一级模块": kwargs.get("l1", "系统管理"),
            "二级模块": kwargs.get("l2", "用户管理"),
            "三级模块": kwargs.get("l3", "用户注册"),
            "客户端类型": kwargs.get("client_type", "后台"),
            "三级模块整体功能描述": "",
            "功能过程": kwargs.get("proc", "注册新用户"),
            "功能过程类型": kwargs.get("proc_type", "新增"),
            "功能过程描述": "用户通过表单注册新账户",
        }

    def test_generates_l3_ui_and_logic_rows(self):
        """三级模块生成 1 条合并界面行 + 每个功能过程 1 条逻辑行。"""
        rows = [self._make_row(proc="添加用户"), self._make_row(proc="查询用户")]
        meta = self._make_meta()
        result = _build_fpa_rule_rows(rows, meta)
        assert len(result) == 3
        assert result[0]["类型"] == "EI"
        assert result[0]["新增/修改功能点"].endswith("-界面开发")
        assert result[1]["类型"] == "ILF"
        assert result[1]["新增/修改功能点"].endswith("-逻辑处理开发")
        assert result[2]["类型"] == "EQ"
        assert result[2]["新增/修改功能点"].endswith("-查询处理开发")

    def test_prefix_generation(self):
        """验证功能点前缀替换正确。"""
        rows = [self._make_row(l1="业务模块", l2="订单模块", l3="创建订单",
                                proc="提交订单", client_type="普通")]
        meta = self._make_meta()
        result = _build_fpa_rule_rows(rows, meta)
        prefix_ui = result[0]["新增/修改功能点"]
        assert "【普通】" in prefix_ui
        assert "业务模块-订单模块-创建订单" in prefix_ui

    def test_sequential_numbering(self):
        """验证序号连续递增。"""
        rows = [self._make_row(proc="p1"), self._make_row(proc="p2"),
                self._make_row(proc="p3")]
        meta = self._make_meta()
        result = _build_fpa_rule_rows(rows, meta)
        assert len(result) == 4  # 1 条三级模块界面 + 3 条逻辑行
        seqs = [r["序号"] for r in result]
        assert seqs == [1, 2, 3, 4]

    def test_empty_rows(self):
        """空输入返回空列表。"""
        result = _build_fpa_rule_rows([], self._make_meta())
        assert result == []

    def test_f_column_starts_empty(self):
        """F 列（计算依据归类）初始为空，待 AI 填充。"""
        rows = [self._make_row()]
        result = _build_fpa_rule_rows(rows, self._make_meta())
        assert result[0]["计算依据归类"] == ""
        assert result[1]["计算依据归类"] == ""

    def test_g_column_has_template(self):
        """G 列（计算依据说明）初始含模板文本。"""
        rows = [self._make_row()]
        result = _build_fpa_rule_rows(rows, self._make_meta())
        g_val = result[0]["计算依据说明"]
        assert "界面开发" in g_val
        assert "具体为以下" in g_val

    def test_adjustment_values(self):
        """界面行调整值=2，非 EI 行调整值=1。"""
        rows = [self._make_row()]
        result = _build_fpa_rule_rows(rows, self._make_meta())
        assert result[0]["调整值"] == 2
        assert result[1]["调整值"] == 1

    def test_uses_meta_subsystem_and_asset(self):
        """子系统名和资产标识取自 meta。"""
        rows = [self._make_row()]
        meta = self._make_meta(subsystem="和乐业", asset="YXGJ-01")
        result = _build_fpa_rule_rows(rows, meta)
        assert result[0]["子系统(模块)"] == "和乐业"
        assert result[0]["资产标识"] == "YXGJ-01"


class TestFpaTotalCalculation:
    def test_fpa_template_workload_formula_matches_python_total_rule(self):
        template_paths = [
            Path("data/out_templates/FPA工作量评估-输出模板.xlsx"),
            FIXTURES / "output_templates" / "FPA工作量评估-输出模板.xlsx",
        ]
        for template in template_paths:
            if not template.exists():
                pytest.skip(f"模板文件缺失: {template}")
            wb = openpyxl.load_workbook(template, data_only=False)
            ws = wb["FPA功能点估算"]
            assert ws.cell(3, 12).value == "=J3*K3"
            assert str(ws.cell(1, 12).value).startswith("=SUM(L3:")
            wb.close()

    def test_calculates_total_from_adjust_and_elements(self):
        rows = [
            {"调整值": 2, "要素数量": 3},
            {"调整值": "1.5", "要素数量": "2"},
            {"调整值": "", "要素数量": 9},
        ]

        assert calculate_fpa_row_workload(rows[0]) == 6
        assert calculate_fpa_total(rows) == 9

    def test_generate_xlsx_preserves_workload_formulas(self, tmp_path):
        template = FIXTURES / "output_templates" / "FPA工作量评估-输出模板.xlsx"
        if not template.exists():
            pytest.skip(f"模板文件缺失: {template}")

        fpa_md = tmp_path / "fpa.md"
        meta_md = tmp_path / "meta.md"
        output = tmp_path / "FPA工作量评估.xlsx"
        fpa_md.write_text(
            "# FPA 工作量评估\n\n"
            "| 序号 | 子系统(模块) | 资产标识 | 新增/修改功能点 | 类型 | 计算依据归类 | 计算依据说明 | 变更状态 | 调整值 | 要素数量 | 生成方式 | 类型理由 | 源功能过程 | 后处理警告 |\n"
            "|------|-------------|---------|----------------|------|-------------|-------------|---------|-------|---------|---------|---------|-----------|-----------|\n"
            "| 1 | 测试系统 | TEST | 界面开发 | EI | 规则一 | 说明 | 新增 | 2 | 3 | fallback | - | - |  |\n"
            "| 2 | 测试系统 | TEST | 查询 | EQ | 规则二 | 说明 | 新增 | 1 | 2 | fallback | - | - |  |\n",
            encoding="utf-8",
        )
        meta_md.write_text("# 元数据\n", encoding="utf-8")

        generate_fpa_xlsx_from_md(str(fpa_md), str(meta_md), str(template), str(output))

        wb = openpyxl.load_workbook(output, data_only=False)
        ws = wb["FPA功能点估算"]
        assert ws.cell(3, 12).value == "=J3*K3"
        assert ws.cell(4, 12).value == "=J4*K4"
        assert ws.cell(1, 12).value == "=SUM(L3:L4)"
        wb.close()

    def test_recalc_validation_passes_when_cached_total_matches(self, monkeypatch, tmp_path):
        xlsx = tmp_path / "fpa.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FPA功能点估算"
        ws.cell(1, 12).value = 9
        wb.save(xlsx)
        wb.close()
        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.gen_fpa._recalculate_with_excel_com",
            lambda path: (True, "mock"),
        )

        assert validate_fpa_excel_recalculation(str(xlsx), 9) == []

    def test_recalc_validation_warns_when_cached_total_differs(self, monkeypatch, tmp_path):
        xlsx = tmp_path / "fpa.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FPA功能点估算"
        ws.cell(1, 12).value = 8
        wb.save(xlsx)
        wb.close()
        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.gen_fpa._recalculate_with_excel_com",
            lambda path: (True, "mock"),
        )

        warnings = validate_fpa_excel_recalculation(str(xlsx), 9)

        assert len(warnings) == 1
        assert "不一致" in warnings[0]

    def test_recalc_validation_warns_when_no_engine_available(self, monkeypatch, tmp_path):
        xlsx = tmp_path / "fpa.xlsx"
        wb = openpyxl.Workbook()
        wb.save(xlsx)
        wb.close()
        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.gen_fpa._recalculate_with_excel_com",
            lambda path: (False, "no excel"),
        )
        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.gen_fpa._recalculate_with_libreoffice",
            lambda path: (False, "no libreoffice"),
        )

        warnings = validate_fpa_excel_recalculation(str(xlsx), 9)

        assert len(warnings) == 1
        assert "未执行 FPA Excel 公式复算校验" in warnings[0]


class TestFormatFpaExplanation:
    def test_adds_newlines_after_keywords(self):
        text = "具体如下：1、接收请求。事件流：用户操作。"
        result = _format_fpa_explanation(text)
        assert "\n" in result

    def test_preserves_original_content(self):
        text = "【后台】系统管理-用户管理-用户注册-界面开发"
        result = _format_fpa_explanation(text)
        assert "后台" in result
        assert "系统管理" in result

    def test_empty_string(self):
        assert _format_fpa_explanation("") == ""
