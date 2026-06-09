import openpyxl

from ai_gen_reimbursement_docs.cosmic_models import CosmicItem, DataMovement
from ai_gen_reimbursement_docs.cosmic_validator import CosmicValidationReport, CosmicValidationResult
from ai_gen_reimbursement_docs.cosmic_writer import write_cosmic_xlsx


def test_write_cosmic_xlsx_uses_manifest_result_sheet_and_rows(monkeypatch, tmp_path):
    monkeypatch.setenv("AI_REIMBURSEMENT_LOG_DIR", str(tmp_path / "log"))
    template = tmp_path / "custom-cosmic.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "自定义功能点"
    headers = [
        "项目",
        "一级模块",
        "二级模块",
        "三级模块",
        "用户",
        "触发事件",
        "功能过程",
        "子过程描述",
        "数据移动类型",
        "数据组",
        "数据属性",
        "复用度",
        "CFP",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(7, col_idx).value = header
        ws.cell(8, col_idx).border = openpyxl.styles.Border(
            left=openpyxl.styles.Side(style="thin"),
            right=openpyxl.styles.Side(style="thin"),
            top=openpyxl.styles.Side(style="thin"),
            bottom=openpyxl.styles.Side(style="thin"),
        )
    wb.save(template)
    wb.close()
    template.with_suffix(".manifest.yaml").write_text(
        "\n".join([
            "kind: cosmic",
            "sheets:",
            "  result:",
            "    name: 自定义功能点",
            "    header_row: 7",
            "    data_start_row: 8",
            "    style_source_row: 8",
            "",
        ]),
        encoding="utf-8",
    )
    item = CosmicItem(
        project="测试项目",
        module_l1="一级",
        module_l2="二级",
        module_l3="三级",
        user="发起者：用户|接收者：系统",
        trigger="用户触发",
        process="提交申请",
        movements=[
            DataMovement(
                order=1,
                sub_process="录入申请信息",
                move_type="E",
                data_group="申请单",
                data_attrs="申请编号,申请人",
            ),
        ],
    )
    report = CosmicValidationReport(
        project="测试项目",
        status="passed",
        results=[CosmicValidationResult(item=item, status="passed")],
        summary={},
    )
    output = tmp_path / "项目功能点拆分表.xlsx"

    write_cosmic_xlsx(
        str(template),
        str(output),
        report,
        cfp_formula='IF(L{row}="复用",1/3,1)',
    )

    wb = openpyxl.load_workbook(output, data_only=False)
    ws = wb["自定义功能点"]
    assert ws.cell(8, 1).value == "测试项目"
    assert ws.cell(8, 8).value == "录入申请信息"
    assert ws.cell(8, 9).value == "E"
    assert ws.cell(8, 13).value == '=IF(L8="复用",1/3,1)'
    wb.close()
