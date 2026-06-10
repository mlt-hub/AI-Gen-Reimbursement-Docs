from pathlib import Path

import openpyxl
from openpyxl.styles import Side
from openpyxl.workbook.defined_name import DefinedName

from ai_gen_reimbursement_docs.gen_list import generate_list_xlsx_from_md


def _write_meta(path: Path) -> None:
    path.write_text(
        """# 文档元数据

| 字段 | 值 |
| --- | --- |
| 项目信息概览-标题 | 项目概览标题 |
| 项目信息概览-项目名称 | 智能报销平台 |
| 项目信息概览-需求部门 | 财务部 |
| 功能清单-标题 | 功能清单标题 |
| 功能清单-项目名称 | 智能报销平台 |
| 功能清单-子系统 | 报销子系统 |
""",
        encoding="utf-8",
    )


def _write_tree(path: Path) -> None:
    path.write_text(
        """# 模块树

| 入口 | 一级模块 | 二级模块 | 三级模块 | 客户端类型 | 三级模块整体功能描述 | 功能过程 | 功能过程描述 | 变更状态 |
|------|----------|----------|----------|------------|----------------------|----------|--------------|--------------|
| Web | 费用管理 | 报销单 | 新建报销单 | PC | 支持员工新建报销单 | 提交报销单 | 员工录入并提交报销单 | 新增 |
""",
        encoding="utf-8",
    )


def _write_template(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "概览自定义"
    ws1.cell(1, 1, "old title")
    for col_idx, header in enumerate(["送审功能点", "项目名称", "送审工作量", "需求部门"], start=1):
        ws1.cell(4, col_idx, header)

    ws2 = wb.create_sheet("清单自定义")
    ws2.cell(1, 1, "old title")
    headers = [
        "类型",
        "三级功能模块名称",
        "一级功能模块名称",
        "二级功能模块名称",
        "送审功能点",
        "送审工作量",
        "项目名称",
        "子系统",
        "序号",
    ]
    for col_idx, header in enumerate(headers, start=1):
        ws2.cell(4, col_idx, header)
    thin = Side(style="thin", color="000000")
    for col_idx in range(1, len(headers) + 1):
        ws2.cell(5, col_idx).border = openpyxl.styles.Border(
            left=thin,
            right=thin,
            top=thin,
            bottom=thin,
        )
    wb.save(path)


def _write_manifest(path: Path) -> None:
    path.write_text(
        """template_id: list_custom_test
kind: list
version: 1
sheets:
  project_info:
    name: 概览自定义
    header_row: 4
    data_start_row: 5
    style_source_row: 5
    columns:
      project_name:
        header: 项目名称
      department:
        header: 需求部门
      workload:
        header: 送审工作量
      cfp:
        header: 送审功能点
  function_list:
    name: 清单自定义
    header_row: 4
    data_start_row: 5
    style_source_row: 5
    columns:
      seq:
        header: 序号
      project_name:
        header: 项目名称
      subsystem:
        header: 子系统
      module_l1:
        header: 一级功能模块名称
      module_l2:
        header: 二级功能模块名称
      module_l3:
        header: 三级功能模块名称
      type:
        header: 类型
      workload:
        header: 送审工作量
      cfp:
        header: 送审功能点
""",
        encoding="utf-8",
    )


def test_generate_list_uses_manifest_sheet_rows_and_headers(tmp_path):
    meta = tmp_path / "meta.md"
    tree = tmp_path / "tree.md"
    template = tmp_path / "custom-list.xlsx"
    output = tmp_path / "out.xlsx"

    _write_meta(meta)
    _write_tree(tree)
    _write_template(template)
    _write_manifest(template.with_suffix(".manifest.yaml"))

    generate_list_xlsx_from_md(
        str(meta),
        str(tree),
        str(template),
        str(output),
        cfp_total=12.5,
        fpa_reduced=8.5,
    )

    wb = openpyxl.load_workbook(output)
    ws1 = wb["概览自定义"]
    assert ws1.cell(5, 1).value == 12.5
    assert ws1.cell(5, 2).value == "智能报销平台"
    assert ws1.cell(5, 3).value == 8.5
    assert ws1.cell(5, 4).value == "财务部"

    ws2 = wb["清单自定义"]
    assert ws2.cell(5, 1).value is None
    assert ws2.cell(5, 2).value == "新建报销单"
    assert ws2.cell(5, 3).value == "费用管理"
    assert ws2.cell(5, 4).value == "报销单"
    assert ws2.cell(5, 5).value == 12.5
    assert ws2.cell(5, 6).value == 8.5
    assert ws2.cell(5, 7).value == "智能报销平台"
    assert ws2.cell(5, 8).value == "报销子系统"
    assert ws2.cell(5, 9).value == 1
    assert ws2.cell(5, 1).border.left.style == "thin"


def test_generate_list_writes_project_info_named_cells_from_manifest(tmp_path):
    meta = tmp_path / "meta.md"
    tree = tmp_path / "tree.md"
    template = tmp_path / "custom-list.xlsx"
    output = tmp_path / "out.xlsx"

    _write_meta(meta)
    _write_tree(tree)
    _write_template(template)

    wb = openpyxl.load_workbook(template)
    ws1 = wb["概览自定义"]
    ws1.cell(2, 1, "项目名称")
    ws1.cell(3, 1, "需求部门")
    ws1.cell(4, 1, "送审工作量")
    ws1.cell(5, 1, "送审功能点")
    for name, cell in {
        "LIST_PROJECT_NAME": "$B$2",
        "LIST_DEPARTMENT": "$B$3",
        "LIST_WORKLOAD": "$B$4",
        "LIST_CFP": "$B$5",
    }.items():
        wb.defined_names.add(DefinedName(name, attr_text=f"'概览自定义'!{cell}"))
    wb.save(template)

    template.with_suffix(".manifest.yaml").write_text(
        """template_id: list_named_cells_test
kind: list
version: 1
sheets:
  project_info:
    name: 概览自定义
    header_row: 4
    data_start_row: 5
    named_cells:
      project_name: LIST_PROJECT_NAME
      department: LIST_DEPARTMENT
      workload: LIST_WORKLOAD
      cfp: LIST_CFP
    columns:
      project_name:
        header: 项目名称
      department:
        header: 需求部门
      workload:
        header: 送审工作量
      cfp:
        header: 送审功能点
  function_list:
    name: 清单自定义
    header_row: 4
    data_start_row: 5
    style_source_row: 5
    columns:
      seq:
        header: 序号
      project_name:
        header: 项目名称
      subsystem:
        header: 子系统
      module_l1:
        header: 一级功能模块名称
      module_l2:
        header: 二级功能模块名称
      module_l3:
        header: 三级功能模块名称
      type:
        header: 类型
      workload:
        header: 送审工作量
      cfp:
        header: 送审功能点
""",
        encoding="utf-8",
    )

    generate_list_xlsx_from_md(
        str(meta),
        str(tree),
        str(template),
        str(output),
        cfp_total=12.5,
        fpa_reduced=8.5,
    )

    wb = openpyxl.load_workbook(output)
    ws1 = wb["概览自定义"]
    assert ws1["B2"].value == "智能报销平台"
    assert ws1["B3"].value == "财务部"
    assert ws1["B4"].value == 8.5
    assert ws1["B5"].value == 12.5

    ws2 = wb["清单自定义"]
    assert ws2.cell(5, 1).value is None
    assert ws2.cell(5, 9).value == 1
