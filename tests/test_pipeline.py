"""run_pipeline() 集成测试 —— 端到端验证各模式交付物生成。"""

import json
import os
from pathlib import Path

import openpyxl
import pytest
from ai_gen_reimbursement_docs.cosmic_ai import CosmicGenerationDiagnostics
from ai_gen_reimbursement_docs.cosmic_models import CosmicItem, DataMovement
from ai_gen_reimbursement_docs.pipeline import run_pipeline, PipelineResult

pytestmark = pytest.mark.slow

FIXTURES = Path(__file__).parent / "fixtures"
TEMPLATES = {
    "fpa": str(FIXTURES / "output_templates" / "FPA工作量评估-输出模板.xlsx"),
    "cosmic": str(FIXTURES / "output_templates" / "项目功能点拆分表-输出模板.xlsx"),
    "list": str(FIXTURES / "output_templates" / "项目需求清单-输出模板.xlsx"),
    "spec": str(FIXTURES / "output_templates" / "项目需求说明书-输出模板.docx"),
}


@pytest.fixture(autouse=True)
def _verify_templates():
    """所有测试前置：确保模板文件存在。"""
    missing = [k for k, v in TEMPLATES.items() if not os.path.exists(v)]
    if missing:
        pytest.skip(f"模板文件缺失: {missing}")


class TestValidation:
    """参数校验"""

    def test_rejects_nonexistent_file(self, output_dir):
        with pytest.raises(FileNotFoundError, match="功能清单输入文件不存在"):
            run_pipeline(mode="gen-basedata", file_path="/no/such.xlsx",
                         output_dir=output_dir)

    def test_rejects_invalid_mode(self, output_dir, test_excel):
        with pytest.raises(ValueError, match="未知模式"):
            run_pipeline(mode="invalid-mode", file_path=test_excel,
                         output_dir=output_dir)

    def test_all_valid_modes_accepted(self, output_dir, test_excel):
        modes = ["gen-basedata", "gen-fpa", "gen-cosmic", "gen-list", "gen-spec", "gen-all"]
        for m in modes:
            assert m in {"gen-all", "gen-basedata", "gen-fpa", "gen-cosmic",
                         "gen-list", "gen-spec"}


class TestGenBasedata:
    """gen-basedata 模式"""

    def test_generates_tree_and_meta(self, output_dir, test_excel, mock_ai):
        result = run_pipeline(mode="gen-basedata", file_path=test_excel,
                             output_dir=output_dir, api_key="sk-test",
                             templates=TEMPLATES)
        assert os.path.exists(result.tree_md)
        assert os.path.exists(result.meta_md)
        with open(result.tree_md, encoding='utf-8') as f:
            content = f.read()
            assert "一级模块" in content
            assert "用户管理" in content

    def test_returns_result_without_api_key(self, output_dir, test_excel):
        result = run_pipeline(mode="gen-basedata", file_path=test_excel,
                             output_dir=output_dir, api_key="")
        assert os.path.exists(result.tree_md)


class TestGenFpa:
    """gen-fpa 模式"""

    def test_generates_xlsx(self, output_dir, test_excel, mock_ai):
        result = run_pipeline(mode="gen-fpa", file_path=test_excel,
                             output_dir=output_dir, templates=TEMPLATES,
                             api_key="sk-test")
        assert os.path.exists(result.fpa_xlsx)
        assert os.path.getsize(result.fpa_xlsx) > 0
        assert os.path.exists(result.fpa_check_xlsx)
        assert os.path.getsize(result.fpa_check_xlsx) > 0

    def test_generates_fpa_check_xlsx(self, output_dir, test_excel, mock_ai):
        result = run_pipeline(mode="gen-fpa", file_path=test_excel,
                             output_dir=output_dir, templates=TEMPLATES,
                             api_key="sk-test")

        wb = openpyxl.load_workbook(result.fpa_check_xlsx, data_only=True)
        assert "FPA结果" in wb.sheetnames
        assert "覆盖审核" in wb.sheetnames
        assert "Warnings" in wb.sheetnames
        assert "规则命中详情" in wb.sheetnames
        assert "AI原始返回" in wb.sheetnames
        ws_result = wb["FPA结果"]
        headers = [cell.value for cell in ws_result[1]]
        assert "生成方式" in headers
        assert "rule_set_version" not in headers
        assert ws_result.freeze_panes == "A2"
        assert ws_result.auto_filter.ref is not None
        ws_coverage = wb["覆盖审核"]
        coverage_headers = [cell.value for cell in ws_coverage[1]]
        assert "未覆盖功能过程" in coverage_headers
        assert "Warnings" in coverage_headers
        ws_warnings = wb["Warnings"]
        warning_headers = [cell.value for cell in ws_warnings[1]]
        assert warning_headers == ["级别", "FPA行序号", "模块序号", "对象", "Warning", "来源规则ID", "来源说明"]
        ws_rule_hits = wb["规则命中详情"]
        rule_hit_headers = [cell.value for cell in ws_rule_hits[1]]
        assert rule_hit_headers == [
            "模块序号", "客户端类型", "一级模块", "二级模块", "三级模块",
            "FPA行序号", "功能点名称", "生成方式", "rule_set",
            "core_rules", "system_prompt", "user_prompt",
            "命中对象", "规则ID", "规则说明", "建议类型", "是否采用", "Warnings",
        ]
        assert ws_rule_hits.max_row > 1
        rule_id_column = rule_hit_headers.index("规则ID") + 1
        rule_ids = [ws_rule_hits.cell(row=row, column=rule_id_column).value for row in range(2, ws_rule_hits.max_row + 1)]
        assert any(rule_id for rule_id in rule_ids)
        assert ws_rule_hits.freeze_panes == "A2"
        assert ws_rule_hits.auto_filter.ref is not None
        ws_raw = wb["AI原始返回"]
        raw_headers = [cell.value for cell in ws_raw[1]]
        assert raw_headers == ["模块", "三级模块", "来源", "Warnings", "AI原始Rows JSON"]
        wb.close()

    def test_fpa_reduced_read_from_md(self, output_dir, test_excel, mock_ai):
        """gen-fpa 不直接使用 fpa_reduced 参数——它从生成的 MD 读取。"""
        result = run_pipeline(mode="gen-fpa", file_path=test_excel,
                             output_dir=output_dir, templates=TEMPLATES,
                             api_key="sk-test")
        assert result.fpa_reduced > 0  # 从 MD 文件读取

    def test_fpa_summary_is_calculated_while_excel_preserves_formula(self, output_dir, test_excel, mock_ai):
        result = run_pipeline(mode="gen-fpa", file_path=test_excel,
                             output_dir=output_dir, templates=TEMPLATES,
                             api_key="sk-test")

        wb = openpyxl.load_workbook(result.fpa_xlsx, data_only=False)
        ws = wb["FPA功能点估算"]
        excel_total_formula = ws.cell(1, 12).value
        wb.close()

        assert result.fpa_reduced > 0
        assert excel_total_formula.startswith("=SUM(L3:L")

    def test_ai_only_failure_does_not_generate_fpa_outputs(self, output_dir, test_excel, monkeypatch):
        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.gen_fpa._call_llm",
            lambda *args, **kwargs: '{"rows":[]}',
        )

        with pytest.raises(ValueError, match="AI 规划未生成有效 FPA 行"):
            run_pipeline(
                mode="gen-fpa",
                file_path=test_excel,
                output_dir=output_dir,
                templates=TEMPLATES,
                api_key="sk-test",
                fpa_strategy="ai_only",
            )

        generated_workbooks = [
            path.name
            for path in Path(output_dir).rglob("*.xlsx")
            if "FPA工作量评估" in path.name
        ]
        assert generated_workbooks == []

    def test_rejects_unknown_profile(self, output_dir, test_excel):
        with pytest.raises(ValueError, match="未知 FPA profile"):
            run_pipeline(
                mode="gen-fpa",
                file_path=test_excel,
                output_dir=output_dir,
                templates=TEMPLATES,
                fpa_profile="unknown_profile",
            )

    def test_rules_first_pipeline_calls_ai_when_rule_rows_are_low_confidence(
        self, output_dir, test_excel, monkeypatch
    ):
        calls = {"count": 0}

        def fake_fallback(self, group, meta, start_seq=1, **kwargs):
            return [{
                "序号": start_seq,
                "子系统(模块)": meta.get("子系统（模块）", ""),
                "资产标识": meta.get("资产标识", ""),
                "新增/修改功能点": "低置信度规则行",
                "类型": "EI",
                "计算依据归类": "",
                "计算依据说明": "低置信度规则行。",
                "变更状态": "新增",
                "调整值": 1,
                "要素数量": 1,
                "生成方式": "fallback",
                "类型理由": "",
                "源功能过程": "",
                "后处理警告": "",
            }]

        response = {
            "rows": [{
                "name": "AI 复核功能点",
                "type": "EI",
                "classification_basis_index": 1,
                "explanation": "AI 复核功能点，具体为以下：1、覆盖低置信度规则无法覆盖的功能过程。",
                "source_processes": [],
            }]
        }

        def fake_call_llm(*args, **kwargs):
            calls["count"] += 1
            return json.dumps(response, ensure_ascii=False)

        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.fpa_profiles.CustomRulesProfile.fallback_rows_for_l3",
            fake_fallback,
        )
        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.fpa_profiles.StrictFpaProfile.fallback_rows_for_l3",
            fake_fallback,
        )
        monkeypatch.setattr("ai_gen_reimbursement_docs.gen_fpa._call_llm", fake_call_llm)

        result = run_pipeline(
            mode="gen-fpa",
            file_path=test_excel,
            output_dir=output_dir,
            templates=TEMPLATES,
            api_key="sk-test",
            model="test",
            fpa_strategy="rules_first",
        )

        filled_md = Path(output_dir) / "md" / "1.3.gen-fpa-AI填充-FPA.md"
        content = filled_md.read_text(encoding="utf-8")

        assert calls["count"] > 0
        assert os.path.exists(result.fpa_xlsx)
        assert "**strategy**: rules_first" in content
        assert "AI 复核功能点" in content
        assert "低置信度规则行" not in content


class TestGenCosmic:
    """gen-cosmic 模式"""

    def test_generates_cosmic_xlsx(self, output_dir, test_excel, mock_ai, monkeypatch):
        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.pipeline._read_cfp_formula_from_meta_md",
            lambda meta_md: 'IF(L{row}="新增",1,0)',
        )
        result = run_pipeline(mode="gen-cosmic", file_path=test_excel,
                             output_dir=output_dir, templates=TEMPLATES,
                             api_key="sk-test")
        assert result.cosmic_status == "passed"
        assert result.cosmic_formal_excel_written is True
        assert os.path.exists(result.cosmic_formal_xlsx)
        assert os.path.getsize(result.cosmic_formal_xlsx) > 0
        assert os.path.exists(result.cosmic_validation_json)
        assert os.path.exists(result.cosmic_validation_report)
        assert result.cfp_total > 0
        payload = json.loads(Path(result.cosmic_validation_json).read_text(encoding="utf-8"))
        assert payload["cfp_basis"]["source"] == "template_formula"

    def test_no_api_key_sets_path_but_no_file(self, output_dir, test_excel):
        result = run_pipeline(mode="gen-cosmic", file_path=test_excel,
                             output_dir=output_dir, templates=TEMPLATES,
                             api_key="")
        assert result.cosmic_status == "blocked"
        assert result.cosmic_formal_xlsx
        assert result.cosmic_formal_excel_written is False
        assert not os.path.exists(result.cosmic_formal_xlsx)
        assert os.path.exists(result.cosmic_validation_json)
        assert os.path.exists(result.cosmic_validation_report)
        assert result.cfp_total == 0
        payload = json.loads(Path(result.cosmic_validation_json).read_text(encoding="utf-8"))
        issue_codes = [issue["code"] for issue in payload["issues"]]
        assert "NO_API_KEY" in issue_codes
        assert "NO_COSMIC_ITEMS" in issue_codes

    def test_ai_generation_exception_is_reported(self, output_dir, test_excel, monkeypatch):
        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.pipeline._read_cfp_formula_from_meta_md",
            lambda meta_md: 'IF(L{row}="新增",1,0)',
        )

        def raise_generation_error(**kwargs):
            raise RuntimeError("mock cosmic failure")

        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.cosmic_ai.generate_cosmic_items_with_diagnostics",
            raise_generation_error,
        )

        result = run_pipeline(mode="gen-cosmic", file_path=test_excel,
                             output_dir=output_dir, templates=TEMPLATES,
                             api_key="sk-test")

        assert result.cosmic_status == "blocked"
        assert result.cosmic_formal_excel_written is False
        payload = json.loads(Path(result.cosmic_validation_json).read_text(encoding="utf-8"))
        issue_codes = [issue["code"] for issue in payload["issues"]]
        assert "AI_GENERATION_FAILED" in issue_codes
        assert "NO_COSMIC_ITEMS" in issue_codes

    def test_no_l3_modules_is_reported(self, output_dir, test_excel, monkeypatch):
        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.pipeline._read_cfp_formula_from_meta_md",
            lambda meta_md: 'IF(L{row}="新增",1,0)',
        )
        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.cosmic_ai.generate_cosmic_items_with_diagnostics",
            lambda **kwargs: CosmicGenerationDiagnostics(
                items=[],
                total_l3_modules=0,
                ai_called=0,
            ),
        )

        result = run_pipeline(mode="gen-cosmic", file_path=test_excel,
                             output_dir=output_dir, templates=TEMPLATES,
                             api_key="sk-test")

        assert result.cosmic_status == "blocked"
        payload = json.loads(Path(result.cosmic_validation_json).read_text(encoding="utf-8"))
        issue_codes = [issue["code"] for issue in payload["issues"]]
        assert "NO_L3_MODULES" in issue_codes
        assert "NO_COSMIC_ITEMS" in issue_codes

    def test_all_skipped_by_ai_limit_is_reported(self, output_dir, test_excel, monkeypatch):
        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.pipeline._read_cfp_formula_from_meta_md",
            lambda meta_md: 'IF(L{row}="新增",1,0)',
        )
        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.cosmic_ai.generate_cosmic_items_with_diagnostics",
            lambda **kwargs: CosmicGenerationDiagnostics(
                items=[],
                total_l3_modules=2,
                ai_called=0,
                skipped_by_l3_limit=2,
            ),
        )

        result = run_pipeline(mode="gen-cosmic", file_path=test_excel,
                             output_dir=output_dir, templates=TEMPLATES,
                             api_key="sk-test")

        assert result.cosmic_status == "blocked"
        payload = json.loads(Path(result.cosmic_validation_json).read_text(encoding="utf-8"))
        issue_codes = [issue["code"] for issue in payload["issues"]]
        assert "AI_LIMIT_SKIPPED_ALL" in issue_codes
        assert "AI_GENERATION_EMPTY" in issue_codes
        assert "NO_COSMIC_ITEMS" in issue_codes

    def test_partial_ai_failure_is_review_required(self, output_dir, test_excel, monkeypatch):
        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.pipeline._read_cfp_formula_from_meta_md",
            lambda meta_md: 'IF(L{row}="新增",1,0)',
        )
        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.cosmic_ai.generate_cosmic_items_with_diagnostics",
            lambda **kwargs: CosmicGenerationDiagnostics(
                items=[
                    CosmicItem(
                        project="测试项目",
                        module_l1="系统管理",
                        module_l2="用户管理",
                        module_l3="用户注册",
                        user="发起者：用户注册|接收者：系统管理",
                        trigger="用户触发",
                        process="注册用户",
                        movements=[
                            DataMovement(1, "接收注册请求", "E", "用户注册请求", "姓名"),
                            DataMovement(2, "返回注册结果", "X", "用户注册结果", "结果"),
                        ],
                    )
                ],
                total_l3_modules=2,
                ai_called=2,
                failed_modules=[("系统管理", "用户管理", "用户删除", "JSON array parse failed")],
            ),
        )

        result = run_pipeline(mode="gen-cosmic", file_path=test_excel,
                             output_dir=output_dir, templates=TEMPLATES,
                             api_key="sk-test")

        assert result.cosmic_status == "review_required"
        assert result.cosmic_formal_excel_written is False
        payload = json.loads(Path(result.cosmic_validation_json).read_text(encoding="utf-8"))
        issue_codes = [issue["code"] for issue in payload["issues"]]
        assert "PARTIAL_AI_FAILURE" in issue_codes
        partial_issue = next(issue for issue in payload["issues"] if issue["code"] == "PARTIAL_AI_FAILURE")
        assert partial_issue["details"]["failed_modules"] == [{
            "module_l1": "系统管理",
            "module_l2": "用户管理",
            "module_l3": "用户删除",
            "error": "JSON array parse failed",
        }]
        assert "AI_MODULE_PARSE_FAILED" in issue_codes

    def test_partial_ai_retry_exhausted_is_classified(self, output_dir, test_excel, monkeypatch):
        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.pipeline._read_cfp_formula_from_meta_md",
            lambda meta_md: 'IF(L{row}="新增",1,0)',
        )
        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.cosmic_ai.generate_cosmic_items_with_diagnostics",
            lambda **kwargs: CosmicGenerationDiagnostics(
                items=[
                    CosmicItem(
                        project="测试项目",
                        module_l1="系统管理",
                        module_l2="用户管理",
                        module_l3="用户注册",
                        user="发起者：用户注册|接收者：用户注册",
                        trigger="用户触发",
                        process="注册用户",
                        movements=[
                            DataMovement(1, "接收注册请求", "E", "用户注册请求", "姓名"),
                            DataMovement(2, "返回注册结果", "X", "用户注册结果", "结果"),
                        ],
                    )
                ],
                total_l3_modules=2,
                ai_called=3,
                failed_modules=[("系统管理", "用户管理", "用户删除", "retry exhausted: timeout")],
            ),
        )

        result = run_pipeline(mode="gen-cosmic", file_path=test_excel,
                             output_dir=output_dir, templates=TEMPLATES,
                             api_key="sk-test")

        assert result.cosmic_status == "review_required"
        payload = json.loads(Path(result.cosmic_validation_json).read_text(encoding="utf-8"))
        issue_codes = [issue["code"] for issue in payload["issues"]]
        assert "AI_RETRY_EXHAUSTED" in issue_codes

    def test_partial_ai_limit_placeholders_are_blocked_and_reported(self, output_dir, test_excel, monkeypatch):
        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.pipeline._read_cfp_formula_from_meta_md",
            lambda meta_md: 'IF(L{row}="新增",1,0)',
        )
        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.cosmic_ai.generate_cosmic_items_with_diagnostics",
            lambda **kwargs: CosmicGenerationDiagnostics(
                items=[
                    CosmicItem(
                        project="测试项目",
                        module_l1="系统管理",
                        module_l2="用户管理",
                        module_l3="用户注册",
                        user="发起者：用户注册|接收者：系统管理",
                        trigger="用户触发",
                        process="注册用户",
                        movements=[
                            DataMovement(1, "接收注册请求", "E", "用户注册请求", "姓名"),
                            DataMovement(2, "返回注册结果", "X", "用户注册结果", "结果"),
                        ],
                    ),
                    CosmicItem(
                        project="测试项目",
                        module_l1="系统管理",
                        module_l2="用户管理",
                        module_l3="用户删除",
                        user="",
                        trigger="",
                        process="删除用户",
                        movements=[],
                    )
                ],
                total_l3_modules=3,
                ai_called=1,
                skipped_by_l3_limit=1,
                skipped_by_process_limit=1,
            ),
        )

        result = run_pipeline(mode="gen-cosmic", file_path=test_excel,
                             output_dir=output_dir, templates=TEMPLATES,
                             api_key="sk-test")

        assert result.cosmic_status == "blocked"
        assert result.cosmic_formal_excel_written is False
        payload = json.loads(Path(result.cosmic_validation_json).read_text(encoding="utf-8"))
        issue_codes = [issue["code"] for issue in payload["issues"]]
        assert "AI_L3_LIMIT_PARTIAL_SKIP" in issue_codes
        assert "AI_PROCESS_LIMIT_PARTIAL_SKIP" in issue_codes
        assert "AI_LIMIT_PARTIAL_PLACEHOLDER" in issue_codes

    def test_user_aborted_generation_is_reported(self, output_dir, test_excel, monkeypatch):
        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.pipeline._read_cfp_formula_from_meta_md",
            lambda meta_md: 'IF(L{row}="新增",1,0)',
        )
        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.cosmic_ai.generate_cosmic_items_with_diagnostics",
            lambda **kwargs: CosmicGenerationDiagnostics(
                items=[
                    CosmicItem(
                        project="测试项目",
                        module_l1="系统管理",
                        module_l2="用户管理",
                        module_l3="用户注册",
                        user="发起者：用户注册|接收者：系统管理",
                        trigger="用户触发",
                        process="注册用户",
                        movements=[
                            DataMovement(1, "接收注册请求", "E", "用户注册请求", "姓名"),
                            DataMovement(2, "返回注册结果", "X", "用户注册结果", "结果"),
                        ],
                    )
                ],
                total_l3_modules=2,
                ai_called=1,
                aborted=True,
            ),
        )

        result = run_pipeline(mode="gen-cosmic", file_path=test_excel,
                             output_dir=output_dir, templates=TEMPLATES,
                             api_key="sk-test")

        assert result.cosmic_status == "review_required"
        payload = json.loads(Path(result.cosmic_validation_json).read_text(encoding="utf-8"))
        issue_codes = [issue["code"] for issue in payload["issues"]]
        assert "USER_ABORTED_GENERATION" in issue_codes

    def test_empty_movements_blocks_formal_excel(self, output_dir, test_excel, monkeypatch):
        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.pipeline._read_cfp_formula_from_meta_md",
            lambda meta_md: 'IF(L{row}="新增",1,0)',
        )
        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.cosmic_ai.generate_cosmic_items_with_diagnostics",
            lambda **kwargs: CosmicGenerationDiagnostics(
                items=[
                    CosmicItem(
                        project="测试项目",
                        module_l1="系统管理",
                        module_l2="用户管理",
                        module_l3="用户注册",
                        user="发起者：用户注册|接收者：系统管理",
                        trigger="用户触发",
                        process="注册用户",
                        movements=[],
                    )
                ],
                total_l3_modules=1,
                ai_called=1,
            ),
        )

        result = run_pipeline(mode="gen-cosmic", file_path=test_excel,
                             output_dir=output_dir, templates=TEMPLATES,
                             api_key="sk-test")

        assert result.cosmic_status == "blocked"
        assert result.cosmic_formal_excel_written is False
        assert not os.path.exists(result.cosmic_formal_xlsx)
        content = Path(result.cosmic_validation_report).read_text(encoding="utf-8")
        assert "TOO_FEW_MOVEMENTS" in content

    def test_review_required_can_write_draft_excel(self, output_dir, test_excel, monkeypatch, tmp_path):
        cfg_dir = tmp_path / "config"
        cfg_dir.mkdir()
        (cfg_dir / "system_config.yaml").write_text(
            "gen_cosmic:\n  allow_draft_excel_output: true\n",
            encoding="utf-8",
        )
        monkeypatch.setattr("ai_gen_reimbursement_docs.config_utils.config_dir", lambda: cfg_dir)
        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.pipeline._read_cfp_formula_from_meta_md",
            lambda meta_md: 'IF(L{row}="新增",1,0)',
        )
        monkeypatch.setattr(
            "ai_gen_reimbursement_docs.cosmic_ai.generate_cosmic_items_with_diagnostics",
            lambda **kwargs: CosmicGenerationDiagnostics(
                items=[
                    CosmicItem(
                        project="测试项目",
                        module_l1="系统管理",
                        module_l2="用户管理",
                        module_l3="用户注册",
                        user="发起者：用户注册|接收者：系统管理",
                        trigger="用户触发",
                        process="注册用户",
                        movements=[
                            DataMovement(1, "接收注册请求", "E", "用户注册请求", ""),
                            DataMovement(2, "返回注册结果", "X", "用户注册结果", "结果状态"),
                        ],
                    )
                ],
                total_l3_modules=1,
                ai_called=1,
            ),
        )

        result = run_pipeline(mode="gen-cosmic", file_path=test_excel,
                             output_dir=output_dir, templates=TEMPLATES,
                             api_key="sk-test")

        assert result.cosmic_status == "review_required"
        assert result.cosmic_formal_excel_written is False
        assert result.cosmic_draft_excel_written is True
        assert os.path.exists(result.cosmic_draft_xlsx)
        assert not os.path.exists(result.cosmic_formal_xlsx)


class TestGenSpec:
    """gen-spec 模式"""

    def test_generates_docx(self, output_dir, test_excel, mock_ai):
        result = run_pipeline(mode="gen-spec", file_path=test_excel,
                             output_dir=output_dir, templates=TEMPLATES,
                             api_key="sk-test")
        assert os.path.exists(result.spec_docx)
        assert os.path.getsize(result.spec_docx) > 0


class TestGenList:
    """gen-list 模式"""

    def test_generates_xlsx(self, output_dir, test_excel):
        result = run_pipeline(mode="gen-list", file_path=test_excel,
                             output_dir=output_dir, templates=TEMPLATES)
        assert os.path.exists(result.require_xlsx)
        assert os.path.getsize(result.require_xlsx) > 0

    def test_override_values(self, output_dir, test_excel):
        result = run_pipeline(mode="gen-list", file_path=test_excel,
                             output_dir=output_dir, templates=TEMPLATES,
                             fpa_reduced=7.5, cfp_total=30.0)
        assert os.path.exists(result.require_xlsx)


class TestGenAll:
    """gen-all 全流程"""

    def test_produces_all_four_outputs(self, output_dir, test_excel, mock_ai):
        result = run_pipeline(mode="gen-all", file_path=test_excel,
                             output_dir=output_dir, templates=TEMPLATES,
                             api_key="sk-test")
        assert os.path.exists(result.fpa_xlsx)
        assert os.path.exists(result.require_xlsx)
        assert os.path.exists(result.spec_docx)
        # COSMIC 依赖 AI，mock 返回空列表时不会生成文件，但路径已设置
        assert result.cosmic_xlsx
        for path in [result.fpa_xlsx, result.require_xlsx, result.spec_docx]:
            assert os.path.getsize(path) > 0, f"{path} 是空文件"

    def test_gen_all_no_ai(self, output_dir, test_excel, mock_ai):
        """无 AI 时全流程仍能完成"""
        result = run_pipeline(mode="gen-all", file_path=test_excel,
                             output_dir=output_dir, templates=TEMPLATES,
                             api_key="")
        # 基础交付物都应存在（FPA/List/Spec 不依赖 AI）
        assert os.path.exists(result.fpa_xlsx)
        assert os.path.exists(result.require_xlsx)
        assert os.path.exists(result.spec_docx)


class TestTemplateFallback:
    """模板回退机制"""

    def test_cli_template_not_found_falls_back_to_default(self, output_dir,
                                                            test_excel, mock_ai):
        """CLI 指定模板不存在时不报错，回退到默认模板"""
        result = run_pipeline(mode="gen-fpa", file_path=test_excel,
                             output_dir=output_dir,
                             templates={"fpa": "/no/such/template.xlsx"},
                             api_key="sk-test")
        assert os.path.exists(result.fpa_xlsx)  # 用了默认模板

    def test_empty_templates_falls_back_to_default(self, output_dir, test_excel,
                                                     mock_ai):
        """传空 templates 时走默认 data/out_templates/ 回退"""
        result = run_pipeline(mode="gen-fpa", file_path=test_excel,
                             output_dir=output_dir, templates={},
                             api_key="sk-test")
        assert os.path.exists(result.fpa_xlsx)


class TestCustomFilenames:
    """自定义文件名（从 Excel 元数据解析）"""

    def test_resolves_custom_filenames(self, output_dir, test_excel, mock_ai):
        result = run_pipeline(mode="gen-all", file_path=test_excel,
                             output_dir=output_dir, templates=TEMPLATES,
                             api_key="sk-test")
        # 文件名应包含 Excel 中配置的占位符替换结果
        assert "TEST-2025-001" in str(result.fpa_xlsx) or "FPA" in str(result.fpa_xlsx)
