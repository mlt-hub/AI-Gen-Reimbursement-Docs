import logging
import json

import openpyxl
import pytest

from ai_gen_reimbursement_docs.config_utils import FpaConfigError, FpaPromptConfigError
from ai_gen_reimbursement_docs.gen_fpa import (
    FPA_PROJECT_DESCRIPTION_MAX_CHARS,
    _build_domain_context,
    _build_fpa_ai_prompt_context,
    _build_fpa_audit_reports_for_groups,
    _extract_json_obj,
    _fpa_ai_cache_key,
    _group_rows_for_audit,
    _group_rows_by_l3,
    _normalize_ai_fpa_rows_for_l3,
    _plan_fpa_rows_with_execution,
    _plan_fpa_rows_with_ai,
    _rules_first_ai_reasons,
    _supplement_ai_rows_with_rules,
    generate_fpa_check_xlsx_from_md,
    preview_fpa_module,
)
from ai_gen_reimbursement_docs.fpa_profiles import (
    CUSTOM_RULES_PROFILE,
    STRICT_FPA_PROFILE,
    UI_API_MAPPING_PROFILE,
    CustomRulesProfile,
    ExternalDataGroupRule,
    FpaCoverageRules,
    FpaProcessRowsPlanningRule,
    FpaRowPlanningRules,
    FpaRuleSetConfig,
    FpaUiRowPlanningRule,
    KeywordTypeRule,
    TypeMappingRule,
    reset_current_fpa_rule_set_config,
    set_current_fpa_rule_set_config,
)
from ai_gen_reimbursement_docs.pipeline_callbacks import PipelineCallbacks
from ai_gen_reimbursement_docs.runtime_context import callbacks_var


def _meta():
    return {"子系统（模块）": "测试系统", "资产标识": "TEST-001"}


def _rows():
    return [
        {
            "客户端类型": "地市后台",
            "一级模块": "垂直行业营销",
            "二级模块": "垂直行业管理",
            "三级模块": "垂直行业管理",
            "三级模块整体功能描述": "维护垂直行业基础信息和管理员。",
            "功能过程": "添加垂直行业",
            "功能过程类型": "新增",
            "功能过程描述": "输入垂直行业名称并保存。",
        },
        {
            "客户端类型": "地市后台",
            "一级模块": "垂直行业营销",
            "二级模块": "垂直行业管理",
            "三级模块": "垂直行业管理",
            "三级模块整体功能描述": "维护垂直行业基础信息和管理员。",
            "功能过程": "查询垂直行业",
            "功能过程类型": "新增",
            "功能过程描述": "按行业名称查询垂直行业列表。",
        },
    ]


def test_build_domain_context_adds_project_description_from_work_order(monkeypatch):
    monkeypatch.setattr("ai_gen_reimbursement_docs.config_utils.load_optional_fpa_domain_context", lambda: {})
    context = _build_domain_context({
        "工单标题": "供应商协同优化",
        "工单内容": "围绕供应商协同关系维护和查询能力建设。",
        "建设目标": "AI 生成目标不应进入 FPA 上下文。",
        "建设必要性": "AI 生成必要性不应进入 FPA 上下文。",
    })

    assert context["project_description"] == "工单标题：供应商协同优化\n工单内容：围绕供应商协同关系维护和查询能力建设。"
    assert "建设目标" not in context["project_description"]
    assert "建设必要性" not in context["project_description"]


def test_build_domain_context_uses_prefixed_work_order_keys(monkeypatch):
    monkeypatch.setattr("ai_gen_reimbursement_docs.config_utils.load_optional_fpa_domain_context", lambda: {})
    context = _build_domain_context({
        "1、工单需求-元数据录入.工单标题": "行业管理建设",
        "1、工单需求-元数据录入.工单内容": "建设行业管理配置能力。",
    })

    assert context["project_description"] == "工单标题：行业管理建设\n工单内容：建设行业管理配置能力。"


def test_build_domain_context_ignores_configured_project_description(monkeypatch):
    monkeypatch.setattr(
        "ai_gen_reimbursement_docs.config_utils.load_optional_fpa_domain_context",
        lambda: {"project_description": "配置文件不作为项目说明来源", "system_boundary": "本系统边界。"},
    )
    context = _build_domain_context({"工单标题": "模板工单标题"})

    assert context["project_description"] == "工单标题：模板工单标题"
    assert context["system_boundary"] == "本系统边界。"


def test_build_domain_context_truncates_long_work_order_content(monkeypatch):
    monkeypatch.setattr("ai_gen_reimbursement_docs.config_utils.load_optional_fpa_domain_context", lambda: {})
    context = _build_domain_context({"工单内容": "长" * (FPA_PROJECT_DESCRIPTION_MAX_CHARS + 100)})

    assert len(context["project_description"]) > FPA_PROJECT_DESCRIPTION_MAX_CHARS
    assert context["project_description"].endswith("（工单内容已截断，完整内容以功能清单录入模板为准。）")


def _custom_default_rule_set() -> FpaRuleSetConfig:
    return FpaRuleSetConfig(
        name="unified_ui_rs",
        row_planning_rules=FpaRowPlanningRules(
            ui_row=FpaUiRowPlanningRule(
                enabled=True,
                scope="l3",
                merge="single_row",
                name_suffix="界面开发",
                fpa_type="EI",
                reason="三级模块兜底合并界面能力。",
                empty_process_text="完成三级模块页面交互能力",
                explanation_template="{name}，具体为以下：\n{items}",
            ),
            process_rows=FpaProcessRowsPlanningRule(
                enabled=True,
                one_row_per_process=True,
                default_name_suffix="逻辑处理开发",
                type_suffixes={"EQ": "查询处理开发", "EO": "导出处理开发", "EI": "导入处理开发"},
                explanation_template="{name}，具体为以下：\n1、{description}",
            ),
        ),
        type_mapping_rules=(
            TypeMappingRule("EI", ("界面开发", "页面")),
            TypeMappingRule("EIF", ("外部应用维护", "外部系统维护", "引用外部数据组", "统一用户中心", "外部主数据")),
            TypeMappingRule("ILF", ("添加", "新增", "编辑", "修改", "删除", "维护", "保存", "启用", "停用", "更新")),
            TypeMappingRule("ILF", ("外部接口", "外部数据")),
        ),
        keyword_rules=(
            KeywordTypeRule("EO", ("导出", "报表输出", "生成文件", "下载", "下载模板", "下载文件")),
            KeywordTypeRule("EQ", ("查询", "查看", "详情", "列表检索", "检索")),
            KeywordTypeRule("EI", ("导入",)),
        ),
    )


def _strict_default_rule_set() -> FpaRuleSetConfig:
    return FpaRuleSetConfig(
        name="strict_fpa_rs",
        keyword_rules=(
            KeywordTypeRule("EO", ("导出", "报表", "下载", "生成文件"), "事务功能产生派生或格式化输出，按 EO。"),
            KeywordTypeRule("EQ", ("查询", "查看", "详情", "检索", "列表"), "事务功能读取数据且无派生输出，按 EQ。"),
            KeywordTypeRule(
                "EI",
                ("新增", "添加", "修改", "编辑", "删除", "维护", "保存", "提交", "审批", "启用", "停用", "导入", "同步", "发起", "写入", "选择", "引用", "关联"),
                "事务功能进入或改变系统边界内数据，按 EI。",
            ),
        ),
        external_data_rules=(
            ExternalDataGroupRule(("统一用户中心", "用户中心"), "统一用户中心账号", ("账号", "账户", "人员", "组织", "机构", "信息")),
        ),
    )


def _structured_explanation(process: str = "添加垂直行业", fpa_type: str = "EI") -> str:
    return (
        f"来源场景：来自“【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-{process}”，"
        "后台用户输入垂直行业名称并保存。"
        "\n业务数据：涉及垂直行业数据，输入字段为垂直行业名称。"
        "\n业务规则：系统根据用户提交动作创建或处理对应垂直行业记录。"
        f"\n计算说明：该功能过程体现后台用户维护垂直行业业务数据，可支撑 FPA 功能点计量，并按 {fpa_type} 识别。"
    )


def _write_fpa_prompt_config(tmp_path, monkeypatch):
    (tmp_path / "fpa_config.yaml").write_text(
        """
default-profile: unified_ui
adjustment_value_method_default: legacy_workload
adjustment_value_methods:
  legacy_workload:
    type_weights:
      EI: 2
      default: 1
profiles:
  unified_ui:
    kind: unified_ui
    strategy: rules_first
    rule_set: unified_ui_rs
    core_rules: unified_ui_cr
    system_prompt: unified_ui_sp
    user_prompt: unified_ui_up
  strict_fpa:
    kind: strict_fpa
    strategy: ai_first
    rule_set: strict_fpa_rs
    core_rules: strict_fpa_cr
    system_prompt: strict_fpa_sp
    user_prompt: strict_fpa_up
core_rules:
  unified_ui_cr: CUSTOM CORE RULES
  strict_fpa_cr: STRICT CORE RULES
system_prompt_sets:
  unified_ui_sp: 系统提示词
  strict_fpa_sp: 严格系统提示词
user_prompt_sets:
  unified_ui_up: |-
    ${core_rules}
    模块输入 JSON：
    ${payload_json}
    判定原则：
    ${judgement_rules}
  strict_fpa_up: |-
    ${core_rules}
    模块输入 JSON：
    ${payload_json}
    判定原则：
    ${judgement_rules}
rule_sets:
  unified_ui_rs:
    row_planning_rules:
      ui_row:
        enabled: true
        scope: l3
        merge: single_row
        name_suffix: "界面开发"
        type: EI
        reason: "三级模块兜底合并界面能力。"
        empty_process_text: "完成三级模块页面交互能力"
        explanation_template: "{name}，具体为以下：\n{items}"
      process_rows:
        enabled: true
        one_row_per_process: true
        default_name_suffix: "逻辑处理开发"
        type_suffixes:
          EQ: "查询处理开发"
          EO: "导出处理开发"
          EI: "导入处理开发"
        explanation_template: "{name}，具体为以下：\n1、{description}"
    type_mapping_rules:
      merge: append
      items:
        - type: EI
          keywords: ["界面开发", "页面"]
        - type: ILF
          keywords: ["添加", "新增", "编辑", "修改", "删除", "维护", "保存", "启用", "停用", "更新"]
    keyword_rules:
      merge: append
      items:
        - type: EO
          keywords: ["导出", "报表输出", "生成文件", "下载", "下载模板", "下载文件"]
        - type: EQ
          keywords: ["查询", "查看", "详情", "列表检索", "检索"]
        - type: EI
          keywords: ["导入"]
    coverage_rules:
      require_process_coverage: true
      require_data_function: true
  strict_fpa_rs:
    coverage_rules:
      require_process_coverage: true
      require_data_function: true
""",
        encoding="utf-8",
    )
    (tmp_path / "fpa_judgement_rules.yaml").write_text(
        "judgement_rules:\n  - 规则一\n  - 规则二\n",
        encoding="utf-8",
    )
    monkeypatch.setattr("ai_gen_reimbursement_docs.config_utils.config_dir", lambda: tmp_path)


def test_fpa_audit_grouping_prefers_source_process_over_l3_substring():
    tree_rows = [
        {
            "客户端类型": "地市后台",
            "一级模块": "垂直行业营销",
            "二级模块": "垂直行业管理",
            "三级模块": "垂直行业管理",
            "三级模块整体功能描述": "",
            "功能过程": "添加垂直行业",
            "功能过程类型": "新增",
            "功能过程描述": "",
        },
        {
            "客户端类型": "地市后台",
            "一级模块": "垂直行业营销",
            "二级模块": "垂直行业管理",
            "三级模块": "合伙商管理",
            "三级模块整体功能描述": "",
            "功能过程": "搜索合作商",
            "功能过程类型": "查询",
            "功能过程描述": "",
        },
    ]
    groups = _group_rows_by_l3(tree_rows)
    fpa_rows = [
        {
            "新增/修改功能点": "【地市后台】垂直行业营销-垂直行业管理-合伙商管理-搜索合作商",
            "计算依据说明": "",
            "源功能过程": "搜索合作商",
        }
    ]

    grouped = _group_rows_for_audit(fpa_rows, groups)

    assert grouped[1] == []
    assert grouped[2] == fpa_rows


def test_markdown_code_block_json_is_parsed():
    data = _extract_json_obj("""```json
{"rows":[{"name":"垂直行业管理界面开发"}]}
```""")
    assert data["rows"][0]["name"] == "垂直行业管理界面开发"


def test_normalize_ai_rows_maps_basis_and_types():
    group = _group_rows_by_l3(_rows())[0]
    token = set_current_fpa_rule_set_config(_custom_default_rule_set())
    try:
        rows, warnings = _normalize_ai_fpa_rows_for_l3(
            group=group,
            meta=_meta(),
            judgement_rules=["规则一", "规则二"],
            start_seq=1,
            ai_rows=[
                {
                    "name": "垂直行业管理界面开发",
                    "type": "EI",
                    "type_reason": "页面交互能力",
                    "classification_basis_index": 1,
                    "explanation": "垂直行业管理界面开发，具体为以下：1、新增列表和查询条件。",
                },
                {
                    "name": "添加垂直行业-逻辑处理开发",
                    "type": "ILF",
                    "classification_basis_index": 2,
                    "explanation": "保存垂直行业基础信息。",
                    "source_processes": ["添加垂直行业"],
                },
            ],
        )
    finally:
        reset_current_fpa_rule_set_config(token)
    assert any("AI 行名称前缀已按源功能清单规范化" in w for w in warnings)
    assert [r["类型"] for r in rows] == ["EI", "ILF"]
    assert rows[0]["新增/修改功能点"] == "【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-垂直行业管理界面开发"
    assert rows[1]["新增/修改功能点"] == "【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-添加垂直行业-逻辑处理开发"
    assert rows[0]["计算依据归类"] == "规则一"
    assert rows[1]["计算依据归类"] == "规则二"


def test_invalid_index_warns_and_leaves_basis_empty():
    group = _group_rows_by_l3(_rows())[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["规则一"],
        start_seq=1,
        ai_rows=[{
            "name": "查询垂直行业-查询处理开发",
            "type": "EQ",
            "classification_basis_index": 99,
            "explanation": "查询垂直行业列表。",
        }],
    )
    assert rows[0]["计算依据归类"] == ""
    assert any("越界" in w for w in warnings)
    rule_hits = rows[0]["_规则命中详情"]
    assert any(hit["rule_id"] == "postprocess.classification_basis_index" for hit in rule_hits)
    assert any("越界" in warning for hit in rule_hits for warning in hit["warnings"])


def test_classification_basis_type_conflict_warns():
    group = _group_rows_by_l3(_rows())[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=[
            "EI: 修改或增加界面的个数",
            "EQ: 提供查询界面输入并展示返回结果",
        ],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[{
            "name": "添加垂直行业",
            "type": "EI",
            "classification_basis_index": 2,
            "explanation": _structured_explanation("添加垂直行业", "EI"),
        }],
    )

    assert rows[0]["类型"] == "EI"
    assert rows[0]["计算依据归类"].startswith("EQ:")
    assert any("计算依据归类指向的类型=EQ" in warning for warning in warnings)
    hit = next(
        hit
        for hit in rows[0]["_规则命中详情"]
        if hit["rule_id"] == "postprocess.classification_basis_type_conflict"
    )
    assert hit["suggested_type"] == "EQ"
    assert hit["adopted"] == "否"


def test_classification_basis_type_match_does_not_warn():
    group = _group_rows_by_l3(_rows())[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=[
            "EI: 修改或增加界面的个数",
            "EQ: 提供查询界面输入并展示返回结果",
        ],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[{
            "name": "查询垂直行业",
            "type": "EQ",
            "classification_basis_index": 2,
            "explanation": _structured_explanation("查询垂直行业", "EQ"),
        }],
    )

    assert rows[0]["类型"] == "EQ"
    assert rows[0]["计算依据归类"].startswith("EQ:")
    assert not any("计算依据归类指向的类型" in warning for warning in warnings)
    assert not any(
        hit["rule_id"] == "postprocess.classification_basis_type_conflict"
        for hit in rows[0]["_规则命中详情"]
    )


def test_structured_explanation_passes_quality_check():
    group = _group_rows_by_l3(_rows())[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["规则一"],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[{
            "name": "添加垂直行业",
            "type": "EI",
            "classification_basis_index": 1,
            "explanation": _structured_explanation("添加垂直行业", "EI"),
        }],
    )

    assert not any("计算依据说明" in warning for warning in warnings)
    assert not any(
        hit["rule_id"] == "postprocess.explanation_quality"
        for hit in rows[0]["_规则命中详情"]
    )


def test_data_function_explanation_accepts_data_group_source_path():
    group = _group_rows_by_l3(_rows())[0]
    explanation = (
        "来源场景：来自“【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-垂直行业数据组”，"
        "本系统维护垂直行业基础信息。"
        "\n业务数据：涉及垂直行业数据，字段包括行业名称。"
        "\n业务规则：系统保存并持续维护垂直行业逻辑数据组。"
        "\n计算说明：该数据组由本系统维护，可支撑 FPA 功能点计量，并按 ILF 识别。"
    )

    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["规则一"],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[{
            "name": "垂直行业数据组",
            "type": "ILF",
            "classification_basis_index": 1,
            "explanation": explanation,
        }],
    )

    assert not any("来源场景未使用完整路径格式" in warning for warning in warnings)
    assert not any(
        hit["rule_id"] == "postprocess.explanation_quality"
        for hit in rows[0]["_规则命中详情"]
    )


def test_data_function_source_path_is_normalized_to_full_row_name():
    group = _group_rows_by_l3(_rows())[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["规则一"],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[{
            "name": "垂直行业数据组",
            "type": "ILF",
            "classification_basis_index": 1,
            "explanation": (
                "来源场景：来自“垂直行业管理”，本系统维护垂直行业基础信息。"
                "\n业务数据：涉及垂直行业数据，字段包括行业名称。"
                "\n业务规则：系统保存并持续维护垂直行业逻辑数据组。"
                "\n计算说明：该数据组由本系统维护，可支撑 FPA 功能点计量，并按 ILF 识别。"
            ),
        }],
    )

    assert rows[0]["计算依据说明"].startswith(
        "来源场景：【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-垂直行业数据组"
    )
    assert not any("来源场景未使用完整路径格式" in warning for warning in warnings)
    assert not any("<功能过程>" in warning for warning in warnings)
    assert not any("<功能点名称>" in warning for warning in warnings)
    quality_hit = next(
        hit for hit in rows[0]["_规则命中详情"]
        if hit["rule_id"] == "postprocess.explanation_source_path"
    )
    assert any("来源场景已按完整功能点路径规范化" in warning for warning in quality_hit["warnings"])


def test_transaction_source_path_is_normalized_to_full_row_name():
    group = _group_rows_by_l3(_rows())[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["规则一"],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[{
            "name": "垂直行业维护",
            "type": "EI",
            "classification_basis_index": 1,
            "explanation": (
                "来源场景：垂直行业维护"
                "\n业务数据：垂直行业基础信息。"
                "\n业务规则：系统保存并持续维护垂直行业信息。"
                "\n计算说明：该事务维护内部逻辑数据组，按 EI 计量。"
            ),
        }],
    )

    assert rows[0]["计算依据说明"].startswith(
        "来源场景：【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-垂直行业维护"
    )
    assert not any("来源场景未使用完整路径格式" in warning for warning in warnings)
    assert any(
        hit["rule_id"] == "postprocess.explanation_source_path"
        for hit in rows[0]["_规则命中详情"]
    )


def test_explanation_quality_accepts_module_or_business_scene_source_text():
    group = _group_rows_by_l3([{
        "客户端类型": "地市后台",
        "一级模块": "权限管理",
        "二级模块": "账号权限",
        "三级模块": "用户中心账号引用",
        "三级模块整体功能描述": "系统引用统一用户中心维护的人员账号，本系统不维护账号主数据。",
        "功能过程": "引用统一用户中心账号",
        "功能过程类型": "新增",
        "功能过程描述": "引用统一用户中心账号基础信息和所属组织。",
    }])[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["规则一"],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[{
            "name": "统一用户中心账号数据组",
            "type": "EIF",
            "classification_basis_index": 1,
            "explanation": (
                "来源场景：模块描述「用户中心账号引用」中明确本系统不维护账号主数据，引用统一用户中心账号。"
                "\n业务数据：统一用户中心账号基础信息和所属组织。"
                "\n业务规则：本系统读取外部数据组，不进行插入、修改或删除。"
                "\n计算说明：按 EIF 评估外部引用的逻辑数据组。"
            ),
        }],
    )

    assert rows[0]["新增/修改功能点"] == "【地市后台】权限管理-账号权限-用户中心账号引用-统一用户中心账号数据组"
    assert not any("来源场景未使用完整路径格式" in warning for warning in warnings)
    assert not any(
        hit["rule_id"] == "postprocess.explanation_quality"
        for hit in rows[0]["_规则命中详情"]
    )


def test_explanation_quality_accepts_source_process_operation_anchor():
    group = _group_rows_by_l3([{
        "客户端类型": "地市后台",
        "一级模块": "消息管理",
        "二级模块": "通知发送",
        "三级模块": "短信通知",
        "三级模块整体功能描述": "运营人员配置短信内容并触发短信发送。",
        "功能过程": "编辑短信模板",
        "功能过程类型": "新增",
        "功能过程描述": "维护短信标题、正文和变量。",
    }])[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["规则一"],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[{
            "name": "短信模板数据组",
            "type": "ILF",
            "classification_basis_index": 1,
            "explanation": (
                "来源场景：编辑短信模板操作维护短信模板数据组"
                "\n业务数据：短信模板数据组，包含标题、正文、变量等字段。"
                "\n业务规则：运营人员编辑短信模板后，系统保存至本系统数据库。"
                "\n计算说明：短信模板为内部逻辑数据组，按 ILF 计量。"
            ),
        }],
    )

    assert rows[0]["新增/修改功能点"] == "【地市后台】消息管理-通知发送-短信通知-短信模板数据组"
    assert not any("来源场景未使用完整路径格式" in warning for warning in warnings)
    assert not any(
        hit["rule_id"] == "postprocess.explanation_quality"
        for hit in rows[0]["_规则命中详情"]
    )


def test_explanation_warns_when_table_count_basis_is_used_as_detail():
    group = _group_rows_by_l3(_rows())[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["按后台数据库变更的表个数计量"],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[{
            "name": "垂直行业数据组",
            "type": "ILF",
            "classification_basis_index": 1,
            "explanation": (
                "来源场景：来自“【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-垂直行业数据组”，"
                "本系统维护垂直行业基础信息。"
                "\n业务数据：涉及垂直行业数据，字段包括行业名称。"
                "\n业务规则：系统保存并持续维护垂直行业逻辑数据组。"
                "\n计算说明：该数据组由本系统维护，数据库表个数=1，按表数量作为详细计量解释。"
            ),
        }],
    )

    assert rows[0]["计算依据归类"] == "按后台数据库变更的表个数计量"
    assert any("应保留在计算依据归类而非计算依据说明" in warning for warning in warnings)
    quality_hit = next(
        hit for hit in rows[0]["_规则命中详情"]
        if hit["rule_id"] == "postprocess.explanation_quality"
    )
    assert any("数据库表个数" in warning for warning in quality_hit["warnings"])


def test_explanation_allows_classification_basis_wording_with_fpa_definition():
    group = _group_rows_by_l3(_rows())[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["按后台数据库变更的表个数计量"],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[{
            "name": "垂直行业数据组",
            "type": "ILF",
            "classification_basis_index": 1,
            "explanation": (
                "来源场景：来自“【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-垂直行业数据组”，"
                "本系统维护垂直行业基础信息。"
                "\n业务数据：涉及垂直行业数据，字段包括行业名称。"
                "\n业务规则：系统保存并持续维护垂直行业逻辑数据组。"
                "\n计算说明：根据系统边界，本系统维护的逻辑数据集合符合 ILF 定义，按后台数据库变更的表个数计量。"
            ),
        }],
    )

    assert rows[0]["计算依据归类"] == "按后台数据库变更的表个数计量"
    assert not any("数据库表个数" in warning for warning in warnings)
    assert not any(
        hit["rule_id"] == "postprocess.explanation_quality"
        for hit in rows[0]["_规则命中详情"]
    )


def test_explanation_normalizes_parenthetical_table_count_detail_without_warning():
    group = _group_rows_by_l3(_rows())[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["按后台数据库变更的表个数计量"],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[{
            "name": "垂直行业数据组",
            "type": "ILF",
            "classification_basis_index": 1,
            "explanation": (
                "来源场景：【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-垂直行业数据组"
                "\n业务数据：垂直行业基础信息。"
                "\n业务规则：本系统内部维护，支持新增、修改、删除等操作。"
                "\n计算说明：本系统维护该逻辑数据组，符合ILF定义，按ILF计量，对应后台数据库变更的1个表。"
            ),
        }],
    )

    assert rows[0]["计算依据归类"] == "按后台数据库变更的表个数计量"
    assert "对应后台数据库变更的1个表" not in rows[0]["计算依据说明"]
    assert not any("数据库表个数" in warning for warning in warnings)
    assert any(
        hit["rule_id"] == "postprocess.explanation_table_count_detail"
        for hit in rows[0]["_规则命中详情"]
    )
    assert not any(
        hit["rule_id"] == "postprocess.explanation_quality"
        for hit in rows[0]["_规则命中详情"]
    )


def test_explanation_quality_warns_for_fabricated_system_elements():
    group = _group_rows_by_l3(_rows())[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["维护业务数据的外部输入，按 EI 识别。"],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[{
            "name": "新增客户",
            "type": "EI",
            "classification_basis_index": 1,
            "explanation": (
                "来源场景：【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-新增客户"
                "\n业务数据：客户名称、证件号码。"
                "\n业务规则：后台用户提交客户资料后系统保存。"
                "\n系统元素：涉及客户信息表，用于保存客户资料。"
                "\n计算说明：该功能体现外部输入事务，按 EI 识别。"
            ),
        }],
    )

    assert any("系统元素疑似包含输入未明确提供" in warning for warning in warnings)
    quality_hit = next(
        hit for hit in rows[0]["_规则命中详情"]
        if hit["rule_id"] == "postprocess.explanation_quality"
    )
    assert any("客户信息表" in warning for warning in quality_hit["warnings"])


def test_explanation_quality_warns_for_multiline_fabricated_system_elements():
    group = _group_rows_by_l3(_rows())[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["维护业务数据的外部输入，按 EI 识别。"],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[{
            "name": "新增客户",
            "type": "EI",
            "classification_basis_index": 1,
            "explanation": (
                "来源场景：【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-新增客户"
                "\n业务数据：客户名称、证件号码。"
                "\n业务规则：后台用户提交客户资料后系统保存。"
                "\n系统元素："
                "\n- 客户信息表，用于保存客户资料。"
                "\n- 客户保存服务。"
                "\n计算说明：该功能体现外部输入事务，按 EI 识别。"
            ),
        }],
    )

    assert any("客户信息表" in warning and "客户保存服务" in warning for warning in warnings)
    quality_hit = next(
        hit for hit in rows[0]["_规则命中详情"]
        if hit["rule_id"] == "postprocess.explanation_quality"
    )
    assert any("客户信息表" in warning and "客户保存服务" in warning for warning in quality_hit["warnings"])


def test_explanation_quality_warns_for_inline_fabricated_system_elements():
    group = _group_rows_by_l3(_rows())[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["维护业务数据的外部输入，按 EI 识别。"],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[{
            "name": "新增客户",
            "type": "EI",
            "classification_basis_index": 1,
            "explanation": (
                "来源场景：【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-新增客户"
                "\n业务数据：客户名称、证件号码。"
                "\n业务规则：后台用户提交客户资料后系统写入客户表。"
                "\n计算说明：该功能体现外部输入事务，按 EI 识别。"
            ),
        }],
    )

    assert any("正文疑似提到输入未明确提供" in warning and "客户表" in warning for warning in warnings)
    quality_hit = next(
        hit for hit in rows[0]["_规则命中详情"]
        if hit["rule_id"] == "postprocess.explanation_quality"
    )
    assert any("客户表" in warning for warning in quality_hit["warnings"])


def test_explanation_quality_accepts_inline_system_elements_from_input():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "客户运营",
            "二级模块": "客户同步",
            "三级模块": "客户资料同步",
            "三级模块整体功能描述": "对接CRM客户查询接口，同步客户基础资料。",
            "功能过程": "同步客户资料",
            "功能过程类型": "新增",
            "功能过程描述": "调用CRM客户查询接口获取客户名称和证件号码。",
        },
    ])[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["维护业务数据的外部输入，按 EI 识别。"],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[{
            "name": "同步客户资料",
            "type": "EI",
            "classification_basis_index": 1,
            "explanation": (
                "来源场景：【地市后台】客户运营-客户同步-客户资料同步-同步客户资料"
                "\n业务数据：客户名称和证件号码。"
                "\n业务规则：系统调用CRM客户查询接口获取客户基础资料并保存。"
                "\n计算说明：该功能体现外部输入事务，按 EI 识别。"
            ),
            "source_process_ids": ["m1_p1"],
        }],
    )

    assert not any("正文疑似提到输入未明确提供" in warning for warning in warnings)
    assert not any(
        hit["rule_id"] == "postprocess.explanation_quality"
        for hit in rows[0]["_规则命中详情"]
    )


def test_explanation_quality_ignores_generic_missing_system_element_wording():
    group = _group_rows_by_l3(_rows())[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["维护业务数据的外部输入，按 EI 识别。"],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[{
            "name": "新增客户",
            "type": "EI",
            "classification_basis_index": 1,
            "explanation": (
                "来源场景：【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-新增客户"
                "\n业务数据：客户名称、证件号码。"
                "\n业务规则：后台用户提交客户资料后系统保存。"
                "\n系统元素：未涉及明确的服务或外部系统。"
                "\n计算说明：该功能体现外部输入事务，按 EI 识别。"
            ),
        }],
    )

    assert not any("系统元素疑似包含输入未明确提供" in warning for warning in warnings)
    assert not any(
        hit["rule_id"] == "postprocess.explanation_quality"
        for hit in rows[0]["_规则命中详情"]
    )


def test_explanation_quality_ignores_development_and_type_wording_as_inline_elements():
    group = _group_rows_by_l3(_rows())[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["接口开发映射为 ILF", "按输出的票据、报表、统计、文件个数计量"],
        start_seq=1,
        profile=UI_API_MAPPING_PROFILE,
        ai_rows=[
            {
                "name": "新增客户-接口开发",
                "type": "ILF",
                "classification_basis_index": 1,
                "explanation": (
                    "来源场景：【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-新增客户-接口开发"
                    "\n业务数据：客户资料。"
                    "\n业务规则：接口开发行按 ILF 识别。"
                    "\n计算说明：该接口开发行固定按 ILF 识别。"
                ),
            },
            {
                "name": "导出客户清单",
                "type": "EO",
                "classification_basis_index": 2,
                "explanation": (
                    "来源场景：【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-导出客户清单"
                    "\n业务数据：客户清单。"
                    "\n业务规则：系统将客户记录整理为文件。"
                    "\n计算说明：输出的票据、报表、统计、文件，按 EO 识别。"
                ),
            },
        ],
    )

    assert not any("接口开发行" in warning for warning in warnings)
    assert not any("外部文件" in warning for warning in warnings)
    assert not any(
        hit["rule_id"] == "postprocess.explanation_quality"
        and any("接口开发行" in warning or "外部文件" in warning for warning in hit["warnings"])
        for row in rows
        for hit in row["_规则命中详情"]
    )


def test_explanation_quality_ignores_list_and_represents_wording_as_inline_elements():
    group = _group_rows_by_l3(_rows())[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["维护业务数据的外部输入，按 EI 识别。"],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[{
            "name": "查询客户",
            "type": "EI",
            "classification_basis_index": 1,
            "explanation": (
                "来源场景：【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-查询客户"
                "\n业务数据：客户名称、证件号码。"
                "\n业务规则：同一三级模块的列表、该行代表查询客户能力。"
                "\n计算说明：该功能体现外部输入事务，按 EI 识别。"
            ),
        }],
    )

    assert not any("正文疑似提到输入未明确提供" in warning for warning in warnings)
    assert not any(
        hit["rule_id"] == "postprocess.explanation_quality"
        and any("列表" in warning or "代表" in warning for warning in hit["warnings"])
        for row in rows
        for hit in row["_规则命中详情"]
    )


def test_explanation_quality_ignores_customer_list_wording_as_inline_element():
    group = _group_rows_by_l3(_rows())[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["查询业务数据且无派生计算，按 EQ 识别。"],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[{
            "name": "查询客户",
            "type": "EQ",
            "classification_basis_index": 1,
            "explanation": (
                "来源场景：【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-查询客户"
                "\n业务数据：客户名称、证件号码。"
                "\n业务规则：按条件展示客户列表，不改变业务数据。"
                "\n计算说明：该功能体现外部查询事务，按 EQ 识别。"
            ),
        }],
    )

    assert not any("正文疑似提到输入未明确提供" in warning for warning in warnings)
    assert not any(
        hit["rule_id"] == "postprocess.explanation_quality"
        and any("客户列表" in warning for warning in hit["warnings"])
        for row in rows
        for hit in row["_规则命中详情"]
    )


def test_explanation_quality_ignores_generated_file_action_as_inline_element():
    group = _group_rows_by_l3(_rows())[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["输出格式化清单或报表，按 EO 识别。"],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[{
            "name": "导出客户清单",
            "type": "EO",
            "classification_basis_index": 1,
            "explanation": (
                "来源场景：【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-导出客户清单"
                "\n业务数据：客户清单。"
                "\n业务规则：导出时触发文件生成。"
                "\n计算说明：该功能体现外部输出事务，按 EO 识别。"
            ),
        }],
    )

    assert not any("正文疑似提到输入未明确提供" in warning for warning in warnings)
    assert not any(
        hit["rule_id"] == "postprocess.explanation_quality"
        and any("触发文件" in warning for warning in hit["warnings"])
        for row in rows
        for hit in row["_规则命中详情"]
    )


def test_explanation_quality_ignores_exported_file_output_as_inline_element():
    group = _group_rows_by_l3(_rows())[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["输出格式化清单或报表，按 EO 识别。"],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[{
            "name": "导出客户清单",
            "type": "EO",
            "classification_basis_index": 1,
            "explanation": (
                "来源场景：【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-导出客户清单"
                "\n业务数据：客户清单。"
                "\n业务规则：导出客户资料为文件。"
                "\n计算说明：此功能为外部输出，输出客户资料文件，按 EO 计量。"
            ),
        }],
    )

    assert not any("正文疑似提到输入未明确提供" in warning for warning in warnings)
    assert not any(
        hit["rule_id"] == "postprocess.explanation_quality"
        and any("客户资料文件" in warning or "客户资料为文件" in warning for warning in hit["warnings"])
        for row in rows
        for hit in row["_规则命中详情"]
    )


def test_explanation_quality_accepts_explicit_system_elements_from_input():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "客户运营",
            "二级模块": "客户同步",
            "三级模块": "客户资料同步",
            "三级模块整体功能描述": "对接CRM客户查询接口，同步客户基础资料。",
            "功能过程": "同步客户资料",
            "功能过程类型": "新增",
            "功能过程描述": "调用CRM客户查询接口获取客户名称和证件号码。",
        },
    ])[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["维护业务数据的外部输入，按 EI 识别。"],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[{
            "name": "同步客户资料",
            "type": "EI",
            "classification_basis_index": 1,
            "explanation": (
                "来源场景：【地市后台】客户运营-客户同步-客户资料同步-同步客户资料"
                "\n业务数据：客户名称和证件号码。"
                "\n业务规则：系统调用接口获取客户基础资料并保存。"
                "\n系统元素：涉及CRM客户查询接口。"
                "\n计算说明：该功能体现外部输入事务，按 EI 识别。"
            ),
            "source_process_ids": ["m1_p1"],
        }],
    )

    assert not any("系统元素疑似包含输入未明确提供" in warning for warning in warnings)
    assert not any(
        hit["rule_id"] == "postprocess.explanation_quality"
        for hit in rows[0]["_规则命中详情"]
    )


def test_ui_api_mapping_keeps_default_ui_rows_without_split_reason():
    group = _group_rows_by_l3([
        {
            "客户端类型": "后台",
            "一级模块": "业务管理",
            "二级模块": "客户管理",
            "三级模块": "客户资料维护",
            "三级模块整体功能描述": "维护客户基础资料。",
            "功能过程": "新增客户",
            "功能过程类型": "新增",
            "功能过程描述": "录入客户资料。",
        },
        {
            "客户端类型": "后台",
            "一级模块": "业务管理",
            "二级模块": "客户管理",
            "三级模块": "客户资料维护",
            "三级模块整体功能描述": "维护客户基础资料。",
            "功能过程": "查询客户",
            "功能过程类型": "查询",
            "功能过程描述": "查询客户列表。",
        },
    ])[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["修改或增加界面的个数"],
        start_seq=1,
        profile=UI_API_MAPPING_PROFILE,
        ai_rows=[
            {
                "name": "新增客户-界面开发",
                "type": "EI",
                "classification_basis_index": 1,
                "explanation": (
                    "来源场景：【后台】业务管理-客户管理-客户资料维护-新增客户-界面开发\n"
                    "业务数据：客户资料。\n"
                    "业务规则：用户录入客户资料。\n"
                    "计算说明：该界面开发行按 EI 识别。"
                ),
                "source_process_ids": ["m1_p1"],
            },
            {
                "name": "查询客户-界面开发",
                "type": "EI",
                "classification_basis_index": 1,
                "explanation": (
                    "来源场景：【后台】业务管理-客户管理-客户资料维护-查询客户-界面开发\n"
                    "业务数据：客户查询条件。\n"
                    "业务规则：用户输入查询条件。\n"
                    "计算说明：该界面开发行按 EI 识别。"
                ),
                "source_process_ids": ["m1_p2"],
            },
        ],
    )

    assert [row["新增/修改功能点"] for row in rows] == [
        "【后台】业务管理-客户管理-客户资料维护-新增客户-界面开发",
        "【后台】业务管理-客户管理-客户资料维护-查询客户-界面开发",
    ]
    assert not any("AI 输出多条界面开发行但缺少 split_reason" in warning for warning in warnings)


def test_ui_api_mapping_normalizes_development_suffix_connectors():
    group = _group_rows_by_l3([
        {
            "客户端类型": "后台",
            "一级模块": "业务管理",
            "二级模块": "客户管理",
            "三级模块": "客户资料维护",
            "三级模块整体功能描述": "维护客户基础资料。",
            "功能过程": "新增客户",
            "功能过程类型": "新增",
            "功能过程描述": "录入客户资料。",
        },
    ])[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["修改或增加界面的个数", "接口开发映射为 ILF"],
        start_seq=1,
        profile=UI_API_MAPPING_PROFILE,
        ai_rows=[
            {
                "name": "新增客户_界面开发",
                "type": "EI",
                "classification_basis_index": 1,
                "explanation": (
                    "来源场景：【后台】业务管理-客户管理-客户资料维护-新增客户-界面开发\n"
                    "业务数据：客户资料。\n"
                    "业务规则：用户录入客户资料。\n"
                    "计算说明：该界面开发行按 EI 识别。"
                ),
                "source_process_ids": ["m1_p1"],
            },
            {
                "name": "新增客户_接口开发",
                "type": "ILF",
                "classification_basis_index": 2,
                "explanation": (
                    "来源场景：【后台】业务管理-客户管理-客户资料维护-新增客户-接口开发\n"
                    "业务数据：客户资料。\n"
                    "业务规则：系统保存客户资料。\n"
                    "计算说明：该接口开发行按 ILF 识别。"
                ),
                "source_process_ids": ["m1_p1"],
            },
        ],
    )

    assert [row["新增/修改功能点"] for row in rows] == [
        "【后台】业务管理-客户管理-客户资料维护-新增客户-界面开发",
        "【后台】业务管理-客户管理-客户资料维护-新增客户-接口开发",
    ]
    assert any("AI 行名称连接符已规范化" in warning and "新增客户_界面开发" in warning for warning in warnings)
    assert any("AI 行名称连接符已规范化" in warning and "新增客户_接口开发" in warning for warning in warnings)
    assert all(
        any(hit["rule_id"] == "postprocess.ai_name_connector" for hit in row["_规则命中详情"])
        for row in rows
    )


def test_ui_api_mapping_keeps_hyphen_development_suffix_connectors():
    group = _group_rows_by_l3([
        {
            "客户端类型": "后台",
            "一级模块": "业务管理",
            "二级模块": "客户管理",
            "三级模块": "客户资料维护",
            "三级模块整体功能描述": "维护客户基础资料。",
            "功能过程": "新增客户",
            "功能过程类型": "新增",
            "功能过程描述": "录入客户资料。",
        },
    ])[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["修改或增加界面的个数"],
        start_seq=1,
        profile=UI_API_MAPPING_PROFILE,
        ai_rows=[{
            "name": "新增客户-界面开发",
            "type": "EI",
            "classification_basis_index": 1,
            "explanation": (
                "来源场景：【后台】业务管理-客户管理-客户资料维护-新增客户-界面开发\n"
                "业务数据：客户资料。\n"
                "业务规则：用户录入客户资料。\n"
                "计算说明：该界面开发行按 EI 识别。"
            ),
            "source_process_ids": ["m1_p1"],
        }],
    )

    assert rows[0]["新增/修改功能点"] == "【后台】业务管理-客户管理-客户资料维护-新增客户-界面开发"
    assert not any("AI 行名称连接符已规范化" in warning for warning in warnings)
    assert not any(
        hit["rule_id"] == "postprocess.ai_name_connector"
        for hit in rows[0]["_规则命中详情"]
    )


def test_explanation_accepts_official_measurement_wording_as_type_evidence():
    group = _group_rows_by_l3(_rows())[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["按修改或增加界面的个数计量", "按输出的票据、报表、统计、文件个数计量"],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[
            {
                "name": "导入客户名单",
                "type": "EI",
                "classification_basis_index": 1,
                "explanation": (
                    "来源场景：【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-导入客户名单"
                    "\n业务数据：客户名单记录。"
                    "\n业务规则：上传文件后校验数据格式，保存有效记录。"
                    "\n计算说明：该过程通过界面操作向ILF插入数据，属于修改或增加界面的个数。"
                ),
            },
            {
                "name": "下载导入模板",
                "type": "EO",
                "classification_basis_index": 5,
                "explanation": (
                    "来源场景：【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-下载导入模板"
                    "\n业务数据：客户名单导入模板文件。"
                    "\n业务规则：系统根据预设模板生成文件并提供下载。"
                    "\n计算说明：输出格式化文件，属于输出的票据、报表、统计、文件。"
                ),
            },
        ],
    )

    assert not any("未明确当前 FPA 类型" in warning for warning in warnings)
    assert not any(
        hit["rule_id"] == "postprocess.explanation_quality"
        for row in rows
        for hit in row["_规则命中详情"]
    )


def test_explanation_accepts_fpa_type_business_aliases():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "消息管理",
            "二级模块": "通知发送",
            "三级模块": "短信通知",
            "三级模块整体功能描述": "运营人员配置短信内容并触发短信发送。",
            "功能过程": "编辑短信模板",
            "功能过程类型": "新增",
            "功能过程描述": "维护短信标题、正文和变量。",
        },
        {
            "客户端类型": "地市后台",
            "一级模块": "消息管理",
            "二级模块": "通知发送",
            "三级模块": "短信通知",
            "三级模块整体功能描述": "运营人员配置短信内容并触发短信发送。",
            "功能过程": "查看发送记录",
            "功能过程类型": "查询",
            "功能过程描述": "查询短信发送状态和失败原因。",
        },
    ])[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["规则一"],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[
            {
                "name": "短信模板维护",
                "type": "EI",
                "classification_basis_index": 1,
                "explanation": (
                    "来源场景：【地市后台】消息管理-通知发送-短信通知-短信模板维护，运营人员编辑短信模板并保存。"
                    "\n业务数据：短信模板的标题、正文、变量等字段。"
                    "\n业务规则：运营人员可新增或修改短信模板内容，系统将数据写入短信模板数据组。"
                    "\n计算说明：该功能对 ILF 执行修改操作，属于外部输入事务，按对 ILF 的插入、修改、删除操作次数计量。"
                ),
                "source_process_ids": ["m1_p1"],
                "source_processes": ["编辑短信模板"],
            },
            {
                "name": "发送记录查询",
                "type": "EQ",
                "classification_basis_index": 1,
                "explanation": (
                    "来源场景：【地市后台】消息管理-通知发送-短信通知-发送记录查询，运营人员通过查询界面输入条件并查看结果。"
                    "\n业务数据：发送记录的状态、失败原因、发送时间等字段。"
                    "\n业务规则：运营人员根据号码、时间等条件筛选并查看发送记录详情。"
                    "\n计算说明：该功能读取发送记录数据组并展示结果，无派生计算，属于外部查询，按提供查询界面输入并展示返回结果计量。"
                ),
                "source_process_ids": ["m1_p2"],
                "source_processes": ["查看发送记录"],
            },
        ],
    )

    assert [row["类型"] for row in rows] == ["EI", "EQ"]
    assert not any("未明确当前 FPA 类型" in warning for warning in warnings)
    assert not any(
        hit["rule_id"] == "postprocess.explanation_quality"
        for row in rows
        for hit in row["_规则命中详情"]
    )


def test_explanation_accepts_ilf_natural_language_aliases():
    group = _group_rows_by_l3(_rows())[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["按后台数据库变更的表个数计量"],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[
            {
                "name": "退款申请数据组",
                "type": "ILF",
                "classification_basis_index": 11,
                "explanation": (
                    "来源场景：【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-退款申请数据组"
                    "\n业务数据：退款申请状态、退款金额、订单号。"
                    "\n业务规则：本系统负责记录和管理退款申请状态。"
                    "\n计算说明：对应本系统维护的逻辑数据集合，按后台数据库变更的表个数计量。"
                ),
            },
            {
                "name": "短信模板数据组",
                "type": "ILF",
                "classification_basis_index": 11,
                "explanation": (
                    "来源场景：【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-短信模板数据组"
                    "\n业务数据：短信标题、正文和变量字段。"
                    "\n业务规则：运营人员可以新增或修改短信模板，系统持久化存储。"
                    "\n计算说明：本系统维护的短信模板数据组属于内部逻辑文件，按后台数据库变更的表个数计量。"
                ),
            },
        ],
    )

    assert [row["类型"] for row in rows] == ["ILF", "ILF"]
    assert not any("未明确当前 FPA 类型" in warning for warning in warnings)
    assert not any(
        hit["rule_id"] == "postprocess.explanation_quality"
        for row in rows
        for hit in row["_规则命中详情"]
    )


def test_explanation_accepts_output_file_and_ilf_insert_type_aliases():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "客户运营",
            "二级模块": "客户数据管理",
            "三级模块": "客户名单导入",
            "三级模块整体功能描述": "运营人员导入客户名单并下载模板。",
            "功能过程": "下载导入模板",
            "功能过程类型": "导出",
            "功能过程描述": "生成并输出Excel模板文件。",
        },
        {
            "客户端类型": "地市后台",
            "一级模块": "客户运营",
            "二级模块": "客户数据管理",
            "三级模块": "客户名单导入",
            "三级模块整体功能描述": "运营人员导入客户名单并下载模板。",
            "功能过程": "导入客户名单",
            "功能过程类型": "新增",
            "功能过程描述": "上传Excel文件并保存有效客户名单。",
        },
    ])[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["规则一"],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[
            {
                "name": "导入模板下载",
                "type": "EO",
                "classification_basis_index": 1,
                "explanation": (
                    "来源场景：【地市后台】客户运营-客户数据管理-客户名单导入-导入模板下载"
                    "\n业务数据：Excel模板文件，包含客户名单导入所需列标题。"
                    "\n业务规则：模板为预定义格式，包含固定列结构。"
                    "\n计算说明：输出的模板文件属于格式化文件输出，按输出的文件个数计量。"
                ),
            },
            {
                "name": "客户名单导入维护",
                "type": "EI",
                "classification_basis_index": 1,
                "explanation": (
                    "来源场景：【地市后台】客户运营-客户数据管理-客户名单导入-客户名单导入维护"
                    "\n业务数据：客户名单记录。"
                    "\n业务规则：用户触发导入操作，系统向本系统内部数据组插入一条或多条客户名单记录。"
                    "\n计算说明：本过程对ILF进行插入操作，按对ILF的插入、修改、删除操作次数计量。"
                ),
            },
        ],
    )

    assert [row["类型"] for row in rows] == ["EO", "EI"]
    assert not any("未明确当前 FPA 类型" in warning for warning in warnings)


def test_explanation_accepts_eif_and_eq_natural_type_aliases_from_trend_samples():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "供应商管理",
            "二级模块": "准入协同",
            "三级模块": "供应商准入协同",
            "三级模块整体功能描述": "维护供应商准入申请，并引用 CRM 客户档案和 OA 审批流程单据。",
            "功能过程": "选择CRM客户档案",
            "功能过程类型": "新增",
            "功能过程描述": "从 CRM 客户档案中选择客户并关联到供应商准入申请。",
        },
        {
            "客户端类型": "地市后台",
            "一级模块": "供应商管理",
            "二级模块": "准入协同",
            "三级模块": "供应商准入协同",
            "三级模块整体功能描述": "维护供应商准入申请，并引用 CRM 客户档案和 OA 审批流程单据。",
            "功能过程": "查看准入申请",
            "功能过程类型": "查询",
            "功能过程描述": "查询供应商准入申请、关联客户档案和审批状态。",
        },
    ])[0]
    rule_set = FpaRuleSetConfig(
        name="strict_fpa_rs",
        external_data_rules=(
            ExternalDataGroupRule(("CRM",), "CRM客户档案", ("档案", "数据组")),
            ExternalDataGroupRule(("OA",), "OA审批流程单据", ("流程单据", "数据组")),
        ),
    )
    token = set_current_fpa_rule_set_config(rule_set)
    try:
        rows, warnings = _normalize_ai_fpa_rows_for_l3(
            group=group,
            meta=_meta(),
            judgement_rules=["规则一"],
            start_seq=1,
            profile=STRICT_FPA_PROFILE,
            strategy="ai_first",
            ai_rows=[
                {
                    "name": "CRM客户档案数据组",
                    "type": "EIF",
                    "classification_basis_index": 1,
                    "explanation": (
                        "来源场景：【地市后台】供应商管理-准入协同-供应商准入协同-CRM客户档案数据组"
                        "\n业务数据：CRM客户档案。"
                        "\n业务规则：本系统引用外部系统数据，不进行维护。"
                        "\n计算说明：按评估范围外相关的表个数计量，外部系统CRM有1个客户档案数据组，纳入FPA数据功能计量。"
                    ),
                },
                {
                    "name": "OA审批流程单据数据组",
                    "type": "EIF",
                    "classification_basis_index": 1,
                    "explanation": (
                        "来源场景：【地市后台】供应商管理-准入协同-供应商准入协同-OA审批流程单据数据组"
                        "\n业务数据：OA审批流程单据。"
                        "\n业务规则：本系统引用外部系统数据，不进行维护。"
                        "\n计算说明：OA系统维护的审批流程单据属于外部接口文件，纳入FPA数据功能计量。"
                    ),
                },
                {
                    "name": "准入申请查询",
                    "type": "EQ",
                    "classification_basis_index": 1,
                    "explanation": (
                        "来源场景：【地市后台】供应商管理-准入协同-供应商准入协同-准入申请查询"
                        "\n业务数据：供应商准入申请、关联客户档案和审批状态。"
                        "\n业务规则：用户输入查询条件后展示结果，不进行数据修改或派生计算。"
                        "\n计算说明：按提供查询界面输入并展示返回结果计量，该查询涉及一个查询界面。"
                    ),
                },
            ],
        )
    finally:
        reset_current_fpa_rule_set_config(token)

    assert [row["类型"] for row in rows] == ["EIF", "EIF", "EQ"]
    assert not any("未明确当前 FPA 类型" in warning for warning in warnings)


def test_unstructured_explanation_records_quality_warning():
    group = _group_rows_by_l3(_rows())[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["规则一"],
        start_seq=1,
        ai_rows=[{
            "name": "添加垂直行业-逻辑处理开发",
            "type": "EI",
            "classification_basis_index": 1,
            "explanation": "保存垂直行业基础信息。",
        }],
    )

    assert any("计算依据说明格式不完整" in warning for warning in warnings)
    assert any("未使用完整路径格式" in warning for warning in warnings)
    assert any("未明确当前 FPA 类型: EI" in warning for warning in warnings)
    rule_hits = rows[0]["_规则命中详情"]
    quality_hit = next(hit for hit in rule_hits if hit["rule_id"] == "postprocess.explanation_quality")
    assert "结构化证据说明规则" in quality_hit["rule_desc"]
    assert "计算依据说明格式不完整" in "；".join(quality_hit["warnings"])


def test_multiple_ui_rows_without_split_reason_are_merged():
    group = _group_rows_by_l3(_rows())[0]
    token = set_current_fpa_rule_set_config(_custom_default_rule_set())
    try:
        rows, warnings = _normalize_ai_fpa_rows_for_l3(
            group=group,
            meta=_meta(),
            judgement_rules=[],
            start_seq=1,
            ai_rows=[
                {"name": "垂直行业列表界面开发", "type": "EI", "explanation": "列表。"},
                {"name": "垂直行业查询界面开发", "type": "EI", "explanation": "查询。"},
                {"name": "添加垂直行业-逻辑处理开发", "type": "ILF", "explanation": "保存。"},
            ],
        )
    finally:
        reset_current_fpa_rule_set_config(token)
    assert sum(1 for r in rows if "界面开发" in r["新增/修改功能点"]) == 1
    assert any("split_reason" in w for w in warnings)


def test_multi_uis_duplicate_ui_rows_are_kept_with_review_metadata():
    group = {
        "client_type": "地市后台",
        "l1": "客户管理",
        "l2": "客户中心",
        "l3": "客户档案",
        "processes": [
            {"name": "维护客户档案", "type": "新增", "desc": "维护客户档案。"},
        ],
    }
    ai_rows = [
        {
            "name": "客户档案-界面开发",
            "type": "EI",
            "classification_basis_index": 1,
            "explanation": _structured_explanation("维护客户档案", "EI"),
            "source_processes": ["维护客户档案"],
            "split_reason": "独立页面：基础信息维护页。",
        },
        {
            "name": "客户档案-界面开发",
            "type": "EI",
            "classification_basis_index": 1,
            "explanation": _structured_explanation("维护客户档案", "EI"),
            "source_processes": ["维护客户档案"],
            "split_reason": "独立业务对象：联系人维护区。",
        },
    ]

    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        ai_rows=ai_rows,
        judgement_rules=["规则一"],
        profile=CustomRulesProfile(name="multi_uis"),
    )

    ui_rows = [row for row in rows if "界面开发" in str(row["新增/修改功能点"])]
    assert len(ui_rows) == 2
    assert "拆分理由" in str(ui_rows[0]["后处理警告"])
    assert "独立页面" in str(ui_rows[0]["后处理警告"])
    assert "同名多界面开发行" in str(ui_rows[1]["后处理警告"])
    assert any("同名多界面开发行" in warning for warning in warnings)


def test_strict_profile_normalizes_ai_development_work_item_names():
    group = _group_rows_by_l3(_rows())[0]
    token = set_current_fpa_rule_set_config(_strict_default_rule_set())
    try:
        rows, warnings = _normalize_ai_fpa_rows_for_l3(
            group=group,
            meta=_meta(),
            judgement_rules=[],
            start_seq=1,
            profile=STRICT_FPA_PROFILE,
            ai_rows=[
                {
                    "name": "添加垂直行业-逻辑处理开发",
                    "type": "ILF",
                    "explanation": "输入垂直行业名称并保存。",
                },
            ],
        )
    finally:
        reset_current_fpa_rule_set_config(token)

    assert rows[0]["新增/修改功能点"] == "【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-添加垂直行业"
    assert rows[0]["类型"] == "EI"
    assert any("已规范化" in w for w in warnings)


def test_ai_name_prefix_is_forced_from_source_module_path():
    group = _group_rows_by_l3(_rows())[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=[],
        start_seq=1,
        ai_rows=[
            {
                "name": "和乐业-垂直行业营销-垂直行业管理-垂直行业管理-查询垂直行业",
                "type": "EQ",
                "explanation": "查询垂直行业。",
            },
            {
                "name": "地市后台-垂直行业营销-垂直行业管理-垂直行业管理-垂直行业数据组",
                "type": "ILF",
                "explanation": "垂直行业数据组。",
            },
            {
                "name": "自定义页面列表查询",
                "type": "EQ",
                "explanation": "查询自定义页面。",
            },
            {
                "name": "【地市后台】垂直行业营销-装修管理-首页装修-行业首页装修数据",
                "type": "ILF",
                "explanation": "行业首页装修数据。",
            },
            {
                "name": "地市后台-垂直行业营销-装修管理-自定义装修-自定义页面信息",
                "type": "ILF",
                "explanation": "自定义页面信息。",
            },
        ],
    )

    assert [row["新增/修改功能点"] for row in rows] == [
        "【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-查询垂直行业",
        "【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-垂直行业数据组",
        "【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-自定义页面列表查询",
        "【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-行业首页装修数据",
        "【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-自定义页面信息",
    ]
    assert len([w for w in warnings if "AI 行名称前缀已按源功能清单规范化" in w]) == 5
    assert all(
        any(hit["rule_id"] == "postprocess.ai_name_prefix" for hit in row["_规则命中详情"])
        for row in rows
    )


def test_ai_name_prefix_overlap_keeps_only_business_suffix():
    group = _group_rows_by_l3([{
        "客户端类型": "地市后台",
        "一级模块": "支付管理",
        "二级模块": "退款处理",
        "三级模块": "退款处理",
        "三级模块整体功能描述": "处理支付退款。",
        "功能过程": "提交退款申请",
        "功能过程类型": "新增",
        "功能过程描述": "提交退款申请并记录退款状态。",
    }])[0]

    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=[],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        strategy="ai_first",
        ai_rows=[
            {
                "name": "【地市后台】支付管理-退款处理-退款处理数据组",
                "type": "ILF",
                "explanation": "来源场景：【地市后台】支付管理-退款处理-退款处理-退款处理数据组\n业务数据：退款申请数据。\n业务规则：记录退款申请。\n计算说明：内部逻辑数据按 ILF 计量。",
            },
            {
                "name": "【地市后台】支付管理-退款处理-退款处理维护",
                "type": "EI",
                "explanation": "来源场景：【地市后台】支付管理-退款处理-退款处理-退款处理维护\n业务数据：退款申请数据。\n业务规则：提交退款申请并更新状态。\n计算说明：外部输入维护内部数据按 EI 计量。",
            },
        ],
    )

    assert [row["新增/修改功能点"] for row in rows] == [
        "【地市后台】支付管理-退款处理-退款处理-退款处理数据组",
        "【地市后台】支付管理-退款处理-退款处理-退款处理维护",
    ]
    assert all("退款处理-【地市后台】" not in row["新增/修改功能点"] for row in rows)
    assert len([w for w in warnings if "AI 行名称前缀已按源功能清单规范化" in w]) == 2
    assert all(
        any(hit["rule_id"] == "postprocess.ai_name_prefix" for hit in row["_规则命中详情"])
        for row in rows
    )


def test_ai_name_prefix_rewrites_three_part_module_path_without_duplicate_prefix():
    group = _group_rows_by_l3([{
        "客户端类型": "地市后台",
        "一级模块": "组织管理",
        "二级模块": "内部组织",
        "三级模块": "内部组织维护",
        "三级模块整体功能描述": "维护内部组织。",
        "功能过程": "查询内部组织",
        "功能过程类型": "查询",
        "功能过程描述": "查询内部组织列表。",
    }])[0]

    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=[],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        strategy="ai_first",
        ai_rows=[{
            "name": "【地市后台】组织管理-内部组织-内部组织查询",
            "type": "EQ",
            "explanation": (
                "来源场景：【地市后台】组织管理-内部组织-内部组织维护-内部组织查询"
                "\n业务数据：内部组织信息。"
                "\n业务规则：查询内部组织列表。"
                "\n计算说明：提供查询界面输入并展示返回结果，符合EQ定义。"
            ),
        }],
    )

    assert rows[0]["新增/修改功能点"] == "【地市后台】组织管理-内部组织-内部组织维护-内部组织查询"
    assert "内部组织维护-【地市后台】" not in rows[0]["新增/修改功能点"]
    assert len([w for w in warnings if "AI 行名称前缀已按源功能清单规范化" in w]) == 1


def test_ai_name_process_suffix_normalization_does_not_duplicate_prefix_warning():
    group = _group_rows_by_l3([{
        "客户端类型": "地市后台",
        "一级模块": "采购管理",
        "二级模块": "供应商协同",
        "三级模块": "ERP订单引用",
        "三级模块整体功能描述": "引用 ERP 订单。",
        "功能过程": "查看ERP订单信息",
        "功能过程类型": "查询",
        "功能过程描述": "查看 ERP 订单的供应商、金额和状态。",
    }])[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=[],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        strategy="ai_first",
        ai_rows=[{
            "name": "【地市后台】采购管理-供应商协同-ERP订单引用-ERP订单信息查询",
            "type": "EQ",
            "explanation": "来源场景：【地市后台】采购管理-供应商协同-ERP订单引用-ERP订单信息查询\n业务数据：ERP订单。\n业务规则：只读取展示。\n计算说明：按 EQ 计量。",
            "source_process_ids": ["m1_p1"],
        }],
    )

    assert rows[0]["新增/修改功能点"] == "【地市后台】采购管理-供应商协同-ERP订单引用-查看ERP订单信息"
    assert not any("AI 行名称末尾已按 source_process_id 规范化" in w for w in warnings)
    assert not any("AI 行名称前缀已按源功能清单规范化" in w for w in warnings)
    suffix_hit = next(
        hit for hit in rows[0]["_规则命中详情"]
        if hit["rule_id"] == "postprocess.ai_name_process_suffix"
    )
    assert "AI 行名称末尾已按 source_process_id 规范化" in suffix_hit["warnings"][0]


def test_strict_profile_corrects_external_service_eif_misclassification():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "消息管理",
            "二级模块": "通知发送",
            "三级模块": "短信通知",
            "三级模块整体功能描述": "系统调用短信平台发送通知短信。",
            "功能过程": "发送测试短信",
            "功能过程类型": "新增",
            "功能过程描述": "调用短信平台发送测试短信。",
        },
    ])[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=[],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        ai_rows=[
            {
                "name": "发送测试短信",
                "type": "EIF",
                "explanation": "调用短信平台发送测试短信。",
            },
        ],
    )

    assert rows[0]["类型"] == "EI"
    conflict_warning = next(w for w in warnings if "明显冲突" in w)
    assert "规则建议 type=EI" in conflict_warning
    assert "普通外部服务调用按触发事务处理，不能直接判 EIF" in conflict_warning
    hit = next(hit for hit in rows[0]["_规则命中详情"] if hit["rule_id"] == "postprocess.keyword_type_conflict")
    assert hit["suggested_type"] == "EI"
    assert "规则建议 type=EI" in hit["rule_desc"]
    assert "普通外部服务调用按触发事务处理，不能直接判 EIF" in hit["rule_desc"]


def test_ai_first_keeps_valid_ai_type_without_keyword_override():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "消息管理",
            "二级模块": "通知发送",
            "三级模块": "短信通知",
            "三级模块整体功能描述": "系统调用短信平台发送通知短信。",
            "功能过程": "发送测试短信",
            "功能过程类型": "新增",
            "功能过程描述": "调用短信平台发送测试短信。",
        },
    ])[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=[],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        strategy="ai_first",
        ai_rows=[
            {
                "name": "发送测试短信",
                "type": "EIF",
                "explanation": "调用短信平台发送测试短信。",
            },
        ],
    )

    assert rows[0]["类型"] == "EIF"
    conflict_warning = next(w for w in warnings if "AI 优先策略下保留 AI type" in w)
    assert "规则建议 type=EI" in conflict_warning
    assert "普通外部服务调用按触发事务处理，不能直接判 EIF" in conflict_warning
    hit = next(hit for hit in rows[0]["_规则命中详情"] if hit["rule_id"] == "postprocess.ai_first_type_conflict")
    assert hit["suggested_type"] == "EI"
    assert "规则建议 type=EI" in hit["rule_desc"]
    assert "普通外部服务调用按触发事务处理，不能直接判 EIF" in hit["rule_desc"]


def test_ai_first_does_not_warn_type_conflict_when_rule_type_matches_ai_type():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "垂直行业营销",
            "二级模块": "垂直行业管理",
            "三级模块": "垂直行业管理",
            "三级模块整体功能描述": "维护垂直行业基础信息和管理员。",
            "功能过程": "添加垂直行业",
            "功能过程类型": "新增",
            "功能过程描述": "点击添加按钮，在弹窗输入垂直行业名称并保存。",
        },
    ])[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=[],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        strategy="ai_first",
        ai_rows=[
            {
                "name": "添加垂直行业",
                "type": "EI",
                "explanation": "点击添加按钮，在弹窗输入垂直行业名称并保存。",
            },
        ],
    )

    assert rows[0]["类型"] == "EI"
    assert not any("AI type=EI 与规则存在冲突" in warning for warning in warnings)
    assert not any(hit["rule_id"] == "postprocess.ai_first_type_conflict" for hit in rows[0]["_规则命中详情"])


def test_ai_first_does_not_warn_type_conflict_when_agent_judgement_supports_eif():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "权限管理",
            "二级模块": "账号权限",
            "三级模块": "用户中心账号引用",
            "三级模块整体功能描述": "系统引用统一用户中心维护的人员账号，本系统不维护账号主数据。",
            "功能过程": "引用统一用户中心账号",
            "功能过程类型": "新增",
            "功能过程描述": "引用统一用户中心账号基础信息和所属组织。",
        },
        {
            "客户端类型": "地市后台",
            "一级模块": "权限管理",
            "二级模块": "账号权限",
            "三级模块": "用户中心账号引用",
            "三级模块整体功能描述": "系统引用统一用户中心维护的人员账号，本系统不维护账号主数据。",
            "功能过程": "选择业务负责人",
            "功能过程类型": "新增",
            "功能过程描述": "从用户中心账号中选择负责人并保存到本系统业务对象。",
        },
    ])[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=[],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        strategy="ai_first",
        ai_rows=[
            {
                "name": "统一用户中心账号数据组",
                "type": "EIF",
                "explanation": "来源场景：【地市后台】权限管理-账号权限-用户中心账号引用-统一用户中心账号数据组。\n业务数据：统一用户中心账号。\n业务规则：统一用户中心维护，本系统引用。\n计算说明：EIF。",
                "source_process_ids": ["m1_p1"],
                "source_processes": ["引用统一用户中心账号"],
            },
        ],
    )

    assert rows[0]["类型"] == "EIF"
    assert not any("AI type=EIF 与规则存在冲突" in warning for warning in warnings)
    assert not any(hit["rule_id"] == "postprocess.ai_first_type_conflict" for hit in rows[0]["_规则命中详情"])


def test_ai_first_data_function_review_uses_agent_external_data_judgement():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "组织管理",
            "二级模块": "归属组织",
            "三级模块": "归属组织选择",
            "三级模块整体功能描述": "引用主数据平台维护的组织主数据，本系统不维护组织主数据。",
            "功能过程": "选择归属组织",
            "功能过程类型": "新增",
            "功能过程描述": "从主数据平台维护的组织主数据中选择记录，并保存到本系统业务对象。",
        },
    ])[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=[],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        strategy="ai_first",
        ai_rows=[
            {
                "name": "组织主数据数据组",
                "type": "EIF",
                "explanation": (
                    "来源场景：【地市后台】组织管理-归属组织-归属组织选择-组织主数据数据组\n"
                    "业务数据：组织主数据。\n"
                    "业务规则：主数据平台维护，本系统引用。\n"
                    "计算说明：外部系统维护且本系统引用的数据组，按 EIF 计量。"
                ),
                "source_process_ids": ["m1_p1"],
                "source_processes": ["选择归属组织"],
            },
        ],
    )

    assert rows[0]["类型"] == "EIF"
    assert not any("AI 数据功能需人工复核" in warning for warning in warnings)
    assert not any(
        hit["rule_id"] == "postprocess.ai_data_group_review"
        for hit in rows[0]["_规则命中详情"]
    )


def test_ai_first_none_judgement_does_not_conflict_with_supported_transaction_row():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "消息管理",
            "二级模块": "通知发送",
            "三级模块": "短信通知",
            "三级模块整体功能描述": "运营人员配置短信内容并触发短信发送，系统调用短信平台完成发送。",
            "功能过程": "编辑短信模板",
            "功能过程类型": "新增",
            "功能过程描述": "维护短信标题、正文和变量。",
        },
    ])[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=[],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        strategy="ai_first",
        ai_rows=[
            {
                "name": "短信模板维护",
                "type": "EI",
                "explanation": (
                    "来源场景：【地市后台】消息管理-通知发送-短信通知-短信模板维护\n"
                    "业务数据：短信模板。\n"
                    "业务规则：运营人员编辑短信模板并保存到本系统。\n"
                    "计算说明：本功能维护内部 ILF，按 EI 计量。"
                ),
                "source_process_ids": ["m1_p1"],
                "source_processes": ["编辑短信模板"],
            },
        ],
    )

    assert rows[0]["类型"] == "EI"
    assert not any("AI type=EI 与规则存在冲突" in warning for warning in warnings)
    assert not any(
        hit["rule_id"] == "postprocess.ai_first_type_conflict"
        for hit in rows[0]["_规则命中详情"]
    )


def test_ai_first_internal_data_group_with_external_reference_ids_keeps_ilf():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "供应商管理",
            "二级模块": "准入协同",
            "三级模块": "供应商准入协同",
            "三级模块整体功能描述": "本系统维护供应商准入申请信息，同时引用 CRM 系统维护的客户档案和 OA 系统维护的审批流程单据。",
            "功能过程": "新增准入申请",
            "功能过程类型": "新增",
            "功能过程描述": "录入供应商准入申请信息并保存。",
        },
        {
            "客户端类型": "地市后台",
            "一级模块": "供应商管理",
            "二级模块": "准入协同",
            "三级模块": "供应商准入协同",
            "三级模块整体功能描述": "本系统维护供应商准入申请信息，同时引用 CRM 系统维护的客户档案和 OA 系统维护的审批流程单据。",
            "功能过程": "选择CRM客户档案",
            "功能过程类型": "新增",
            "功能过程描述": "从 CRM 客户档案中选择客户并关联到供应商准入申请。",
        },
    ])[0]
    rule_set = FpaRuleSetConfig(
        name="strict_fpa_rs",
        external_data_rules=(
            ExternalDataGroupRule(("CRM",), "CRM客户档案", ("档案", "ID")),
            ExternalDataGroupRule(("OA",), "OA审批流程单据", ("单据", "ID")),
        ),
    )
    token = set_current_fpa_rule_set_config(rule_set)
    try:
        rows, warnings = _normalize_ai_fpa_rows_for_l3(
            group=group,
            meta=_meta(),
            judgement_rules=[],
            start_seq=1,
            profile=STRICT_FPA_PROFILE,
            strategy="ai_first",
            ai_rows=[
                {
                    "name": "供应商准入申请数据组",
                    "type": "ILF",
                    "explanation": (
                        "来源场景：【地市后台】供应商管理-准入协同-供应商准入协同-供应商准入申请数据组，本系统维护准入申请信息。"
                        "\n业务数据：供应商准入申请信息，包括申请单号、供应商名称、关联CRM客户档案ID、关联OA审批单ID、申请状态等。"
                        "\n业务规则：本系统内部维护该数据组，支持新增、修改、关联外部数据。"
                        "\n计算说明：作为本系统内部维护的逻辑数据组，按 ILF 计量。"
                    ),
                    "source_process_ids": [],
                    "source_processes": [],
                },
            ],
        )
    finally:
        reset_current_fpa_rule_set_config(token)

    assert rows[0]["类型"] == "ILF"
    assert not any("AI type=ILF 与规则存在冲突" in warning for warning in warnings)
    assert not any(
        hit["rule_id"] == "postprocess.ai_first_type_conflict"
        for hit in rows[0]["_规则命中详情"]
    )


def test_ai_first_oa_approval_reference_keeps_external_doc_and_internal_relation_types():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "审批管理",
            "二级模块": "OA审批协同",
            "三级模块": "OA审批单关联",
            "三级模块整体功能描述": "系统引用 OA 系统维护的审批流程单据，本系统只保存业务对象与审批单的关联关系。",
            "功能过程": "关联审批单",
            "功能过程类型": "新增",
            "功能过程描述": "从 OA 流程单据中选择审批单并关联到当前业务申请。",
        },
        {
            "客户端类型": "地市后台",
            "一级模块": "审批管理",
            "二级模块": "OA审批协同",
            "三级模块": "OA审批单关联",
            "三级模块整体功能描述": "系统引用 OA 系统维护的审批流程单据，本系统只保存业务对象与审批单的关联关系。",
            "功能过程": "查看审批进度",
            "功能过程类型": "查询",
            "功能过程描述": "查看 OA 审批流程状态、当前审批人和审批意见。",
        },
    ])[0]

    strict_rule_set = _strict_default_rule_set()
    oa_rule_set = FpaRuleSetConfig(
        name=strict_rule_set.name,
        keyword_rules=strict_rule_set.keyword_rules,
        external_data_rules=(
            *strict_rule_set.external_data_rules,
            ExternalDataGroupRule(("OA", "OA 系统", "OA系统"), "OA流程单据", ("流程单据", "审批单", "单据")),
        ),
    )
    token = set_current_fpa_rule_set_config(oa_rule_set)
    try:
        rows, warnings = _normalize_ai_fpa_rows_for_l3(
            group=group,
            meta=_meta(),
            judgement_rules=[],
            start_seq=1,
            profile=STRICT_FPA_PROFILE,
            strategy="ai_first",
            ai_rows=[
                {
                    "name": "【地市后台】审批管理-OA审批协同-OA审批单关联-OA流程单据",
                    "type": "EIF",
                    "explanation": (
                        "来源场景：【地市后台】审批管理-OA审批协同-OA审批单关联-OA流程单据。\n"
                        "业务数据：OA 系统维护的审批流程单据。\n"
                        "业务规则：本系统只引用审批单本体，不维护审批单内容。\n"
                        "计算说明：外部系统维护且本系统引用的数据组，按 EIF 计量。"
                    ),
                },
                {
                    "name": "【地市后台】审批管理-OA审批协同-OA审批单关联-OA审批单关联数据组",
                    "type": "ILF",
                    "explanation": (
                        "来源场景：【地市后台】审批管理-OA审批协同-OA审批单关联-OA审批单关联数据组。\n"
                        "业务数据：业务对象ID、审批单ID、关联关系创建时间等字段。\n"
                        "业务规则：本系统只保存关联关系数据，不包含审批单本身内容。\n"
                        "计算说明：该关联关系数据组由本系统维护，按 ILF 计量。"
                    ),
                },
                {
                    "name": "【地市后台】审批管理-OA审批协同-OA审批单关联-OA审批单关联维护",
                    "type": "EI",
                    "explanation": (
                        "来源场景：【地市后台】审批管理-OA审批协同-OA审批单关联-OA审批单关联维护。\n"
                        "业务数据：业务对象ID、选中的审批单ID。\n"
                        "业务规则：用户选择已有审批单并保存关联关系。\n"
                        "计算说明：该事务维护本系统 OA 审批单关联数据组，按 EI 计量。"
                    ),
                    "source_process_ids": ["m1_p1"],
                    "source_processes": ["关联审批单"],
                },
                {
                    "name": "【地市后台】审批管理-OA审批协同-OA审批单关联-审批进度查询",
                    "type": "EQ",
                    "explanation": (
                        "来源场景：【地市后台】审批管理-OA审批协同-OA审批单关联-审批进度查询。\n"
                        "业务数据：审批流程状态、当前审批人和审批意见。\n"
                        "业务规则：读取 OA 审批进度并展示，不修改本系统数据。\n"
                        "计算说明：只读取并展示数据，按 EQ 计量。"
                    ),
                    "source_process_ids": ["m1_p2"],
                    "source_processes": ["查看审批进度"],
                },
            ],
        )
    finally:
        reset_current_fpa_rule_set_config(token)

    assert [row["类型"] for row in rows] == ["EIF", "ILF", "EI", "EQ"]
    assert not any("AI type=" in warning and "与规则存在冲突" in warning for warning in warnings)
    assert not any("AI 数据功能需人工复核" in warning for warning in warnings)
    assert not any(
        hit["rule_id"] == "postprocess.ai_first_type_conflict"
        for row in rows
        for hit in row["_规则命中详情"]
    )


def test_ai_first_external_master_data_maintenance_keeps_ei_transaction_type():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "组织管理",
            "二级模块": "外部组织引用",
            "三级模块": "外部组织引用",
            "三级模块整体功能描述": "引用主数据平台维护的组织主数据，本系统不维护组织主数据。",
            "功能过程": "选择外部组织",
            "功能过程类型": "新增",
            "功能过程描述": "从主数据平台组织主数据中选择外部组织并关联到当前业务对象。",
        },
    ])[0]

    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=[],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        strategy="ai_first",
        ai_rows=[{
            "name": "【地市后台】组织管理-外部组织引用-外部组织引用-组织主数据维护",
            "type": "EI",
            "explanation": (
                "来源场景：【地市后台】组织管理-外部组织引用-外部组织引用-组织主数据维护。\n"
                "业务数据：组织主数据、当前业务对象关联关系。\n"
                "业务规则：将外部组织关联到当前业务对象，内部数据变化。\n"
                "计算说明：涉及修改或增加界面的输入操作，改变内部数据，按 EI 计量。"
            ),
            "source_process_ids": ["m1_p1"],
            "source_processes": ["选择外部组织"],
        }],
    )

    assert rows[0]["类型"] == "EI"
    assert not any("AI type=EI 与规则存在冲突" in warning for warning in warnings)
    assert not any(
        hit["rule_id"] == "postprocess.ai_first_type_conflict"
        for hit in rows[0]["_规则命中详情"]
    )


def test_ai_first_external_data_group_with_system_reference_text_keeps_eif():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "供应商管理",
            "二级模块": "准入协同",
            "三级模块": "供应商准入协同",
            "三级模块整体功能描述": "本系统维护供应商准入申请信息，同时引用 CRM 系统维护的客户档案。",
            "功能过程": "选择CRM客户档案",
            "功能过程类型": "新增",
            "功能过程描述": "从 CRM 客户档案中选择客户并关联到供应商准入申请。",
        },
    ])[0]
    rule_set = FpaRuleSetConfig(
        name="strict_fpa_rs",
        external_data_rules=(ExternalDataGroupRule(("CRM",), "CRM客户档案", ("档案", "ID")),),
    )
    token = set_current_fpa_rule_set_config(rule_set)
    try:
        rows, warnings = _normalize_ai_fpa_rows_for_l3(
            group=group,
            meta=_meta(),
            judgement_rules=[],
            start_seq=1,
            profile=STRICT_FPA_PROFILE,
            strategy="ai_first",
            ai_rows=[
                {
                    "name": "CRM客户档案数据组",
                    "type": "EIF",
                    "explanation": (
                        "来源场景：【地市后台】供应商管理-准入协同-供应商准入协同-CRM客户档案数据组，作为外部引用数据组识别。"
                        "\n业务数据：CRM客户档案信息。"
                        "\n业务规则：本系统引用CRM系统维护的客户档案。"
                        "\n计算说明：外部维护本系统引用的数据组为 EIF。"
                    ),
                },
            ],
        )
    finally:
        reset_current_fpa_rule_set_config(token)

    assert rows[0]["类型"] == "EIF"
    assert not any("AI type=EIF 与规则存在冲突" in warning for warning in warnings)
    assert not any(
        hit["rule_id"] == "postprocess.ai_first_type_conflict"
        for hit in rows[0]["_规则命中详情"]
    )


def test_ai_first_internal_data_group_suffix_keeps_ilf_when_module_path_has_maintenance():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "组织管理",
            "二级模块": "内部组织",
            "三级模块": "内部组织维护",
            "三级模块整体功能描述": "本系统维护内部组织信息，包括组织名称、组织编码和启停状态。",
            "功能过程": "新增内部组织",
            "功能过程类型": "新增",
            "功能过程描述": "录入内部组织名称、编码并保存。",
        },
        {
            "客户端类型": "地市后台",
            "一级模块": "组织管理",
            "二级模块": "内部组织",
            "三级模块": "内部组织维护",
            "三级模块整体功能描述": "本系统维护内部组织信息，包括组织名称、组织编码和启停状态。",
            "功能过程": "查询内部组织",
            "功能过程类型": "查询",
            "功能过程描述": "按组织名称和状态查询内部组织列表。",
        },
    ])[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=[],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        strategy="ai_first",
        ai_rows=[
            {
                "name": "【地市后台】组织管理-内部组织-内部组织维护-内部组织数据组",
                "type": "ILF",
                "explanation": (
                    "来源场景：【地市后台】组织管理-内部组织-内部组织维护-内部组织数据组\n"
                    "业务数据：内部组织，包含组织名称、组织编码和启停状态。\n"
                    "业务规则：本系统内部维护该数据组。\n"
                    "计算说明：本系统维护的逻辑数据组，按 ILF 计量。"
                ),
                "source_process_ids": ["m1_p1", "m1_p2"],
                "source_processes": ["新增内部组织", "查询内部组织"],
            },
        ],
    )

    assert rows[0]["类型"] == "ILF"
    assert not any("AI type=ILF 与规则存在冲突" in warning for warning in warnings)
    assert not any("AI 数据功能需人工复核" in warning for warning in warnings)
    assert not any(
        hit["rule_id"] in {"postprocess.ai_first_type_conflict", "postprocess.ai_data_group_review"}
        for hit in rows[0]["_规则命中详情"]
    )


def test_ai_first_data_function_supplement_warning_does_not_report_zero_missing_processes():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "客户管理",
            "二级模块": "客户档案",
            "三级模块": "客户档案",
            "三级模块整体功能描述": "维护客户档案。",
            "功能过程": "添加客户档案",
            "功能过程类型": "新增",
            "功能过程描述": "保存客户档案。",
        },
    ])[0]
    ai_rows, _ = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=[],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        strategy="ai_first",
        ai_rows=[
            {
                "name": "添加客户档案",
                "type": "EI",
                "explanation": "保存客户档案。",
                "source_processes": ["添加客户档案"],
            },
        ],
    )

    combined, warnings = _supplement_ai_rows_with_rules(
        group=group,
        meta=_meta(),
        ai_rows=ai_rows,
        profile=STRICT_FPA_PROFILE,
        strategy="ai_first",
    )

    assert len(combined) == 2
    assert any(row["生成方式"] == "rules_fallback" and row["类型"] == "ILF" for row in combined)
    assert any("AI 结果未包含数据功能行" in warning for warning in warnings)
    assert not any("未覆盖 0 个功能过程" in warning for warning in warnings)


def test_ai_first_supplements_missing_eif_even_when_ai_has_internal_ilf():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "审批管理",
            "二级模块": "OA审批协同",
            "三级模块": "OA审批单关联",
            "三级模块整体功能描述": "系统引用 OA 系统维护的审批流程单据，本系统只保存业务对象与审批单的关联关系。",
            "功能过程": "关联审批单",
            "功能过程类型": "新增",
            "功能过程描述": "从 OA 流程单据中选择审批单并关联到当前业务申请。",
        },
        {
            "客户端类型": "地市后台",
            "一级模块": "审批管理",
            "二级模块": "OA审批协同",
            "三级模块": "OA审批单关联",
            "三级模块整体功能描述": "系统引用 OA 系统维护的审批流程单据，本系统只保存业务对象与审批单的关联关系。",
            "功能过程": "查看审批进度",
            "功能过程类型": "查询",
            "功能过程描述": "查看 OA 审批流程状态、当前审批人和审批意见。",
        },
    ])[0]

    oa_rule_set = FpaRuleSetConfig(
        name="strict_fpa_rs",
        external_data_rules=(ExternalDataGroupRule(("OA", "OA 系统", "OA系统"), "OA流程单据", ("流程单据", "审批单", "单据")),),
    )
    token = set_current_fpa_rule_set_config(oa_rule_set)
    try:
        ai_rows, _ = _normalize_ai_fpa_rows_for_l3(
            group=group,
            meta=_meta(),
            judgement_rules=[],
            start_seq=1,
            profile=STRICT_FPA_PROFILE,
            strategy="ai_first",
            ai_rows=[{
                "name": "OA审批单关联数据组",
                "type": "ILF",
                "explanation": "来源场景：【地市后台】审批管理-OA审批协同-OA审批单关联-OA审批单关联数据组\n业务数据：审批单关联关系。\n业务规则：本系统保存业务对象与审批单的关联关系。\n计算说明：本系统维护的逻辑数据组，按 ILF 计量。",
            }],
        )

        combined, warnings = _supplement_ai_rows_with_rules(
            group=group,
            meta=_meta(),
            ai_rows=ai_rows,
            profile=STRICT_FPA_PROFILE,
            strategy="ai_first",
        )
    finally:
        reset_current_fpa_rule_set_config(token)

    assert any(row["类型"] == "ILF" and row["生成方式"] == "ai" for row in combined)
    assert any(row["类型"] == "EIF" and row["生成方式"] == "rules_fallback" for row in combined)
    assert any("AI 结果未包含数据功能行" in warning for warning in warnings)


def test_ai_first_process_id_covers_near_name_difference_without_rules_fallback():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "垂直行业营销",
            "二级模块": "订单管理",
            "三级模块": "卡券订单",
            "三级模块整体功能描述": "查询卡券订单。",
            "功能过程": "搜索卡劵订单",
            "功能过程类型": "新增",
            "功能过程描述": "按条件搜索卡劵订单。",
        },
    ])[0]
    process_id = group["processes"][0]["process_id"]

    ai_rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=[],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        strategy="ai_first",
        ai_rows=[{
            "name": "【地市后台】垂直行业营销-订单管理-卡券订单-搜索卡券订单",
            "type": "EQ",
            "explanation": "来源场景：按条件搜索卡券订单。\n业务数据：读取卡劵订单数据。\n业务规则：不改变系统状态。\n计算说明：查询条件输入并返回订单列表，按 EQ 识别。",
            "source_process_ids": [process_id],
            "source_processes": ["搜索卡券订单"],
        }],
    )

    combined, supplement_warnings = _supplement_ai_rows_with_rules(
        group=group,
        meta=_meta(),
        ai_rows=ai_rows,
        profile=STRICT_FPA_PROFILE,
        strategy="ai_first",
        rule_set_config=FpaRuleSetConfig(
            name="process_only",
            coverage_rules=FpaCoverageRules(require_data_function=False),
        ),
    )
    audit = _build_fpa_audit_reports_for_groups(
        groups=[group],
        rows_by_module={1: combined},
        warnings_by_module={1: warnings + supplement_warnings},
        profile=STRICT_FPA_PROFILE,
        profile_version=STRICT_FPA_PROFILE.version,
        strategy="ai_first",
        rule_set="strict_fpa_rs",
    )[0]

    assert len(combined) == 1
    assert supplement_warnings == []
    assert combined[0]["新增/修改功能点"] == "【地市后台】垂直行业营销-订单管理-卡券订单-搜索卡劵订单"
    assert combined[0]["源功能过程"] == "搜索卡劵订单"
    assert combined[0]["source_process_ids"] == [process_id]
    assert audit.covered_processes == ["搜索卡劵订单"]
    assert audit.missing_processes == []


def test_ai_first_unified_ui_supplements_profile_required_process_rows_even_when_ui_covers_process():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "客户管理",
            "二级模块": "客户查询",
            "三级模块": "客户查询",
            "三级模块整体功能描述": "查询客户并展示列表。",
            "功能过程": "查询客户",
            "功能过程类型": "新增",
            "功能过程描述": "按客户名称查询客户列表。",
        },
    ])[0]
    process_id = group["processes"][0]["process_id"]
    rule_set = _custom_default_rule_set()
    token = set_current_fpa_rule_set_config(rule_set)
    try:
        ai_rows, _ = _normalize_ai_fpa_rows_for_l3(
            group=group,
            meta=_meta(),
            judgement_rules=[],
            start_seq=1,
            profile=CUSTOM_RULES_PROFILE,
            strategy="ai_first",
            ai_rows=[{
                "name": "客户查询-界面开发",
                "type": "EI",
                "explanation": "客户查询页面提供查询条件和结果列表。",
                "source_process_ids": [process_id],
                "source_processes": ["查询客户"],
            }],
        )

        combined, warnings = _supplement_ai_rows_with_rules(
            group=group,
            meta=_meta(),
            ai_rows=ai_rows,
            profile=CUSTOM_RULES_PROFILE,
            strategy="ai_first",
            rule_set_config=rule_set,
        )
    finally:
        reset_current_fpa_rule_set_config(token)

    assert any(
        row["新增/修改功能点"] == "【地市后台】客户管理-客户查询-客户查询-查询客户-查询处理开发"
        and row["生成方式"] == "rules_fallback"
        for row in combined
    )
    assert any("contract 强制行" in warning for warning in warnings)


def test_ai_first_ui_api_mapping_supplements_required_default_rows_and_normalizes_type():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "消息管理",
            "二级模块": "发送记录",
            "三级模块": "发送记录查询",
            "三级模块整体功能描述": "查询消息发送记录。",
            "功能过程": "查看发送记录",
            "功能过程类型": "新增",
            "功能过程描述": "按手机号查询发送记录并返回详情。",
        },
    ])[0]
    tag = "【地市后台】消息管理-发送记录-发送记录查询"
    api_name = f"{tag}-查看发送记录-接口开发"
    ui_name = f"{tag}-查看发送记录-界面开发"

    ai_rows, _ = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=[],
        start_seq=1,
        profile=UI_API_MAPPING_PROFILE,
        strategy="ai_first",
        ai_rows=[{
            "name": api_name,
            "type": "EQ",
            "explanation": "查询发送记录接口返回详情。",
            "source_processes": ["查看发送记录"],
        }],
    )

    combined, warnings = _supplement_ai_rows_with_rules(
        group=group,
        meta=_meta(),
        ai_rows=ai_rows,
        profile=UI_API_MAPPING_PROFILE,
        strategy="ai_first",
    )

    api_row = next(row for row in combined if row["新增/修改功能点"] == api_name)
    assert api_row["类型"] == "ILF"
    assert api_row["生成方式"] == "ai"
    assert any(hit["rule_id"] == "ui_api_mapping.contract_required_type" for hit in api_row["_规则命中详情"])
    assert sum(1 for row in combined if row["新增/修改功能点"] == api_name) == 1
    assert any(row["新增/修改功能点"] == ui_name and row["类型"] == "EI" and row["生成方式"] == "rules_fallback" for row in combined)
    assert any("默认行类型已修正" in warning and "contract 强制行" in warning for warning in warnings)


def test_ai_first_ui_api_mapping_normalizes_type_without_reporting_added_rows():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "消息管理",
            "二级模块": "发送记录",
            "三级模块": "发送记录查询",
            "三级模块整体功能描述": "查询消息发送记录。",
            "功能过程": "查看发送记录",
            "功能过程类型": "新增",
            "功能过程描述": "按手机号查询发送记录并返回详情。",
        },
    ])[0]
    tag = "【地市后台】消息管理-发送记录-发送记录查询"
    api_name = f"{tag}-查看发送记录-接口开发"
    ui_name = f"{tag}-查看发送记录-界面开发"

    ai_rows, _ = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=[],
        start_seq=1,
        profile=UI_API_MAPPING_PROFILE,
        strategy="ai_first",
        ai_rows=[
            {
                "name": api_name,
                "type": "EQ",
                "explanation": "查询发送记录接口返回详情。",
                "source_processes": ["查看发送记录"],
            },
            {
                "name": ui_name,
                "type": "EI",
                "explanation": "发送记录查询页面提供查询条件。",
                "source_processes": ["查看发送记录"],
            },
        ],
    )

    combined, warnings = _supplement_ai_rows_with_rules(
        group=group,
        meta=_meta(),
        ai_rows=ai_rows,
        profile=UI_API_MAPPING_PROFILE,
        strategy="ai_first",
        rule_set_config=FpaRuleSetConfig(
            name="contract_only",
            coverage_rules=FpaCoverageRules(require_data_function=False, require_process_coverage=False),
        ),
    )

    api_row = next(row for row in combined if row["新增/修改功能点"] == api_name)
    assert api_row["类型"] == "ILF"
    assert len(combined) == 2
    assert any("已修正 1 条 ai 行类型" in warning for warning in warnings)
    assert not any("rules_fallback" in warning for warning in warnings)


def test_ai_first_missing_source_process_ids_warns_and_falls_back_to_source_names():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "客户管理",
            "二级模块": "客户查询",
            "三级模块": "客户查询",
            "三级模块整体功能描述": "查询客户。",
            "功能过程": "查询客户",
            "功能过程类型": "新增",
            "功能过程描述": "按客户名称查询客户列表。",
        },
    ])[0]

    ai_rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=[],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        strategy="ai_first",
        ai_rows=[{
            "name": "查询客户",
            "type": "EQ",
            "explanation": "按客户名称查询客户列表。",
            "source_processes": ["查询客户"],
        }],
    )
    combined, supplement_warnings = _supplement_ai_rows_with_rules(
        group=group,
        meta=_meta(),
        ai_rows=ai_rows,
        profile=STRICT_FPA_PROFILE,
        strategy="ai_first",
        rule_set_config=FpaRuleSetConfig(
            name="process_only",
            coverage_rules=FpaCoverageRules(require_data_function=False),
        ),
    )

    assert len(combined) == 1
    assert supplement_warnings == []
    assert any("AI 未返回合法 source_process_ids" in warning for warning in warnings)


def test_ai_first_unknown_source_process_ids_warns_and_does_not_count_unknown_id():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "客户管理",
            "二级模块": "客户查询",
            "三级模块": "客户查询",
            "三级模块整体功能描述": "查询和导出客户。",
            "功能过程": "查询客户",
            "功能过程类型": "新增",
            "功能过程描述": "按客户名称查询客户列表。",
        },
        {
            "客户端类型": "地市后台",
            "一级模块": "客户管理",
            "二级模块": "客户查询",
            "三级模块": "客户查询",
            "三级模块整体功能描述": "查询和导出客户。",
            "功能过程": "导出客户",
            "功能过程类型": "新增",
            "功能过程描述": "导出客户列表。",
        },
    ])[0]

    ai_rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=[],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        strategy="ai_first",
        ai_rows=[{
            "name": "查询客户",
            "type": "EQ",
            "explanation": "按客户名称查询客户列表。",
            "source_process_ids": [group["processes"][0]["process_id"], "m999_p9"],
            "source_processes": ["查询客户"],
        }],
    )

    audit = _build_fpa_audit_reports_for_groups(
        groups=[group],
        rows_by_module={1: ai_rows},
        warnings_by_module={1: warnings},
        profile=STRICT_FPA_PROFILE,
        profile_version=STRICT_FPA_PROFILE.version,
        strategy="ai_first",
        rule_set="strict_fpa_rs",
    )[0]

    assert any("AI 返回未知 source_process_ids: m999_p9" in warning for warning in warnings)
    assert ai_rows[0]["source_process_ids"] == [group["processes"][0]["process_id"]]
    assert audit.covered_processes == ["查询客户"]
    assert audit.missing_processes == ["导出客户"]


def test_ai_first_retries_once_when_validator_finds_high_confidence_issue(monkeypatch, tmp_path):
    source_rows = [{
        "客户端类型": "地市后台",
        "一级模块": "客户管理",
        "二级模块": "客户查询",
        "三级模块": "客户查询",
        "三级模块整体功能描述": "查询客户。",
        "功能过程": "查询客户",
        "功能过程类型": "新增",
        "功能过程描述": "按客户名称查询客户列表。",
    }]
    calls = []
    responses = [
        {
            "rows": [{
                "name": "查询客户",
                "type": "EI",
                "explanation": "来源场景：查询客户。\n业务数据：客户列表。\n业务规则：只读取并展示列表。\n计算说明：误判为 EI。",
                "source_process_ids": ["m1_p1"],
                "source_processes": ["查询客户"],
            }]
        },
        {
            "rows": [{
                "name": "查询客户",
                "type": "EQ",
                "explanation": "来源场景：查询客户。\n业务数据：客户列表。\n业务规则：只读取并展示列表。\n计算说明：查询条件输入并返回客户列表，按 EQ。",
                "source_process_ids": ["m1_p1"],
                "source_processes": ["查询客户"],
            }]
        },
    ]

    def fake_call_llm(prompt, *args, **kwargs):
        calls.append(prompt)
        return json.dumps(responses[len(calls) - 1], ensure_ascii=False)

    monkeypatch.setattr("ai_gen_reimbursement_docs.gen_fpa._call_llm", fake_call_llm)
    audit_trace = tmp_path / "trace.json"

    result = _plan_fpa_rows_with_execution(
        source_rows,
        _meta(),
        [],
        api_key="sk-test",
        model="mock-model",
        base_url="",
        profile=STRICT_FPA_PROFILE,
        strategy="ai_first",
        rule_set="strict_fpa_rs",
        rule_set_config=FpaRuleSetConfig(
            name="process_only",
            coverage_rules=FpaCoverageRules(require_data_function=False),
        ),
        audit_trace_path=str(audit_trace),
    )

    assert len(calls) == 2
    assert "上一次 FPA JSON 输出未通过项目口径校验" in calls[1]
    assert any(row["类型"] == "EQ" for row in result)
    trace = json.loads(audit_trace.read_text(encoding="utf-8"))
    assert trace["modules"][0]["retry_trigger_source"] == "validator"
    warnings = trace["modules"][0]["warnings"]
    assert any("AI 输出稳定性校验触发一次重试" in warning for warning in warnings)


def test_ai_first_retries_once_when_quality_review_finds_high_confidence_issue(monkeypatch, tmp_path):
    source_rows = [{
        "客户端类型": "地市后台",
        "一级模块": "客户管理",
        "二级模块": "客户查询",
        "三级模块": "客户查询",
        "三级模块整体功能描述": "查询客户。",
        "功能过程": "查询客户",
        "功能过程类型": "新增",
        "功能过程描述": "按客户名称查询客户列表。",
    }]
    calls = []
    responses = [
        {
            "rows": [{
                "name": "查询客户",
                "type": "EO",
                "explanation": "来源场景：查询客户。\n业务数据：客户列表。\n业务规则：只读取并展示列表。\n计算说明：误判为 EO。",
                "source_process_ids": ["m1_p1"],
                "source_processes": ["查询客户"],
            }]
        },
        {
            "rows": [{
                "name": "查询客户",
                "type": "EQ",
                "explanation": "来源场景：查询客户。\n业务数据：客户列表。\n业务规则：只读取并展示列表。\n计算说明：查询条件输入并返回客户列表，按 EQ。",
                "source_process_ids": ["m1_p1"],
                "source_processes": ["查询客户"],
            }]
        },
    ]

    def fake_call_llm(prompt, *args, **kwargs):
        calls.append(prompt)
        return json.dumps(responses[len(calls) - 1], ensure_ascii=False)

    monkeypatch.setattr("ai_gen_reimbursement_docs.gen_fpa._call_llm", fake_call_llm)
    audit_trace = tmp_path / "trace.json"

    result = _plan_fpa_rows_with_execution(
        source_rows,
        _meta(),
        [],
        api_key="sk-test",
        model="mock-model",
        base_url="",
        profile=STRICT_FPA_PROFILE,
        strategy="ai_first",
        rule_set="strict_fpa_rs",
        rule_set_config=FpaRuleSetConfig(
            name="process_only",
            coverage_rules=FpaCoverageRules(require_data_function=False),
        ),
        audit_trace_path=str(audit_trace),
    )

    assert len(calls) == 2
    assert "上一次 FPA JSON 输出未通过质量审核" in calls[1]
    assert "type_judgement" in calls[1]
    assert any(row["类型"] == "EQ" for row in result)
    trace = json.loads(audit_trace.read_text(encoding="utf-8"))
    assert trace["modules"][0]["retry_trigger_source"] == "quality_review"
    warnings = trace["modules"][0]["warnings"]
    assert any("AI 输出稳定性校验触发一次重试" in warning for warning in warnings)


def test_ai_first_keeps_first_parseable_rows_when_retry_parse_fails(monkeypatch, tmp_path):
    source_rows = [{
        "客户端类型": "地市后台",
        "一级模块": "客户管理",
        "二级模块": "客户查询",
        "三级模块": "客户查询",
        "三级模块整体功能描述": "查询客户。",
        "功能过程": "查询客户",
        "功能过程类型": "新增",
        "功能过程描述": "按客户名称查询客户列表。",
    }]
    calls = []
    responses = [
        json.dumps({
            "rows": [{
                "name": "查询客户",
                "type": "EO",
                "explanation": "来源场景：查询客户。\n业务数据：客户列表。\n业务规则：只读取并展示列表。\n计算说明：误判为 EO。",
                "source_process_ids": ["m1_p1"],
                "source_processes": ["查询客户"],
            }]
        }, ensure_ascii=False),
        '{"rows":[{"name":"查询客户","type":"EQ"',
    ]

    def fake_call_llm(prompt, *args, **kwargs):
        calls.append(prompt)
        return responses[len(calls) - 1]

    monkeypatch.setattr("ai_gen_reimbursement_docs.gen_fpa._call_llm", fake_call_llm)
    audit_trace = tmp_path / "trace.json"

    result = _plan_fpa_rows_with_execution(
        source_rows,
        _meta(),
        [],
        api_key="sk-test",
        model="mock-model",
        base_url="",
        profile=STRICT_FPA_PROFILE,
        strategy="ai_first",
        rule_set="strict_fpa_rs",
        rule_set_config=FpaRuleSetConfig(
            name="process_only",
            coverage_rules=FpaCoverageRules(require_data_function=False),
        ),
        audit_trace_path=str(audit_trace),
    )

    assert len(calls) == 2
    assert [row["生成方式"] for row in result] == ["ai"]
    assert result[0]["类型"] == "EO"
    trace = json.loads(audit_trace.read_text(encoding="utf-8"))
    assert trace["modules"][0]["source"] == "ai"
    warnings = trace["modules"][0]["warnings"]
    assert any("重试调用或解析失败，已保留首次可解析 AI 输出" in warning for warning in warnings)
    assert not any("AI 调用或解析失败" in warning for warning in warnings)


def test_strict_fpa_fallback_rows_fill_classification_basis_from_judgement_rules(tmp_path):
    source_rows = [
        {
            "客户端类型": "地市后台",
            "一级模块": "营销管理",
            "二级模块": "签到管理",
            "三级模块": "签到奖品配置",
            "三级模块整体功能描述": "配置签到奖品并查询奖品列表。",
            "功能过程": "添加签到奖品数据",
            "功能过程类型": "新增",
            "功能过程描述": "录入签到奖品名称和数量并保存。",
        },
        {
            "客户端类型": "地市后台",
            "一级模块": "营销管理",
            "二级模块": "签到管理",
            "三级模块": "签到奖品配置",
            "三级模块整体功能描述": "配置签到奖品并查询奖品列表。",
            "功能过程": "查询奖品数据",
            "功能过程类型": "新增",
            "功能过程描述": "按奖品名称查询奖品列表并展示结果。",
        },
    ]
    judgement_rules = [
        "后台数据库变更/内部逻辑数据组相关原则",
        "修改或增加界面的个数，或进入/改变系统边界数据的事务原则",
        "提供查询界面输入并展示返回结果",
        "输出票据、报表、统计、文件等",
        "外部应用维护数据组/外部接口文件相关原则",
    ]

    result = _plan_fpa_rows_with_execution(
        source_rows,
        _meta(),
        judgement_rules,
        api_key="",
        model="mock-model",
        base_url="",
        profile=STRICT_FPA_PROFILE,
        strategy="rules_only",
        rule_set="strict_fpa_rs",
        rule_set_config=FpaRuleSetConfig(name="strict_fpa_rs"),
        audit_trace_path=str(tmp_path / "trace.json"),
    )

    fallback_rows = [row for row in result if row["生成方式"] == "fallback"]
    assert fallback_rows
    assert all(row["计算依据归类"] in judgement_rules for row in fallback_rows)
    assert any(row["类型"] == "ILF" and row["计算依据归类"] == judgement_rules[0] for row in fallback_rows)
    assert any(row["类型"] == "EI" and row["计算依据归类"] == judgement_rules[1] for row in fallback_rows)
    assert any(row["类型"] == "EQ" and row["计算依据归类"] == judgement_rules[2] for row in fallback_rows)
    assert any(
        hit["rule_id"] == "strict_fpa.fallback_classification_basis"
        for row in fallback_rows
        for hit in row.get("_规则命中详情", [])
    )


def test_fpa_prompt_context_appends_json_only_reasoning_constraint(monkeypatch, tmp_path):
    _write_fpa_prompt_config(tmp_path, monkeypatch)
    group = _group_rows_by_l3(_rows())[0]

    context = _build_fpa_ai_prompt_context(
        group=group,
        judgement_rules=["提供查询界面输入并展示返回结果"],
        domain_context={},
        profile=STRICT_FPA_PROFILE,
    )

    assert "不要输出 reasoning" in context.system_prompt
    assert "debug_summary" in context.system_prompt
    assert "rows 必须完整" in context.system_prompt


def test_coverage_rules_can_disable_data_function_supplement():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "客户管理",
            "二级模块": "客户档案",
            "三级模块": "客户档案",
            "三级模块整体功能描述": "维护客户档案。",
            "功能过程": "添加客户档案",
            "功能过程类型": "新增",
            "功能过程描述": "保存客户档案。",
        },
    ])[0]
    ai_rows, _ = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=[],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        strategy="ai_first",
        ai_rows=[{
            "name": "添加客户档案",
            "type": "EI",
            "explanation": "保存客户档案。",
            "source_processes": ["添加客户档案"],
        }],
    )

    combined, warnings = _supplement_ai_rows_with_rules(
        group=group,
        meta=_meta(),
        ai_rows=ai_rows,
        profile=STRICT_FPA_PROFILE,
        strategy="ai_first",
        rule_set_config=FpaRuleSetConfig(
            name="no_data_supplement",
            coverage_rules=FpaCoverageRules(require_data_function=False),
        ),
    )

    assert combined == ai_rows
    assert warnings == []


def test_coverage_rules_can_disable_missing_process_supplement():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "客户管理",
            "二级模块": "客户查询",
            "三级模块": "客户查询",
            "三级模块整体功能描述": "查询和导出客户。",
            "功能过程": "查询客户",
            "功能过程类型": "新增",
            "功能过程描述": "按客户名称查询客户列表。",
        },
        {
            "客户端类型": "地市后台",
            "一级模块": "客户管理",
            "二级模块": "客户查询",
            "三级模块": "客户查询",
            "三级模块整体功能描述": "查询和导出客户。",
            "功能过程": "导出客户",
            "功能过程类型": "新增",
            "功能过程描述": "导出客户列表。",
        },
    ])[0]
    token = set_current_fpa_rule_set_config(_custom_default_rule_set())
    try:
        ai_rows, _ = _normalize_ai_fpa_rows_for_l3(
            group=group,
            meta=_meta(),
            judgement_rules=[],
            start_seq=1,
            profile=CUSTOM_RULES_PROFILE,
            strategy="ai_first",
            ai_rows=[{
                "name": "查询客户-查询处理开发",
                "type": "EQ",
                "explanation": "按客户名称查询客户列表。",
                "source_processes": ["查询客户"],
            }],
        )
    finally:
        reset_current_fpa_rule_set_config(token)

    combined, warnings = _supplement_ai_rows_with_rules(
        group=group,
        meta=_meta(),
        ai_rows=ai_rows,
        profile=CUSTOM_RULES_PROFILE,
        strategy="ai_first",
        rule_set_config=FpaRuleSetConfig(
            name="no_process_supplement",
            coverage_rules=FpaCoverageRules(require_process_coverage=False, require_data_function=False),
        ),
    )

    assert combined == ai_rows
    assert warnings == []


def test_strict_profile_keeps_real_external_data_group_eif():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "权限管理",
            "二级模块": "账号权限",
            "三级模块": "用户中心账号引用",
            "三级模块整体功能描述": "系统引用统一用户中心维护的人员账号。",
            "功能过程": "同步账号",
            "功能过程类型": "新增",
            "功能过程描述": "统一用户中心维护的人员账号，本系统只引用。",
        },
    ])[0]
    token = set_current_fpa_rule_set_config(_strict_default_rule_set())
    try:
        rows, warnings = _normalize_ai_fpa_rows_for_l3(
            group=group,
            meta=_meta(),
            judgement_rules=[],
            start_seq=1,
            profile=STRICT_FPA_PROFILE,
            ai_rows=[
                {
                    "name": "统一用户中心账号",
                    "type": "EIF",
                    "explanation": "统一用户中心维护的人员账号，本系统只引用。",
                },
            ],
        )
    finally:
        reset_current_fpa_rule_set_config(token)

    assert rows[0]["类型"] == "EIF"
    assert not any("明显冲突" in w for w in warnings)


def test_strict_profile_warns_when_ai_complex_eif_needs_manual_review():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "风控管理",
            "二级模块": "风险画像",
            "三级模块": "风险画像",
            "三级模块整体功能描述": "结合多源业务信息生成风险画像。",
            "功能过程": "查看画像",
            "功能过程类型": "查询",
            "功能过程描述": "查看企业风险画像详情。",
        },
    ])[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=[],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        strategy="ai_first",
        ai_rows=[
            {
                "name": "跨域企业风险画像",
                "type": "EIF",
                "explanation": "AI 判断该风险画像由外部风控平台维护，本系统只引用。",
            },
        ],
    )

    assert rows[0]["类型"] == "EIF"
    assert any("AI 数据功能需人工复核" in warning for warning in warnings)
    review_hits = [
        hit
        for hit in rows[0]["_规则命中详情"]
        if hit["rule_id"] == "postprocess.ai_data_group_review"
    ]
    assert review_hits
    assert review_hits[0]["suggested_type"] == "EIF"


def test_strict_profile_warns_when_ai_complex_ilf_needs_manual_review():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "营销管理",
            "二级模块": "客群洞察",
            "三级模块": "客群洞察",
            "三级模块整体功能描述": "生成客户洞察结果。",
            "功能过程": "刷新洞察",
            "功能过程类型": "新增",
            "功能过程描述": "刷新客户洞察。",
        },
    ])[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=[],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        strategy="ai_first",
        ai_rows=[
            {
                "name": "客户生命周期洞察",
                "type": "ILF",
                "explanation": "AI 判断本系统保存并持续维护客户生命周期洞察结果。",
            },
        ],
    )

    assert rows[0]["类型"] == "ILF"
    assert any("AI 数据功能需人工复核" in warning for warning in warnings)
    assert any(
        hit["rule_id"] == "postprocess.ai_data_group_review"
        for hit in rows[0]["_规则命中详情"]
    )


def test_strict_profile_data_group_review_survives_unrelated_ai_warning():
    group = _group_rows_by_l3([
        {
            "客户端类型": "地市后台",
            "一级模块": "风控管理",
            "二级模块": "风险画像",
            "三级模块": "风险画像",
            "三级模块整体功能描述": "结合多源业务信息生成风险画像。",
            "功能过程": "查看画像",
            "功能过程类型": "查询",
            "功能过程描述": "查看企业风险画像详情。",
        },
    ])[0]
    rows, warnings = _normalize_ai_fpa_rows_for_l3(
        group=group,
        meta=_meta(),
        judgement_rules=["有效规则"],
        start_seq=1,
        profile=STRICT_FPA_PROFILE,
        strategy="ai_first",
        ai_rows=[
            {
                "name": "跨域企业风险画像",
                "type": "EIF",
                "explanation": "AI 判断该风险画像由外部风控平台维护，本系统只引用。",
                "classification_basis_index": 99,
            },
        ],
    )

    assert rows[0]["类型"] == "EIF"
    assert any("classification_basis_index 越界" in warning for warning in warnings)
    assert any("AI 数据功能需人工复核" in warning for warning in warnings)


def test_keyword_type_fallbacks():
    token = set_current_fpa_rule_set_config(_custom_default_rule_set())
    try:
        assert CUSTOM_RULES_PROFILE.infer_type("客户界面开发")[0] == "EI"
        assert CUSTOM_RULES_PROFILE.infer_type("添加客户-逻辑处理开发")[0] == "ILF"
        assert CUSTOM_RULES_PROFILE.infer_type("查询客户-查询处理开发")[0] == "EQ"
        assert CUSTOM_RULES_PROFILE.infer_type("导出客户-导出处理开发")[0] == "EO"
        assert CUSTOM_RULES_PROFILE.infer_type("导入客户-导入处理开发")[0] == "EI"
        assert CUSTOM_RULES_PROFILE.infer_type("同步外部接口数据-逻辑处理开发")[0] == "ILF"
        assert CUSTOM_RULES_PROFILE.infer_type("引用统一用户中心账号-外部接口处理开发")[0] == "EIF"
    finally:
        reset_current_fpa_rule_set_config(token)


class LowConfidenceRulesProfile(CustomRulesProfile):
    def fallback_rows_for_l3(self, group, meta, start_seq=1):
        return [{
            "序号": start_seq,
            "子系统(模块)": meta.get("子系统（模块）", ""),
            "资产标识": meta.get("资产标识", ""),
            "新增/修改功能点": "低置信度规则行",
            "类型": "",
            "计算依据归类": "",
            "计算依据说明": "低置信度规则行。",
            "变更状态": "新增",
            "调整值": 1,
            "要素数量": 1,
            "生成方式": "fallback",
            "类型理由": "",
            "源功能过程": "",
            "后处理警告": "",
        }]


def test_rules_first_keeps_rules_when_rule_rows_are_usable(monkeypatch, tmp_path):
    _write_fpa_prompt_config(tmp_path, monkeypatch)
    monkeypatch.setattr(
        "ai_gen_reimbursement_docs.gen_fpa._call_llm",
        lambda *args, **kwargs: pytest.fail("rules_first should not call AI when rules are usable"),
    )

    result = _plan_fpa_rows_with_ai(
        _rows(),
        _meta(),
        ["规则一"],
        api_key="sk-test",
        model="test",
        base_url="",
        strategy="rules_first",
    )

    assert result
    assert {row["生成方式"] for row in result} == {"fallback"}
    assert _rules_first_ai_reasons(_group_rows_by_l3(_rows())[0], result) == []


def test_rules_first_calls_ai_when_rule_rows_are_low_confidence(monkeypatch, tmp_path):
    _write_fpa_prompt_config(tmp_path, monkeypatch)
    response = {
        "rows": [{
            "name": "AI 复核功能点",
            "type": "EI",
            "classification_basis_index": 1,
            "explanation": "AI 复核功能点，具体为以下：1、覆盖低置信度规则无法覆盖的功能过程。",
            "source_processes": ["添加垂直行业", "查询垂直行业"],
        }]
    }
    calls = {"count": 0}

    def fake_call_llm(*args, **kwargs):
        calls["count"] += 1
        return json.dumps(response, ensure_ascii=False)

    monkeypatch.setattr("ai_gen_reimbursement_docs.gen_fpa._call_llm", fake_call_llm)

    result = _plan_fpa_rows_with_execution(
        _rows(),
        _meta(),
        ["规则一"],
        api_key="sk-test",
        model="test",
        base_url="",
        profile=LowConfidenceRulesProfile(),
        strategy="rules_first",
        rule_set="unified_ui_rs",
    )

    assert calls["count"] == 1
    assert result[0]["生成方式"] == "ai"
    assert result[0]["新增/修改功能点"] == "【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-AI 复核功能点"


def test_rules_first_without_api_keeps_low_confidence_rules_and_audit_warning(monkeypatch, tmp_path):
    _write_fpa_prompt_config(tmp_path, monkeypatch)
    audit_trace = tmp_path / "audit.json"

    result = _plan_fpa_rows_with_execution(
        _rows(),
        _meta(),
        ["规则一"],
        api_key="",
        model="test",
        base_url="",
        profile=LowConfidenceRulesProfile(),
        strategy="rules_first",
        rule_set="unified_ui_rs",
        audit_trace_path=str(audit_trace),
    )

    assert result[0]["生成方式"] == "fallback"
    trace = json.loads(audit_trace.read_text(encoding="utf-8"))
    warnings = trace["modules"][0]["warnings"]
    assert any("规则结果需要 AI 复核但未配置 API Key" in warning for warning in warnings)
    assert any("类型无效" in warning for warning in warnings)
    assert trace["stability_report"]["summary"]["warning_count"] == len(warnings)
    assert trace["stability_report"]["summary"]["source_counts"] == {"rules": 1}


def test_ai_parse_failure_falls_back(monkeypatch, tmp_path, caplog):
    _write_fpa_prompt_config(tmp_path, monkeypatch)
    monkeypatch.setattr(
        "ai_gen_reimbursement_docs.gen_fpa._call_llm",
        lambda *args, **kwargs: "这里不是 JSON",
    )
    caplog.set_level(logging.WARNING)
    result = _plan_fpa_rows_with_ai(
        _rows(),
        _meta(),
        ["规则一"],
        api_key="sk-test",
        model="test",
        base_url="",
        strategy="ai_first",
    )
    assert result[0]["生成方式"] == "fallback"
    assert any("FPA AI 响应解析失败" in r.message for r in caplog.records)


def test_missing_fpa_config_does_not_fall_back(monkeypatch, tmp_path):
    monkeypatch.setattr("ai_gen_reimbursement_docs.config_utils.config_dir", lambda: tmp_path)
    monkeypatch.setattr(
        "ai_gen_reimbursement_docs.gen_fpa._call_llm",
        lambda *args, **kwargs: pytest.fail("missing prompt config must stop before LLM call"),
    )

    with pytest.raises(FpaConfigError, match="未找到 FPA 配置文件"):
        _plan_fpa_rows_with_ai(
            _rows(),
            _meta(),
            ["规则一"],
            api_key="sk-test",
            model="test",
            base_url="",
            strategy="ai_first",
        )


def test_ai_first_requires_api_key():
    with pytest.raises(ValueError, match="需要 API Key"):
        _plan_fpa_rows_with_ai(
            _rows(),
            _meta(),
            ["规则一"],
            api_key="",
            model="test",
            base_url="",
            profile=STRICT_FPA_PROFILE,
            strategy="ai_first",
        )


def test_fpa_preview_returns_ai_debug(monkeypatch, tmp_path):
    _write_fpa_prompt_config(tmp_path, monkeypatch)
    xlsx = tmp_path / "功能清单.xlsx"
    xlsx.write_bytes(b"placeholder")
    response = {
        "rows": [
            {
                "name": "垂直行业数据维护",
                "type": "ILF",
                "type_reason": "内部逻辑文件",
                "classification_basis_index": 1,
                "explanation": "垂直行业数据维护，具体如下：触发事件：管理员维护数据；事件流：系统保存数据。",
            }
        ]
    }

    monkeypatch.setattr(
        "ai_gen_reimbursement_docs.gen_fpa.read_base_data_from_excel",
        lambda path: {"tree_rows": _rows(), "meta": _meta()},
    )
    def fake_call_llm(*args, **kwargs):
        assert kwargs.get("return_thinking") is True
        return json.dumps(response, ensure_ascii=False), "思考过程"

    monkeypatch.setattr("ai_gen_reimbursement_docs.gen_fpa._call_llm", fake_call_llm)

    result = preview_fpa_module(
        file_path=str(xlsx),
        module_name="垂直行业管理",
        api_key="sk-test",
        model="test-model",
        base_url="",
        strategy="ai_only",
    )

    debug = result["debug"]
    assert debug["ai_called"] is True
    assert debug["model"] == "test-model"
    assert debug["system_prompt"].startswith("系统提示词")
    assert "不要输出 reasoning" in debug["system_prompt"]
    assert "debug_summary" in debug["system_prompt"]
    assert debug["core_rules_source"] == "用户配置（配置目录/fpa_config.yaml: core_rules.unified_ui_cr）"
    assert debug["system_prompt_source"] == "用户配置（配置目录/fpa_config.yaml: system_prompt_sets.unified_ui_sp）"
    assert debug["user_prompt_source"] == "用户配置（配置目录/fpa_config.yaml: user_prompt_sets.unified_ui_up）"
    assert "垂直行业管理" in debug["user_prompt"]
    assert "1) 规则一" in debug["user_prompt"]
    assert "2) 规则二" in debug["user_prompt"]
    assert "[system]" in debug["ai_prompt"]
    assert "垂直行业数据维护" in debug["raw_response"]
    assert debug["thinking"] == "思考过程"
    assert debug["parsed_rows"] == response["rows"]
    assert result["rows"][0]["classification_basis"] == "规则一"
    assert debug["final_rows"][0]["name"] == "【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-垂直行业数据维护"
    assert debug["agent_review"]["summary"]["pending_agent_roles"] == []
    assert debug["agent_review"]["roles"][1]["status"] == "completed"
    assert debug["quality_review"]["summary"]["issue_count"] >= 0
    assert result["audit"]["raw_ai"]["source"] == "ai"
    assert result["audit"]["agent_review"]["summary"]["pending_agent_roles"] == []
    assert result["audit"]["raw_ai"]["raw_rows"] == response["rows"]


def test_ai_only_preview_empty_rows_does_not_fallback(monkeypatch, tmp_path):
    _write_fpa_prompt_config(tmp_path, monkeypatch)
    xlsx = tmp_path / "功能清单.xlsx"
    xlsx.write_bytes(b"placeholder")
    monkeypatch.setattr(
        "ai_gen_reimbursement_docs.gen_fpa.read_base_data_from_excel",
        lambda path: {"tree_rows": _rows(), "meta": _meta()},
    )
    monkeypatch.setattr(
        "ai_gen_reimbursement_docs.gen_fpa._call_llm",
        lambda *args, **kwargs: (json.dumps({"rows": []}), "") if kwargs.get("return_thinking") else json.dumps({"rows": []}),
    )

    def fail_if_fallback(*args, **kwargs):
        pytest.fail("ai_only 失败时不应使用 rules_fallback")

    monkeypatch.setattr(CustomRulesProfile, "fallback_rows_for_l3", fail_if_fallback)

    with pytest.raises(ValueError, match="AI 规划未生成有效 FPA 行"):
        preview_fpa_module(
            file_path=str(xlsx),
            module_name="垂直行业管理",
            api_key="sk-test",
            model="test-model",
            base_url="",
            strategy="ai_only",
        )


def test_fpa_preview_returns_confirmation_questions_for_high_risk_validator_issue(monkeypatch, tmp_path):
    _write_fpa_prompt_config(tmp_path, monkeypatch)
    xlsx = tmp_path / "功能清单.xlsx"
    xlsx.write_bytes(b"placeholder")
    monkeypatch.setattr(
        "ai_gen_reimbursement_docs.gen_fpa.read_base_data_from_excel",
        lambda path: {"tree_rows": _rows(), "meta": _meta()},
    )
    monkeypatch.setattr(
        "ai_gen_reimbursement_docs.gen_fpa._call_llm",
        lambda *args, **kwargs: (
            json.dumps({
                "rows": [{
                    "name": "查询垂直行业",
                    "type": "EI",
                    "explanation": "来源场景：查询垂直行业。\n业务数据：垂直行业列表。\n业务规则：只读取并展示列表。\n计算说明：误判为 EI。",
                    "source_process_ids": ["m1_p2"],
                    "source_processes": ["查询垂直行业"],
                }]
            }, ensure_ascii=False),
            "",
        ),
    )

    result = preview_fpa_module(
        file_path=str(xlsx),
        module_name="垂直行业管理",
        api_key="sk-test",
        model="test-model",
        base_url="",
        profile_name="strict_fpa",
        strategy="ai_only",
        fpa_confirmation_mode="cautious",
    )

    assert result["status"] == "needs_confirmation"
    assert result["confirmation_mode"] == "cautious"
    assert result["confirmation_questions"][0]["topic"] == "类型判定"
    assert result["confirmation_questions"][0]["recommendation"] == "eq"


def test_fpa_preview_confirmed_decisions_enter_prompt_and_suppress_question(monkeypatch, tmp_path):
    _write_fpa_prompt_config(tmp_path, monkeypatch)
    xlsx = tmp_path / "功能清单.xlsx"
    xlsx.write_bytes(b"placeholder")
    prompts = []
    monkeypatch.setattr(
        "ai_gen_reimbursement_docs.gen_fpa.read_base_data_from_excel",
        lambda path: {"tree_rows": _rows(), "meta": _meta()},
    )

    def fake_call_llm(prompt, *args, **kwargs):
        prompts.append(prompt)
        return (
            json.dumps({
                "rows": [{
                    "name": "查询垂直行业",
                    "type": "EI",
                    "explanation": "来源场景：查询垂直行业。\n业务数据：垂直行业列表。\n业务规则：只读取并展示列表。\n计算说明：误判为 EI。",
                    "source_process_ids": ["m1_p2"],
                    "source_processes": ["查询垂直行业"],
                }]
            }, ensure_ascii=False),
            "",
        )

    monkeypatch.setattr("ai_gen_reimbursement_docs.gen_fpa._call_llm", fake_call_llm)

    first = preview_fpa_module(
        file_path=str(xlsx),
        module_name="垂直行业管理",
        api_key="sk-test",
        model="test-model",
        base_url="",
        profile_name="strict_fpa",
        strategy="ai_only",
        fpa_confirmation_mode="cautious",
    )
    decision_id = first["confirmation_questions"][0]["id"]
    second = preview_fpa_module(
        file_path=str(xlsx),
        module_name="垂直行业管理",
        api_key="sk-test",
        model="test-model",
        base_url="",
        profile_name="strict_fpa",
        strategy="ai_only",
        fpa_confirmation_mode="cautious",
        confirmed_decisions={decision_id: {"value": "eq", "scope": "current_run"}},
    )

    assert "必须作为硬约束执行" in prompts[1]
    assert second["confirmed_decision_count"] == 1
    assert second["status"] == "ok"
    assert second["confirmation_questions"] == []


def test_fpa_batch_generation_pauses_for_confirmation_and_replans(monkeypatch, tmp_path):
    _write_fpa_prompt_config(tmp_path, monkeypatch)
    prompts: list[str] = []
    confirmation_payloads: list[dict] = []

    def fake_call_llm(prompt, *args, **kwargs):
        prompts.append(prompt)
        fpa_type = "EQ" if "必须作为硬约束执行" in prompt else "EI"
        return json.dumps({
            "rows": [{
                "name": "查询垂直行业",
                "type": fpa_type,
                "explanation": "来源场景：查询垂直行业。\n业务数据：垂直行业列表。\n业务规则：只读取并展示列表。\n计算说明：按查询类 EQ 计量。",
                "source_process_ids": ["m1_p2"],
                "source_processes": ["查询垂直行业"],
            }]
        }, ensure_ascii=False)

    def fake_wait_for_fpa_confirmation(payload):
        confirmation_payloads.append(payload)
        question_id = payload["confirmation_questions"][0]["id"]
        return {question_id: {"value": "eq", "scope": "current_run"}}

    monkeypatch.setattr("ai_gen_reimbursement_docs.gen_fpa._call_llm", fake_call_llm)
    token = callbacks_var.set(PipelineCallbacks(
        wait_for_fpa_confirmation=fake_wait_for_fpa_confirmation,
    ))
    try:
        rows = _plan_fpa_rows_with_execution(
            _rows(),
            _meta(),
            ["规则一"],
            api_key="sk-test",
            model="test",
            base_url="",
            profile=STRICT_FPA_PROFILE,
            strategy="ai_only",
            rule_set="strict_fpa_rs",
            fpa_confirmation_mode="cautious",
        )
    finally:
        callbacks_var.reset(token)

    assert confirmation_payloads
    assert confirmation_payloads[0]["module"]["l3"] == "垂直行业管理"
    assert confirmation_payloads[0]["confirmation_questions"][0]["topic"] == "类型判定"
    assert len(prompts) == 2
    assert "必须作为硬约束执行" in prompts[1]
    assert rows[0]["类型"] == "EQ"


def test_fpa_preview_prompt_includes_project_domain_context(monkeypatch, tmp_path):
    _write_fpa_prompt_config(tmp_path, monkeypatch)
    (tmp_path / "domain_context.json").write_text(
        """
{
  "system_boundary": "本系统维护供应商协同关系，不维护供应商主档。",
  "internal_data_groups": [{"name": "供应商协同关系"}],
  "external_data_groups": [{"name": "供应商档案", "source": "供应商平台"}],
  "external_services": [{"name": "短信平台"}]
}
""",
        encoding="utf-8",
    )
    xlsx = tmp_path / "功能清单.xlsx"
    xlsx.write_bytes(b"placeholder")
    monkeypatch.setattr(
        "ai_gen_reimbursement_docs.gen_fpa.read_base_data_from_excel",
        lambda path: {
            "tree_rows": _rows(),
            "meta": {
                **_meta(),
                "工单标题": "供应商协同优化",
                "工单内容": "围绕供应商协同关系维护和查询能力建设。",
            },
        },
    )
    monkeypatch.setattr(
        "ai_gen_reimbursement_docs.gen_fpa._call_llm",
        lambda *args, **kwargs: (
            json.dumps(
                {
                    "rows": [{
                        "name": "领域上下文验证功能点",
                        "type": "EI",
                        "classification_basis_index": 1,
                        "explanation": "领域上下文验证功能点，具体为以下：1、覆盖上下文传入。",
                    }]
                },
                ensure_ascii=False,
            ),
            "",
        ),
    )

    result = preview_fpa_module(
        file_path=str(xlsx),
        module_name="垂直行业管理",
        api_key="sk-test",
        model="test-model",
        base_url="",
        strategy="ai_only",
    )

    prompt = result["debug"]["user_prompt"]
    assert '"子系统（模块）": "测试系统"' in prompt
    assert '"project_description": "工单标题：供应商协同优化\\n工单内容：围绕供应商协同关系维护和查询能力建设。"' in prompt
    assert '"system_boundary": "本系统维护供应商协同关系，不维护供应商主档。"' in prompt
    assert '"name": "供应商协同关系"' in prompt
    assert '"name": "供应商档案"' in prompt
    assert '"source": "供应商平台"' in prompt
    assert '"name": "短信平台"' in prompt


def test_rules_first_preview_calls_ai_when_rule_rows_are_low_confidence(monkeypatch, tmp_path):
    _write_fpa_prompt_config(tmp_path, monkeypatch)
    xlsx = tmp_path / "功能清单.xlsx"
    xlsx.write_bytes(b"placeholder")
    monkeypatch.setattr(
        "ai_gen_reimbursement_docs.gen_fpa.read_base_data_from_excel",
        lambda path: {"tree_rows": _rows(), "meta": _meta()},
    )

    def fake_fallback(self, group, meta, start_seq=1):
        return [{
            "序号": start_seq,
            "子系统(模块)": meta.get("子系统（模块）", ""),
            "资产标识": meta.get("资产标识", ""),
            "新增/修改功能点": "低置信度规则行",
            "类型": "",
            "计算依据归类": "",
            "计算依据说明": "低置信度规则行。",
            "变更状态": "新增",
            "调整值": 1,
            "要素数量": 1,
            "生成方式": "fallback",
            "类型理由": "",
            "源功能过程": "",
            "后处理警告": "",
        }]

    monkeypatch.setattr(
        "ai_gen_reimbursement_docs.fpa_profiles.CustomRulesProfile.fallback_rows_for_l3",
        fake_fallback,
    )
    response = {
        "rows": [{
            "name": "AI 预览复核功能点",
            "type": "EI",
            "classification_basis_index": 1,
            "explanation": "AI 预览复核功能点，具体为以下：1、覆盖低置信度规则输出。",
            "source_processes": ["添加垂直行业", "查询垂直行业"],
        }]
    }

    def fake_call_llm(*args, **kwargs):
        assert kwargs.get("return_thinking") is True
        return json.dumps(response, ensure_ascii=False), "思考过程"

    monkeypatch.setattr("ai_gen_reimbursement_docs.gen_fpa._call_llm", fake_call_llm)

    result = preview_fpa_module(
        file_path=str(xlsx),
        module_name="垂直行业管理",
        api_key="sk-test",
        model="test-model",
        strategy="rules_first",
    )

    assert result["used_ai"] is True
    assert result["debug"]["reason"] == "rules_first_needs_ai"
    assert result["rows"][0]["name"] == "【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-AI 预览复核功能点"
    assert any("规则结果触发 AI 复核" in warning for warning in result["warnings"])


def test_ai_cache_hit_skips_llm(monkeypatch, tmp_path, caplog):
    _write_fpa_prompt_config(tmp_path, monkeypatch)
    cache_path = tmp_path / "fpa_ai_cache.json"
    audit_trace_path = tmp_path / "fpa_audit_trace.json"
    response = {
        "rows": [
            {
                "name": "垂直行业管理界面开发",
                "type": "EI",
                "classification_basis_index": 1,
                "explanation": "垂直行业管理界面开发，具体为以下：1、新增列表和查询条件。",
            }
        ]
    }
    calls = {"count": 0}

    def first_call(*args, **kwargs):
        calls["count"] += 1
        return json.dumps(response, ensure_ascii=False)

    monkeypatch.setattr("ai_gen_reimbursement_docs.gen_fpa._call_llm", first_call)
    first = _plan_fpa_rows_with_ai(
        _rows(),
        _meta(),
        ["规则一"],
        api_key="sk-test",
        model="test-model",
        base_url="",
        cache_path=str(cache_path),
        strategy="ai_first",
        audit_trace_path=str(audit_trace_path),
    )
    assert calls["count"] == 1
    assert cache_path.exists()
    assert first[0]["生成方式"] == "ai"
    cache = json.loads(cache_path.read_text(encoding="utf-8"))
    entry = next(iter(cache["entries"].values()))
    assert entry["profile"] == "unified_ui"
    assert entry["profile_version"] == "1"
    assert entry["strategy"] == "ai_first"
    assert entry["rule_set"] == "unified_ui_rs"
    trace = json.loads(audit_trace_path.read_text(encoding="utf-8"))
    assert trace["modules"][0]["source"] == "ai"
    assert trace["modules"][0]["raw_rows"] == response["rows"]

    def fail_if_called(*args, **kwargs):
        raise AssertionError("LLM should not be called on cache hit")

    monkeypatch.setattr("ai_gen_reimbursement_docs.gen_fpa._call_llm", fail_if_called)
    caplog.set_level(logging.INFO)
    second = _plan_fpa_rows_with_ai(
        _rows(),
        _meta(),
        ["规则一"],
        api_key="sk-test",
        model="test-model",
        base_url="",
        cache_path=str(cache_path),
        strategy="ai_first",
        audit_trace_path=str(audit_trace_path),
    )

    assert second[0]["新增/修改功能点"] == "【地市后台】垂直行业营销-垂直行业管理-垂直行业管理-垂直行业管理界面开发"
    trace = json.loads(audit_trace_path.read_text(encoding="utf-8"))
    assert trace["modules"][0]["source"] == "ai_cache"
    assert any("缓存命中" in r.message for r in caplog.records)


def test_ai_cache_is_invalidated_when_project_domain_context_changes(monkeypatch, tmp_path):
    _write_fpa_prompt_config(tmp_path, monkeypatch)
    domain_context_path = tmp_path / "domain_context.json"
    domain_context_path.write_text(
        """
{
  "system_boundary": "本系统维护供应商关系。",
  "internal_data_groups": [{"name": "供应商关系"}],
  "external_data_groups": [],
  "external_services": []
}
""",
        encoding="utf-8",
    )
    cache_path = tmp_path / "fpa_ai_cache.json"
    response = {
        "rows": [{
            "name": "供应商关系维护",
            "type": "ILF",
            "classification_basis_index": 1,
            "explanation": "供应商关系维护，具体为以下：1、保存供应商关系。",
        }]
    }
    calls = {"count": 0}

    def fake_call(*args, **kwargs):
        calls["count"] += 1
        return json.dumps(response, ensure_ascii=False)

    monkeypatch.setattr("ai_gen_reimbursement_docs.gen_fpa._call_llm", fake_call)
    for _ in range(2):
        _plan_fpa_rows_with_ai(
            _rows(),
            _meta(),
            ["规则一"],
            api_key="sk-test",
            model="test-model",
            base_url="",
            cache_path=str(cache_path),
            strategy="ai_first",
        )
    assert calls["count"] == 1

    domain_context_path.write_text(
        domain_context_path.read_text(encoding="utf-8").replace("供应商关系。", "供应商协同关系。"),
        encoding="utf-8",
    )
    _plan_fpa_rows_with_ai(
        _rows(),
        _meta(),
        ["规则一"],
        api_key="sk-test",
        model="test-model",
        base_url="",
        cache_path=str(cache_path),
        strategy="ai_first",
    )

    assert calls["count"] == 2


def test_ai_cache_key_includes_configured_core_rules():
    base_kwargs = {
        "group": _group_rows_by_l3(_rows())[0],
        "judgement_rules": ["规则一"],
        "domain_context": {},
        "model": "test-model",
        "profile": CUSTOM_RULES_PROFILE,
        "strategy": "ai_first",
        "rule_set": "unified_ui_rs",
        "rule_set_config": {},
        "system_prompt": "system",
        "user_prompt": "user",
    }

    first = _fpa_ai_cache_key(core_rules="custom core v1", **base_kwargs)
    second = _fpa_ai_cache_key(core_rules="custom core v2", **base_kwargs)

    assert first != second


def test_fpa_check_xlsx_columns_can_be_configured(monkeypatch, tmp_path):
    (tmp_path / "system_config.yaml").write_text(
        """
fpa_check_columns:
  FPA结果: ["新增/修改功能点", "类型", "后处理警告"]
  Warnings: ["对象", "Warning", "来源规则ID"]
  规则命中详情: ["功能点名称", "规则ID", "是否采用"]
""",
        encoding="utf-8",
    )
    monkeypatch.setattr("ai_gen_reimbursement_docs.config_utils.config_dir", lambda: tmp_path)

    fpa_md = tmp_path / "fpa.md"
    fpa_md.write_text(
        """# FPA 工作量评估

**profile**: unified_ui
**strategy**: rules_first
**rule_set**: unified_ui_rs

| 序号 | 子系统(模块) | 资产标识 | 新增/修改功能点 | 类型 | 计算依据归类 | 计算依据说明 | 变更状态 | 调整值 | 要素数量 | 生成方式 | 类型理由 | 源功能过程 | 后处理警告 |
|------|-------------|---------|----------------|------|-------------|-------------|---------|-------|---------|---------|---------|-----------|-----------|
| 1 | 测试系统 | TEST | 查询客户-查询处理开发 | EQ |  | 查询客户。 | 新增 | 2 | 1 | ai | AI 根据功能点名称和业务说明判定。 | 查询客户 | 查询客户 classification_basis_index 越界: 99 |
""",
        encoding="utf-8",
    )
    tree_md = tmp_path / "tree.md"
    tree_md.write_text(
        """| 入口 | 一级模块 | 二级模块 | 三级模块 | 客户端类型 | 三级模块整体功能描述 | 功能过程 | 功能过程类型 | 功能过程描述 |
|------|---------|---------|---------|----------|----------------------|----------|--------------|--------------|
| 后台 | 客户管理 | 客户查询 | 客户查询 | 地市后台 | 查询客户。 | 查询客户 | 查询 | 按条件查询客户。 |
""",
        encoding="utf-8",
    )
    audit_trace = tmp_path / "trace.json"
    audit_trace.write_text(
        json.dumps({
            "modules": [{
                "rule_hits": [{
                    "fpa_seq": 1,
                    "name": "查询客户-查询处理开发",
                    "generation": "ai",
                    "hit_object": "查询客户-查询处理开发",
                    "rule_id": "postprocess.classification_basis_index",
                    "rule_desc": "classification_basis_index 必须落在模板判定原则范围内。",
                    "suggested_type": "EQ",
                    "adopted": "是",
                    "warnings": ["查询客户 classification_basis_index 越界: 99"],
                }],
            }],
        }, ensure_ascii=False),
        encoding="utf-8",
    )

    output = tmp_path / "check.xlsx"
    generate_fpa_check_xlsx_from_md(str(fpa_md), str(tree_md), str(output), str(audit_trace))

    wb = openpyxl.load_workbook(output, data_only=True)
    assert [cell.value for cell in wb["FPA结果"][1]] == ["新增/修改功能点", "类型", "后处理警告"]
    assert [cell.value for cell in wb["Warnings"][1]] == ["对象", "Warning", "来源规则ID"]
    assert [cell.value for cell in wb["规则命中详情"][1]] == ["功能点名称", "规则ID", "是否采用"]
    assert wb["Warnings"].cell(2, 3).value == "postprocess.classification_basis_index"
    wb.close()


def test_fpa_check_xlsx_includes_rule_set_config_warning(tmp_path):
    fpa_md = tmp_path / "fpa.md"
    fpa_md.write_text(
        """# FPA 工作量评估

**profile**: unified_ui
**strategy**: rules_first
**rule_set**: sms_service_rules

| 序号 | 子系统(模块) | 资产标识 | 新增/修改功能点 | 类型 | 计算依据归类 | 计算依据说明 | 变更状态 | 调整值 | 要素数量 | 生成方式 | 类型理由 | 源功能过程 | 后处理警告 |
|------|-------------|---------|----------------|------|-------------|-------------|---------|-------|---------|---------|---------|-----------|-----------|
| 1 | 测试系统 | TEST | 发送短信-逻辑处理开发 | EI |  | 调用短信平台发送通知。 | 新增 | 2 | 1 | fallback | 普通外部服务调用按事务处理。 | 发送短信 |  |
""",
        encoding="utf-8",
    )
    tree_md = tmp_path / "tree.md"
    tree_md.write_text(
        """| 入口 | 一级模块 | 二级模块 | 三级模块 | 客户端类型 | 三级模块整体功能描述 | 功能过程 | 功能过程类型 | 功能过程描述 |
|------|---------|---------|---------|----------|----------------------|----------|--------------|--------------|
| 后台 | 通知 | 短信通知 | 短信通知 | 地市后台 | 发送短信通知。 | 发送短信 | 新增 | 调用短信平台发送通知。 |
""",
        encoding="utf-8",
    )
    warning = (
        "FPA 配置 warning: rule_set sms_service_rules 的 external_data_rules "
        "将普通外部服务「短信平台」配置为外部数据组「短信平台消息记录」。"
    )
    audit_trace = tmp_path / "trace.json"
    audit_trace.write_text(
        json.dumps({
            "modules": [{
                "module": "【地市后台】通知-短信通知-短信通知",
                "l3": "短信通知",
                "source": "rules",
                "raw_rows": [],
                "warnings": [warning, "规则优先策略未调用 AI"],
                "rule_hits": [],
            }],
        }, ensure_ascii=False),
        encoding="utf-8",
    )

    output = tmp_path / "check.xlsx"
    generate_fpa_check_xlsx_from_md(str(fpa_md), str(tree_md), str(output), str(audit_trace))

    wb = openpyxl.load_workbook(output, data_only=True)
    warning_rows = [
        [cell.value for cell in row]
        for row in wb["Warnings"].iter_rows(min_row=2)
    ]
    assert any(row[4] == warning for row in warning_rows)
    assert any(row[5] == "config.external_data_rules.external_service" for row in warning_rows)
    coverage_headers = [cell.value for cell in wb["覆盖审核"][1]]
    coverage_warning = wb["覆盖审核"].cell(2, coverage_headers.index("Warnings") + 1).value
    assert warning in coverage_warning
    raw_headers = [cell.value for cell in wb["AI原始返回"][1]]
    raw_warning = wb["AI原始返回"].cell(2, raw_headers.index("Warnings") + 1).value
    assert warning in raw_warning
    wb.close()
