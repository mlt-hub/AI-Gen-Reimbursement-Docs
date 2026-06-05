import json
from pathlib import Path

import openpyxl

from ai_gen_reimbursement_docs.excel_source import generate_md_files
from ai_gen_reimbursement_docs.gen_fpa import (
    calculate_fpa_excel_formula_projection,
    _read_fpa_rows_md_for_audit,
    generate_fpa_check_xlsx_from_md,
    generate_fpa_xlsx_from_md,
    plan_fpa_md_from_tree,
    preview_fpa_module,
)


FIXTURE_DIR = Path(__file__).parent / "fixtures" / "fpa_golden_cases"


def _write_meta_md(path: Path, meta: dict[str, str]) -> None:
    rows = "\n".join(f"| {key} | {value} |" for key, value in meta.items())
    path.write_text(f"# 文档元数据\n\n{rows}\n", encoding="utf-8")


def _write_tree_md(path: Path, rows: list[dict[str, str]]) -> None:
    lines = [
        "| 入口 | 一级模块 | 二级模块 | 三级模块 | 客户端类型 | 三级模块整体功能描述 | 功能过程 | 功能过程类型 | 功能过程描述 |",
        "|------|---------|---------|---------|----------|----------------------|----------|--------------|--------------|",
    ]
    for row in rows:
        lines.append(
            "| "
            + " | ".join([
                "后台",
                row.get("一级模块", ""),
                row.get("二级模块", ""),
                row.get("三级模块", ""),
                row.get("客户端类型", ""),
                row.get("三级模块整体功能描述", ""),
                row.get("功能过程", ""),
                row.get("功能过程类型", ""),
                row.get("功能过程描述", ""),
            ])
            + " |"
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def _headers(ws) -> list[str]:
    return [cell.value for cell in ws[1]]


def _sheet_records(ws) -> list[dict[str, object]]:
    headers = _headers(ws)
    return [
        {header: ws.cell(row=row, column=idx + 1).value for idx, header in enumerate(headers)}
        for row in range(2, ws.max_row + 1)
    ]


def _summary_total(summary_md: Path) -> float:
    text = summary_md.read_text(encoding="utf-8")
    marker = "FPA工作量（人/天）:"
    assert marker in text
    return float(text.split(marker, 1)[1].strip().split()[0])


def _write_minimal_excel(path: Path, rows: list[dict[str, str]], meta: dict[str, str]) -> None:
    wb = openpyxl.Workbook()
    ws_meta = wb.active
    ws_meta.title = "1、工单需求-元数据录入"
    ws_meta.append(["项目", "内容"])
    for key, value in meta.items():
        ws_meta.append([key, value])
    for title in [
        "2、功能清单-内容录入",
        "3、FPA工作量评估-元数据录入",
        "4、项目需求说明书-元数据录入",
        "5、项目功能点拆分表-元数据录入",
        "6、项目需求清单-元数据录入",
        "7、测试-元数据自动统计",
    ]:
        wb.create_sheet(title)
    ws_func = wb["2、功能清单-内容录入"]
    ws_func.append([
        "入口", "一级模块", "二级模块", "三级模块", "客户端类型",
        "三级模块整体功能描述", "功能过程", "功能过程类型", "功能过程描述",
    ])
    for row in rows:
        ws_func.append([
            "后台",
            row.get("一级模块", ""),
            row.get("二级模块", ""),
            row.get("三级模块", ""),
            row.get("客户端类型", ""),
            row.get("三级模块整体功能描述", ""),
            row.get("功能过程", ""),
            row.get("功能过程类型", ""),
            row.get("功能过程描述", ""),
        ])
    wb.save(path)
    wb.close()


def test_fpa_acceptance_formula_projection_matches_summary_across_type_strategies(tmp_path):
    case = json.loads((FIXTURE_DIR / "vertical_industry_management.json").read_text(encoding="utf-8"))
    template = Path(__file__).parent / "fixtures" / "output_templates" / "FPA工作量评估-输出模板.xlsx"
    tree_md = tmp_path / "tree.md"
    meta_md = tmp_path / "meta.md"
    _write_tree_md(tree_md, case["rows"])
    _write_meta_md(meta_md, case["meta"])

    totals: dict[str, float] = {}
    projections: dict[str, float] = {}
    for profile_name in ("unified_ui", "strict_fpa"):
        fpa_md = tmp_path / f"{profile_name}.md"
        summary_md = tmp_path / f"{profile_name}-summary.md"
        fpa_xlsx = tmp_path / f"{profile_name}.xlsx"

        plan_fpa_md_from_tree(
            str(tree_md),
            str(meta_md),
            str(fpa_md),
            summary_md_path=str(summary_md),
            profile_name=profile_name,
            strategy="rules_only",
            rule_set={"unified_ui": "unified_ui_rs", "strict_fpa": "strict_fpa_rs"}[profile_name],
        )
        generate_fpa_xlsx_from_md(str(fpa_md), str(meta_md), str(template), str(fpa_xlsx))

        totals[profile_name] = _summary_total(summary_md)
        projections[profile_name] = calculate_fpa_excel_formula_projection(str(fpa_xlsx))

    assert totals == {"unified_ui": 41.0, "strict_fpa": 32.0}
    assert projections == totals


def test_fpa_acceptance_strict_rules_formal_check_workbook_from_golden_case(tmp_path):
    case = json.loads((FIXTURE_DIR / "vertical_industry_management.json").read_text(encoding="utf-8"))
    tree_md = tmp_path / "tree.md"
    meta_md = tmp_path / "meta.md"
    fpa_md = tmp_path / "fpa.md"
    summary_md = tmp_path / "summary.md"
    audit_trace = tmp_path / "trace.json"
    check_xlsx = tmp_path / "check.xlsx"

    _write_tree_md(tree_md, case["rows"])
    _write_meta_md(meta_md, case["meta"])

    plan_fpa_md_from_tree(
        str(tree_md),
        str(meta_md),
        str(fpa_md),
        summary_md_path=str(summary_md),
        profile_name="strict_fpa",
        strategy="rules_only",
        rule_set="strict_fpa_rs",
        audit_trace_path=str(audit_trace),
    )
    generate_fpa_check_xlsx_from_md(str(fpa_md), str(tree_md), str(check_xlsx), str(audit_trace))

    _, fpa_rows = _read_fpa_rows_md_for_audit(str(fpa_md))
    actual = [
        {"name": row["新增/修改功能点"], "type": row["类型"]}
        for row in fpa_rows
    ]
    assert actual == case["expected"]["strict_fpa"]
    assert "FPA工作量（人/天）: 32.0" in summary_md.read_text(encoding="utf-8")

    wb = openpyxl.load_workbook(check_xlsx, data_only=True)
    assert wb.sheetnames == ["FPA结果", "覆盖审核", "Warnings", "规则命中详情", "AI原始返回"]
    result_headers = _headers(wb["FPA结果"])
    for header in ["复杂度", "DET", "RET", "FTR", "复杂度说明", "调整值计算方式"]:
        assert header in result_headers

    ws_coverage = wb["覆盖审核"]
    coverage_headers = _headers(ws_coverage)
    assert ws_coverage.cell(2, coverage_headers.index("功能过程总数") + 1).value == 6
    assert ws_coverage.cell(2, coverage_headers.index("未覆盖数") + 1).value == 0
    assert ws_coverage.cell(2, coverage_headers.index("已覆盖数") + 1).value == 6

    ws_rule_hits = wb["规则命中详情"]
    rule_ids = [
        ws_rule_hits.cell(row=row, column=_headers(ws_rule_hits).index("规则ID") + 1).value
        for row in range(2, ws_rule_hits.max_row + 1)
    ]
    assert "strict_fpa.internal_data_group" in rule_ids
    assert "strict_fpa.transaction.ei" in rule_ids
    assert "strict_fpa.transaction.eq" in rule_ids

    ws_raw = wb["AI原始返回"]
    raw_headers = _headers(ws_raw)
    assert ws_raw.cell(2, raw_headers.index("来源") + 1).value == "rules"
    assert ws_raw.cell(2, raw_headers.index("Warnings") + 1).value == "仅规则策略未调用 AI"
    wb.close()


def test_fpa_acceptance_preview_and_formal_rules_use_same_rows(tmp_path):
    case = json.loads((FIXTURE_DIR / "sms_notification_service.json").read_text(encoding="utf-8"))
    input_xlsx = tmp_path / "input.xlsx"
    tree_md = tmp_path / "tree.md"
    meta_md = tmp_path / "meta.md"
    fpa_md = tmp_path / "fpa.md"
    check_xlsx = tmp_path / "check.xlsx"

    _write_minimal_excel(input_xlsx, case["rows"], case["meta"])
    _write_tree_md(tree_md, case["rows"])
    _write_meta_md(meta_md, case["meta"])

    preview = preview_fpa_module(
        file_path=str(input_xlsx),
        module_index=1,
        profile_name="unified_ui",
        strategy="rules_only",
        work_dir=str(tmp_path / "preview"),
    )
    plan_fpa_md_from_tree(
        str(tree_md),
        str(meta_md),
        str(fpa_md),
        profile_name="unified_ui",
        strategy="rules_only",
    )
    _, formal_rows = _read_fpa_rows_md_for_audit(str(fpa_md))

    preview_summary = [
        {"name": row["name"], "type": row["type"]}
        for row in preview["rows"]
    ]
    formal_summary = [
        {"name": row["新增/修改功能点"], "type": row["类型"]}
        for row in formal_rows
    ]
    assert preview_summary == formal_summary == case["expected"]["unified_ui"]
    assert preview["audit"]["coverage"]["missing_count"] == 0

    generate_fpa_check_xlsx_from_md(str(fpa_md), str(tree_md), str(check_xlsx))
    wb = openpyxl.load_workbook(check_xlsx, data_only=True)
    try:
        ws_coverage = wb["覆盖审核"]
        coverage_headers = _headers(ws_coverage)
        assert ws_coverage.cell(2, coverage_headers.index("功能过程总数") + 1).value == preview["audit"]["coverage"]["process_total"]
        assert ws_coverage.cell(2, coverage_headers.index("已覆盖数") + 1).value == preview["audit"]["coverage"]["covered_count"]
        assert ws_coverage.cell(2, coverage_headers.index("未覆盖数") + 1).value == preview["audit"]["coverage"]["missing_count"]
        generation_counts = json.loads(ws_coverage.cell(2, coverage_headers.index("生成方式统计") + 1).value)
        assert generation_counts == preview["audit"]["generation_counts"]
    finally:
        wb.close()
    assert not list((tmp_path / "preview").glob("**/FPA工作量评估.xlsx"))


def test_fpa_acceptance_real_excel_to_md_to_formal_check_workbook(tmp_path):
    case = json.loads((FIXTURE_DIR / "mixed_internal_external_data_functions.json").read_text(encoding="utf-8"))
    input_xlsx = tmp_path / "mixed.xlsx"
    md_dir = tmp_path / "md"
    fpa_md = tmp_path / "fpa.md"
    summary_md = tmp_path / "summary.md"
    audit_trace = tmp_path / "trace.json"
    check_xlsx = tmp_path / "check.xlsx"

    _write_minimal_excel(input_xlsx, case["rows"], case["meta"])
    md_paths = generate_md_files(str(input_xlsx), str(md_dir))

    plan_fpa_md_from_tree(
        md_paths["module_tree_md"],
        md_paths["doc_meta_md"],
        str(fpa_md),
        summary_md_path=str(summary_md),
        profile_name="strict_fpa",
        strategy="rules_only",
        audit_trace_path=str(audit_trace),
    )
    generate_fpa_check_xlsx_from_md(str(fpa_md), md_paths["module_tree_md"], str(check_xlsx), str(audit_trace))

    _, fpa_rows = _read_fpa_rows_md_for_audit(str(fpa_md))
    actual = [
        {"name": row["新增/修改功能点"], "type": row["类型"]}
        for row in fpa_rows
    ]
    assert actual == case["expected"]["strict_fpa"]
    assert "FPA工作量（人/天）: 29.0" in summary_md.read_text(encoding="utf-8")

    wb = openpyxl.load_workbook(check_xlsx, data_only=True)
    assert wb.sheetnames == ["FPA结果", "覆盖审核", "Warnings", "规则命中详情", "AI原始返回"]
    ws_coverage = wb["覆盖审核"]
    coverage_headers = _headers(ws_coverage)
    assert ws_coverage.cell(2, coverage_headers.index("功能过程总数") + 1).value == 4
    assert ws_coverage.cell(2, coverage_headers.index("已覆盖数") + 1).value == 4
    assert ws_coverage.cell(2, coverage_headers.index("未覆盖数") + 1).value == 0

    ws_rule_hits = wb["规则命中详情"]
    rule_headers = _headers(ws_rule_hits)
    rule_ids = [
        ws_rule_hits.cell(row=row, column=rule_headers.index("规则ID") + 1).value
        for row in range(2, ws_rule_hits.max_row + 1)
    ]
    assert rule_ids.count("strict_fpa.external_data_group") >= 2
    assert "strict_fpa.internal_data_group" in rule_ids
    wb.close()


def test_fpa_acceptance_mock_ai_warning_source_reaches_check_workbook(monkeypatch, tmp_path):
    rows = [
        {
            "客户端类型": "地市后台",
            "一级模块": "客户管理",
            "二级模块": "客户查询",
            "三级模块": "客户查询",
            "三级模块整体功能描述": "按条件查询客户。",
            "功能过程": "查询客户",
            "功能过程类型": "新增",
            "功能过程描述": "按客户名称查询客户列表。",
        }
    ]
    tree_md = tmp_path / "tree.md"
    meta_md = tmp_path / "meta.md"
    fpa_md = tmp_path / "fpa.md"
    audit_trace = tmp_path / "trace.json"
    check_xlsx = tmp_path / "check.xlsx"

    _write_tree_md(tree_md, rows)
    _write_meta_md(meta_md, {"子系统（模块）": "测试系统", "资产标识": "TEST-001"})
    monkeypatch.setattr(
        "ai_gen_reimbursement_docs.gen_fpa._call_llm",
        lambda *args, **kwargs: json.dumps({
            "rows": [{
                "name": "查询客户-查询处理开发",
                "type": "EQ",
                "classification_basis_index": 99,
                "explanation": "按客户名称查询客户列表。",
                "source_processes": ["查询客户"],
            }]
        }, ensure_ascii=False),
    )

    plan_fpa_md_from_tree(
        str(tree_md),
        str(meta_md),
        str(fpa_md),
        api_key="sk-test",
        model="mock-model",
        profile_name="unified_ui",
        strategy="ai_first",
        audit_trace_path=str(audit_trace),
    )
    generate_fpa_check_xlsx_from_md(str(fpa_md), str(tree_md), str(check_xlsx), str(audit_trace))

    wb = openpyxl.load_workbook(check_xlsx, data_only=True)
    ws_warnings = wb["Warnings"]
    warning_headers = _headers(ws_warnings)
    warning_values = {
        header: ws_warnings.cell(2, idx + 1).value
        for idx, header in enumerate(warning_headers)
    }
    assert "classification_basis_index 越界" in warning_values["Warning"]
    assert warning_values["来源规则ID"] == "postprocess.classification_basis_index"
    warning_source_ids = [
        ws_warnings.cell(row=row, column=warning_headers.index("来源规则ID") + 1).value
        for row in range(2, ws_warnings.max_row + 1)
    ]
    assert "coverage.missing_process" not in warning_source_ids

    ws_rule_hits = wb["规则命中详情"]
    rule_headers = _headers(ws_rule_hits)
    rule_ids = [
        ws_rule_hits.cell(row=row, column=rule_headers.index("规则ID") + 1).value
        for row in range(2, ws_rule_hits.max_row + 1)
    ]
    assert "postprocess.ai_type_validation" in rule_ids
    assert "postprocess.classification_basis_index" in rule_ids
    wb.close()


def test_fpa_acceptance_check_workbook_does_not_report_type_conflict_when_types_match(monkeypatch, tmp_path):
    rows = [
        {
            "客户端类型": "地市后台",
            "一级模块": "垂直行业营销",
            "二级模块": "垂直行业管理",
            "三级模块": "垂直行业管理",
            "三级模块整体功能描述": "维护垂直行业基础信息和管理员。",
            "功能过程": "添加垂直行业",
            "功能过程类型": "新增",
            "功能过程描述": "点击添加按钮，在弹窗输入垂直行业名称并保存。",
        }
    ]
    tree_md = tmp_path / "tree.md"
    meta_md = tmp_path / "meta.md"
    fpa_md = tmp_path / "fpa.md"
    audit_trace = tmp_path / "trace.json"
    check_xlsx = tmp_path / "check.xlsx"

    _write_tree_md(tree_md, rows)
    _write_meta_md(meta_md, {"子系统（模块）": "测试系统", "资产标识": "TEST-001"})
    monkeypatch.setattr(
        "ai_gen_reimbursement_docs.gen_fpa._call_llm",
        lambda *args, **kwargs: json.dumps({
            "rows": [{
                "name": "添加垂直行业",
                "type": "EI",
                "explanation": "点击添加按钮，在弹窗输入垂直行业名称并保存。",
                "source_processes": ["添加垂直行业"],
            }]
        }, ensure_ascii=False),
    )

    plan_fpa_md_from_tree(
        str(tree_md),
        str(meta_md),
        str(fpa_md),
        api_key="sk-test",
        model="mock-model",
        profile_name="strict_fpa",
        strategy="ai_first",
        audit_trace_path=str(audit_trace),
    )
    generate_fpa_check_xlsx_from_md(str(fpa_md), str(tree_md), str(check_xlsx), str(audit_trace))

    wb = openpyxl.load_workbook(check_xlsx, data_only=True)
    try:
        warning_text = "；".join(str(row["Warning"] or "") for row in _sheet_records(wb["Warnings"]))
        assert "AI type=EI 与规则存在冲突" not in warning_text

        rule_hits = _sheet_records(wb["规则命中详情"])
        assert not any(row["规则ID"] == "postprocess.ai_first_type_conflict" for row in rule_hits)
    finally:
        wb.close()


def test_fpa_acceptance_check_workbook_type_conflict_metadata_is_consistent(monkeypatch, tmp_path):
    rows = [
        {
            "客户端类型": "地市后台",
            "一级模块": "消息管理",
            "二级模块": "通知发送",
            "三级模块": "短信通知",
            "三级模块整体功能描述": "系统调用短信平台发送通知短信。",
            "功能过程": "发送测试短信",
            "功能过程类型": "新增",
            "功能过程描述": "调用短信平台发送测试短信。",
        }
    ]
    tree_md = tmp_path / "tree.md"
    meta_md = tmp_path / "meta.md"
    fpa_md = tmp_path / "fpa.md"
    audit_trace = tmp_path / "trace.json"
    check_xlsx = tmp_path / "check.xlsx"

    _write_tree_md(tree_md, rows)
    _write_meta_md(meta_md, {"子系统（模块）": "测试系统", "资产标识": "TEST-001"})
    monkeypatch.setattr(
        "ai_gen_reimbursement_docs.gen_fpa._call_llm",
        lambda *args, **kwargs: json.dumps({
            "rows": [{
                "name": "发送测试短信",
                "type": "EIF",
                "explanation": "调用短信平台发送测试短信。",
                "source_processes": ["发送测试短信"],
            }]
        }, ensure_ascii=False),
    )

    plan_fpa_md_from_tree(
        str(tree_md),
        str(meta_md),
        str(fpa_md),
        api_key="sk-test",
        model="mock-model",
        profile_name="strict_fpa",
        strategy="ai_first",
        audit_trace_path=str(audit_trace),
    )
    generate_fpa_check_xlsx_from_md(str(fpa_md), str(tree_md), str(check_xlsx), str(audit_trace))

    wb = openpyxl.load_workbook(check_xlsx, data_only=True)
    try:
        result_rows = _sheet_records(wb["FPA结果"])
        ai_row = next(row for row in result_rows if row["生成方式"] == "ai")
        assert ai_row["类型"] == "EIF"

        warnings = _sheet_records(wb["Warnings"])
        assert any("AI type=EIF 与规则存在冲突" in str(row["Warning"]) for row in warnings)

        rule_hits = _sheet_records(wb["规则命中详情"])
        conflict_hit = next(row for row in rule_hits if row["规则ID"] == "postprocess.ai_first_type_conflict")
        assert conflict_hit["建议类型"] == "EI"
        assert conflict_hit["是否采用"] == "否"
        assert "规则建议 type=EI" in conflict_hit["规则说明"]
    finally:
        wb.close()


def test_fpa_acceptance_check_workbook_distinguishes_data_function_supplement_from_missing_process(monkeypatch, tmp_path):
    rows = [
        {
            "客户端类型": "地市后台",
            "一级模块": "客户管理",
            "二级模块": "客户档案",
            "三级模块": "客户档案",
            "三级模块整体功能描述": "维护客户档案。",
            "功能过程": "添加客户档案",
            "功能过程类型": "新增",
            "功能过程描述": "保存客户档案。",
        }
    ]
    tree_md = tmp_path / "tree.md"
    meta_md = tmp_path / "meta.md"
    fpa_md = tmp_path / "fpa.md"
    audit_trace = tmp_path / "trace.json"
    check_xlsx = tmp_path / "check.xlsx"

    _write_tree_md(tree_md, rows)
    _write_meta_md(meta_md, {"子系统（模块）": "测试系统", "资产标识": "TEST-001"})
    monkeypatch.setattr(
        "ai_gen_reimbursement_docs.gen_fpa._call_llm",
        lambda *args, **kwargs: json.dumps({
            "rows": [{
                "name": "添加客户档案",
                "type": "EI",
                "explanation": "保存客户档案。",
                "source_processes": ["添加客户档案"],
            }]
        }, ensure_ascii=False),
    )

    plan_fpa_md_from_tree(
        str(tree_md),
        str(meta_md),
        str(fpa_md),
        api_key="sk-test",
        model="mock-model",
        profile_name="strict_fpa",
        strategy="ai_first",
        audit_trace_path=str(audit_trace),
    )
    generate_fpa_check_xlsx_from_md(str(fpa_md), str(tree_md), str(check_xlsx), str(audit_trace))

    wb = openpyxl.load_workbook(check_xlsx, data_only=True)
    try:
        coverage_rows = _sheet_records(wb["覆盖审核"])
        coverage_warning = str(coverage_rows[0]["Warnings"] or "")
        assert "AI 结果未包含数据功能行" in coverage_warning
        assert "未覆盖 0 个功能过程" not in coverage_warning

        warning_rows = _sheet_records(wb["Warnings"])
        warning_text = "；".join(str(row["Warning"] or "") for row in warning_rows)
        assert "AI 结果未包含数据功能行" in warning_text
        assert "未覆盖 0 个功能过程" not in warning_text

        rule_hits = _sheet_records(wb["规则命中详情"])
        assert any(
            row["生成方式"] == "rules_fallback"
            and row["建议类型"] == "ILF"
            and "AI 结果未包含数据功能行" in str(row["Warnings"])
            for row in rule_hits
        )
    finally:
        wb.close()


def test_fpa_acceptance_check_workbook_reports_missing_process_supplement(monkeypatch, tmp_path):
    rows = [
        {
            "客户端类型": "地市后台",
            "一级模块": "客户管理",
            "二级模块": "客户查询",
            "三级模块": "客户查询",
            "三级模块整体功能描述": "查询和导出客户。",
            "功能过程": "查询客户",
            "功能过程类型": "新增",
            "功能过程描述": "按客户名称查询客户列表。",
        },
        {
            "客户端类型": "地市后台",
            "一级模块": "客户管理",
            "二级模块": "客户查询",
            "三级模块": "客户查询",
            "三级模块整体功能描述": "查询和导出客户。",
            "功能过程": "导出客户",
            "功能过程类型": "新增",
            "功能过程描述": "导出客户列表。",
        },
    ]
    tree_md = tmp_path / "tree.md"
    meta_md = tmp_path / "meta.md"
    fpa_md = tmp_path / "fpa.md"
    audit_trace = tmp_path / "trace.json"
    check_xlsx = tmp_path / "check.xlsx"

    _write_tree_md(tree_md, rows)
    _write_meta_md(meta_md, {"子系统（模块）": "测试系统", "资产标识": "TEST-001"})
    monkeypatch.setattr(
        "ai_gen_reimbursement_docs.gen_fpa._call_llm",
        lambda *args, **kwargs: json.dumps({
            "rows": [{
                "name": "查询客户-查询处理开发",
                "type": "EQ",
                "explanation": "按客户名称查询客户列表。",
                "source_processes": ["查询客户"],
            }]
        }, ensure_ascii=False),
    )

    plan_fpa_md_from_tree(
        str(tree_md),
        str(meta_md),
        str(fpa_md),
        api_key="sk-test",
        model="mock-model",
        profile_name="unified_ui",
        strategy="ai_first",
        audit_trace_path=str(audit_trace),
    )
    generate_fpa_check_xlsx_from_md(str(fpa_md), str(tree_md), str(check_xlsx), str(audit_trace))

    wb = openpyxl.load_workbook(check_xlsx, data_only=True)
    try:
        coverage_warning = str(_sheet_records(wb["覆盖审核"])[0]["Warnings"] or "")
        assert "AI 结果未覆盖 1 个功能过程" in coverage_warning
        assert "导出客户" not in str(_sheet_records(wb["覆盖审核"])[0]["未覆盖功能过程"] or "")

        rule_hits = _sheet_records(wb["规则命中详情"])
        assert any(
            row["生成方式"] == "rules_fallback"
            and "AI 结果未覆盖该功能过程" in str(row["Warnings"])
            for row in rule_hits
        )
    finally:
        wb.close()


def test_fpa_acceptance_rules_first_check_workbook_explains_rules_without_ai(monkeypatch, tmp_path):
    rows = [
        {
            "客户端类型": "地市后台",
            "一级模块": "客户管理",
            "二级模块": "客户查询",
            "三级模块": "客户查询",
            "三级模块整体功能描述": "按条件查询客户。",
            "功能过程": "查询客户",
            "功能过程类型": "新增",
            "功能过程描述": "按客户名称查询客户列表。",
        }
    ]
    tree_md = tmp_path / "tree.md"
    meta_md = tmp_path / "meta.md"
    fpa_md = tmp_path / "fpa.md"
    audit_trace = tmp_path / "trace.json"
    check_xlsx = tmp_path / "check.xlsx"

    _write_tree_md(tree_md, rows)
    _write_meta_md(meta_md, {"子系统（模块）": "测试系统", "资产标识": "TEST-001"})
    monkeypatch.setattr(
        "ai_gen_reimbursement_docs.gen_fpa._call_llm",
        lambda *args, **kwargs: (_ for _ in ()).throw(AssertionError("rules_first should not call AI")),
    )

    plan_fpa_md_from_tree(
        str(tree_md),
        str(meta_md),
        str(fpa_md),
        api_key="sk-test",
        model="mock-model",
        profile_name="unified_ui",
        strategy="rules_first",
        audit_trace_path=str(audit_trace),
    )
    generate_fpa_check_xlsx_from_md(str(fpa_md), str(tree_md), str(check_xlsx), str(audit_trace))

    wb = openpyxl.load_workbook(check_xlsx, data_only=True)
    try:
        raw_rows = _sheet_records(wb["AI原始返回"])
        assert raw_rows[0]["来源"] == "rules"
        assert "规则优先策略未调用 AI：规则结果覆盖完整且基础字段有效" in str(raw_rows[0]["Warnings"])
    finally:
        wb.close()


def test_fpa_acceptance_rules_first_low_confidence_check_workbook_explains_ai_review(monkeypatch, tmp_path):
    rows = [
        {
            "客户端类型": "地市后台",
            "一级模块": "垂直行业营销",
            "二级模块": "垂直行业管理",
            "三级模块": "垂直行业管理",
            "三级模块整体功能描述": "维护垂直行业基础信息。",
            "功能过程": "添加垂直行业",
            "功能过程类型": "新增",
            "功能过程描述": "输入垂直行业名称并保存。",
        }
    ]
    tree_md = tmp_path / "tree.md"
    meta_md = tmp_path / "meta.md"
    fpa_md = tmp_path / "fpa.md"
    audit_trace = tmp_path / "trace.json"
    check_xlsx = tmp_path / "check.xlsx"

    def fake_fallback(self, group, meta, start_seq=1):
        return [{
            "序号": start_seq,
            "子系统(模块)": meta.get("子系统（模块）", ""),
            "资产标识": meta.get("资产标识", ""),
            "新增/修改功能点": "低置信度规则行",
            "类型": "",
            "计算依据归类": "",
            "计算依据说明": "低置信度规则行。",
            "变更状态": "新增",
            "调整值": 1,
            "要素数量": 1,
            "生成方式": "fallback",
            "类型理由": "",
            "源功能过程": "",
            "后处理警告": "",
        }]

    _write_tree_md(tree_md, rows)
    _write_meta_md(meta_md, {"子系统（模块）": "测试系统", "资产标识": "TEST-001"})
    monkeypatch.setattr(
        "ai_gen_reimbursement_docs.fpa_profiles.CustomRulesProfile.fallback_rows_for_l3",
        fake_fallback,
    )
    monkeypatch.setattr(
        "ai_gen_reimbursement_docs.gen_fpa._call_llm",
        lambda *args, **kwargs: json.dumps({
            "rows": [{
                "name": "AI 复核功能点",
                "type": "EI",
                "explanation": "AI 复核功能点，具体为以下：1、覆盖低置信度规则无法覆盖的功能过程。",
                "source_processes": ["添加垂直行业"],
            }]
        }, ensure_ascii=False),
    )

    plan_fpa_md_from_tree(
        str(tree_md),
        str(meta_md),
        str(fpa_md),
        api_key="sk-test",
        model="mock-model",
        profile_name="unified_ui",
        strategy="rules_first",
        audit_trace_path=str(audit_trace),
    )
    generate_fpa_check_xlsx_from_md(str(fpa_md), str(tree_md), str(check_xlsx), str(audit_trace))

    wb = openpyxl.load_workbook(check_xlsx, data_only=True)
    try:
        raw_rows = _sheet_records(wb["AI原始返回"])
        assert raw_rows[0]["来源"] == "ai"
        assert "规则结果触发 AI 复核" in str(raw_rows[0]["Warnings"])

        result_rows = _sheet_records(wb["FPA结果"])
        assert result_rows[0]["生成方式"] == "ai"
        assert "AI 复核功能点" in str(result_rows[0]["新增/修改功能点"])
    finally:
        wb.close()


def test_fpa_acceptance_ai_cache_hit_is_visible_in_audit_and_check(monkeypatch, tmp_path):
    rows = [
        {
            "客户端类型": "地市后台",
            "一级模块": "客户管理",
            "二级模块": "客户查询",
            "三级模块": "客户查询",
            "三级模块整体功能描述": "按条件查询客户。",
            "功能过程": "查询客户",
            "功能过程类型": "新增",
            "功能过程描述": "按客户名称查询客户列表。",
        }
    ]
    tree_md = tmp_path / "tree.md"
    meta_md = tmp_path / "meta.md"
    first_fpa_md = tmp_path / "first.md"
    second_fpa_md = tmp_path / "second.md"
    audit_trace = tmp_path / "trace.json"
    check_xlsx = tmp_path / "check.xlsx"
    calls = {"count": 0}

    _write_tree_md(tree_md, rows)
    _write_meta_md(meta_md, {"子系统（模块）": "测试系统", "资产标识": "TEST-001"})

    def fake_llm(*args, **kwargs):
        calls["count"] += 1
        return json.dumps({
            "rows": [{
                "name": "查询客户-查询处理开发",
                "type": "EQ",
                "classification_basis_index": 1,
                "explanation": "按客户名称查询客户列表。",
                "source_processes": ["查询客户"],
            }]
        }, ensure_ascii=False)

    monkeypatch.setattr("ai_gen_reimbursement_docs.gen_fpa._call_llm", fake_llm)
    plan_fpa_md_from_tree(
        str(tree_md),
        str(meta_md),
        str(first_fpa_md),
        api_key="sk-test",
        model="mock-model",
        profile_name="unified_ui",
        strategy="ai_first",
        audit_trace_path=str(audit_trace),
    )
    plan_fpa_md_from_tree(
        str(tree_md),
        str(meta_md),
        str(second_fpa_md),
        api_key="sk-test",
        model="mock-model",
        profile_name="unified_ui",
        strategy="ai_first",
        audit_trace_path=str(audit_trace),
    )
    assert calls["count"] == 1

    trace = json.loads(audit_trace.read_text(encoding="utf-8"))
    assert trace["modules"][0]["source"] == "ai_cache"
    generate_fpa_check_xlsx_from_md(str(second_fpa_md), str(tree_md), str(check_xlsx), str(audit_trace))
    wb = openpyxl.load_workbook(check_xlsx, data_only=True)
    ws_raw = wb["AI原始返回"]
    raw_headers = _headers(ws_raw)
    assert ws_raw.cell(2, raw_headers.index("来源") + 1).value == "ai_cache"
    wb.close()
