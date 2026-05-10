"""生成 FPA工作量评估.xlsx 和 项目需求清单.xlsx"""

import logging
import os
import re
from copy import copy
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment

from cosmic_tool.excel_source import replace_placeholders, strip_ai_marker
from cosmic_tool.md_table import parse_md_table_row

logger = logging.getLogger('cosmic_tool.gen_xlsx')


# ============================================================
#  公用
# ============================================================

def _load_meta_md(meta_md_path: str) -> dict:
    """解析文档元数据.md 为扁平字典。"""
    meta = {}
    current_sheet = ""
    with open(meta_md_path, encoding='utf-8') as f:
        for line in f:
            line = line.rstrip()
            m = re.match(r'^##\s+(.+)$', line)
            if m:
                current_sheet = m.group(1).strip()
                continue
            if '|' in line and line.startswith('|'):
                cells = parse_md_table_row(line, min_cols=2)
                if cells is not None and cells[0] and cells[1]:
                    meta[cells[0]] = cells[1]
                    if current_sheet:
                        meta[f"{current_sheet}.{cells[0]}"] = cells[1]
    return meta


def _load_module_rows(tree_md_path: str) -> list[dict]:
    """解析功能清单模块树.md 为行字典列表。"""
    rows = []
    with open(tree_md_path, encoding='utf-8') as f:
        in_table = False
        for line in f:
            line = line.rstrip()
            if "| 入口 | 一级模块" in line:
                in_table = True
                continue
            if "|------" in line and in_table:
                continue
            if in_table:
                cells = parse_md_table_row(line, min_cols=9)
                if cells is not None:
                    rows.append({
                        "入口": cells[0],
                        "一级模块": cells[1],
                        "二级模块": cells[2],
                        "三级模块": cells[3],
                        "客户端类型": cells[4],
                        "三级模块整体功能描述": cells[5],
                        "功能过程": cells[6],
                        "功能过程类型": cells[7],
                        "功能过程描述": cells[8],
                    })
    return rows


def _receiver_from_client_type(client_type: str, rules_text: str) -> str:
    """根据客户端类型判定接收者。"""
    # 默认值
    default = "操作员"
    if not rules_text:
        if "后台" in client_type:
            return "后台管理员"
        if "前台" in client_type:
            return "普通用户"
        if "渠道" in client_type:
            return "渠道人员"
        return default

    # 解析规则文本：后台：后台管理员\n前台：普通用户\n...
    for line in rules_text.split('\n'):
        line = line.strip()
        if '：' in line:
            keyword, receiver = line.split('：', 1)
            if keyword in client_type:
                return receiver.strip()
    return default


# ============================================================
#  FPA工作量评估.xlsx
# ============================================================

def _save_ai_log(tag: str, model: str, prompt: str, response: str, log_type: str = "prompt"):
    """保存 AI 请求或响应到日志目录。"""
    try:
        base_log = os.environ.get('COSMIC_LOG_DIR', '') or os.path.join(
            os.path.dirname(os.path.dirname(__file__)), 'log'
        )
        sub_dir = 'ai_prompts' if log_type == 'prompt' else 'ai_responses'
        log_dir = os.path.join(base_log, sub_dir)
        os.makedirs(log_dir, exist_ok=True)
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        fname = f'{ts}_{tag}_{log_type}.txt'
        with open(os.path.join(log_dir, fname), 'w', encoding='utf-8') as f:
            f.write(f"# {log_type}: {tag}\n")
            f.write(f"# Model: {model}\n")
            f.write(f"# Time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
            f.write(prompt if log_type == 'prompt' else response)
    except Exception:
        pass


def _call_llm(prompt: str, system_prompt: str, api_key: str, model: str,
              base_url: str, tag: str = "") -> str:
    """调用 LLM。"""
    logger.info(f"AI 生成请求 [{tag}] 模型: {model}")
    _save_ai_log(tag, model, f"{system_prompt}\n\n---\n\n{prompt}", "", "prompt")

    try:
        import anthropic
        client_kwargs = {"api_key": api_key}
        if base_url:
            client_kwargs["base_url"] = base_url
        client = anthropic.Anthropic(**client_kwargs)
        from cosmic_tool.config_utils import load_max_tokens
        msg = client.messages.create(
            model=model or "deepseek-v4-flash",
            max_tokens=load_max_tokens(),
            system=system_prompt,
            messages=[{"role": "user", "content": prompt}],
        )
        # 取最后一个 TextBlock（跳过 ThinkingBlock）
        resp_text = ""
        for block in msg.content:
            if hasattr(block, 'text'):
                resp_text = block.text.strip()
                break

        _save_ai_log(tag, model, "", resp_text, "response")
        logger.info(f"AI 生成完成 [{tag}] 长度: {len(resp_text)} 字")
        return resp_text
    except Exception as e:
        logger.warning(f"LLM 调用失败 [{tag}]: {e}")
        return ""


def _build_fpa_rule_rows(rows: list[dict], meta: dict) -> list[dict]:
    """从功能清单行构建 FPA 模板行（规则骨架）。"""
    prefix_rule = meta.get("新增/修改功能点前缀生成规则",
                           "【客户端类型】一级模块-二级模块-三级模块-功能过程")
    receiver_rules = meta.get("功能用户-接收者判定", "")
    subsystem = meta.get("子系统（模块）", "")
    asset = meta.get("资产标识", "")

    fpa_rows = []
    seq = 0

    for r in rows:
        seq += 1
        client_type = r["客户端类型"]
        l1 = r["一级模块"]
        l2 = r["二级模块"]
        l3 = r["三级模块"]
        proc = r["功能过程"]
        proc_desc = r["功能过程描述"]
        proc_type = r["功能过程类型"]

        # 接收者
        receiver = _receiver_from_client_type(client_type, receiver_rules)

        # 功能点前缀
        fp_prefix = prefix_rule \
            .replace("【客户端类型】", client_type) \
            .replace("一级模块", l1) \
            .replace("二级模块", l2) \
            .replace("三级模块", l3) \
            .replace("功能过程", proc)

        # 界面行
        fpa_rows.append({
            "序号": seq,
            "子系统(模块)": subsystem,
            "资产标识": asset,
            "新增/修改功能点": f"{fp_prefix}-界面开发",
            "类型": "EI",
            "计算依据归类": "",
            "计算依据说明": f"【{client_type}】{l1}-{l2}-{l3}-{proc}-界面开发，具体如下：\n1、",
            "变更状态": proc_type,
            "基准值": "",
            "调整值": 2,
            "要素数量": 1,
            "FPA工作量": "",
            "核减后工作量": "",
            "备注说明": "",
        })

        # 接口行
        fpa_rows.append({
            "序号": seq,
            "子系统(模块)": subsystem,
            "资产标识": asset,
            "新增/修改功能点": f"{fp_prefix}-接口开发",
            "类型": "ILF",
            "计算依据归类": "",
            "计算依据说明": f"【{client_type}】{l1}-{l2}-{l3}-{proc}-接口开发，具体如下：\n1、",
            "变更状态": proc_type,
            "基准值": "",
            "调整值": 1,
            "要素数量": 1,
            "FPA工作量": "",
            "核减后工作量": "",
            "备注说明": "",
        })

    return fpa_rows


def _ai_fill_fpa(
    fpa_rows: list[dict],
    judgement_rules: str,
    api_key: str,
    model: str,
    base_url: str,
) -> list[dict]:
    """AI 填充 FPA 行的 F/G 列。"""
    if not api_key:
        logger.warning("未设置 API Key，跳过 AI 填充 FPA")
        return fpa_rows

    from cosmic_tool.config_utils import load_ai_system_prompt
    system_prompt = load_ai_system_prompt("fpa_eval") or (
        "你是一个 FPA 功能点评估助手。根据功能过程描述和类型，"
        "从判定原则中选择最匹配的一项，并展开计算依据说明。"
        "直接输出结果，不要输出其他内容。"
    )

    total = len(fpa_rows)
    from cosmic_tool.config_utils import load_flow_max_ai
    _max_ai = load_flow_max_ai("gen_fpa")
    for idx, row in enumerate(fpa_rows, 1):
        if _max_ai > 0 and idx > _max_ai:
            logger.info(f"  [{idx}/{total}] 跳过（超过 AI 限制 {_max_ai}）")
            continue
        if not row["计算依据说明"]:
            continue

        row_tag = f"fpa_{row['类型']}_{row['新增/修改功能点'][:30]}"
        prompt = (
            f"功能过程描述：{row['计算依据说明']}\n"
            f"类型：{row['类型']}\n\n"
            f"判定原则列表：\n{judgement_rules}\n\n"
            f"请按以下格式输出（直接输出，不要解释）：\n"
            f"计算依据归类：<选中的判定原则>\n"
            f"计算依据说明：<展开的详细说明>"
        )

        logger.info(f"  FPA AI 填充 [{idx}/{total}] {row['新增/修改功能点'][:40]}...")
        resp = _call_llm(prompt, system_prompt, api_key, model, base_url, tag=row_tag)
        if resp:
            m_cat = re.search(r'计算依据归类[：:]\s*(.+?)(?:\n|$)', resp)
            m_desc = re.search(r'计算依据说明[：:]\s*(.+)', resp, re.DOTALL)
            if m_cat:
                row["计算依据归类"] = m_cat.group(1).strip()
            if m_desc:
                row["计算依据说明"] = m_desc.group(1).strip()

    return fpa_rows


def init_fpa_template_md(
    tree_md_path: str,
    meta_md_path: str,
    output_md_path: str,
    summary_md_path: str = "",
) -> str:
    """生成 FPA 模板 MD（规则骨架，F/G 列留空待 AI 填充）。

    Args:
        summary_md_path: 非空时同步写入 FPA工作量-计算结果.md（调整值×要素数量 的求和）
    """
    logger.info("生成 FPA 模板 MD...")
    meta = _load_meta_md(meta_md_path)
    rows = _load_module_rows(tree_md_path)
    fpa_rows = _build_fpa_rule_rows(rows, meta)

    total = 0.0
    with open(output_md_path, 'w', encoding='utf-8') as f:
        f.write("# FPA 工作量评估\n\n")
        f.write(f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        f.write("| 序号 | 子系统(模块) | 资产标识 | 新增/修改功能点 | 类型 | 计算依据归类 | 计算依据说明 | 变更状态 | 调整值 | 要素数量 |\n")
        f.write("|------|-------------|---------|----------------|------|-------------|-------------|---------|-------|---------|\n")
        for row in fpa_rows:
            vals = [
                str(row["序号"]),
                row["子系统(模块)"],
                row["资产标识"],
                row["新增/修改功能点"].replace('|', '\\|'),
                row["类型"],
                row["计算依据归类"],
                row["计算依据说明"].replace('|', '\\|').replace('\n', ' '),
                row["变更状态"],
                str(row["调整值"]),
                str(row["要素数量"]),
            ]
            f.write("| " + " | ".join(vals) + " |\n")
            try:
                total += float(row["调整值"]) * float(row["要素数量"])
            except (ValueError, TypeError):
                pass

    logger.info(f"FPA 模板 MD 已生成: {output_md_path} ({len(fpa_rows)} 行)")

    if summary_md_path:
        os.makedirs(os.path.dirname(summary_md_path) or '.', exist_ok=True)
        with open(summary_md_path, 'w', encoding='utf-8') as f:
            f.write("# FPA 工作量\n\n")
            f.write(f"FPA工作量（人/天）: {total}\n")
        logger.info(f"FPA 工作量已写入: {summary_md_path} ({total})")

    return output_md_path


def ai_fill_fpa_md(
    fpa_md_path: str,
    meta_md_path: str,
    api_key: str = "",
    model: str = "",
    base_url: str = "",
) -> str:
    """读取 FPA 模板 MD，AI 填充 F/G 列，写回 MD。"""
    logger.info("AI 填充 FPA 数据...")
    logger.info(f"AI 模型: {model}  端点: {base_url or '默认'}  API Key: {'已设置' if api_key else '未设置'}")

    meta = _load_meta_md(meta_md_path)
    judgement_rules = meta.get("计算依据归类判定原则", "")

    # 解析 MD 表格
    fpa_rows = []
    with open(fpa_md_path, encoding='utf-8') as f:
        in_table = False
        for line in f:
            line = line.rstrip()
            if "| 序号 | 子系统" in line:
                in_table = True
                continue
            if "|------|" in line and in_table:
                continue
            if in_table:
                cells = parse_md_table_row(line, min_cols=10)
                if cells is not None:
                    fpa_rows.append({
                        "序号": cells[0],
                        "子系统(模块)": cells[1],
                        "资产标识": cells[2],
                        "新增/修改功能点": cells[3],
                        "类型": cells[4],
                        "计算依据归类": cells[5],
                        "计算依据说明": cells[6],
                        "变更状态": cells[7],
                        "调整值": cells[8],
                        "要素数量": cells[9],
                    })

    # AI 填充
    fpa_rows = _ai_fill_fpa(fpa_rows, judgement_rules, api_key, model, base_url)

    # 写回 MD
    with open(fpa_md_path, 'w', encoding='utf-8') as f:
        f.write("# FPA 工作量评估\n\n")
        f.write(f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"**AI 填充**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        f.write("| 序号 | 子系统(模块) | 资产标识 | 新增/修改功能点 | 类型 | 计算依据归类 | 计算依据说明 | 变更状态 | 调整值 | 要素数量 |\n")
        f.write("|------|-------------|---------|----------------|------|-------------|-------------|---------|-------|---------|\n")
        for row in fpa_rows:
            vals = [
                str(row["序号"]),
                row["子系统(模块)"],
                row["资产标识"],
                row["新增/修改功能点"].replace('|', '\\|'),
                row["类型"],
                row["计算依据归类"],
                row["计算依据说明"].replace('|', '\\|').replace('\n', ' '),
                row["变更状态"],
                str(row["调整值"]),
                str(row["要素数量"]),
            ]
            f.write("| " + " | ".join(vals) + " |\n")

    logger.info(f"FPA MD 已填充: {fpa_md_path}")
    return fpa_md_path


def generate_fpa_xlsx_from_md(
    fpa_md_path: str,
    meta_md_path: str,
    template_path: str,
    output_path: str,
) -> str:
    """从已填充的 FPA MD 生成 FPA工作量评估.xlsx。"""
    logger.info("从 FPA MD 生成 Excel...")

    meta = _load_meta_md(meta_md_path)
    base_formula = meta.get("基准值公式", "")
    workload_formula = meta.get("FPA工作量公式", "J{row}*K{row}")

    # 从 MD 解析 FPA 行
    fpa_rows = []
    with open(fpa_md_path, encoding='utf-8') as f:
        in_table = False
        for line in f:
            line = line.rstrip()
            if "| 序号 | 子系统" in line:
                in_table = True
                continue
            if "|------|" in line and in_table:
                continue
            if in_table:
                cells = parse_md_table_row(line, min_cols=10)
                if cells is not None:
                    fpa_rows.append({
                        "序号": cells[0],
                        "子系统(模块)": cells[1],
                        "资产标识": cells[2],
                        "新增/修改功能点": cells[3],
                        "类型": cells[4],
                        "计算依据归类": cells[5],
                        "计算依据说明": cells[6],
                        "变更状态": cells[7],
                        "调整值": cells[8],
                        "要素数量": cells[9],
                    })

    # 填充模板
    wb = openpyxl.load_workbook(template_path)
    ws = wb['FPA功能点估算']

        # 保存模板第3行的格式作为参照
    tmpl_format = {}
    for col_idx in range(1, 15):
        c = ws.cell(3, col_idx)
        tmpl_format[col_idx] = {
            'font': c.font.copy() if c.font else None,
            'fill': c.fill.copy() if c.fill else None,
            'border': c.border.copy() if c.border else None,
            'number_format': c.number_format,
            'alignment': c.alignment.copy() if c.alignment else None,
        }

    # 清除旧数据（从第3行开始，保留格式）
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        for cell in row:
            cell.value = None

    # 写入数据
    for i, fpa_row in enumerate(fpa_rows):
        excel_row = i + 3
        ws.cell(excel_row, 1, fpa_row["序号"])
        ws.cell(excel_row, 2, fpa_row["子系统(模块)"])
        ws.cell(excel_row, 3, fpa_row["资产标识"])
        ws.cell(excel_row, 4, fpa_row["新增/修改功能点"])
        ws.cell(excel_row, 5, fpa_row["类型"])
        ws.cell(excel_row, 6, fpa_row["计算依据归类"])
        ws.cell(excel_row, 7, fpa_row["计算依据说明"])
        ws.cell(excel_row, 8, fpa_row["变更状态"])
        if base_formula:
            formula = base_formula.replace("E3", f"E{excel_row}")                 .replace("H3", f"H{excel_row}").replace("I3", f"I{excel_row}")                 .replace("J3", f"J{excel_row}").replace("K3", f"K{excel_row}")
            ws.cell(excel_row, 9).value = f"={formula}" if not formula.startswith('=') else formula
        try:
            ws.cell(excel_row, 10, int(fpa_row["调整值"]))
        except (ValueError, TypeError):
            ws.cell(excel_row, 10, fpa_row["调整值"])
        try:
            ws.cell(excel_row, 11, int(fpa_row["要素数量"]))
        except (ValueError, TypeError):
            ws.cell(excel_row, 11, fpa_row["要素数量"])
        if workload_formula:
            formula = workload_formula.replace("J{row}", f"J{excel_row}")                 .replace("K{row}", f"K{excel_row}")                 .replace("J3", f"J{excel_row}").replace("K3", f"K{excel_row}")
            ws.cell(excel_row, 12).value = f"={formula}" if not formula.startswith('=') else formula
        ws.cell(excel_row, 13, "")
        ws.cell(excel_row, 14, "")

        # 从模板第3行复制格式
        for col_idx in range(1, 15):
            c = ws.cell(excel_row, col_idx)
            fmt = tmpl_format.get(col_idx, {})
            if fmt.get('font'):
                c.font = fmt['font']
            if fmt.get('fill'):
                c.fill = fmt['fill']
            if fmt.get('border'):
                c.border = fmt['border']
            if fmt.get('number_format'):
                c.number_format = fmt['number_format']
            if col_idx in (4, 7):
                orig_align = fmt.get('alignment')
                if orig_align:
                    c.alignment = Alignment(
                        wrap_text=True,
                        vertical='center',
                        horizontal=orig_align.horizontal or 'center',
                    )
                else:
                    c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            else:
                if fmt.get('alignment'):
                    c.alignment = fmt['alignment']

    # 更新第1行合计公式
    last_data_row = len(fpa_rows) + 2
    for col_idx in [9, 10, 11, 12, 13]:
        cell = ws.cell(1, col_idx)
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        cell.value = f"=SUM({col_letter}3:{col_letter}{last_data_row})"
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    wb.save(output_path)
    logger.info(f"FPA工作量评估已生成: {output_path} ({len(fpa_rows)} 行)")

    return output_path


# ============================================================
#  项目需求清单.xlsx
# ============================================================

def generate_require_xlsx(
    meta_md_path: str,
    tree_md_path: str,
    template_path: str,
    output_path: str,
    cfp_total: float = 0,
) -> str:
    """生成项目需求清单.xlsx。"""
    logger.info("开始生成项目需求清单.xlsx...")

    meta = _load_meta_md(meta_md_path)
    rows = _load_module_rows(tree_md_path)

    wb = openpyxl.load_workbook(template_path)

    # ====== Sheet 1: 项目信息概览 ======
    ws1 = wb['项目信息概览']

    # 替换标题
    title = meta.get("项目信息概览-标题", "")
    ws1.cell(1, 1, title)

    # 替换数据行（第3行，第2行为表头）
    ws1.cell(3, 2, meta.get("项目信息概览-项目名称", ""))
    ws1.cell(3, 3, meta.get("项目信息概览-子系统名称", ""))
    ws1.cell(3, 4, meta.get("项目信息概览-项目类型", ""))
    ws1.cell(3, 5, meta.get("项目信息概览-所属域", ""))
    ws1.cell(3, 6, meta.get("项目信息概览-所属系统", ""))
    ws1.cell(3, 7, meta.get("项目信息概览-需求部门", ""))
    ws1.cell(3, 8, meta.get("项目信息概览-需求负责人", ""))
    ws1.cell(3, 9, meta.get("项目信息概览-需求负责人联系方式", ""))

    # 送审工作量（保留公式）
    sw_formula = meta.get("项目信息概览-送审工作量", "")
    if sw_formula:
        ws1.cell(3, 10).value = sw_formula if sw_formula.startswith('=') else float(sw_formula)

    # 送审功能点 = CFP 总量
    if cfp_total > 0:
        ws1.cell(3, 11, cfp_total)

    # ====== Sheet 2: 功能清单 ======
    ws2 = wb['功能清单']

    # 替换标题
    fl_title = meta.get("功能清单-标题", "")
    ws2.cell(1, 1, fl_title)

    # 取消所有合并单元格，避免逐行写入时遇到 MergedCell
    for merge_range in list(ws2.merged_cells.ranges):
        ws2.unmerge_cells(str(merge_range))

    project_name = meta.get("功能清单-项目名称", "")
    subsystem = meta.get("功能清单-子系统", "")
    sw_formula2 = meta.get("功能清单-送审工作量", "")

    # 按三级模块去重写入（先存为列表，之后合并单元格）
    data_rows_data = []
    seen_modules = set()
    seq = 0

    from openpyxl.styles import Alignment
    _center = Alignment(horizontal='center', vertical='center')

    for r in rows:
        key = (r["一级模块"], r["二级模块"], r["三级模块"])
        if key not in seen_modules:
            seen_modules.add(key)
            seq += 1
            row_idx = seq + 2
            for col_idx in range(1, 10):
                ws2.cell(row_idx, col_idx).alignment = _center
            ws2.cell(row_idx, 1, seq)
            ws2.cell(row_idx, 2, project_name)
            ws2.cell(row_idx, 3, subsystem)
            ws2.cell(row_idx, 4, r["一级模块"])
            ws2.cell(row_idx, 5, r["二级模块"])
            ws2.cell(row_idx, 6, r["三级模块"])
            ws2.cell(row_idx, 7, r["功能过程类型"])
            if sw_formula2:
                ws2.cell(row_idx, 8).value = sw_formula2 if sw_formula2.startswith('=') else float(sw_formula2)
            if cfp_total > 0:
                ws2.cell(row_idx, 9, cfp_total)
            data_rows_data.append({
                "row": row_idx,
                "project_name": project_name,
                "subsystem": subsystem,
                "module_l1": r["一级模块"],
                "module_l2": r["二级模块"],
            })

    # 合并行1标题居中（A~I列）
    if seq > 0:
        from openpyxl.styles import Alignment
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
        ws2.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')

    # 合并相同值的列：B(项目名称), C(子系统), D(一级模块)
    for col_idx in [2, 3, 4]:
        i = 0
        while i < len(data_rows_data):
            val_key = ["project_name", "subsystem", "module_l1"][col_idx - 2]
            curr_val = data_rows_data[i][val_key]
            j = i
            while j < len(data_rows_data) and data_rows_data[j][val_key] == curr_val:
                j += 1
            count = j - i
            if count > 1:
                ws2.merge_cells(
                    start_row=data_rows_data[i]["row"],
                    start_column=col_idx,
                    end_row=data_rows_data[j - 1]["row"],
                    end_column=col_idx
                )
            i = j

    # 合并送审工作量(H/8列)和送审功能点(I/9列) — 所有行相同值
    if len(data_rows_data) > 1:
        ws2.merge_cells(
            start_row=data_rows_data[0]["row"],
            start_column=8,
            end_row=data_rows_data[-1]["row"],
            end_column=8
        )
        ws2.merge_cells(
            start_row=data_rows_data[0]["row"],
            start_column=9,
            end_row=data_rows_data[-1]["row"],
            end_column=9
        )

    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    wb.save(output_path)
    logger.info(f"项目需求清单已生成: {output_path} ({len(seen_modules)} 模块)")

    return output_path
