"""Write COSMIC decompositions to Excel template."""

import copy
import logging

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from cosmic_tool.models import CosmicItem
from cosmic_tool.config_utils import load_cfp_formula

logger = logging.getLogger('cosmic_tool.excel_writer')


# Read a reference cell style from the template to apply to new cells
_REF_STYLE = Font(name='微软雅黑', size=11, color='FF000000')
_REF_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
_REF_BORDER = Border(
    left=Side(style='thin', color='FF000000'),
    right=Side(style='thin', color='FF000000'),
    top=Side(style='thin', color='FF000000'),
    bottom=Side(style='thin', color='FF000000'),
)
_CENTER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
_LEFT_ALIGN = Alignment(horizontal='left', vertical='center', wrap_text=True)


def _get_ref_style(ws, row: int, col: int) -> dict:
    """Extract style from a reference cell."""
    cell = ws.cell(row=row, column=col)
    return {
        'font': copy.copy(cell.font),
        'fill': copy.copy(cell.fill),
        'alignment': copy.copy(cell.alignment),
        'border': copy.copy(cell.border),
        'number_format': cell.number_format,
    }


def _apply_style(cell, style: dict, skip_fill: bool = False) -> None:
    """Apply style dict to a cell."""
    cell.font = style['font']
    if not skip_fill and style.get('fill'):
        cell.fill = style['fill']
    cell.alignment = style['alignment']
    cell.border = style['border']
    cell.number_format = style['number_format']


def write_to_template(
    template_path: str,
    output_path: str,
    items: list[CosmicItem],
) -> None:
    """Write COSMIC items to Excel template.

    Preserves header rows (1-5) and formatting, fills data starting at row 6.
    Also preserves any existing footer rows from the template.
    """
    wb = openpyxl.load_workbook(template_path)
    ws = wb['2、功能点拆分表']

    # --- Save template row 6 format as reference ---
    tmpl_format_row6 = {}
    for col_idx in range(1, 14):
        tmpl_format_row6[col_idx] = _get_ref_style(ws, 6, col_idx)
    _CFP_FILL_TMPL = copy.copy(ws.cell(row=6, column=13).fill)

# --- Save existing footer notes (rows below header rows, before clearing) ---
    footer_saved = []  # list of (merge_range_string, {col: (value, style_dict)})
# Collect rows from max_row upward that are NOT sample data (footer notes)
    seen_merges = list(ws.merged_cells.ranges)
    for row_num in range(ws.max_row, 5, -1):
        cell_a = ws.cell(row=row_num, column=1).value
        if cell_a and isinstance(cell_a, str) and len(cell_a) > 20:
            # Check if rest of the row is empty (footer note, not data)
            has_data = any(
                ws.cell(row=row_num, column=c).value
                for c in range(5, 14)
            )
            if not has_data:
                # Save merged cells for this row
                row_merges = []
                for mr in seen_merges:
                    if mr.min_row <= row_num <= mr.max_row:
                        row_merges.append(str(mr))
                # Save cell values and styles
                row_vals = {}
                for col in range(1, 14):
                    c = ws.cell(row=row_num, column=col)
                    if c.value is not None or col == 1:
                        row_vals[col] = (c.value, _get_ref_style(ws, row_num, col))
                # Save row heights for merged range
                _saved_heights: dict[int, float] = {}
                for _mr in row_merges:
                    _parts = _mr.split(':')
                    _r1 = int(''.join(c for c in _parts[0] if c.isdigit()))
                    _r2 = int(''.join(c for c in _parts[1] if c.isdigit()))
                    for _r in range(_r1, _r2 + 1):
                        _h = ws.row_dimensions[_r].height
                        if _h is not None:
                            _saved_heights[_r] = _h
                if row_vals:
                    footer_saved.insert(0, (row_merges, row_vals, _saved_heights))
                continue
        # Stop at first non-footer row (contiguous from bottom)
        if footer_saved:
            break

# --- Clear existing data (rows 6+) ---
    # 1. Remove merged cells in data area
    merged_to_remove = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row >= 6:
            merged_to_remove.append(str(mr))
    for mr_str in merged_to_remove:
        ws.unmerge_cells(mr_str)

    # 2. 删除数据区域所有行（第6行起），彻底清除旧数据及格式
    if ws.max_row >= 6:
        ws.delete_rows(6, ws.max_row - 5)

    # --- Flatten all rows ---
    all_rows = []
    for item in items:
        all_rows.extend(item.to_rows())

    # Save source data for debugging
    _save_source_data(all_rows)

    if not all_rows:
        logger.warning("No data rows to write.")
        wb.save(output_path)
        return

    # Use template row 6 format for data cells
    _DATA_STYLE = None
    _CFP_FILL = _CFP_FILL_TMPL

    # Column K (数据属性) should be left-aligned
    # Column H (子过程描述) should be left-aligned

    # --- Write data rows ---
    start_row = 6
    for i, row_data in enumerate(all_rows):
        row_num = start_row + i
        for col_idx in range(1, 14):
            cell = ws.cell(row=row_num, column=col_idx)
            col_letter = get_column_letter(col_idx)
            key_map = {
                1: 'project', 2: 'module_l1', 3: 'module_l2', 4: 'module_l3',
                5: 'user', 6: 'trigger', 7: 'process', 8: 'sub_process',
                9: 'move_type', 10: 'data_group', 11: 'data_attrs',
                12: 'reuse', 13: 'cfp'
            }
            if col_idx == 13:
                # CFP 列用公式（从配置文件读取）
                cfp_formula = load_cfp_formula()
                cell.value = '=' + cfp_formula.replace('{row}', str(row_num))
            else:
                cell.value = row_data.get(key_map[col_idx], '')

            # Apply template row 6 format（跳过 fill，避免空单元格带底色）
            _apply_style(cell, tmpl_format_row6[col_idx], skip_fill=True)

            # CFP 列：绿色底色 + 分数格式（复用=1/3时显示分数）
            if col_idx == 13:
                cell.fill = _CFP_FILL
                cell.number_format = '[=1]0;[=0]0;# ?/?'

            # Long text columns: left-align
            if col_idx in (8, 10, 11):
                cell.alignment = _LEFT_ALIGN

    # --- Apply merged cells for repeating values ---
    total_rows = len(all_rows)

    # Merge project name (column A) - all rows
    if total_rows > 1:
        ws.merge_cells(start_row=start_row, start_column=1,
                       end_row=start_row + total_rows - 1, end_column=1)

    # Merge L1 module (column B)
    _merge_column_groups(ws, start_row, total_rows, 2, all_rows, 'module_l1')

    # Merge L2 module (column C)
    _merge_column_groups(ws, start_row, total_rows, 3, all_rows, 'module_l2')

    # Merge L3 module (column D)
    _merge_column_groups(ws, start_row, total_rows, 4, all_rows, 'module_l3')

    # Merge user (column E), trigger (column F), process name (column G)
    _merge_column_groups(ws, start_row, total_rows, 5, all_rows, 'process')
    _merge_column_groups(ws, start_row, total_rows, 6, all_rows, 'process')
    _merge_column_groups(ws, start_row, total_rows, 7, all_rows, 'process')

    # Keep the already-set data validations for reuse dropdown
    # Add new data validation for L column (复用度) - dropdown
    _add_reuse_validation(ws, start_row, total_rows)

    # --- Restore saved footer notes (from template) below the new data ---
    for i, (merges, vals, row_heights) in enumerate(footer_saved):
            note_row = start_row + total_rows + i
            for mr_str in merges:
                # Translate merged cell range to the new row number
                parts = mr_str.split(':')
                old_min = int(''.join(c for c in parts[0] if c.isdigit()))
                old_max = int(''.join(c for c in parts[1] if c.isdigit()))
                span = old_max - old_min
                new_min_str = parts[0].rstrip('0123456789') + str(note_row)
                new_max_str = parts[1].rstrip('0123456789') + str(note_row + span)
                ws.merge_cells(f'{new_min_str}:{new_max_str}')
            for col, (val, style) in vals.items():
                cell = ws.cell(row=note_row, column=col)
                cell.value = val
                _apply_style(cell, style)
            # 合并单元格后补全四周边框（WPS 需要四个角都设才渲染完整）
            for mr_str in merges:
                parts = mr_str.split(':')
                old_min = int(''.join(c for c in parts[0] if c.isdigit()))
                old_max = int(''.join(c for c in parts[1] if c.isdigit()))
                span = old_max - old_min
                new_min_row = note_row
                new_max_row = note_row + span
                new_min_col_str = parts[0].rstrip('0123456789')
                new_max_col_str = parts[1].rstrip('0123456789')
                import openpyxl.utils as _ou
                new_min_col = _ou.column_index_from_string(new_min_col_str)
                new_max_col = _ou.column_index_from_string(new_max_col_str)
                # 从左上角单元格获取边框样式（强制黑色，WPS 对 auto 颜色渲染为浅灰）
                _tl = ws.cell(new_min_row, new_min_col)
                _b = _tl.border
                _sides = {}
                _black = 'FF000000'
                for _side_name, _attr in [('left', 'left'), ('right', 'right'), ('top', 'top'), ('bottom', 'bottom')]:
                    _s = getattr(_b, _attr)
                    if _s and _s.style:
                        _sides[_side_name] = Side(style=_s.style, color=_black)
                if _sides:
                    ws.cell(new_min_row, new_min_col).border = Border(
                        left=_sides.get('left'), right=_sides.get('right'),
                        top=_sides.get('top'), bottom=_sides.get('bottom'))
                    if new_max_col > new_min_col:
                        ws.cell(new_min_row, new_max_col).border = Border(
                            left=None, right=_sides.get('right'),
                            top=_sides.get('top'), bottom=None)
                        # 顶行中间单元格也设上边框（WPS 需要顶行全部有 top=thin 才渲染实线）
                        for _mid_col in range(new_min_col + 1, new_max_col):
                            ws.cell(new_min_row, _mid_col).border = Border(
                                left=None, right=None,
                                top=_sides.get('top'), bottom=None)
                    if new_max_row > new_min_row:
                        ws.cell(new_max_row, new_min_col).border = Border(
                            left=_sides.get('left'), right=None,
                            top=None, bottom=_sides.get('bottom'))
                    if new_max_col > new_min_col and new_max_row > new_min_row:
                        ws.cell(new_max_row, new_max_col).border = Border(
                            left=None, right=_sides.get('right'),
                            top=None, bottom=_sides.get('bottom'))
                        # 底行中间单元格也设下边框
                        for _mid_col in range(new_min_col + 1, new_max_col):
                            ws.cell(new_max_row, _mid_col).border = Border(
                                left=None, right=None,
                                top=None, bottom=_sides.get('bottom'))
            # 恢复合并单元格的行高
            for _mr_str in merges:
                _parts = _mr_str.split(':')
                _r1 = int(''.join(c for c in _parts[0] if c.isdigit()))
                _r2 = int(''.join(c for c in _parts[1] if c.isdigit()))
                for _off in range(_r2 - _r1 + 1):
                    _old_r = _r1 + _off
                    if _old_r in row_heights:
                        ws.row_dimensions[note_row + _off].height = row_heights[_old_r]
            logger.debug(f"Restored footer note at row {note_row}")



    # 合并后补回边框
    for row_num in range(start_row, start_row + total_rows):
        for col_idx in range(1, 14):
            ws.cell(row=row_num, column=col_idx).border = tmpl_format_row6[col_idx]['border']


    # --- Apply warning indicators (after merges, so they don't get overwritten) ---
    from cosmic_tool.config_utils import load_cosmic_warn_marker
    _warn_enabled = load_cosmic_warn_marker()
    if not _warn_enabled:
        _warn_marker_disabled_skip = True
    else:
        _warn_marker_disabled_skip = False

    _WARN_FILL = PatternFill(start_color='FFFFF2CC', end_color='FFFFF2CC', fill_type='solid')
    for i, row_data in enumerate(all_rows):
        row_num = start_row + i
        row_warnings = row_data.get('warnings', [])
        if _warn_marker_disabled_skip:
            row_warnings = []
        move_flagged = row_data.get('move_type_flagged', False)

        if row_warnings:
            # Yellow highlight on the first visible data cell (col 8, sub_process)
            ws.cell(row=row_num, column=8).fill = _WARN_FILL
            # Excel comment with warning text
            from openpyxl.comments import Comment
            ws.cell(row=row_num, column=8).comment = Comment(
                "\n".join(f"⚠ {w}" for w in row_warnings), "COSMIC Tool"
            )

        if move_flagged:
            # Yellow on fuzzy-matched move_type cell (col 9)
            ws.cell(row=row_num, column=9).fill = _WARN_FILL

    # --- Auto-fit column widths and row heights ---
    _auto_fit(ws, start_row, start_row + total_rows - 1)

    wb.save(output_path)
    logger.info(f"Written {total_rows} rows to {output_path}")
    if footer_saved:
        logger.info(f"Restored {len(footer_saved)} footer note(s) from template")


def _merge_column_groups(ws, start_row, total_rows, col, all_rows, key):
    """Merge cells in a column for groups of consecutive same values."""
    i = 0
    while i < total_rows:
        val = all_rows[i].get(key, '')
        j = i
        while j < total_rows and all_rows[j].get(key, '') == val:
            j += 1
        count = j - i
        if count > 1:
            ws.merge_cells(
                start_row=start_row + i,
                start_column=col,
                end_row=start_row + j - 1,
                end_column=col
            )
        i = j


def _add_reuse_validation(ws, start_row, total_rows):
    """Add dropdown validation for 复用度 column (L)."""
    from openpyxl.worksheet.datavalidation import DataValidation
    if total_rows > 0:
        dv = DataValidation(
            type="list",
            formula1='"新增,复用,利旧"',
            allow_blank=True
        )
        dv.sqref = f"L{start_row}:L{start_row + total_rows - 1}"
        ws.add_data_validation(dv)


def _save_source_data(rows: list[dict]) -> None:
    """Save flattened source data to log/source_data/ for debugging."""
    import json, os
    from datetime import datetime
    base_log = os.environ.get('COSMIC_LOG_DIR', '') or os.path.join(
        os.path.dirname(os.path.dirname(__file__)), 'log'
    )
    log_dir = os.path.join(base_log, 'source_data')
    os.makedirs(log_dir, exist_ok=True)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filepath = os.path.join(log_dir, f'{timestamp}_excel_source.json')
    # Remove non-serializable keys before saving
    clean = [{k: v for k, v in row.items() if k in (
        'project', 'module_l1', 'module_l2', 'module_l3',
        'user', 'trigger', 'process', 'sub_process',
        'move_type', 'data_group', 'data_attrs', 'reuse', 'cfp'
    )} for row in rows]
    with open(filepath, 'w', encoding='utf-8') as f:
        json.dump(clean, f, ensure_ascii=False, indent=2)
    logger.info(f"源数据已保存: {filepath}")


def _auto_fit(ws, start_row: int, end_row: int) -> None:
    """Auto-fit row heights based on content."""
    from openpyxl.utils import get_column_letter
    import math
    for r in range(start_row, end_row + 1):
        max_lines = 1
        for col_idx in range(1, 14):
            cell = ws.cell(row=r, column=col_idx)
            if cell.value:
                text = str(cell.value)
                for line in text.split('\n'):
                    visual_len = sum(2 if ord(c) > 127 else 1 for c in line)
                    col_w = ws.column_dimensions[get_column_letter(col_idx)].width or 10
                    wraps = math.ceil(visual_len / max(col_w, 1))
                    max_lines = max(max_lines, wraps)
        ws.row_dimensions[r].height = max(15, min(max_lines * 18, 200))


def write_environment_sheet(
    template_path: str,
    output_path: str,
    project_name: str,
    target: str,
    necessity: str
) -> None:
    """Write construction goals and necessity to the environment sheet."""
    wb = openpyxl.load_workbook(template_path)
    ws = wb['1、环境图']

    # Find and update the target/necessity cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=27):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if '建设目标' in cell.value and len(cell.value) < 20:
                    # The cell to the right is where the target text goes
                    target_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    # Find the merged range containing this cell
                    # Actually just write to the cell - it's part of a merged range
                    if target:
                        target_cell.value = target
                if '建设必要性' in cell.value and len(cell.value) < 20:
                    necessity_cell = ws.cell(row=cell.row, column=cell.column + 1)
                    if necessity:
                        necessity_cell.value = necessity

    wb.save(output_path)
    logger.info(f"Updated 1、环境图 sheet with project info")


def copy_template_sheets(
    template_path: str,
    output_path: str,
    func_point_sheet_only: bool = False
) -> None:
    """Copy the complete template preserving all sheets."""
    wb = openpyxl.load_workbook(template_path)
    wb.save(output_path)
    logger.info(f"Template copied to {output_path}")
