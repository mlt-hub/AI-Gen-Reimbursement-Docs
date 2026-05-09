"""Excel 功能清单解析器 — 读取功能清单.xlsx，生成中间 MD 文件"""

import logging
import os
import re
from datetime import datetime

import openpyxl

logger = logging.getLogger('cosmic_tool.excel_source')


def _cell_val(cell) -> str:
    """读取单元格值，None 转空字符串。"""
    return str(cell).strip() if cell is not None else ""


def _resolve_inherited_rows(ws):
    """将合并单元格/空单元格的值继承上一行同列的值。返回二维列表。"""
    rows = []
    prev_vals = [""] * ws.max_column
    for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过表头
        vals = [_cell_val(c) for c in row]
        for i in range(len(vals)):
            if not vals[i] and prev_vals[i]:
                vals[i] = prev_vals[i]
        prev_vals = vals
        rows.append(vals)
    return rows


def _key_value_sheet(ws):
    """读取 key-value 格式 sheet（2列：项目 | 内容），返回 dict。"""
    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = _cell_val(row[0])
        val = _cell_val(row[1]) if len(row) > 1 else ""
        if key:
            data[key] = val
    return data


def _calc_cfp_limit(wb) -> float:
    """从 sheet 5 读取 CFP 限制值（如果含公式则计算）。"""
    ws = wb['5、预估工作量-元数据录入']
    b1 = _cell_val(ws.cell(2, 2).value)  # 预估工作量
    b2 = _cell_val(ws.cell(3, 2).value)  # FPA核减后的工作量
    b3 = _cell_val(ws.cell(4, 2).value)  # CFP数量限制倍数
    b4 = _cell_val(ws.cell(5, 2).value)  # CFP数量限制值（公式）
    try:
        limit = float(b2) * float(b3) if b4.startswith('=') else float(b4)
    except (ValueError, ZeroDivisionError):
        limit = 0
    return limit


def generate_md_files(excel_path: str, output_dir: str = "") -> dict:
    """读取功能清单.xlsx，生成功能清单模块树.md 和 文档元数据模板.md。

    Args:
        excel_path: 功能清单.xlsx 路径
        output_dir: 输出目录（空则取 excel 所在目录）

    Returns:
        {"module_tree_md": 路径, "doc_meta_md": 路径}
    """
    if not output_dir:
        output_dir = os.path.dirname(os.path.abspath(excel_path))
    os.makedirs(output_dir, exist_ok=True)

    wb = openpyxl.load_workbook(excel_path, data_only=False)

    # ========== 解析各个 sheet ==========

    # 1、工单需求-内容录入
    ws1 = wb['1、工单需求-内容录入']
    project_info = _key_value_sheet(ws1)

    # 2、功能清单-内容录入 — 模块树 + 功能过程
    ws2 = wb['2、功能清单-内容录入']
    func_rows = _resolve_inherited_rows(ws2)

    # 3、FPA工作量评估-元数据录入
    ws3 = wb['3、FPA工作量评估-元数据录入']
    fpa_meta = _key_value_sheet(ws3)

    # 4、项目需求说明书-元数据录入
    ws4 = wb['4、项目需求说明书-元数据录入']
    docx_meta = _key_value_sheet(ws4)

    # 5、预估工作量-元数据录入
    ws5 = wb['5、预估工作量-元数据录入']
    workload_meta = _key_value_sheet(ws5)

    # 6、项目功能点拆分表-元数据录入
    ws6 = wb['6、项目功能点拆分表-元数据录入']
    cosmic_meta = _key_value_sheet(ws6)

    # 7、项目需求清单-元数据录入
    ws7 = wb['7、项目需求清单-元数据录入']
    require_meta = _key_value_sheet(ws7)

    wb.close()

    # data_only=True 还原公式单元格的计算值
    wb_val = openpyxl.load_workbook(excel_path, data_only=True)
    for row in wb_val['5、预估工作量-元数据录入'].iter_rows(min_row=2, values_only=True):
        k, v = row[0], row[1]
        if k and str(v).strip():
            wk = str(k).strip()
            wv = str(v).strip()
            if wk in workload_meta and workload_meta[wk].startswith('='):
                workload_meta[wk] = wv
    for row in wb_val['7、项目需求清单-元数据录入'].iter_rows(min_row=2, values_only=True):
        k, v = row[0], row[1]
        if k and str(v).strip():
            wk = str(k).strip()
            wv = str(v).strip()
            if wk in require_meta and require_meta[wk].startswith('='):
                require_meta[wk] = wv
    wb_val.close()

    # 9、测试元数据自动统计（从解析结果计算唯一值，替代 COUNTA 公式）
    stats_meta = {
        "入口（个数）": str(len({r[0] for r in func_rows if r[0]})),
        "一级模块（个数）": str(len({r[1] for r in func_rows if r[1]})),
        "二级模块（个数）": str(len({r[2] for r in func_rows if r[2]})),
        "三级模块（个数）": str(len({r[3] for r in func_rows if r[3]})),
        "功能过程（个数）": str(len({r[6] for r in func_rows if r[6]})),
    }

    # ========== 生成 功能清单模块树.md ==========

    md_tree_path = os.path.join(output_dir, '功能清单模块树.md')
    with open(md_tree_path, 'w', encoding='utf-8') as f:
        f.write("# 功能清单模块树\n\n")
        f.write(f"**来源文件**：{os.path.basename(excel_path)}\n")
        f.write(f"**生成日期**：{datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n")
        f.write("| 入口 | 一级模块 | 二级模块 | 三级模块 | 客户端类型 | 三级模块整体功能描述 | 功能过程 | 功能过程类型 | 功能过程描述 |\n")
        f.write("|------|---------|---------|---------|-----------|-------------------|---------|-----------|-------------|\n")
        for row in func_rows:
            # row: [入口, 一级模块, 二级模块, 三级模块, 客户端类型, 三级模块整体功能描述, 功能过程, 功能过程类型, 功能过程描述]
            vals = [cell.replace('|', '\\|').replace('\n', ' ') for cell in row[:9]]
            line = " | ".join(vals)
            f.write(f"| {line} |\n")

    logger.info(f"功能清单模块树已生成: {md_tree_path}")

    # ========== 生成 文档元数据模板.md ==========

    md_meta_path = os.path.join(output_dir, '文档元数据模板.md')
    with open(md_meta_path, 'w', encoding='utf-8') as f:
        f.write("# 文档元数据\n\n")
        f.write(f"**来源文件**：{os.path.basename(excel_path)}\n")
        f.write(f"**生成日期**：{datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n")

        # 按 sheet 分组写入
        sections = [
            ("1、工单需求-内容录入", project_info),
            ("3、FPA工作量评估-元数据录入", fpa_meta),
            ("4、项目需求说明书-元数据录入", docx_meta),
            ("5、预估工作量-元数据录入", workload_meta),
            ("6、项目功能点拆分表-元数据录入", cosmic_meta),
            ("7、项目需求清单-元数据录入", require_meta),
            ("9、测试元数据自动统计", stats_meta),
        ]

        for sheet_name, kv in sections:
            f.write(f"## {sheet_name}\n\n")
            f.write("| 项目 | 内容 |\n")
            f.write("|------|------|\n")
            for key, val in kv.items():
                # 替换 【占位符】 为实际值
                v = val
                for ph, src in [("${工单编号}", project_info.get("工单编号", "")),
                                 ("${工单名称}", project_info.get("工单标题", "")),
                                 ("${工单标题}", project_info.get("工单标题", "")),
                                 ("${工单内容}", project_info.get("工单内容", "")),
                                 ("${子系统（模块）}", fpa_meta.get("子系统（模块）", ""))]:
                    v = v.replace(ph, src)
                val_escaped = v.replace('|', '\\|')
                f.write(f"| {key} | {val_escaped} |\n")
            f.write("\n")

        # 附加：CFP 限制值
        cfp_limit = _calc_cfp_limit(wb)
        f.write("## 计算值\n\n")
        f.write("| 项目 | 值 |\n")
        f.write("|------|-----|\n")
        f.write(f"| CFP数量限制值 | {cfp_limit} |\n")
        f.write(f"| CFP数量限制倍数 | {workload_meta.get('CFP数量限制倍数', '1.5')} |\n")
        f.write("\n")

    logger.info(f"文档元数据已生成: {md_meta_path}")

    return {"module_tree_md": md_tree_path, "doc_meta_md": md_meta_path}


def read_template_config(excel_path: str) -> dict[str, str]:
    """读取 功能清单-录入-模板.xlsx → sheet 8，返回模板名→路径的映射。

    返回示例:
        {"FPA工作量评估-模板": "data/templates/FPA工作量评估-模板.xlsx",
         "项目需求说明书-模板": "data/templates/项目需求说明书-模板.docx",
         "项目功能点拆分表-模板": "data/templates/项目功能点拆分表-模板.xlsx",
         "项目需求清单-模板": "data/templates/项目需求清单-模板.xlsx"}
    """
    result: dict[str, str] = {}
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        if '8、各文档-模板路径录入' in wb.sheetnames:
            ws = wb['8、各文档-模板路径录入']
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row,
                                    min_col=1, max_col=2, values_only=True):
                name, path = row
                if name:
                    val = str(path).strip() if path else ""
                    if val:
                        result[str(name).strip()] = val.replace('/', os.sep).replace('\\', os.sep)
        wb.close()
    except Exception as e:
        logger.warning(f"读取模板配置失败: {e}")
    return result


def read_fpa_xlsx_sum(fpa_xlsx_path: str) -> float:
    """读取 FPA工作量评估.xlsx 中核减后工作量列的求和。"""
    try:
        wb = openpyxl.load_workbook(fpa_xlsx_path, data_only=True)
        ws = wb['FPA功能点估算']
        total = 0.0
        for row in ws.iter_rows(min_row=3, values_only=True):
            val = row[12]  # M列（核减后工作量），0-based index
            if val is not None:
                try:
                    total += float(val)
                except (ValueError, TypeError):
                    pass
        wb.close()
        logger.info(f"FPA核减后工作量求和: {total}")
        return total
    except Exception as e:
        logger.warning(f"读取FPA求和失败: {e}")
        return 0.0


def verify_module_tree_stats(tree_md_path: str, meta_md_path: str) -> bool:
    """验证功能清单模块树.md 的统计信息与文档元数据模板.md 中的期望值一致。

    读取模块树 MD 表格，统计入口/L1/L2/L3/功能过程数，
    与文档元数据中 ## 9、测试元数据自动统计 进行对比。

    Returns:
        True 全部通过, False 有差异
    """
    from cosmic_tool.md_table import parse_md_table_row

    # 从模块树 MD 统计
    entries: set[str] = set()
    l1s: set[str] = set()
    l2s: set[str] = set()
    l3s: set[str] = set()
    procs: set[str] = set()

    with open(tree_md_path, encoding='utf-8') as f:
        in_table = False
        for line in f:
            if "| 入口 | 一级模块" in line:
                in_table = True
                continue
            if "|------" in line and in_table:
                continue
            if in_table:
                cells = parse_md_table_row(line, min_cols=9)
                if cells is not None:
                    if cells[0]:
                        entries.add(cells[0])
                    if cells[1]:
                        l1s.add(cells[1])
                    if cells[2]:
                        l2s.add(cells[2])
                    if cells[3]:
                        l3s.add(cells[3])
                    if cells[6]:
                        procs.add(cells[6])

    actual = {
        "入口（个数）": len(entries),
        "一级模块（个数）": len(l1s),
        "二级模块（个数）": len(l2s),
        "三级模块（个数）": len(l3s),
        "功能过程（个数）": len(procs),
    }

    # 从文档元数据读取期望值
    expected: dict[str, int] = {}
    in_stats = False
    with open(meta_md_path, encoding='utf-8') as f:
        for line in f:
            line = line.rstrip()
            if line.startswith("## 9、测试元数据自动统计"):
                in_stats = True
                continue
            if in_stats:
                if line.startswith("## "):
                    break
                if line.startswith("|") and not line.startswith("|--"):
                    cells = [c.strip() for c in line.split("|")]
                    cells = [c for c in cells if c]
                    if len(cells) >= 2:
                        try:
                            expected[cells[0]] = int(cells[1])
                        except ValueError:
                            pass

    # 对比输出
    all_ok = True
    logger.info("═══ 模块树统计验证 ═══")
    for key in ["入口（个数）", "一级模块（个数）", "二级模块（个数）",
                 "三级模块（个数）", "功能过程（个数）"]:
        exp = expected.get(key)
        act = actual[key]
        if exp is not None:
            status = "✓" if act == exp else "✗"
            all_ok = all_ok and (act == exp)
            logger.info(f"  {status} {key}: 期望={exp}, 实际={act}")
        else:
            logger.info(f"  ? {key}: 期望=未配置, 实际={act}")

    if all_ok:
        logger.info("═══ 全部通过 ═══")
    else:
        logger.warning("═══ 存在差异，请检查数据 ═══")
    return all_ok


def replace_placeholders(text: str, project_info: dict, fpa_meta: dict) -> str:
    """替换 【占位符】 为实际值。"""
    placeholders = {
        "${工单编号}": project_info.get("工单编号", ""),
        "${工单名称}": project_info.get("工单标题", ""),
        "${工单标题}": project_info.get("工单标题", ""),
        "${工单内容}": project_info.get("工单内容", ""),
        "${子系统（模块）}": fpa_meta.get("子系统（模块）", ""),
    }
    for ph, val in placeholders.items():
        text = text.replace(ph, val)
    return text


def strip_ai_marker(text: str) -> tuple[str, bool]:
    """判断是否含 #AI生成# 或 #AI生成-XXX# 标记，去掉标记返回纯净文本。"""
    if text.startswith("#AI生成#"):
        return text[len("#AI生成#"):], True
    m = re.match(r'^#AI生成-(.+?)#\s*(.*)', text)
    if m:
        # #AI生成-工单内容# → 保留 "基于工单内容" 作为提示词
        hint = m.group(1)
        rest = m.group(2)
        prompt = f"基于{hint}" if not rest else rest
        return prompt, True
    return text, False
