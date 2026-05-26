"""
Web UI for ai-gen-reimbursement-docs.
启动: python -m uvicorn web_app.server:app --host 0.0.0.0 --port 8080
"""

import asyncio
import contextvars
import json
import logging
import os
import queue
import shutil
import tempfile
import threading
import uuid
from datetime import datetime
from pathlib import Path

from ai_gen_reimbursement_docs.exceptions import CancelledError
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles

# ── 日志隔离 ──────────────────────────────────────────────

session_var: contextvars.ContextVar[str | None] = contextvars.ContextVar(
    "session_id", default=None
)
session_queues: dict[str, queue.Queue] = {}
session_outputs: dict[str, Path] = {}  # session_id →交付物目录（本机模式）
session_zips: dict[str, Path] = {}  # session_id → zip 文件路径（远程服务模式）
session_dirs: dict[str, Path] = {}  # session_id → 临时工作目录（远程服务模式）
session_cancelled: dict[str, bool] = {}  # session_id → 是否已取消
session_input_events: dict[str, threading.Event] = {}  # session_id → 等待用户输入的事件
session_input_results: dict[str, dict] = {}  # session_id → 用户输入的结果


def emit_session_event(data: dict) -> None:
    """向当前 session 的 SSE 队列发送结构化事件。data 必须含 'type' 字段。"""
    sid = session_var.get()
    if not sid:
        return
    q = session_queues.get(sid)
    if q:
        q.put(json.dumps(data, ensure_ascii=False))


def wait_for_fpa_input(default_fpa: float) -> float:
    """在 pipeline 线程中调用，通过 SSE 通知前端弹输入框，等待用户确认送审工作量。"""
    sid = session_var.get()
    if not sid:
        return default_fpa

    event = threading.Event()
    session_input_events[sid] = event

    emit_session_event({
        "type": "prompt",
        "field": "fpa_reduced",
        "default": default_fpa,
        "msg": f"请输入送审工作量（直接确认则使用默认值：{default_fpa}）",
    })

    event.wait(timeout=1800)
    session_input_events.pop(sid, None)
    result = session_input_results.pop(sid, {})
    return float(result.get("fpa_reduced", default_fpa))


class SessionHandler(logging.Handler):
    """将日志路由到对应 session 的队列，实现多会话日志隔离。"""

    def emit(self, record):
        sid = session_var.get()
        if sid and sid in session_queues:
            msg = json.dumps(
                {
                    "type": "log",
                    "level": record.levelname,
                    "msg": self.format(record),
                    "time": datetime.fromtimestamp(record.created).strftime(
                        "%H:%M:%S"
                    ),
                },
                ensure_ascii=False,
            )
            try:
                session_queues[sid].put_nowait(msg)
            except queue.Full:
                pass  # 队列满时丢弃旧日志，优先保留关键信息


# 确保父 logger 级别足够低，子 logger 的 INFO/DEBUG 能传播过来
_parent = logging.getLogger("ai_gen_reimbursement_docs")
_parent.setLevel(logging.DEBUG)

from ai_gen_reimbursement_docs.cli.logging import PathShortener

_handler = SessionHandler()
_handler.setLevel(logging.DEBUG)
_handler.setFormatter(logging.Formatter("%(message)s"))
_handler.addFilter(PathShortener())
_parent.addHandler(_handler)

# 添加全局日志 handler（与 CLI 一致）
from ai_gen_reimbursement_docs.cli.logging import init_global_logging
init_global_logging()

# ── 常量 ──────────────────────────────────────────────────

BASE_DIR = Path(__file__).resolve().parent.parent
TEMPLATE_DIR = BASE_DIR / "data" / "out_templates"

_log = logging.getLogger("ai_gen_reimbursement_docs")
_log.info("[Web UI] AI生成项目报账文档 v%s（FastAPI 服务启动）",
          __import__('tomllib').load(open(BASE_DIR / 'pyproject.toml', 'rb'))['project']['version'])

MODE_INFO: dict[str, dict[str, str]] = {
    "from-excel-gen-all": {"label": "gen-all → 全套报账文档", "desc": "生成全套报账文档"},
    "from-excel-gen-basedata": {
        "label": "gen-basedata → 基础数据：模块树+元数据",
        "desc": "仅解析 功能清单Excel 生成中间 MD",
    },
    "from-excel-gen-fpa": {"label": "gen-fpa → FPA工作量评估", "desc": "生成FPA工作量评估.xlsx"},
    "from-excel-gen-spec": {"label": "gen-spec → 项目需求说明书", "desc": "生成项目需求说明书.docx"},
    "from-excel-gen-cosmic": {
        "label": "gen-cosmic → 项目功能点拆分表",
        "desc": "生成项目功能点拆分表.xlsx",
    },
    "from-excel-gen-list": {"label": "gen-list → 项目需求清单", "desc": "生成项目需求清单.xlsx"},
    
}

_MODE_MAP: dict[str, str] = {
    "from-excel-gen-all": "gen-all",
    "from-excel-gen-basedata": "gen-basedata",
    "from-excel-gen-fpa": "gen-fpa",
    "from-excel-gen-cosmic": "gen-cosmic",
    "from-excel-gen-list": "gen-list",
    "from-excel-gen-spec": "gen-spec",
}

def _spa_index():
    """SPA 入口：返回 Vite 构建产物。"""
    dist_index = Path(__file__).parent / "static" / "dist" / "index.html"
    if dist_index.exists():
        return HTMLResponse(dist_index.read_text(encoding="utf-8"))
    return HTMLResponse("<html><body>前端未构建，请运行 npm run dev 或 npm run build</body></html>")


# ── FastAPI App ───────────────────────────────────────────

app = FastAPI(title="AI生成项目报账文档")


@app.on_event("shutdown")
async def _on_shutdown():
    """服务关闭时标记所有 session 为已取消，避免后台 AI 调用继续重试。"""
    for sid in list(session_queues.keys()):
        session_cancelled[sid] = True

# 静态文件：Vite 构建产物
_dist_dir = Path(__file__).parent / "static" / "dist"
if _dist_dir.exists():
    app.mount("/static/dist", StaticFiles(directory=str(_dist_dir)), name="static_dist")
# 保留旧 static/ 挂载以支持旧文件和资源
_static_dir = Path(__file__).parent / "static"
_static_dir.mkdir(parents=True, exist_ok=True)
app.mount("/static", StaticFiles(directory=str(_static_dir)), name="static")


@app.get("/")
async def index():
    """SPA 入口：生产环境返回 dist/index.html，开发环境回退 static/index.html。"""
    return _spa_index()


@app.get("/config")
async def config_page():
    return _spa_index()


@app.get("/prompt-debug")
async def prompt_debug():
    return _spa_index()


@app.get("/api/is-local")
async def is_local(request: Request):
    """判断请求是否来自本机。"""
    host = request.client.host if request.client else ""
    return {"local": host in ("127.0.0.1", "::1", "localhost")}


@app.get("/api/log-level")
async def get_log_level():
    """返回当前日志级别。"""
    from ai_gen_reimbursement_docs.config_utils import load_log_level
    return {"level": load_log_level()}


@app.post("/api/log-level")
async def set_log_level(data: dict):
    """运行时设置日志级别。"""
    level = data.get("level", "INFO").strip().upper()
    if level not in ("DEBUG", "INFO", "WARNING", "ERROR"):
        raise HTTPException(400, f"无效的日志级别: {level}")
    lv = getattr(logging, level, logging.INFO)
    # 仅更新 SSE SessionHandler 级别（文件日志始终 DEBUG）
    _handler.setLevel(lv)
    return {"ok": True, "level": level}


@app.get("/api/modes")
async def get_modes():
    """返回操作模式列表，供前端动态渲染下拉框。"""
    return MODE_INFO


@app.post("/api/cancel/{session_id}")
async def cancel_session(session_id: str):
    """停止指定 session 的执行。"""
    session_cancelled[session_id] = True
    ev = session_input_events.get(session_id)
    if ev:
        ev.set()  # 唤醒等待用户输入的 pipeline 线程
    return {"ok": True}


@app.post("/api/continue/{session_id}")
async def api_continue(session_id: str, data: dict):
    """接收前端交互输入（送审工作量），唤醒等待中的 pipeline。"""
    ev = session_input_events.get(session_id)
    if not ev:
        raise HTTPException(404, "会话不存在或无需输入")
    session_input_results[session_id] = data
    ev.set()
    return {"ok": True}


def check_cancelled():
    """检查当前 session 是否已被取消，若是则抛出 CancelledError。"""
    from ai_gen_reimbursement_docs.exceptions import CancelledError
    sid = session_var.get()
    if sid and session_cancelled.get(sid):
        raise CancelledError("任务已被用户停止")


@app.get("/api/version")
async def get_version():
    """返回当前版本号（从 pyproject.toml 读取）。"""
    try:
        import tomllib
        toml = BASE_DIR / "pyproject.toml"
        if toml.exists():
            return {"version": tomllib.load(toml.open("rb"))["project"]["version"]}
    except Exception:
        pass
    return {"version": "unknown"}


# ── 系统配置 ──────────────────────────────────────────────


def _config_dir() -> Path:
    return Path(os.path.expanduser("~")) / ".ai-gen-reimbursement-docs"


def _read_config() -> dict:
    """读取所有配置文件，返回合并后的 dict。"""
    cfg_dir = _config_dir()
    result: dict = {"_env": {}, "_system": {}, "_biz": {}}

    env_path = cfg_dir / ".env"
    if env_path.exists():
        env: dict[str, str] = {}
        for line in env_path.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
        result["_env"] = env

    for key, filename in [
        ("_system", "system_config.yaml"),
        ("_biz", "business_rules.yaml"),
    ]:
        path = cfg_dir / filename
        if path.exists():
            try:
                import yaml
                result[key] = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
            except Exception:
                pass

    return result


@app.get("/api/config")
async def get_config():
    return _read_config()


@app.post("/api/config")
async def save_config(data: dict):
    """保存配置。data 含 _env / _system / _biz 三个 key。"""
    cfg_dir = _config_dir()
    cfg_dir.mkdir(parents=True, exist_ok=True)

    if "_env" in data and data["_env"]:
        lines = []
        for k, v in data["_env"].items():
            lines.append(f"{k}={v}")
        (cfg_dir / ".env").write_text("\n".join(lines) + "\n", encoding="utf-8")

    for key, filename in [("_system", "system_config.yaml"), ("_biz", "business_rules.yaml")]:
        if key in data and data[key]:
            import yaml
            path = cfg_dir / filename
            path.write_text(yaml.dump(data[key], allow_unicode=True, default_flow_style=False), encoding="utf-8")

    return {"ok": True}


# ── 提示词调试 ────────────────────────────────────────────


@app.post("/api/test-prompt")
async def test_prompt(data: dict):
    """提交系统提示词和用户提示词，返回 AI 生成结果。"""
    system_prompt = data.get("system_prompt", "").strip()
    user_prompt = data.get("user_prompt", "").strip()
    if not user_prompt and not system_prompt:
        raise HTTPException(400, "系统提示词和用户提示词不能同时为空")

    from ai_gen_reimbursement_docs.config_utils import (
        load_api_key, load_base_url, load_model_name,
    )
    from ai_gen_reimbursement_docs.llm_client import call_llm

    api_key = data.get("api_key", "").strip() or load_api_key()
    model = data.get("model", "").strip() or load_model_name()
    base_url = data.get("base_url", "").strip() or load_base_url()

    if not api_key:
        raise HTTPException(400, "未配置 API Key")

    if api_key:
        os.environ["ANTHROPIC_API_KEY"] = api_key
    if base_url:
        os.environ["ANTHROPIC_BASE_URL"] = base_url

    try:
        result, thinking = call_llm(
            prompt=user_prompt,
            system=system_prompt,
            api_key=api_key,
            model=model,
            base_url=base_url,
            tag="prompt_debug",
            return_thinking=True,
        )
        return {
            "result": result,
            "thinking": thinking,
            "system_prompt": system_prompt,
            "user_prompt": user_prompt,
        }
    except Exception as e:
        raise HTTPException(500, f"AI 调用失败: {e}")


# ── 本机模式 ──────────────────────────────────────────────


@app.post("/api/play-notify")
async def play_notify(request: Request):
    """播放完成提示音（仅本机模式生效）。"""
    host = request.client.host if request.client else ""
    if host not in ("127.0.0.1", "::1", "localhost"):
        raise HTTPException(403, "仅本机模式支持提示音")
    from ai_gen_reimbursement_docs.cli.notify import play_notify_sound
    play_notify_sound()
    return {"ok": True}


@app.post("/api/run-local")
async def api_run_local(
    xlsx_path: str = Form(...),
    output_dir: str = Form(""),
    mode: str = Form(...),
    api_key: str = Form(""),
    model: str = Form(""),
    base_url: str = Form(""),
    max_tokens: str = Form(""),
    project_name: str = Form(""),
    clean: str = Form(""),
    fpa_template: UploadFile | None = File(None),
    cosmic_template: UploadFile | None = File(None),
    list_template: UploadFile | None = File(None),
    spec_template: UploadFile | None = File(None),
):
    """本机模式：接受文件路径或目录路径，目录则自动搜索功能清单 xlsx。"""
    if mode not in MODE_INFO:
        raise HTTPException(400, f"未知模式: {mode}")

    xlsx_input = Path(xlsx_path)
    if not xlsx_input.exists():
        raise HTTPException(400, f"路径不存在: {xlsx_path}")

    from_dir = ""
    if xlsx_input.is_dir():
        from_dir = str(xlsx_input)
        import glob
        from ai_gen_reimbursement_docs.excel_source import is_valid_input_xlsx
        xlsx_files = [f for f in glob.glob(os.path.join(from_dir, "*.xlsx"))
                      if is_valid_input_xlsx(f)]
        if not xlsx_files:
            raise HTTPException(400, f"目录中未找到符合规范的功能清单 .xlsx: {xlsx_path}")
        # 优先匹配常见命名
        preferred = [f for f in xlsx_files
                     if os.path.basename(f) in ("功能清单-录入模板.xlsx", "功能清单.xlsx")]
        xlsx = Path(preferred[0] if preferred else xlsx_files[0])
    else:
        xlsx = xlsx_input

    import re
    from ai_gen_reimbursement_docs.pipeline import _try_read_project_name

    if output_dir:
        out = Path(output_dir)
    elif project_name:
        root = Path(from_dir) if from_dir else xlsx.parent
        safe = re.sub(r'[\/:*?"<>|]', '_', project_name)
        out = root / safe
    else:
        root = Path(from_dir) if from_dir else xlsx.parent
        auto_name = _try_read_project_name(str(xlsx))
        if auto_name:
            safe = re.sub(r'[\/:*?"<>|]', '_', auto_name)
            out = root / safe
        else:
            out = root
    out.mkdir(parents=True, exist_ok=True)

    custom_t_dir = await _save_custom_templates(
        out, fpa_template, cosmic_template, list_template, spec_template
    )

    session_id = uuid.uuid4().hex[:8]
    log_queue: queue.Queue = queue.Queue(maxsize=2000)
    session_queues[session_id] = log_queue
    session_outputs[session_id] = out

    def run():
        session_var.set(session_id)
        _result = None
        try:
            _result = _execute_mode(
                mode, str(xlsx), str(out), custom_t_dir,
                api_key, model, base_url, project_name,
                max_tokens=max_tokens, clean=bool(clean),
            )
        except CancelledError as e:
            logging.getLogger("ai_gen_reimbursement_docs").info(f"任务已停止: {e}")
            emit_session_event({"type": "cancelled"})
        except Exception as e:
            logging.getLogger("ai_gen_reimbursement_docs").error(f"执行失败: {e}")
            emit_session_event({"type": "error", "msg": f"执行失败: {e}"})
            session_cancelled[session_id] = True
        finally:
            if not session_cancelled.get(session_id):
                emit_session_event({
                    "type": "done",
                    "files": _build_file_summary(_result) if _result else [],
                })
            session_cancelled.pop(session_id, None)
            session_var.set(None)

    asyncio.create_task(asyncio.to_thread(run))
    return {"session_id": session_id, "output_dir": str(out)}


# ── 远程服务模式 ──────────────────────────────────────────────


@app.post("/api/run-upload")
async def api_run_upload(
    file: UploadFile = File(...),
    mode: str = Form(...),
    api_key: str = Form(""),
    model: str = Form(""),
    base_url: str = Form(""),
    max_tokens: str = Form(""),
    project_name: str = Form(""),
    clean: str = Form(""),
    fpa_template: UploadFile | None = File(None),
    cosmic_template: UploadFile | None = File(None),
    list_template: UploadFile | None = File(None),
    spec_template: UploadFile | None = File(None),
):
    """远程服务模式：上传文件，交付物打包 ZIP 下载。"""
    if mode not in MODE_INFO:
        raise HTTPException(400, f"未知模式: {mode}")

    if not file.filename:
        raise HTTPException(400, "未选择文件")

    session_id = uuid.uuid4().hex[:8]
    work_dir = Path(tempfile.mkdtemp(prefix=f"ard_web_{session_id}_"))
    input_dir = work_dir / "input"
    output_dir = work_dir / "output"
    custom_t_dir = work_dir / "custom_templates"
    for d in [input_dir, output_dir, custom_t_dir]:
        d.mkdir(parents=True)

    # 保存上传的 xlsx
    safe_name = Path(file.filename).name
    file_path = input_dir / safe_name
    content = await file.read()
    file_path.write_bytes(content)

    # 保存自定义模板
    await _save_custom_templates_into(
        custom_t_dir, fpa_template, cosmic_template, list_template, spec_template
    )

    log_queue: queue.Queue = queue.Queue(maxsize=2000)
    session_queues[session_id] = log_queue
    session_dirs[session_id] = work_dir

    def run():
        session_var.set(session_id)
        _result = None
        try:
            _result = _execute_mode(
                mode, str(file_path), str(output_dir), str(custom_t_dir),
                api_key, model, base_url, project_name,
                max_tokens=max_tokens, clean=bool(clean),
            )
            # 打包交付物 ZIP
            zip_path = work_dir / f"交付物_{session_id}.zip"
            shutil.make_archive(
                str(zip_path.with_suffix("")), "zip", str(output_dir)
            )
            session_zips[session_id] = zip_path
        except CancelledError as e:
            logging.getLogger("ai_gen_reimbursement_docs").info(f"任务已停止: {e}")
            emit_session_event({"type": "cancelled"})
        except Exception as e:
            logging.getLogger("ai_gen_reimbursement_docs").error(f"执行失败: {e}")
            emit_session_event({"type": "error", "msg": f"执行失败: {e}"})
            session_cancelled[session_id] = True
        finally:
            if not session_cancelled.get(session_id):
                emit_session_event({
                    "type": "done",
                    "files": _build_file_summary(_result) if _result else [],
                })
            session_cancelled.pop(session_id, None)
            session_var.set(None)

    asyncio.create_task(asyncio.to_thread(run))
    return {"session_id": session_id, "has_download": True}


# ── SSE 日志流 ────────────────────────────────────────────


@app.get("/api/log-stream")
async def log_stream(session: str):
    q = session_queues.get(session)
    if q is None:
        raise HTTPException(404, "未知会话")

    async def generate():
        while True:
            try:
                msg = await asyncio.get_event_loop().run_in_executor(
                    None, lambda: q.get(timeout=0.1)
                )
                data = json.loads(msg)
                yield f"data: {json.dumps(data, ensure_ascii=False)}\n\n"
                if data.get("type") in ("done", "cancelled", "error"):
                    break
            except queue.Empty:
                yield ": heartbeat\n\n"

        session_queues.pop(session, None)

    return StreamingResponse(
        generate(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
        },
    )


# ── 交付物操作 ──────────────────────────────────────────────


@app.get("/api/download/{session_id}")
async def download(session_id: str):
    """远程服务模式：下载交付物 ZIP。"""
    zip_path = session_zips.get(session_id)
    if zip_path is None:
        raise HTTPException(404, "交付物不存在或会话已过期")
    if not zip_path.exists():
        raise HTTPException(404, "交付物文件已被清理")

    return FileResponse(
        zip_path,
        filename=f"交付物_{datetime.now():%Y%m%d_%H%M%S}.zip",
        media_type="application/zip",
        background=_cleanup_after_download(session_id),
    )


@app.get("/api/open-folder")
async def open_folder(session: str):
    """本机模式：在资源管理器中打开交付物目录。"""
    out_dir = session_outputs.get(session)
    if out_dir is None:
        raise HTTPException(404, "未知会话")
    if not out_dir.exists():
        raise HTTPException(404, "交付物目录不存在")
    os.startfile(str(out_dir))
    return {"ok": True}


# ── AI 交互日志 ────────────────────────────────────────────


def _find_log_dir(session_id: str) -> Path | None:
    """根据 session 找到日志目录。"""
    out_dir = session_outputs.get(session_id)
    if out_dir is None:
        work_dir = session_dirs.get(session_id)
        if work_dir:
            out_dir = work_dir / "output"
    if out_dir is None or not out_dir.exists():
        return None

    # 搜索「日志」目录
    for log_dir in out_dir.rglob("日志"):
        if log_dir.is_dir():
            return log_dir
    return None


@app.get("/api/ai-log/{session_id}")
async def get_ai_log(session_id: str):
    """返回 AI 对话日志内容。"""
    log_dir = _find_log_dir(session_id)
    if log_dir is None:
        raise HTTPException(404, "未找到日志目录")

    combined = log_dir / "ai_对话日志.md"
    if not combined.exists():
        raise HTTPException(404, "AI 对话日志尚未生成")

    content = combined.read_text(encoding="utf-8")
    return {"content": content, "filename": combined.name}


@app.get("/api/ai-interactions/{session_id}")
async def list_ai_interactions(session_id: str):
    """列出 AI prompts 和 responses 文件清单及内容。"""
    log_dir = _find_log_dir(session_id)
    if log_dir is None:
        raise HTTPException(404, "未找到日志目录")

    prompts_dir = log_dir / "ai_prompts"
    responses_dir = log_dir / "ai_responses"

    files: list[dict] = []

    if prompts_dir.is_dir():
        for fname in sorted(os.listdir(prompts_dir)):
            if fname.endswith(".txt"):
                path = prompts_dir / fname
                files.append({
                    "name": fname,
                    "type": "prompt",
                    "content": path.read_text(encoding="utf-8"),
                })

    if responses_dir.is_dir():
        for fname in sorted(os.listdir(responses_dir)):
            if fname.endswith(".txt"):
                path = responses_dir / fname
                files.append({
                    "name": fname,
                    "type": "response",
                    "content": path.read_text(encoding="utf-8"),
                })

    return {"interactions": files, "count": len(files)}


# ── 清理 ──────────────────────────────────────────────────


async def _cleanup_after_download(session_id: str):
    """下载完成后延迟 5 分钟清理，避免干扰。"""
    await asyncio.sleep(300)
    work_dir = session_dirs.pop(session_id, None)
    session_queues.pop(session_id, None)
    session_zips.pop(session_id, None)
    if work_dir and work_dir.exists():
        shutil.rmtree(work_dir, ignore_errors=True)


async def _save_custom_templates_into(
    target_dir: Path,
    fpa_template: UploadFile | None,
    cosmic_template: UploadFile | None,
    list_template: UploadFile | None,
    spec_template: UploadFile | None,
):
    """将自定义模板保存到指定目录。"""
    for tpl_file, tpl_name in [
        (fpa_template, "FPA工作量评估-输出模板.xlsx"),
        (cosmic_template, "项目功能点拆分表-输出模板.xlsx"),
        (list_template, "项目需求清单-输出模板.xlsx"),
        (spec_template, "项目需求说明书-输出模板.docx"),
    ]:
        if tpl_file is not None and tpl_file.filename:
            tpl_content = await tpl_file.read()
            (target_dir / tpl_name).write_bytes(tpl_content)


async def _save_custom_templates(
    parent_dir: Path,
    fpa_template: UploadFile | None,
    cosmic_template: UploadFile | None,
    list_template: UploadFile | None,
    spec_template: UploadFile | None,
) -> str:
    """将自定义模板保存到临时目录，返回目录路径。"""
    custom_t_dir = parent_dir / "custom_templates"
    custom_t_dir.mkdir(parents=True, exist_ok=True)
    await _save_custom_templates_into(
        custom_t_dir, fpa_template, cosmic_template, list_template, spec_template
    )
    return str(custom_t_dir)


# ── 执行分发 ──────────────────────────────────────────────


def _build_file_summary(result) -> list[dict]:
    """从 PipelineResult 构建文件摘要。标注 _TEMP 文件。"""
    files = []
    labels = [
        ("FPA 工作量评估", getattr(result, 'fpa_xlsx', '')),
        ("项目功能点拆分表", getattr(result, 'cosmic_xlsx', '')),
        ("项目需求清单", getattr(result, 'require_xlsx', '')),
        ("项目需求说明书", getattr(result, 'spec_docx', '')),
    ]
    for label, path in labels:
        if path and os.path.exists(path):
            size = os.path.getsize(path)
            is_temp = '_TEMP' in os.path.basename(path)
            files.append({
                "label": label,
                "path": path,
                "size_kb": round(size / 1024),
                "is_temp": is_temp,
            })
    return files


def _execute_mode(
    mode: str,
    file_path: str,
    output_dir: str,
    custom_t_dir: str,
    api_key: str,
    model: str,
    base_url: str,
    project_name: str = "",
    max_tokens: str = "",
    clean: bool = False,
):
    """一站式管道入口，CLI / Web UI 共享。"""
    from ai_gen_reimbursement_docs.pipeline import run_pipeline_simple

    if max_tokens:
        os.environ["AI_REIMBURSEMENT_MAX_TOKENS"] = max_tokens
    if clean and os.path.isdir(output_dir):
        for name in os.listdir(output_dir):
            if name not in ("custom_templates",):
                path = os.path.join(output_dir, name)
                if os.path.isfile(path):
                    os.remove(path)
                else:
                    shutil.rmtree(path, ignore_errors=True)

    logger = logging.getLogger("ai_gen_reimbursement_docs")
    logger.info(f"操作模式: {MODE_INFO.get(mode, {}).get('label', mode)}")

    pipeline_mode = _MODE_MAP[mode]
    templates = _build_templates_dict(custom_t_dir)

    return run_pipeline_simple(
        mode=pipeline_mode,
        file_path=file_path,
        output_dir=output_dir,
        api_key=api_key,
        model=model,
        base_url=base_url,
        project_name=project_name,
        templates=templates or None,
    )


def _build_templates_dict(custom_t_dir: str) -> dict[str, str]:
    """构建 templates dict：自定义模板优先。"""
    import glob as _glob

    templates: dict[str, str] = {}
    for key, glob_pat in [
        ("fpa", "FPA*评估*模板*.xlsx"),
        ("cosmic", "*功能点拆分表*模板*.xlsx"),
        ("list", "*需求清单*模板*.xlsx"),
        ("spec", "*需求说明书*模板*.docx"),
    ]:
        matches = _glob.glob(os.path.join(custom_t_dir, glob_pat))
        if matches:
            templates[key] = matches[0]
    return templates


# ── 配置读取 API（Config 页面用） ──────────────────────────

@app.get("/api/config-read")
async def config_read():
    """读取 ~/.ai-gen-reimbursement-docs/ 下的三个配置文件内容。"""
    cfg_dir = Path(os.path.expanduser("~")) / ".ai-gen-reimbursement-docs"
    result: dict = {}
    for key, fname in [("env", ".env"), ("system_config", "system_config.yaml"),
                        ("business_rules", "business_rules.yaml")]:
        fp = cfg_dir / fname
        result[key] = fp.read_text(encoding="utf-8") if fp.exists() else ""
    return result


# ── 提示词调试 API（PromptDebug 页面用） ────────────────────

@app.post("/api/test-ai-reliability-desc")
async def test_reliability_desc(xlsx_path: str = Form("")):
    """测试调整因子中的可靠性描述 AI 生成。"""
    import glob
    excel_path = xlsx_path.strip()
    if not excel_path:
        for name in ["功能清单-录入模板.xlsx", "功能清单.xlsx"]:
            matches = glob.glob(name)
            if matches:
                excel_path = matches[0]
                break
    if not excel_path or not os.path.exists(excel_path):
        raise HTTPException(400, "未找到功能清单 .xlsx 文件")

    from ai_gen_reimbursement_docs.config_utils import load_ai_system_prompt, load_sheet_names, load_api_key
    import openpyxl

    api_key = load_api_key()
    if not api_key:
        raise HTTPException(400, "未配置 API Key，请先在配置页设置")

    _s = load_sheet_names()
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[_s["func_list"]]
    descriptions: list[str] = []
    seen: set[str] = set()
    prev = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        desc = str(row[5]).strip() if len(row) > 5 and row[5] else ""
        desc = desc or prev
        if desc:
            prev = desc
        if desc and desc not in seen:
            seen.add(desc)
            descriptions.append(desc)
    wb.close()

    user_prompt = (
        "根据功能清单，提取其中涉及与可靠性方面的模块，生成一句关于可靠性业务描述。不少于50字。\n"
        "功能清单：\n" + '\n'.join(f'- {d}' for d in descriptions)
    )
    system_prompt = load_ai_system_prompt("reliability_desc")

    from ai_gen_reimbursement_docs.llm_client import call_llm
    try:
        result_text = call_llm(
            prompt=user_prompt, system=system_prompt,
            api_key=api_key, model="", base_url="", tag="web_reliability_desc",
        )
        return {"result": result_text}
    except Exception as e:
        raise HTTPException(500, f"AI 调用失败: {e}")


@app.post("/api/test-ai-metadata")
async def test_metadata(xlsx_path: str = Form(""), field_key: str = Form("")):
    """测试元数据中指定字段的 #AI生成# 效果。"""
    import glob, re
    excel_path = xlsx_path.strip()
    if not excel_path:
        for name in ["功能清单-录入模板.xlsx", "功能清单.xlsx"]:
            matches = glob.glob(name)
            if matches:
                excel_path = matches[0]
                break
    if not excel_path or not os.path.exists(excel_path):
        raise HTTPException(400, "未找到功能清单 .xlsx 文件")
    if not field_key.strip():
        raise HTTPException(400, "请提供 field_key")

    from ai_gen_reimbursement_docs.config_utils import load_api_key, load_sheet_names, load_ai_system_prompt
    from ai_gen_reimbursement_docs.excel_source import strip_ai_marker
    import openpyxl

    api_key = load_api_key()
    if not api_key:
        raise HTTPException(400, "未配置 API Key")

    _s = load_sheet_names()
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    raw_value = ""
    for sheet_key in ["meta", "fpa_meta", "spec_meta", "cosmic_meta", "list_meta"]:
        sn = _s.get(sheet_key, "")
        if not sn or sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for row in ws.iter_rows(min_row=2, values_only=True):
            k = str(row[0]).strip() if row[0] else ""
            v = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if k == field_key:
                raw_value = v
                break
        if raw_value:
            break
    wb.close()

    if not raw_value:
        raise HTTPException(400, f"未找到字段「{field_key}」")

    prompt_template, needs_ai = strip_ai_marker(raw_value)
    if not needs_ai:
        return {"result": f"字段「{field_key}」不含 #AI生成# 标记，当前值: {raw_value}"}

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    project_info: dict[str, str] = {}
    for row in wb[_s["work_order_meta"]].iter_rows(min_row=2, values_only=True):
        k2 = str(row[0]).strip() if row[0] else ""
        v2 = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        if k2:
            project_info[k2] = v2
    fpa_meta: dict[str, str] = {}
    for row in wb[_s["fpa_meta"]].iter_rows(min_row=2, values_only=True):
        k2 = str(row[0]).strip() if row[0] else ""
        v2 = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        if k2:
            fpa_meta[k2] = v2
    wb.close()

    user_prompt = prompt_template
    user_prompt = user_prompt.replace('${工单编号}', project_info.get('工单编号', ''))
    user_prompt = user_prompt.replace('${工单名称}', project_info.get('工单标题', ''))
    user_prompt = user_prompt.replace('${工单标题}', project_info.get('工单标题', ''))
    user_prompt = user_prompt.replace('${工单内容}', project_info.get('工单内容', ''))
    user_prompt = user_prompt.replace('${子系统（模块）}', fpa_meta.get('子系统（模块）', ''))

    system_prompt = load_ai_system_prompt("metadata_gen")

    from ai_gen_reimbursement_docs.llm_client import call_llm
    try:
        result_text = call_llm(
            prompt=user_prompt, system=system_prompt,
            api_key=api_key, model="", base_url="", tag="web_metadata_test",
        )
        return {"result": result_text}
    except Exception as e:
        raise HTTPException(500, f"AI 调用失败: {e}")
