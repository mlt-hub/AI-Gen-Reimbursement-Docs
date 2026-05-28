import os

from fastapi import APIRouter, Form, HTTPException


router = APIRouter()


@router.post("/api/test-prompt")
async def test_prompt(data: dict):
    """提交系统提示词和用户提示词，返回 AI 生成结果。"""
    system_prompt = data.get("system_prompt", "").strip()
    user_prompt = data.get("user_prompt", "").strip()
    if not user_prompt and not system_prompt:
        raise HTTPException(400, "系统提示词和用户提示词不能同时为空")

    from ai_gen_reimbursement_docs.config_utils import (
        load_api_key,
        load_base_url,
        load_model_name,
    )
    from ai_gen_reimbursement_docs.llm_client import call_llm

    api_key = data.get("api_key", "").strip() or load_api_key()
    model = data.get("model", "").strip() or load_model_name()
    base_url = data.get("base_url", "").strip() or load_base_url()

    if not api_key:
        raise HTTPException(400, "未配置 API Key")

    if api_key:
        os.environ["ANTHROPIC_API_KEY"] = api_key
    if base_url:
        os.environ["ANTHROPIC_BASE_URL"] = base_url

    try:
        result, thinking = call_llm(
            prompt=user_prompt,
            system=system_prompt,
            api_key=api_key,
            model=model,
            base_url=base_url,
            tag="prompt_debug",
            return_thinking=True,
        )
        return {
            "result": result,
            "thinking": thinking,
            "system_prompt": system_prompt,
            "user_prompt": user_prompt,
        }
    except Exception as e:
        raise HTTPException(500, f"AI 调用失败: {e}")


@router.post("/api/test-ai-reliability-desc")
async def test_reliability_desc(xlsx_path: str = Form("")):
    """测试调整因子中的可靠性描述 AI 生成。"""
    import glob

    excel_path = xlsx_path.strip()
    if not excel_path:
        for name in ["功能清单-录入模板.xlsx", "功能清单.xlsx"]:
            matches = glob.glob(name)
            if matches:
                excel_path = matches[0]
                break
    if not excel_path or not os.path.exists(excel_path):
        raise HTTPException(400, "未找到功能清单 .xlsx 文件")

    import openpyxl

    from ai_gen_reimbursement_docs.config_utils import (
        load_ai_system_prompt,
        load_api_key,
        load_sheet_names,
    )

    api_key = load_api_key()
    if not api_key:
        raise HTTPException(400, "未配置 API Key，请先在配置页设置")

    _s = load_sheet_names()
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[_s["func_list"]]
    descriptions: list[str] = []
    seen: set[str] = set()
    prev = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        desc = str(row[5]).strip() if len(row) > 5 and row[5] else ""
        desc = desc or prev
        if desc:
            prev = desc
        if desc and desc not in seen:
            seen.add(desc)
            descriptions.append(desc)
    wb.close()

    user_prompt = (
        "根据功能清单，提取其中涉及与可靠性方面的模块，生成一句关于可靠性业务描述。不少于50字。\n"
        "功能清单：\n" + "\n".join(f"- {d}" for d in descriptions)
    )
    system_prompt = load_ai_system_prompt("reliability_desc")

    from ai_gen_reimbursement_docs.llm_client import call_llm

    try:
        result_text = call_llm(
            prompt=user_prompt,
            system=system_prompt,
            api_key=api_key,
            model="",
            base_url="",
            tag="web_reliability_desc",
        )
        return {"result": result_text}
    except Exception as e:
        raise HTTPException(500, f"AI 调用失败: {e}")


@router.post("/api/test-ai-metadata")
async def test_metadata(xlsx_path: str = Form(""), field_key: str = Form("")):
    """测试元数据中指定字段的 #AI生成# 效果。"""
    import glob

    excel_path = xlsx_path.strip()
    if not excel_path:
        for name in ["功能清单-录入模板.xlsx", "功能清单.xlsx"]:
            matches = glob.glob(name)
            if matches:
                excel_path = matches[0]
                break
    if not excel_path or not os.path.exists(excel_path):
        raise HTTPException(400, "未找到功能清单 .xlsx 文件")
    if not field_key.strip():
        raise HTTPException(400, "请提供 field_key")

    import openpyxl

    from ai_gen_reimbursement_docs.config_utils import (
        load_ai_system_prompt,
        load_api_key,
        load_sheet_names,
    )
    from ai_gen_reimbursement_docs.excel_source import strip_ai_marker

    api_key = load_api_key()
    if not api_key:
        raise HTTPException(400, "未配置 API Key")

    _s = load_sheet_names()
    wb = openpyxl.load_workbook(excel_path, data_only=True)

    raw_value = ""
    for sheet_key in ["meta", "fpa_meta", "spec_meta", "cosmic_meta", "list_meta"]:
        sn = _s.get(sheet_key, "")
        if not sn or sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        for row in ws.iter_rows(min_row=2, values_only=True):
            k = str(row[0]).strip() if row[0] else ""
            v = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if k == field_key:
                raw_value = v
                break
        if raw_value:
            break
    wb.close()

    if not raw_value:
        raise HTTPException(400, f"未找到字段「{field_key}」")

    prompt_template, needs_ai = strip_ai_marker(raw_value)
    if not needs_ai:
        return {"result": f"字段「{field_key}」不含 #AI生成# 标记，当前值: {raw_value}"}

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    project_info: dict[str, str] = {}
    for row in wb[_s["work_order_meta"]].iter_rows(min_row=2, values_only=True):
        k2 = str(row[0]).strip() if row[0] else ""
        v2 = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        if k2:
            project_info[k2] = v2
    fpa_meta: dict[str, str] = {}
    for row in wb[_s["fpa_meta"]].iter_rows(min_row=2, values_only=True):
        k2 = str(row[0]).strip() if row[0] else ""
        v2 = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        if k2:
            fpa_meta[k2] = v2
    wb.close()

    user_prompt = prompt_template
    user_prompt = user_prompt.replace("${工单编号}", project_info.get("工单编号", ""))
    user_prompt = user_prompt.replace("${工单名称}", project_info.get("工单标题", ""))
    user_prompt = user_prompt.replace("${工单标题}", project_info.get("工单标题", ""))
    user_prompt = user_prompt.replace("${工单内容}", project_info.get("工单内容", ""))
    user_prompt = user_prompt.replace("${子系统（模块）}", fpa_meta.get("子系统（模块）", ""))

    system_prompt = load_ai_system_prompt("metadata_gen")

    from ai_gen_reimbursement_docs.llm_client import call_llm

    try:
        result_text = call_llm(
            prompt=user_prompt,
            system=system_prompt,
            api_key=api_key,
            model="",
            base_url="",
            tag="web_metadata_test",
        )
        return {"result": result_text}
    except Exception as e:
        raise HTTPException(500, f"AI 调用失败: {e}")
