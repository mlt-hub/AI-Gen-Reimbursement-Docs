"""Excel 功能清单解析器 — 读取功能清单.xlsx，生成中间 MD 文件"""

import logging
import os
import re
import sys
from datetime import datetime

import openpyxl

from ai_gen_reimbursement_docs.config_utils import load_sheet_names

logger = logging.getLogger('ai_gen_reimbursement_docs.excel_source')


def _cell_val(cell: object) -> str:
    """读取单元格值，None 转空字符串。"""
    return str(cell).strip() if cell is not None else ""


def _resolve_inherited_rows(ws: "openpyxl.worksheet.worksheet.Worksheet") -> list[list[str]]:
    """将合并单元格/空单元格的值继承上一行同列的值。返回二维列表。"""
    rows = []
    prev_vals = [""] * ws.max_column
    for row in ws.iter_rows(min_row=2, values_only=True):  # 跳过表头
        vals = [_cell_val(c) for c in row]
        for i in range(len(vals)):
            if not vals[i] and prev_vals[i]:
                vals[i] = prev_vals[i]
        prev_vals = vals
        rows.append(vals)
    return rows


def _key_value_sheet(ws: "openpyxl.worksheet.worksheet.Worksheet") -> dict[str, str]:
    """读取 key-value 格式 sheet（2列：项目 | 内容），返回 dict。"""
    data = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        key = _cell_val(row[0])
        val = _cell_val(row[1]) if len(row) > 1 else ""
        if key:
            data[key] = val
    return data


def generate_md_files(excel_path: str, output_dir: str = "") -> dict[str, str]:
    """读取功能清单.xlsx，生成gen-basedata-功能清单-模块树.md 和 gen-basedata-录入文档元数据-模板.md。

    Args:
        excel_path: 功能清单.xlsx 路径
        output_dir: 交付物输出目录（空则取 excel 所在目录）

    Returns:
        {"module_tree_md": 路径, "doc_meta_md": 路径}
    """
    _s = load_sheet_names()
    if not output_dir:
        output_dir = os.path.dirname(os.path.abspath(excel_path))
    os.makedirs(output_dir, exist_ok=True)

    try:
        wb = openpyxl.load_workbook(excel_path, data_only=False)
    except FileNotFoundError:
        raise FileNotFoundError(f"功能清单文件不存在: {excel_path}")
    except Exception as e:
        raise ValueError(
            f"无法打开功能清单文件，请确认是有效的 .xlsx 格式: {excel_path}\n"
            f"内部错误: {e}"
        ) from e

    # ========== 解析各个 sheet ==========

    # 1、工单需求-元数据录入
    ws1 = wb[_s["work_order_meta"]]
    project_info = _key_value_sheet(ws1)

    # 2、功能清单-内容录入 — 模块树 + 功能过程
    ws2 = wb[_s["func_list"]]
    func_rows = _resolve_inherited_rows(ws2)

    # 3、FPA工作量评估-元数据录入
    ws3 = wb[_s["fpa_meta"]]
    fpa_meta = _key_value_sheet(ws3)

    # 4、项目需求说明书-元数据录入
    ws4 = wb[_s["spec_meta"]]
    docx_meta = _key_value_sheet(ws4)

    # 6、项目功能点拆分表-元数据录入
    ws6 = wb[_s["cosmic_meta"]]
    cosmic_meta = _key_value_sheet(ws6)

    # 7、项目需求清单-元数据录入
    ws7 = wb[_s["list_meta"]]
    list_meta = _key_value_sheet(ws7)

    wb.close()

    # data_only=True 还原公式单元格的计算值
    try:
        wb_val = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        raise ValueError(
            f"无法以 data_only 模式读取功能清单，文件可能已损坏: {excel_path}\n"
            f"内部错误: {e}"
        ) from e
    for row in wb_val[_s["list_meta"]].iter_rows(min_row=2, values_only=True):
        k, v = row[0], row[1]
        if k and str(v).strip():
            wk = str(k).strip()
            wv = str(v).strip()
            if wk in list_meta and list_meta[wk].startswith('='):
                list_meta[wk] = wv
    wb_val.close()

    # 统计元数据自动统计（层级去重，与 Excel 展示一致）
    stats_meta = {
        "入口（个数）": str(len({r[0] for r in func_rows if r[0]})),
        "一级模块（个数）": str(len({(r[0], r[1]) for r in func_rows if r[1]})),
        "二级模块（个数）": str(len({(r[0], r[1], r[2]) for r in func_rows if r[2]})),
        "三级模块（个数）": str(len({(r[0], r[1], r[2], r[3]) for r in func_rows if r[3]})),
        "功能过程（个数）": str(len({r[6] for r in func_rows if r[6]})),
    }

    # ========== 生成 gen-basedata-功能清单-模块树.md ==========

    md_tree_path = os.path.join(output_dir, 'gen-basedata-功能清单-模块树.md')
    with open(md_tree_path, 'w', encoding='utf-8') as f:
        f.write("# 功能清单模块树\n\n")
        f.write(f"**来源文件**：{os.path.basename(excel_path)}\n")
        f.write(f"**生成日期**：{datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n")
        f.write("| 入口 | 一级模块 | 二级模块 | 三级模块 | 客户端类型 | 三级模块整体功能描述 | 功能过程 | 功能过程类型 | 功能过程描述 |\n")
        f.write("|------|---------|---------|---------|-----------|-------------------|---------|-----------|-------------|\n")
        for row in func_rows:
            # row: [入口, 一级模块, 二级模块, 三级模块, 客户端类型, 三级模块整体功能描述, 功能过程, 功能过程类型, 功能过程描述]
            vals = [cell.replace('|', '\\|').replace('\n', ' ') for cell in row[:9]]
            line = " | ".join(vals)
            f.write(f"| {line} |\n")

    logger.info(f"第0.1步：功能清单模块树已生成: {md_tree_path}")

    # ========== 生成 gen-basedata-录入文档元数据-模板.md ==========

    md_meta_path = os.path.join(output_dir, 'gen-basedata-录入文档元数据-模板.md')
    with open(md_meta_path, 'w', encoding='utf-8') as f:
        f.write("# 文档元数据\n\n")
        f.write(f"**来源文件**：{os.path.basename(excel_path)}\n")
        f.write(f"**生成日期**：{datetime.now().strftime('%Y-%m-%d %H:%M')}\n\n")

        # 按 sheet 分组写入
        sections = [
            (_s["work_order_meta"], project_info),
            (_s["fpa_meta"], fpa_meta),
            (_s["spec_meta"], docx_meta),
            (_s["cosmic_meta"], cosmic_meta),
            (_s["list_meta"], list_meta),
            (_s["stats_meta"], stats_meta),
        ]

        for sheet_name, kv in sections:
            f.write(f"## {sheet_name}\n\n")
            f.write("| 项目 | 内容 |\n")
            f.write("|------|------|\n")
            for key, val in kv.items():
                # 替换 【占位符】 为实际值
                v = val
                for ph, src in [("${工单编号}", project_info.get("工单编号", "")),
                                 ("${工单名称}", project_info.get("工单标题", "")),
                                 ("${工单标题}", project_info.get("工单标题", "")),
                                 ("${子系统（模块）}", fpa_meta.get("子系统（模块）", ""))]:
                    v = v.replace(ph, src)
                val_escaped = v.replace('|', '\\|').replace('\n', ' ').replace('\r', ' ')
                f.write(f"| {key} | {val_escaped} |\n")
            f.write("\n")

    logger.info(f"第0.2步：录入文档元数据-模板已生成: {md_meta_path}")

    return {"module_tree_md": md_tree_path, "doc_meta_md": md_meta_path}


def read_fpa_xlsx_sum(fpa_xlsx_path: str) -> float:
    """读取 FPA工作量评估.xlsx 中 FPA工作量列（L列）的求和。"""
    try:
        wb = openpyxl.load_workbook(fpa_xlsx_path, data_only=True)
        from ai_gen_reimbursement_docs.config_utils import _get_system_config_value
        _fpa_sheet = _get_system_config_value('fpa_sheet', 'FPA功能点估算')
        ws = wb[_fpa_sheet]
        total = 0.0
        for row in ws.iter_rows(min_row=3, values_only=True):
            val = row[11]  # L列（FPA工作量），0-based index
            if val is not None:
                try:
                    total += float(val)
                except (ValueError, TypeError):
                    pass
        wb.close()
        logger.info(f"FPA工作量求和: {total}")
        return total
    except Exception as e:
        logger.warning(f"读取FPA工作量求和失败: {e}")
        return 0.0


def verify_module_tree_stats(tree_md_path: str, meta_md_path: str) -> bool:
    """验证gen-basedata-功能清单-模块树.md 的统计信息与gen-basedata-录入文档元数据-模板.md 中的期望值一致。

    读取模块树 MD 表格，统计入口/L1/L2/L3/功能过程数，
    与文档元数据中 stats_meta 段进行对比。

    Returns:
        True 全部通过, False 有差异
    """
    from ai_gen_reimbursement_docs.md_table import parse_md_table_row
    from ai_gen_reimbursement_docs.config_utils import load_sheet_names
    _stats_section = load_sheet_names().get("stats_meta", "9、测试元数据自动统计")

    # 从模块树 MD 统计（层级去重，与 Excel 展示一致）
    entries: set[str] = set()
    l1s: set[tuple[str, str]] = set()
    l2s: set[tuple[str, str, str]] = set()
    l3s: set[tuple[str, str, str, str]] = set()
    procs: set[str] = set()

    with open(tree_md_path, encoding='utf-8') as f:
        in_table = False
        for line in f:
            if "| 入口 | 一级模块" in line:
                in_table = True
                continue
            if "|------" in line and in_table:
                continue
            if in_table:
                cells = parse_md_table_row(line, min_cols=9)
                if cells is not None:
                    if cells[0]:
                        entries.add(cells[0])
                    if cells[1]:
                        l1s.add((cells[0], cells[1]))
                    if cells[2]:
                        l2s.add((cells[0], cells[1], cells[2]))
                    if cells[3]:
                        l3s.add((cells[0], cells[1], cells[2], cells[3]))
                    if cells[6]:
                        procs.add(cells[6])

    actual = {
        "入口（个数）": len(entries),
        "一级模块（个数）": len(l1s),
        "二级模块（个数）": len(l2s),
        "三级模块（个数）": len(l3s),
        "功能过程（个数）": len(procs),
    }

    # 从文档元数据读取期望值
    expected: dict[str, int] = {}
    in_stats = False
    with open(meta_md_path, encoding='utf-8') as f:
        for line in f:
            line = line.rstrip()
            if line.startswith(f"## {_stats_section}"):
                in_stats = True
                continue
            if in_stats:
                if line.startswith("## "):
                    break
                if line.startswith("|") and not line.startswith("|--"):
                    cells = [c.strip() for c in line.split("|")]
                    cells = [c for c in cells if c]
                    if len(cells) >= 2:
                        try:
                            expected[cells[0]] = int(cells[1])
                        except ValueError:
                            pass

    # 对比输出
    if not expected:
        logger.info("第0.3步：模块树统计验证: 跳过（元数据中未找到 %s 段）", _stats_section)
        return True

    all_ok = True
    logger.info("第0.3步：模块树统计验证")
    for key in ["入口（个数）", "一级模块（个数）", "二级模块（个数）",
                 "三级模块（个数）", "功能过程（个数）"]:
        exp = expected.get(key)
        act = actual[key]
        if exp is not None:
            status = "✓" if act == exp else "✗"
            all_ok = all_ok and (act == exp)
            logger.info(f"  {status} {key}: 期望={exp}, 实际={act}")
        else:
            logger.info(f"  ? {key}: 期望=未配置, 实际={act}")

    if all_ok:
        logger.info("═══ 全部通过 ═══")
    else:
        logger.warning("═══ 存在差异，请检查数据 ═══")
    return all_ok


def replace_placeholders(text: str, project_info: dict[str, str], fpa_meta: dict[str, str]) -> str:
    """替换 【占位符】 为实际值。"""
    placeholders = {
        "${工单编号}": project_info.get("工单编号", ""),
        "${工单名称}": project_info.get("工单标题", ""),
        "${工单标题}": project_info.get("工单标题", ""),
        "${工单内容}": project_info.get("工单内容", ""),
        "${子系统（模块）}": fpa_meta.get("子系统（模块）", ""),
    }
    for ph, val in placeholders.items():
        text = text.replace(ph, val)
    return text


def strip_ai_marker(text: str) -> tuple[str, bool]:
    """判断是否含 #AI生成# 或 #AI生成-XXX# 标记，去掉标记返回纯净文本。"""
    if text.startswith("#AI生成#"):
        return text[len("#AI生成#"):], True
    m = re.match(r'^#AI生成-(.+?)#\s*(.*)', text)
    if m:
        # #AI生成-工单内容# → 保留 "基于工单内容" 作为提示词
        hint = m.group(1)
        rest = m.group(2)
        prompt = f"基于{hint}" if not rest else rest
        return prompt, True
    return text, False


def parse_module_tree_md(tree_md_path: str) -> list[dict[str, str]]:
    """解析 功能清单-模块树.md 表格为行字典列表。

    表格列：入口 | 一级模块 | 二级模块 | 三级模块 | 客户端类型 |
            三级模块整体功能描述 | 功能过程 | 功能过程类型 | 功能过程描述
    """
    from ai_gen_reimbursement_docs.md_table import parse_md_table_row
    rows = []
    with open(tree_md_path, encoding='utf-8') as f:
        in_table = False
        for line in f:
            if "| 入口 | 一级模块" in line:
                in_table = True
                continue
            if "|------" in line and in_table:
                continue
            if in_table:
                cells = parse_md_table_row(line, min_cols=9)
                if cells is not None:
                    rows.append({
                        "入口": cells[0],
                        "一级模块": cells[1],
                        "二级模块": cells[2],
                        "三级模块": cells[3],
                        "客户端类型": cells[4],
                        "三级模块整体功能描述": cells[5],
                        "功能过程": cells[6],
                        "功能过程类型": cells[7],
                        "功能过程描述": cells[8],
                    })
    return rows


def safe_load_workbook(path: str, label: str):
    """安全加载 xlsx 模板，失败时抛出可读的错误。"""
    try:
        return openpyxl.load_workbook(path)
    except FileNotFoundError:
        raise FileNotFoundError(f"「{label}」模板文件不存在: {path}")
    except Exception as e:
        raise ValueError(
            f"「{label}」模板无法打开，请检查文件是否为有效的 .xlsx 格式: {path}\n"
            f"内部错误: {e}"
        ) from e


def _collect_l3_names(tree_md: str) -> list[str]:
    """从 功能清单-模块树.md 收集所有去重的三级模块名（保持原始顺序）。"""
    from ai_gen_reimbursement_docs.md_table import parse_md_table_row
    names: list[str] = []
    seen: set[str] = set()
    if not os.path.exists(tree_md):
        return names
    with open(tree_md, encoding='utf-8') as f:
        for line in f:
            cells = parse_md_table_row(line, min_cols=4)
            if cells is not None and cells[3]:
                name = cells[3].strip()
                if name and name not in seen:
                    seen.add(name)
                    names.append(name)
    return names


def _call_llm_once(prompt: str, api_key: str, model: str, base_url: str,
                   tag: str = "") -> str:
    """单次 LLM 调用，返回文本（委托至 llm_client 公共模块）。"""
    if not api_key:
        return ""
    from ai_gen_reimbursement_docs.config_utils import load_ai_system_prompt
    system_prompt = load_ai_system_prompt("metadata_gen")
    from ai_gen_reimbursement_docs.llm_client import call_llm
    try:
        return call_llm(
            prompt=prompt, system=system_prompt,
            api_key=api_key, model=model, base_url=base_url, tag=tag,
        )
    except Exception as e:
        logging.getLogger('ai_gen_reimbursement_docs.excel_source').warning(
            "AI 调用失败 [%s]: %s", tag, e)
        return ""


def ai_fill_meta_md(src_md: str, dst_md: str, api_key: str, model: str, base_url: str,
                     tree_md: str = "") -> str:
    """AI 填充元数据 MD 中的 #AI生成# 标记，写入目标文件。"""
    from ai_gen_reimbursement_docs.md_table import parse_md_table_row

    meta_data = {}
    with open(src_md, encoding='utf-8') as f:
        for line in f:
            cells = parse_md_table_row(line, min_cols=2)
            if cells is not None and cells[0]:
                meta_data[cells[0]] = cells[1]

    project_info = {}
    for k, v in meta_data.items():
        key = k.replace("1、工单需求-元数据录入.", "")
        if key in ("工单编号", "工单标题", "工单内容", "总体描述",
                    "建设目标", "建设必要性", "系统概况"):
            project_info[key] = v
    fpa_meta = {}
    for k, v in meta_data.items():
        key = k.replace("3、FPA工作量评估-元数据录入.", "")
        if key in ("子系统（模块）",):
            fpa_meta[key] = v

    with open(src_md, encoding='utf-8') as f:
        lines = f.readlines()

    new_lines = []
    for line in lines:
        cells = parse_md_table_row(line, min_cols=2)
        if cells is not None:
            key, val = cells[0], cells[1]
            if val and ('#AI生成#' in val or '#AI生成-' in val):
                prompt_raw, needs_ai = strip_ai_marker(val)
                if needs_ai:
                    if not prompt_raw:
                        prompt_raw = f"基于{key}"
                    elif '${' in prompt_raw:
                        prompt_raw = replace_placeholders(prompt_raw, project_info, fpa_meta)
                        if '${三级模块}' in prompt_raw and tree_md:
                            _l3_names = _collect_l3_names(tree_md)
                            prompt_raw = prompt_raw.replace('${三级模块}', '、'.join(_l3_names))
                    resp = _call_llm_once(prompt_raw, api_key, model, base_url,
                                          tag=f"meta_{key}")
                    if resp:
                        new_lines.append(f"| {key} | {resp} |\n")
                        continue
        new_lines.append(line)

    with open(dst_md, 'w', encoding='utf-8') as f:
        f.writelines(new_lines)
    return dst_md


# ═══════════════════════════════════════════════════════════
#  共享工具函数
# ═══════════════════════════════════════════════════════════

def project_root() -> str:
    """项目根目录（兼容源码和 PyInstaller exe）。"""
    if getattr(sys, 'frozen', False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.dirname(__file__))


def read_md_value(path: str, pattern: str) -> float:
    """从 MD 文件中按正则提取数值，文件不存在返回 0。"""
    import re
    if not os.path.exists(path):
        return 0.0
    with open(path, encoding='utf-8') as f:
        for line in f:
            m = re.search(pattern, line)
            if m:
                return float(m.group(1))
    return 0.0


def write_cfp_sum(md_dir: str, total: float) -> None:
    """将 CFP 总和写入 gen-cosmic-CFP-总和.md。"""
    logger.info(f"第3.5步：写入 CFP 总和")
    path = os.path.join(md_dir, 'gen-cosmic-CFP-总和.md')
    with open(path, 'w', encoding='utf-8') as f:
        f.write("# CFP 总和\n\n")
        f.write(f"CFP 总和: {total}\n")
    logger.info(f"CFP 总和已写入: {path}（{total}）")


def read_project_name(meta_md_path: str) -> str:
    """从 录入文档元数据.md 读取项目名称（工单标题）。"""
    try:
        with open(meta_md_path, encoding='utf-8') as f:
            for line in f:
                if '| 工单标题' in line and '|' in line:
                    parts = [c.strip() for c in line.split('|')]
                    if len(parts) >= 3 and parts[2]:
                        return parts[2]
    except Exception:
        pass
    return ""


def build_modules_from_tree_md(md_path: str) -> list:
    """从 功能清单-模块树.md 构建 FunctionModule 列表。

    自动去重并构建 L1→L2→L3 层级，功能过程作为 L3 的 children。
    """
    from ai_gen_reimbursement_docs.models import FunctionModule
    rows = parse_module_tree_md(md_path)

    if not rows:
        return []

    modules = []
    seen_l1 = set()
    seen_l2 = {}
    seen_l3 = {}
    l3_procs = {}

    for r in rows:
        l1 = r["一级模块"]
        l2 = r["二级模块"]
        l3 = r["三级模块"]
        proc = r["功能过程"]
        desc = r["三级模块整体功能描述"]

        if l1 not in seen_l1:
            seen_l1.add(l1)
            modules.append(FunctionModule(name=l1, level=1))
        if l1 not in seen_l2:
            seen_l2[l1] = set()
        l2_parent = f"{l1}/{l2}" if l2 else ""
        if l2 and l2 not in seen_l2[l1]:
            seen_l2[l1].add(l2)
            modules.append(FunctionModule(name=l2, level=2, parent=l1))
        l3_key = (l1, l2)
        if l3_key not in seen_l3:
            seen_l3[l3_key] = set()
        if l3 not in seen_l3[l3_key]:
            seen_l3[l3_key].add(l3)
            modules.append(FunctionModule(name=l3, level=3, parent=l2_parent,
                                          description=desc))
        procs_key = (l1, l2, l3)
        if procs_key not in l3_procs:
            l3_procs[procs_key] = []
        if proc and proc not in l3_procs[procs_key]:
            l3_procs[procs_key].append(proc)

    for m in modules:
        if m.level == 3:
            parent_parts = m.parent.split("/") if m.parent else []
            l1_name = parent_parts[0] if len(parent_parts) >= 1 else ""
            l2_name = parent_parts[1] if len(parent_parts) >= 2 else m.parent
            procs_key = (l1_name, l2_name, m.name)
            m.children = l3_procs.get(procs_key, [])

    l3_count = len([m for m in modules if m.level == 3])
    logger.debug(f"从表格解析到模块层级: {len(seen_l1)}个一级模块, "
                f"{sum(len(v) for v in seen_l2.values())}个二级模块, "
                f"{l3_count}个三级模块")
    return modules



def is_valid_input_xlsx(xlsx_path: str) -> bool:
    """检查 xlsx 是否为符合规范的功能清单录入文档（至少包含核心功能清单 Sheet）。"""
    import openpyxl

    sheets = load_sheet_names()
    func_sheet = sheets.get("func_list", "")

    try:
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception:
        return False

    valid = func_sheet in wb.sheetnames
    wb.close()
    return valid
