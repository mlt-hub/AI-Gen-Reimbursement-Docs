"""from-excel 管道 —— CLI 和 Web UI 共享的执行编排。

抽取自 main.py 的 --gen-* 逻辑，不依赖 CLI 参数解析或环境变量。
"""

import logging
import os
import re
import shutil
from dataclasses import dataclass, field

from ai_gen_reimbursement_docs.config_utils import (
    load_enable_ai_fill_meta,
    load_fpa_profile,
    load_fpa_rule_set,
    load_fpa_strategy,
    load_spec_remind_update_toc, load_spec_auto_update_toc,
    load_out_templates,
)
from ai_gen_reimbursement_docs.fpa_profiles import resolve_fpa_execution_config
from ai_gen_reimbursement_docs.pipeline_callbacks import PipelineCallbacks
from ai_gen_reimbursement_docs.runtime_context import callbacks_var, current_callbacks
from ai_gen_reimbursement_docs.excel_source import (
    generate_md_files, verify_module_tree_stats
)
from ai_gen_reimbursement_docs.gen_spec import (
    generate_spec_docx_from_md, ai_fill_spec_md, init_spec_template_md, parse_meta_md
)
from ai_gen_reimbursement_docs.gen_fpa import (
    generate_fpa_check_xlsx_from_md,
    generate_fpa_xlsx_from_md,
    init_fpa_template_md,
    plan_fpa_md_from_tree,
)
from ai_gen_reimbursement_docs.gen_list import generate_list_xlsx_from_md
from ai_gen_reimbursement_docs.gen_cosmic import (
    init_cosmic_template_md, ai_fill_cosmic_md, generate_cosmic_xlsx_from_md,
)

logger = logging.getLogger('ai_gen_reimbursement_docs.pipeline')


def _is_web() -> bool:
    """当前线程是否在 Web UI pipeline 中运行。"""
    return current_callbacks().is_web_mode()


def _check_cancelled():
    """Web UI 模式下检查是否被停止，CLI 模式跳过。"""
    if _is_web():
        current_callbacks().check_cancelled()


def _prompt_fpa_reduced(default_value: float) -> float:
    """若 fpa_reduced_use_workload=false，弹输入框让用户确认送审工作量。Web UI / CLI 通用。"""
    from ai_gen_reimbursement_docs.config_utils import load_fpa_reduced_use_workload
    if load_fpa_reduced_use_workload():
        return default_value
    if _is_web():
        return current_callbacks().wait_for_fpa_input(default_value)
    # CLI fallback
    if default_value > 0:
        prompt = f"\n请输入FPA核减后的工作量（人/天）（直接回车使用默认值：{default_value}）: "
    else:
        prompt = "\n请输入FPA核减后的工作量（人/天）: "
    try:
        inp = input(prompt).strip()
        if inp:
            return float(inp)
    except (EOFError, OSError, ValueError, KeyboardInterrupt):
        print()
    return default_value


def _prompt_list_values(md_dir: str, cfp_total: float, fpa_reduced: float) -> tuple[float, float]:
    """gen-list 步骤：让用户确认送审工作量和送审功能点。Web UI / CLI 通用。"""
    from ai_gen_reimbursement_docs.config_utils import load_fpa_reduced_use_workload
    if load_fpa_reduced_use_workload():
        return cfp_total, fpa_reduced
    if _is_web():
        return current_callbacks().wait_for_list_input(cfp_total, fpa_reduced)
    # CLI fallback
    from ai_gen_reimbursement_docs.cli.interactive import prompt_list_values
    return prompt_list_values(md_dir)


def _step(key: str):
    """向前端发送步骤进度事件。key: basedata | fpa | spec | cosmic | list"""
    if _is_web():
        current_callbacks().emit_event({"type": "step", "key": key})


VALID_MODES = {"gen-all", "gen-basedata", "gen-fpa", "gen-cosmic", "gen-list", "gen-spec"}


@dataclass
class PipelineResult:
    """管道执行结果，各字段在对应步骤完成后填充。"""
    tree_md: str = ""
    meta_md: str = ""
    fpa_xlsx: str = ""
    fpa_check_xlsx: str = ""
    cosmic_xlsx: str = ""
    require_xlsx: str = ""
    spec_docx: str = ""
    cfp_total: float = 0.0
    fpa_reduced: float = 0.0
    errors: list[str] = field(default_factory=list)


def run_pipeline(
    *,
    mode: str,
    file_path: str,
    output_dir: str,
    api_key: str = "",
    model: str = "",
    base_url: str = "",
    project_name: str = "",
    templates: dict[str, str] | None = None,
    fpa_reduced: float | None = None,
    cfp_total: float | None = None,
    fpa_profile: str = "",
    fpa_strategy: str = "",
    fpa_rule_set: str = "",
    callbacks: PipelineCallbacks | None = None,
) -> PipelineResult:
    """from-excel 管道总入口。

    Args:
        mode: "gen-all" | "gen-basedata" | "gen-fpa" | "gen-cosmic" | "gen-list" | "gen-spec"
        file_path: 输入 Excel 路径（功能清单）
        output_dir: 交付物输出目录（自动创建）
        api_key: Anthropic API Key
        model: 模型名
        base_url: API 端点
        project_name: 项目名，为空时从 Excel 自动读取
        templates: {"fpa": path, "cosmic": path, "list": path, "spec": path}
        fpa_reduced: FPA 核减后工作量，None 则从 MD 文件自动读取
        cfp_total: 送审功能点数，None 则从 MD 文件自动读取

    Returns:
        PipelineResult：各交付物路径和统计值
    """
    if callbacks is not None and callbacks_var.get() is not callbacks:
        token = callbacks_var.set(callbacks)
        try:
            return run_pipeline(
                mode=mode,
                file_path=file_path,
                output_dir=output_dir,
                api_key=api_key,
                model=model,
                base_url=base_url,
                project_name=project_name,
                templates=templates,
                fpa_reduced=fpa_reduced,
                cfp_total=cfp_total,
                fpa_profile=fpa_profile,
                fpa_strategy=fpa_strategy,
                fpa_rule_set=fpa_rule_set,
            )
        finally:
            callbacks_var.reset(token)

    # 校验
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"功能清单输入文件不存在: {file_path}")
    if mode not in VALID_MODES:
        raise ValueError(f"未知模式: {mode}，支持: {', '.join(sorted(VALID_MODES))}")

    # 日志设置（per-run 文件 handler，CLI / Web UI 共享）
    log_dir = os.path.join(output_dir, '日志')
    os.makedirs(log_dir, exist_ok=True)
    os.environ['AI_REIMBURSEMENT_LOG_DIR'] = log_dir
    os.environ['AI_REIMBURSEMENT_OUTPUT_DIR'] = output_dir
    try:
        from ai_gen_reimbursement_docs.cli.logging import setup_logging
        setup_logging(log_dir, 'AI生成项目报账文档')
    except Exception as e:
        logger.warning("per-run 日志设置失败: %s", e)

    # 准备目录
    os.makedirs(output_dir, exist_ok=True)
    doc_dir = os.path.join(output_dir, 'cosmic文档')
    md_dir = os.path.join(output_dir, 'md')
    os.makedirs(doc_dir, exist_ok=True)
    os.makedirs(md_dir, exist_ok=True)

    # 路径常量（中间文件）
    tree_md = os.path.join(md_dir, '0.1.gen-basedata-功能清单-模块树.md')
    meta_md_tpl = os.path.join(md_dir, '0.2.gen-basedata-录入文档元数据-模板.md')
    meta_filled_md = os.path.join(md_dir, '0.4.gen-basedata-AI填充-录入文档元数据.md')
    fpa_sum_md = os.path.join(md_dir, '1.2.gen-fpa-FPA工作量-总和.md')

    # 项目名称
    if not project_name:
        project_name = read_project_name_from_excel(file_path)
    if not project_name:
        from ai_gen_reimbursement_docs.excel_source import read_project_name
        project_name = read_project_name(meta_md_tpl) if os.path.exists(meta_md_tpl) else ""
    if not project_name:
        project_name = "products"

    from ai_gen_reimbursement_docs.config_utils import load_sheet_names
    _s = load_sheet_names()

    # 模板解析（优先级：传入 > Sheet 8 > 默认 data/templates/）
    templates_dict = _resolve_templates(file_path, templates)

    result = PipelineResult()

    # 交付物文件路径（初始默认值，后续由 _resolve_output_filename 覆盖）
    _default_fpa = os.path.join(output_dir, 'FPA工作量评估.xlsx')
    _default_cosmic = os.path.join(doc_dir, '项目功能点拆分表.xlsx')
    _default_require = os.path.join(doc_dir, '项目需求清单.xlsx')
    _default_spec = os.path.join(doc_dir, '项目需求说明书.docx')

    # 确保基础数据存在，以便读取元数据中的自定义文件名
    if mode == "gen-all":
        logger.info("gen-all全流程模式：按依赖顺序执行...")
    _ensure_basedata_impl(file_path, md_dir, tree_md, meta_md_tpl)
    _fill_meta_if_needed(meta_md_tpl, meta_filled_md, tree_md, api_key, model, base_url)
    _current_meta = meta_filled_md if os.path.exists(meta_filled_md) else meta_md_tpl

    # 从元数据解析输出文件名（Excel Sheet 中配置的「文件名」字段）
    fpa_xlsx = _resolve_output_filename(_current_meta, _s['fpa_meta'],
                                         _default_fpa, output_dir)
    cosmic_xlsx = _resolve_output_filename(_current_meta, _s['cosmic_meta'],
                                            _default_cosmic, doc_dir)
    require_xlsx = _resolve_output_filename(_current_meta, _s['list_meta'],
                                             _default_require, doc_dir)
    spec_docx = _resolve_output_filename(_current_meta, _s['spec_meta'],
                                          _default_spec, doc_dir)

    # ── 模式分发 ──
    if mode == "gen-all":
        result = _generate_all(
            file_path, output_dir, doc_dir, md_dir,
            tree_md, meta_md_tpl, meta_filled_md, fpa_sum_md,
            fpa_xlsx, cosmic_xlsx, require_xlsx, spec_docx,
            templates_dict, api_key, model, base_url, project_name, result,
            fpa_reduced, cfp_total,
            fpa_profile, fpa_strategy, fpa_rule_set,
        )
    elif mode == "gen-basedata":
        result.tree_md = tree_md
        result.meta_md = _current_meta
    elif mode == "gen-fpa":
        meta_md = _current_meta
        result = _generate_fpa(
            file_path, output_dir, md_dir, tree_md, meta_md, fpa_sum_md, fpa_xlsx,
            templates_dict, api_key, model, base_url, result, fpa_profile,
            fpa_strategy, fpa_rule_set,
        )
    elif mode == "gen-cosmic":
        meta_md = _current_meta
        result = _generate_cosmic(
            file_path, md_dir, tree_md, meta_md, fpa_sum_md, doc_dir, cosmic_xlsx,
            templates_dict, api_key, model, base_url, project_name, result,
            fpa_reduced,
        )
    elif mode == "gen-list":
        meta_md = _current_meta
        result = _generate_list(
            md_dir, tree_md, meta_md, doc_dir, require_xlsx,
            templates_dict, result, fpa_reduced, cfp_total,
        )
    elif mode == "gen-spec":
        meta_md = _current_meta
        result = _generate_spec(
            file_path, md_dir, tree_md, meta_md, meta_md_tpl, meta_filled_md,
            doc_dir, spec_docx, templates_dict, api_key, model, base_url, result
        )

    try:
        from ai_gen_reimbursement_docs.cli.logging import write_combined_ai_log
        write_combined_ai_log(mode)
    except Exception as e:
        logger.warning("AI 对话日志生成失败: %s", e)
    return result


# ═══════════════════════════════════════════════════════════════
#  内部辅助
# ═══════════════════════════════════════════════════════════════

def _read_fpa_sum(fpa_sum_md_path: str) -> float:
    """从 FPA工作量-总和.md 读取值，文件不存在返回 0。"""
    import re
    if not os.path.exists(fpa_sum_md_path):
        return 0.0
    with open(fpa_sum_md_path, encoding='utf-8') as f:
        for line in f:
            m = re.search(r'FPA工作量（人/天）[：:]\s*([\d.]+)', line)
            if m:
                return float(m.group(1))
    return 0.0


def _save_fpa_reduced_md(md_dir: str, fpa_reduced: float) -> str:
    """保存 FPA 核减后的工作量到 MD 文件。"""
    path = os.path.join(md_dir, '3.1.gen-cosmic-FPA核减后的工作量-总和.md')
    os.makedirs(md_dir, exist_ok=True)
    with open(path, 'w', encoding='utf-8') as f:
        f.write(f"# FPA 核减后的工作量\n\n")
        f.write(f"FPA核减后的工作量（人/天）: {fpa_reduced}\n")
    logger.info(f"FPA核减后的工作量已保存: {path} ({fpa_reduced})")
    return path


def _read_fpa_reduced_md(md_dir: str) -> float:
    """读取 FPA 核减后工作量，读不到返回 0。"""
    import re
    reduced_md = os.path.join(md_dir, '3.1.gen-cosmic-FPA核减后的工作量-总和.md')
    if os.path.exists(reduced_md):
        with open(reduced_md, encoding='utf-8') as f:
            for line in f:
                m = re.search(r'FPA核减后的工作量（人/天）[：:]\s*([\d.]+)', line)
                if m:
                    return float(m.group(1))
    return 0.0


def _resolve_templates(file_path: str, cli_templates: dict | None) -> dict:
    """解析模板路径，优先级：CLI 参数/Web UI指定 > 配置文件"""
    from ai_gen_reimbursement_docs.excel_source import project_root

    cfg_templates = load_out_templates()
    templates = {}

    for key, config_name, default_filename in [
        ('fpa',    'fpa_out_template',     'FPA工作量评估-输出模板.xlsx'),
        ('spec',   'spec_out_template',     '项目需求说明书-输出模板.docx'),
        ('cosmic', 'cosmic_out_template',   '项目功能点拆分表-输出模板.xlsx'),
        ('list',   'list_out_template',       '项目需求清单-输出模板.xlsx'),
    ]:
        # 1. CLI 参数（最高优先级）
        if cli_templates and key in cli_templates and cli_templates[key]:
            path = cli_templates[key]
            if os.path.exists(path):
                templates[key] = path
                continue

        # 2. 配置文件（system_config.yaml → out_templates）
        cfg_path = cfg_templates.get(config_name, '')
        if cfg_path:
            if not os.path.isabs(cfg_path):
                cfg_path = os.path.join(project_root(), cfg_path)
            if os.path.exists(cfg_path):
                templates[key] = cfg_path
                continue

    # 检查缺失的模板，给出明确提示
    missing = [k for k in ['fpa', 'cosmic', 'list', 'spec'] if not templates.get(k)]
    if missing:
        _names = {'fpa': 'FPA工作量评估', 'cosmic': '项目功能点拆分表',
                   'list': '项目需求清单', 'spec': '项目需求说明书'}
        logger.error(
            f"未找到 {len(missing)} 个输出模板: {', '.join(_names[k] for k in missing)}。"
            f"请在 ~/.ai-gen-reimbursement-docs/system_config.yaml 中配置 out_templates，"
            f"或通过 CLI --{missing[0]}-out-template 参数指定路径。"
        )
    return templates


def _ensure_basedata_impl(file_path: str, md_dir: str,
                          tree_md: str, meta_md_tpl: str) -> None:
    """生成 gen-basedata-*.md 数据源文件。"""
    _step("basedata")
    logger.info("第0步: 生成基础数据")
    generate_md_files(file_path, md_dir)
    verify_module_tree_stats(tree_md, meta_md_tpl)


def _fill_meta_if_needed(meta_md_tpl: str, meta_filled_md: str, tree_md: str,
                         api_key: str, model: str, base_url: str) -> None:
    """AI 填充元数据。"""
    logger.info("第0.4步：AI填充文档元数据...")
    if not api_key:
        logger.info("未设置 API Key，跳过 AI 填充文档元数据")
        return
    from ai_gen_reimbursement_docs.excel_source import ai_fill_meta_md
    if load_enable_ai_fill_meta():
        ai_fill_meta_md(meta_md_tpl, meta_filled_md, api_key, model, base_url, tree_md=tree_md)
    else:
        logger.info("enable_ai_fill_meta=false，跳过 AI 填充，直接复制模板")
        shutil.copy2(meta_md_tpl, meta_filled_md)


def _resolve_output_filename(meta_md_path: str, sheet_name: str,
                             default_path: str, target_dir: str) -> str:
    """从元数据 MD 中读取自定义输出文件名。

    Excel Sheet 各段含「文件名」字段，支持 ${工单编号}、${工单标题} 等占位符。
    若 meta_md 未配置或解析失败，回退到 default_path。
    """
    if not os.path.exists(meta_md_path):
        return default_path
    with open(meta_md_path, encoding='utf-8') as f:
        content = f.read()

    # 定位 sheet 段
    section = re.search(
        rf'##\s*{re.escape(sheet_name)}.*?(?=##|\Z)', content, re.DOTALL
    )
    if not section:
        return default_path

    # 提取「文件名」字段
    m = re.search(r'\|\s*文件名\s*\|\s*(.+?)\s*(?:\||$)', section.group())
    if not m:
        return default_path

    name = m.group(1).strip()

    # 占位符替换
    for placeholder, key in [
        ('${工单编号}', '工单编号'),
        ('${工单名称}', '工单标题'),
        ('${工单标题}', '工单标题'),
        ('${子系统（模块）}', '子系统（模块）'),
    ]:
        pm = re.search(rf'{key}\s*\|\s*(.+?)(?:\s*\||$)', content)
        if pm:
            name = name.replace(placeholder, pm.group(1).strip())

    return os.path.join(target_dir, name)


def read_project_name_from_excel(file_path: str) -> str:
    """从 Excel 的「1、工单需求-元数据录入」sheet 读取工单标题。"""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb['1、工单需求-元数据录入']
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]).strip() == '工单标题':
                name = str(row[1]).strip() if row[1] else ''
                wb.close()
                return name
        wb.close()
    except Exception:
        pass
    return ""


# ═══════════════════════════════════════════════════════════════
#  各子模式实现
# ═══════════════════════════════════════════════════════════════

def _check_template(templates_dict: dict, key: str, name: str):
    """检查模板路径，为空时抛出可读的错误。"""
    path = templates_dict.get(key, '')
    if not path:
        raise FileNotFoundError(
            f"未找到「{name}」模板。"
            f"请在 ~/.ai-gen-reimbursement-docs/system_config.yaml 的 out_templates 中配置，"
            f"或通过 CLI 参数指定。"
        )
    if not os.path.exists(path):
        raise FileNotFoundError(f"「{name}」模板文件不存在: {path}")
    return path


def _generate_fpa(file_path, output_dir, md_dir, tree_md, meta_md,
             fpa_sum_md, fpa_xlsx, templates_dict, api_key, model, base_url, result,
             fpa_profile="", fpa_strategy="", fpa_rule_set=""):
    """第1步：FPA 工作量评估。"""
    _check_cancelled()
    _step("fpa")
    logger.info("第1步: 生成FPA工作量评估...")
    fpa_src = _check_template(templates_dict, 'fpa', 'FPA工作量评估')

    fpa_md = os.path.join(md_dir, '1.1.gen-fpa-FPA-模板.md')
    fpa_filled_md = os.path.join(md_dir, '1.3.gen-fpa-AI填充-FPA.md')
    execution = resolve_fpa_execution_config(
        fpa_profile or load_fpa_profile(),
        fpa_strategy or load_fpa_strategy(),
        fpa_rule_set or load_fpa_rule_set(),
    )
    profile_name = execution.profile.name
    init_fpa_template_md(
        tree_md,
        meta_md,
        fpa_md,
        summary_md_path=fpa_sum_md,
        profile_name=profile_name,
        rule_set=execution.rule_set,
    )

    if execution.strategy in {"ai_first", "ai_only"}:
        plan_fpa_md_from_tree(
            tree_md,
            meta_md,
            fpa_filled_md,
            template_path=fpa_src,
            api_key=api_key,
            model=model,
            base_url=base_url,
            summary_md_path=fpa_sum_md,
            profile_name=profile_name,
            strategy=execution.strategy,
            rule_set=execution.rule_set,
        )
    else:
        fpa_filled_md = fpa_md

    fpa_xlsx = generate_fpa_xlsx_from_md(fpa_filled_md, meta_md, fpa_src, fpa_xlsx)
    fpa_check_xlsx = os.path.splitext(fpa_xlsx)[0] + "-check.xlsx"
    fpa_check_xlsx = generate_fpa_check_xlsx_from_md(fpa_filled_md, tree_md, fpa_check_xlsx)

    from ai_gen_reimbursement_docs.excel_source import read_md_value
    result.fpa_reduced = read_md_value(fpa_sum_md, r'FPA工作量（人/天）[：:]\s*([\d.]+)') or 0.0
    result.fpa_xlsx = fpa_xlsx
    result.fpa_check_xlsx = fpa_check_xlsx
    logger.info(f"FPA工作量评估已生成: {fpa_xlsx}")
    return result


def _generate_cosmic(file_path, md_dir, tree_md, meta_md, fpa_sum_md,
                doc_dir, cosmic_xlsx, templates_dict, api_key, model, base_url,
                project_name, result, fpa_reduced=None):
    """第2步：COSMIC 功能点拆分表。"""
    _check_cancelled()
    _step("cosmic")
    logger.info("第3步：生成项目功能点拆分表...")

    from ai_gen_reimbursement_docs.excel_source import read_project_name, read_md_value

    cosmic_src = _check_template(templates_dict, 'cosmic', '项目功能点拆分表')
    project = read_project_name(meta_md)

    logger.info("第3.1步：生成FPA核减后的工作量 MD ...")
    if fpa_reduced is not None:
        result.fpa_reduced = fpa_reduced
    else:
        result.fpa_reduced = _read_fpa_sum(fpa_sum_md)
    result.fpa_reduced = _prompt_fpa_reduced(result.fpa_reduced)

    # 保存 FPA 核减后的工作量
    _save_fpa_reduced_md(md_dir, result.fpa_reduced)

    init_md_path = os.path.join(md_dir, '3.2.gen-cosmic-COSMIC-模板.md')
    filled_md_path = os.path.join(md_dir, '3.3.gen-cosmic-AI填充-COSMIC.md')
    _, _cosmic_modules = init_cosmic_template_md(tree_md, project, init_md_path)

    if api_key:
        shutil.copy2(init_md_path, filled_md_path)
        ai_fill_cosmic_md(filled_md_path, tree_md, project, api_key, model, base_url, meta_md,
                          modules=_cosmic_modules)
        _cosmic_cfp = _read_cfp_formula_from_meta_md(meta_md)
        cosmic_xlsx = generate_cosmic_xlsx_from_md(filled_md_path, cosmic_src, cosmic_xlsx, meta_md,
                                                   md_dir=md_dir, project_name=project_name,
                                                   cfp_formula=_cosmic_cfp)
        result.cfp_total = read_md_value(
            os.path.join(md_dir, '3.5.gen-cosmic-CFP-总和.md'),
            r'CFP 总和[：:]\s*([\d.]+)') or 0
    else:
        logger.warning("未设置 API Key，无法生成 COSMIC 拆分数据")

    result.cosmic_xlsx = cosmic_xlsx
    return result


def _generate_list(md_dir, tree_md, meta_md,
              doc_dir, require_xlsx, templates_dict, result,
              fpa_reduced=None, cfp_total=None):
    """第3步：需求清单。fpa_reduced/cfp_total 为 None 时从 MD 文件读取默认值并弹窗确认。"""
    _check_cancelled()
    _step("list")
    logger.info("第4步：生成项目需求清单...")
    require_src = _check_template(templates_dict, 'list', '项目需求清单')

    from ai_gen_reimbursement_docs.excel_source import read_md_value
    if cfp_total is None:
        cfp_total = read_md_value(
            os.path.join(md_dir, '3.5.gen-cosmic-CFP-总和.md'),
            r'CFP 总和[：:]\s*([\d.]+)') or 0
    if fpa_reduced is None:
        fpa_reduced = _read_fpa_reduced_md(md_dir)

    cfp_total, fpa_reduced = _prompt_list_values(
        md_dir, float(cfp_total or 0), float(fpa_reduced or 0))

    require_xlsx = generate_list_xlsx_from_md(meta_md, tree_md, require_src, require_xlsx,
                                              cfp_total=cfp_total, fpa_reduced=fpa_reduced)
    result.require_xlsx = require_xlsx
    return result


def _generate_spec(file_path, md_dir, tree_md, meta_md, meta_md_tpl, meta_filled_md,
              doc_dir, spec_docx, templates_dict, api_key, model, base_url, result):
    """需求说明书（无固定顺序依赖）。"""
    _check_cancelled()
    _step("spec")
    logger.info("第2步：生成项目需求说明书...")
    spec_src = _check_template(templates_dict, 'spec', '项目需求说明书')

    # 确保元数据已填充
    if api_key and not os.path.exists(meta_filled_md):
        if load_enable_ai_fill_meta():
            from ai_gen_reimbursement_docs.excel_source import ai_fill_meta_md
            ai_fill_meta_md(meta_md_tpl, meta_filled_md, api_key, model, base_url, tree_md=tree_md)
    if os.path.exists(meta_filled_md):
        meta_md = meta_filled_md

    spec_md = os.path.join(md_dir, '2.1.gen-spec-SPEC-功能需求章节-模板.md')
    spec_filled_md = os.path.join(md_dir, '2.2.gen-spec-AI填充-SPEC-功能需求章节.md')
    init_spec_template_md(tree_md, meta_md, spec_md)
    if api_key:
        ai_fill_spec_md(spec_md, spec_filled_md, api_key, model, base_url)
    else:
        shutil.copy2(spec_md, spec_filled_md)

    spec_docx = generate_spec_docx_from_md(spec_src, spec_docx, meta_md, tree_md, filled_md_path=spec_filled_md)

    # 自动更新目录（Word COM）
    _toc_updated = False
    if load_spec_auto_update_toc():
        from ai_gen_reimbursement_docs.gen_spec import auto_update_docx_toc
        _toc_updated = auto_update_docx_toc(spec_docx)

    # 未自动更新时，按配置添加提醒前缀
    if not _toc_updated and load_spec_remind_update_toc():
        _doc_dir, _doc_name = os.path.split(spec_docx)
        if not _doc_name.startswith("【提醒】请手动更新整个目录"):
            _new_path = os.path.join(_doc_dir, f"【提醒】请手动更新整个目录 {_doc_name}")
            os.rename(spec_docx, _new_path)
            spec_docx = _new_path

    result.spec_docx = spec_docx
    logger.info(f"项目需求说明书已生成: {spec_docx}")
    return result


def _generate_all(file_path, output_dir, doc_dir, md_dir,
             tree_md, meta_md_tpl, meta_filled_md, fpa_sum_md,
             fpa_xlsx, cosmic_xlsx, require_xlsx, spec_docx,
             templates_dict, api_key, model, base_url, project_name, result,
             fpa_reduced=None, cfp_total=None, fpa_profile="", fpa_strategy="", fpa_rule_set=""):
    """全流程：basedata → fpa → spec → cosmic → list（委托独立函数按依赖顺序编排）。"""

    # 入口检查所有模板（提前发现模板缺失）
    _check_template(templates_dict, 'fpa', 'FPA工作量评估')
    _check_template(templates_dict, 'cosmic', '项目功能点拆分表')
    _check_template(templates_dict, 'list', '项目需求清单')
    _check_template(templates_dict, 'spec', '项目需求说明书')

    # 元数据已在 run_pipeline 入口处由 _fill_meta_if_needed 处理
    meta_md = meta_filled_md if os.path.exists(meta_filled_md) else meta_md_tpl

    # Step 1: FPA
    result = _generate_fpa(file_path, output_dir, md_dir, tree_md, meta_md,
                           fpa_sum_md, fpa_xlsx, templates_dict, api_key, model,
                           base_url, result, fpa_profile, fpa_strategy, fpa_rule_set)

    # Step 2: SPEC
    result = _generate_spec(file_path, md_dir, tree_md, meta_md, meta_md_tpl,
                            meta_filled_md, doc_dir, spec_docx, templates_dict,
                            api_key, model, base_url, result)

    # Step 3: COSMIC
    result = _generate_cosmic(file_path, md_dir, tree_md, meta_md, fpa_sum_md,
                              doc_dir, cosmic_xlsx, templates_dict, api_key, model,
                              base_url, project_name, result,
                              fpa_reduced=result.fpa_reduced)

    # Step 4: LIST
    result = _generate_list(md_dir, tree_md, meta_md, doc_dir, require_xlsx,
                            templates_dict, result,
                            cfp_total=cfp_total,
                            fpa_reduced=result.fpa_reduced)

    logger.info("全流程完成")
    return result


def _read_cfp_formula_from_meta_md(meta_md: str) -> str:
    """从 gen-basedata-AI填充-录入文档元数据.md 读取 CFP 计算公式。
    在「6、项目功能点拆分表-元数据录入」section 中查找 key 为「CFP计算公式」的行。
    未配置时返回空字符串。
    """
    if not meta_md or not os.path.exists(meta_md):
        return ""
    with open(meta_md, 'r', encoding='utf-8') as f:
        content = f.read()
    # 定位到 COSMIC 元数据 section
    marker = "## 6、项目功能点拆分表-元数据录入"
    idx = content.find(marker)
    if idx < 0:
        return ""
    # 在 section 内查找表格行 | CFP计算公式 | xxx |
    section = content[idx:]
    for line in section.split('\n'):
        line = line.strip()
        if line.startswith('|') and ('CFP计算公式' in line or 'cfp_formula' in line):
            parts = [p.strip() for p in line.split('|')]
            if len(parts) >= 3:
                return parts[2].replace('\\\"', '\"')
    return ""


def _try_read_project_name(excel_path: str) -> str:
    """从 Excel 功能清单的元数据 Sheet 读取工单标题。"""
    import openpyxl

    from ai_gen_reimbursement_docs.config_utils import load_sheet_names

    sheets = load_sheet_names()
    meta_sheet = sheets.get("work_order_meta", "1、工单需求-元数据录入")
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        if meta_sheet not in wb.sheetnames:
            wb.close()
            return ""
        ws = wb[meta_sheet]
        for row in ws.iter_rows(min_row=2, values_only=True):
            key = str(row[0]).strip() if row[0] else ""
            val = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if key == "工单标题":
                wb.close()
                return val
        wb.close()
    except Exception:
        pass
    return ""


def run_pipeline_simple(
    *,
    mode: str,
    file_path: str,
    output_dir: str = "",
    api_key: str = "",
    model: str = "",
    base_url: str = "",
    project_name: str = "",
    templates: dict | None = None,
    fpa_profile: str = "",
    fpa_strategy: str = "",
    fpa_rule_set: str = "",
    callbacks: PipelineCallbacks | None = None,
) -> PipelineResult:
    """一站式管道入口，CLI / Web UI / 零参数 共享。

    自动处理：配置回退、工单标题读取、交付物输出目录创建、日志设置。
    """
    from ai_gen_reimbursement_docs.config_utils import (
        load_api_key, load_base_url, load_model_name,
    )

    api_key = api_key or load_api_key()
    model = model or load_model_name()
    base_url = base_url or load_base_url()

    if api_key:
        os.environ["ANTHROPIC_API_KEY"] = api_key
    if base_url:
        os.environ["ANTHROPIC_BASE_URL"] = base_url

    excel_dir = os.path.dirname(os.path.abspath(file_path))
    if output_dir:
        out_dir = output_dir
    elif project_name:
        safe = re.sub(r'[\/:*?"<>|]', '_', project_name)
        out_dir = os.path.join(excel_dir, safe)
    else:
        auto_name = _try_read_project_name(file_path)
        if auto_name:
            safe = re.sub(r'[\/:*?"<>|]', '_', auto_name)
            out_dir = os.path.join(excel_dir, safe)
            project_name = auto_name
        else:
            out_dir = excel_dir

    logger.info(f"功能清单输入文件: {os.path.basename(file_path)}")
    logger.info(f"运行模式: {'Web UI' if _is_web() else 'CLI'}")
    if project_name:
        logger.info(f"项目名称: {project_name}")
    logger.info(f"交付物输出目录: {out_dir}")

    return run_pipeline(
        mode=mode,
        file_path=file_path,
        output_dir=out_dir,
        api_key=api_key,
        model=model,
        base_url=base_url,
        project_name=project_name,
        templates=templates,
        fpa_profile=fpa_profile or load_fpa_profile(),
        fpa_strategy=fpa_strategy or load_fpa_strategy(),
        fpa_rule_set=fpa_rule_set or load_fpa_rule_set(),
        callbacks=callbacks,
    )
