"""生成 FPA工作量评估.xlsx"""

import json as _json
import logging
import os
import re
import shutil
import subprocess
import tempfile
import hashlib
from difflib import SequenceMatcher
from copy import copy
from dataclasses import dataclass, field
from datetime import datetime
from typing import Any

import openpyxl
from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles import Alignment, Font, Border
from openpyxl.styles import PatternFill

from ai_gen_reimbursement_docs.constants import (
    FPA_COL_SEQ, FPA_COL_SUBSYSTEM, FPA_COL_ASSET, FPA_COL_FUNC_POINT,
    FPA_COL_TYPE, FPA_COL_CLASSIFICATION, FPA_COL_EXPLANATION, FPA_COL_STATUS,
    FPA_COL_FORMULA_BASE, FPA_COL_ADJUST, FPA_COL_ELEMENTS, FPA_COL_FORMULA_WORKLOAD,
    FPA_TOTAL_COLS, FPA_COL_KEY_MAP,
)
from ai_gen_reimbursement_docs.excel_source import (
    replace_placeholders, strip_ai_marker, parse_module_tree_md,
    read_base_data_from_excel, safe_load_workbook,
)
from ai_gen_reimbursement_docs.fpa_confirmation import (
    build_fpa_confirmation_questions,
    confirmation_feedback,
    confirmed_decision_count,
    normalize_confirmed_decisions,
    normalize_confirmation_mode,
)
from ai_gen_reimbursement_docs.fpa_agent_review import build_fpa_agent_review
from ai_gen_reimbursement_docs.fpa_profiles import (
    CUSTOM_RULES_PROFILE,
    CustomRulesProfile,
    FpaRuleSetConfig,
    adjust_value_for_type as _adjust_value_for_type,
    basis_for_fpa_type as _basis_for_fpa_type,
    calculate_fpa_adjustment_for_row,
    current_fpa_rule_set_config,
    get_fpa_profile,
    group_tag as _group_tag,
    module_change_status as _module_change_status,
    reset_current_fpa_rule_set_config,
    resolve_fpa_execution_config,
    set_current_fpa_rule_set_config,
)
from ai_gen_reimbursement_docs.template_manifest import load_template_manifest
from ai_gen_reimbursement_docs.fpa_quality_review import (
    build_fpa_quality_review,
    quality_feedback,
    retryable_quality_issues,
)
from ai_gen_reimbursement_docs.fpa_stability_report import build_fpa_stability_report
from ai_gen_reimbursement_docs.fpa_type_judgement import build_fpa_type_judgement
from ai_gen_reimbursement_docs.fpa_validator import (
    FpaValidationIssue,
    retryable_validation_issues,
    validate_fpa_rows,
    validation_feedback,
)
from ai_gen_reimbursement_docs.md_table import parse_md_table_row
from ai_gen_reimbursement_docs.runtime_context import current_callbacks

logger = logging.getLogger('ai_gen_reimbursement_docs.gen_fpa')

VALID_FPA_TYPES = {"EI", "ILF", "EQ", "EO", "EIF"}
FPA_PROFILE = CUSTOM_RULES_PROFILE
RULE_HITS_KEY = "_规则命中详情"
CONFIG_WARNING_PREFIX = "FPA 配置 warning:"
EXPLANATION_REQUIRED_LABELS = ("来源场景：", "业务数据：", "业务规则：", "计算说明：")
EXPLANATION_MISSING_HINTS = ("未识别到", "未明确说明", "需求未明确说明")
FPA_PROJECT_DESCRIPTION_MAX_CHARS = 5000
EXPLANATION_TABLE_COUNT_DETAIL_HINTS = ("数据库表个数=", "表个数=", "表数量", "1张表", "1 张表", "1个表", "1 个表")
EXPLANATION_SYSTEM_ELEMENT_MARKERS = ("表", "服务", "接口", "文件", "系统", "平台")
EXPLANATION_INLINE_SYSTEM_ELEMENT_MARKERS = ("表", "服务", "接口", "文件", "平台")
EXPLANATION_SYSTEM_ELEMENT_SKIP_HINTS = ("未识别到", "未明确", "无明确", "没有明确", "未涉及")
EXPLANATION_INLINE_SYSTEM_ELEMENT_SKIP_HINTS = (
    "按后台数据库变更的表个数计量",
    "按数据库表个数计量",
    "表个数计量",
    "输出的票据、报表、统计、文件",
    "输出格式化文件",
    "生成文件",
    "文件生成",
    "内部逻辑文件",
    "外部接口文件",
    "接口开发行",
    "界面开发行",
)
BROAD_SYSTEM_ELEMENT_CANDIDATES = (
    "服务或外部系统",
    "外部接口",
    "外部文件",
    "接口开发",
    "界面开发",
)
NON_SYSTEM_ELEMENT_CANDIDATE_SUFFIXES = (
    "列表",
    "代表",
    "表示",
)
BASIS_TYPE_HINTS = {
    "EI": ("外部输入", "修改或增加界面", "插入、修改、删除", "输入界面"),
    "EQ": ("外部查询", "查询界面"),
    "EO": ("外部输出", "输出结果", "报表", "统计"),
    "ILF": ("后台数据库变更", "内部逻辑文件", "内部逻辑数据"),
    "EIF": ("外部接口文件", "外部逻辑文件", "外部数据组", "外部系统维护"),
}
EXPLANATION_STRUCTURED_LABELS = (
    "来源场景：", "来源场景:",
    "业务数据：", "业务数据:",
    "业务规则：", "业务规则:",
    "系统元素：", "系统元素:",
    "计算说明：", "计算说明:",
)
FPA_TYPE_EXPLANATION_ALIASES = {
    "EI": (
        "外部输入", "输入事务", "维护类事务", "维护类EI", "维护类 EI",
        "对ILF进行插入", "对 ILF 进行插入", "对ILF的插入", "对 ILF 的插入",
        "维护内部数据", "改变本系统内部数据", "修改或增加界面",
    ),
    "EQ": (
        "外部查询", "查询类事务", "查询类EQ", "查询类 EQ",
        "提供查询界面输入并展示返回结果", "查询界面输入并展示返回结果",
        "查询界面", "展示返回结果",
    ),
    "EO": (
        "外部输出", "输出类事务", "输出类EO", "输出类 EO",
        "文件输出", "输出文件", "输出的文件", "格式化文件输出", "生成并输出",
        "输出的票据、报表、统计、文件",
    ),
    "ILF": (
        "内部逻辑数据", "内部逻辑文件", "内部数据功能", "内部数据组",
        "逻辑数据集合", "本系统维护的数据组", "本系统维护的逻辑数据集合",
    ),
    "EIF": (
        "外部逻辑数据", "外部数据功能", "外部数据组", "本系统引用但不维护",
        "外部接口文件", "外部引用数据组", "外部系统数据", "外部系统维护",
        "评估范围外相关的表", "外部引用数据",
    ),
}


@dataclass(frozen=True)
class FpaPromptContext:
    """Rendered FPA prompts plus user-safe source labels."""

    system_prompt: str
    user_prompt: str
    core_rules: str
    core_rules_source: str
    system_prompt_source: str
    user_prompt_source: str


@dataclass
class FpaAuditReport:
    """FPA 预览/审核的结构化信息。"""

    profile: str
    profile_version: str
    strategy: str
    rule_set: str
    module: dict[str, object]
    process_total: int
    covered_processes: list[str]
    missing_processes: list[str]
    generation_counts: dict[str, int]
    warnings: list[str] = field(default_factory=list)
    raw_source: str = ""
    raw_rows: list[object] = field(default_factory=list)
    raw_warnings: list[str] = field(default_factory=list)
    rule_hits: list[dict[str, object]] = field(default_factory=list)
    agent_review: dict[str, object] = field(default_factory=dict)

    def to_dict(self) -> dict[str, object]:
        return {
            "profile": self.profile,
            "profile_version": self.profile_version,
            "strategy": self.strategy,
            "rule_set": self.rule_set,
            "module": self.module,
            "coverage": {
                "process_total": self.process_total,
                "covered_count": len(self.covered_processes),
                "missing_count": len(self.missing_processes),
                "covered_processes": self.covered_processes,
                "missing_processes": self.missing_processes,
            },
            "generation_counts": self.generation_counts,
            "warnings": self.warnings,
            "raw_ai": {
                "source": self.raw_source,
                "warnings": self.raw_warnings,
                "raw_rows": self.raw_rows,
            },
            "rule_hits": self.rule_hits,
            "agent_review": self.agent_review,
        }


def parse_meta_md(meta_md_path: str) -> dict[str, str]:
    """解析文档元数据.md 为扁平字典。支持跨多行的表格值。"""
    from ai_gen_reimbursement_docs.gen_spec import parse_meta_md
    return parse_meta_md(meta_md_path)


def _explanation_quality_warnings(
    *,
    group: dict[str, object],
    name: str,
    fpa_type: str,
    explanation: str,
    profile_name: str,
) -> list[str]:
    """Check formal FPA calculation explanation quality without changing output."""
    text = str(explanation or "").strip()
    if not text:
        return []

    warnings: list[str] = []
    if profile_name == "strict_fpa":
        missing_labels = [label.rstrip("：") for label in EXPLANATION_REQUIRED_LABELS if label not in text]
        if missing_labels:
            warnings.append(
                f"{name} 计算依据说明格式不完整，缺少结构化项: {'、'.join(missing_labels)}"
            )

    source_prefix = (
        f"【{group.get('client_type', '')}】"
        f"{group.get('l1', '')}-{group.get('l2', '')}-{group.get('l3', '')}-"
    )
    if (
        source_prefix.strip("-")
        and source_prefix not in text
        and not _explanation_has_source_anchor(group=group, explanation=text, name=name)
    ):
        expected_tail = "<数据组名称>" if fpa_type in {"ILF", "EIF"} else "<功能点名称>"
        warnings.append(
            f"{name} 计算依据说明来源场景未使用完整路径格式: {source_prefix}{expected_tail}"
        )

    type_aliases = FPA_TYPE_EXPLANATION_ALIASES.get(fpa_type, ())
    if fpa_type and fpa_type not in text and not any(alias in text for alias in type_aliases):
        warnings.append(f"{name} 计算依据说明的计算说明未明确当前 FPA 类型: {fpa_type}")

    if any(hint in text for hint in EXPLANATION_MISSING_HINTS):
        warnings.append(f"{name} 正式计算依据说明包含缺失提示，应移入 check/debug 输出")

    if any(hint in text for hint in EXPLANATION_TABLE_COUNT_DETAIL_HINTS):
        warnings.append(
            f"{name} 计算依据说明疑似将数据库表个数作为详细计量解释，应保留在计算依据归类而非计算依据说明"
        )

    fabricated_elements = _suspected_fabricated_system_elements(
        group=group,
        name=name,
        explanation=text,
    )
    if fabricated_elements:
        warnings.append(
            f"{name} 计算依据说明的系统元素疑似包含输入未明确提供的表、服务、接口、文件或外部系统: "
            + "、".join(fabricated_elements)
        )

    inline_fabricated_elements = _suspected_inline_fabricated_system_elements(
        group=group,
        name=name,
        explanation=text,
    )
    if inline_fabricated_elements:
        warnings.append(
            f"{name} 计算依据说明正文疑似提到输入未明确提供的表、服务、接口、文件或平台，需人工复核: "
            + "、".join(inline_fabricated_elements)
        )

    return warnings


def _source_text_for_explanation_system_elements(
    *,
    group: dict[str, object],
    name: str,
) -> str:
    parts = [
        str(name or ""),
        str(group.get("client_type", "") or ""),
        str(group.get("l1", "") or ""),
        str(group.get("l2", "") or ""),
        str(group.get("l3", "") or ""),
        str(group.get("l3_desc", "") or ""),
    ]
    processes = group.get("processes", [])
    if isinstance(processes, list):
        for process in processes:
            if not isinstance(process, dict):
                continue
            parts.extend([
                str(process.get("id", "") or ""),
                str(process.get("process_id", "") or ""),
                str(process.get("name", "") or ""),
                str(process.get("process_name", "") or ""),
                str(process.get("desc", "") or ""),
                str(process.get("description", "") or ""),
            ])
    return "\n".join(part for part in parts if part)


def _system_element_lines(explanation: str) -> list[str]:
    lines = []
    in_system_element_block = False
    for line in str(explanation or "").splitlines():
        clean = line.strip()
        if clean.startswith("系统元素：") or clean.startswith("系统元素:"):
            in_system_element_block = True
            value = clean.split("：", 1)[-1] if "：" in clean else clean.split(":", 1)[-1]
            if value.strip():
                lines.append(value)
            continue
        if not in_system_element_block:
            continue
        if any(clean.startswith(label) for label in EXPLANATION_STRUCTURED_LABELS):
            in_system_element_block = False
            continue
        if clean:
            lines.append(re.sub(r"^[\-*•·\d.、）)\s]+", "", clean).strip())
    return lines


def _candidate_system_elements(system_element_text: str) -> list[str]:
    candidates: list[str] = []
    for chunk in re.split(r"[、，,；;\n]+", system_element_text):
        item = chunk.strip(" 。.：:（）()[]【】")
        if not item or any(hint in item for hint in EXPLANATION_SYSTEM_ELEMENT_SKIP_HINTS):
            continue
        if not any(marker in item for marker in EXPLANATION_SYSTEM_ELEMENT_MARKERS):
            continue
        item = re.sub(r"^(?:涉及|调用|对接|访问|使用|通过|依赖|外部|本系统|系统元素)\s*", "", item).strip()
        item = re.sub(r"(?:用于|支撑|完成|返回|提供).*$", "", item).strip(" 。.：:（）()[]【】")
        cn_match = re.search(r"[\u4e00-\u9fffA-Za-z0-9_（）()·-]{2,40}(?:表|服务|接口|文件|系统|平台)", item)
        if cn_match:
            candidate = _normalize_system_element_candidate(cn_match.group(0))
            if candidate:
                candidates.append(candidate)
            continue
        latin_match = re.search(r"[A-Za-z][A-Za-z0-9_./-]{1,}", item)
        if latin_match:
            candidate = _normalize_system_element_candidate(latin_match.group(0))
            if candidate:
                candidates.append(candidate)
    return list(dict.fromkeys(candidates))


def _normalize_system_element_candidate(candidate: str) -> str:
    clean = str(candidate or "").strip(" 。.：:（）()[]【】")
    if not clean:
        return ""
    clean = re.sub(r"[、,，/]*(?:EI|EQ|EO|ILF|EIF)\b.*$", "", clean, flags=re.IGNORECASE).strip(" 、,，/。.")
    if not clean or clean.upper() in VALID_FPA_TYPES:
        return ""
    if clean.endswith(NON_SYSTEM_ELEMENT_CANDIDATE_SUFFIXES):
        return ""
    if any(hint in clean for hint in BROAD_SYSTEM_ELEMENT_CANDIDATES):
        return ""
    if "或" in clean and any(marker in clean for marker in EXPLANATION_SYSTEM_ELEMENT_MARKERS):
        return ""
    return clean


def _non_system_element_lines(explanation: str) -> list[str]:
    lines: list[str] = []
    in_system_element_block = False
    for line in str(explanation or "").splitlines():
        clean = line.strip()
        if clean.startswith("系统元素：") or clean.startswith("系统元素:"):
            in_system_element_block = True
            continue
        if any(clean.startswith(label) for label in EXPLANATION_STRUCTURED_LABELS):
            in_system_element_block = False
        if (
            not in_system_element_block
            and clean
            and (clean.startswith("业务规则：") or clean.startswith("业务规则:") or clean.startswith("计算说明：") or clean.startswith("计算说明:"))
        ):
            lines.append(clean)
    return lines


def _candidate_inline_system_elements(explanation_text: str) -> list[str]:
    candidates: list[str] = []
    for chunk in re.split(r"[、，,；;。\n]+", explanation_text):
        item = chunk.strip(" 。.：:（）()[]【】")
        if not item or any(hint in item for hint in EXPLANATION_SYSTEM_ELEMENT_SKIP_HINTS):
            continue
        if any(hint in item for hint in EXPLANATION_INLINE_SYSTEM_ELEMENT_SKIP_HINTS):
            continue
        if "文件" in item and any(action in item for action in ("导出", "输出", "生成", "下载")):
            continue
        if not any(marker in item for marker in EXPLANATION_INLINE_SYSTEM_ELEMENT_MARKERS):
            continue
        item = re.sub(r"^(?:来源场景|业务数据|业务规则|计算说明)[:：]\s*", "", item).strip()
        item = re.sub(
            r"^.*?(?:涉及|调用|对接|访问|使用|通过|依赖|写入|读取|保存到|同步到|上传|下载|导入|导出|生成)",
            "",
            item,
        ).strip()
        item = re.sub(r"(?:用于|支撑|完成|返回|提供|并|以|，|,).*$", "", item).strip(" 。.：:（）()[]【】")
        cn_match = re.search(
            r"[\u4e00-\u9fffA-Za-z0-9_（）()·-]{2,30}(?:表|服务|接口|文件|平台)",
            item,
        )
        if cn_match:
            candidate = _normalize_system_element_candidate(cn_match.group(0))
            if candidate:
                candidates.append(candidate)
            continue
        latin_match = re.search(r"[A-Za-z][A-Za-z0-9_./-]{1,}", item)
        if latin_match:
            candidate = _normalize_system_element_candidate(latin_match.group(0))
            if candidate:
                candidates.append(candidate)
    return list(dict.fromkeys(candidates))


def _suspected_fabricated_system_elements(
    *,
    group: dict[str, object],
    name: str,
    explanation: str,
) -> list[str]:
    source_text = _source_text_for_explanation_system_elements(group=group, name=name)
    if not source_text.strip():
        return []
    suspicious: list[str] = []
    for line in _system_element_lines(explanation):
        for candidate in _candidate_system_elements(line):
            if candidate and candidate not in source_text:
                suspicious.append(candidate)
    return list(dict.fromkeys(suspicious))


def _suspected_inline_fabricated_system_elements(
    *,
    group: dict[str, object],
    name: str,
    explanation: str,
) -> list[str]:
    source_text = _source_text_for_explanation_system_elements(group=group, name=name)
    if not source_text.strip():
        return []
    suspicious: list[str] = []
    for line in _non_system_element_lines(explanation):
        for candidate in _candidate_inline_system_elements(line):
            if candidate and candidate not in source_text:
                suspicious.append(candidate)
    return list(dict.fromkeys(suspicious))


def _explanation_needs_full_source_path(
    *,
    group: dict[str, object],
    name: str,
    explanation: str,
) -> bool:
    text = str(explanation or "").strip()
    if not text:
        return False
    source_prefix = (
        f"【{group.get('client_type', '')}】"
        f"{group.get('l1', '')}-{group.get('l2', '')}-{group.get('l3', '')}-"
    )
    return (
        bool(source_prefix.strip("-"))
        and source_prefix not in text
        and not _explanation_has_source_anchor(group=group, explanation=text, name=name)
    )


def _normalize_explanation_source_path(
    *,
    explanation: str,
    source_name: str,
) -> str:
    text = str(explanation or "").strip()
    clean_source = str(source_name or "").strip()
    if not text or not clean_source or "来源场景：" not in text:
        return text
    lines = text.splitlines()
    for index, line in enumerate(lines):
        if not line.startswith("来源场景："):
            continue
        lines[index] = f"来源场景：{clean_source}"
        return "\n".join(lines)
    return text


def _normalize_explanation_table_count_detail(explanation: str) -> str:
    text = str(explanation or "").strip()
    if not text:
        return text
    # Keep the official classification wording in 计算依据归类, but remove
    # low-value parenthetical table-count details from the formal explanation.
    text = re.sub(
        r"（(?:保守按)?1\s*(?:个|张)表(?:/数据组)?(?:对应1\s*个(?:ILF|EIF|EI|EO|EQ))?）",
        "",
        text,
    )
    return re.sub(
        r"[，,]\s*对应后台数据库变更的1\s*个表",
        "",
        text,
    )


def _explanation_has_source_anchor(
    *,
    group: dict[str, object],
    explanation: str,
    name: str,
) -> bool:
    """Accept source-scene text that names the module or source processes."""
    text = str(explanation or "")
    processes = group.get("processes", [])
    if isinstance(processes, list):
        process_anchors = [
            str(process.get("name", "") or "").strip()
            for process in processes
            if isinstance(process, dict)
        ]
        if any(anchor and anchor in text for anchor in process_anchors):
            return True
    source_anchor_markers = ("模块描述", "业务场景", "功能过程", "流程", "场景「", "操作")
    if not any(marker in text for marker in source_anchor_markers):
        return False
    anchors = [
        str(group.get("l3", "") or "").strip(),
        str(name or "").rsplit("-", maxsplit=1)[-1].strip(),
    ]
    return any(anchor and anchor in text for anchor in anchors)


def _receiver_from_client_type(client_type: str, rules_text: str) -> str:
    """根据客户端类型判定接收者。"""
    default = "操作员"
    if not rules_text:
        logger.warning(
            "Excel 模板未配置「功能用户-接收者判定」，接收者将使用默认值，请在模板 Sheet 6 中补充"
        )
        return default

    for line in rules_text.split('\n'):
        line = line.strip()
        if '：' in line:
            keyword, receiver = line.split('：', 1)
            if keyword in client_type:
                return receiver.strip()
    return default


def _group_rows_by_l3(rows: list[dict[str, str]]) -> list[dict[str, object]]:
    """按客户端/一二三级模块聚合功能过程，供 FPA 三级模块规划使用。"""
    groups: list[dict[str, object]] = []
    index: dict[tuple[str, str, str, str], dict[str, object]] = {}
    for r in rows:
        key = (
            str(r.get("客户端类型", "")).strip(),
            str(r.get("一级模块", "")).strip(),
            str(r.get("二级模块", "")).strip(),
            str(r.get("三级模块", "")).strip(),
        )
        group = index.get(key)
        if group is None:
            group = {
                "client_type": key[0],
                "l1": key[1],
                "l2": key[2],
                "l3": key[3],
                "l3_desc": str(r.get("三级模块整体功能描述", "") or "").strip(),
                "processes": [],
            }
            index[key] = group
            groups.append(group)
        elif not group.get("l3_desc") and r.get("三级模块整体功能描述"):
            group["l3_desc"] = str(r.get("三级模块整体功能描述", "")).strip()
        processes = group["processes"]
        if isinstance(processes, list):
            process_id = f"m{groups.index(group) + 1}_p{len(processes) + 1}"
            process_name = str(r.get("功能过程", "") or "").strip()
            process_desc = str(r.get("功能过程描述", "") or "").strip()
            processes.append({
                "process_id": process_id,
                "process_name": process_name,
                "description": process_desc,
                "name": process_name,
                "change_status": str(r.get("变更状态", "") or "").strip(),
                "desc": process_desc,
            })
    return groups


def _call_llm(
    prompt: str,
    system_prompt: str,
    api_key: str,
    model: str,
    base_url: str,
    tag: str = "",
    return_thinking: bool = False,
) -> str | tuple[str, str]:
    """调用 LLM（委托至 llm_client 公共模块）。"""
    from ai_gen_reimbursement_docs.llm_client import call_llm

    try:
        return call_llm(
            prompt=prompt, system=system_prompt,
            api_key=api_key, model=model, base_url=base_url, tag=tag,
            return_thinking=return_thinking,
        )
    except Exception as e:
        logger.warning("LLM 调用失败 [%s]: %s", tag, e)
        if return_thinking:
            return "", ""
        return ""


def _build_fpa_rule_rows(
    rows: list[dict[str, str]],
    meta: dict[str, str],
    profile: CustomRulesProfile = FPA_PROFILE,
) -> list[dict[str, object]]:
    """从功能清单行构建 FPA 模板行（三级模块兜底骨架）。"""
    if current_fpa_rule_set_config() is None:
        execution = resolve_fpa_execution_config(profile.name)
        token = set_current_fpa_rule_set_config(execution.rule_set_config)
        try:
            return _build_fpa_rule_rows(rows, meta, profile=profile)
        finally:
            reset_current_fpa_rule_set_config(token)

    fpa_rows: list[dict[str, object]] = []
    seq = 1
    for group in _group_rows_by_l3(rows):
        group_rows = profile.fallback_rows_for_l3(group, meta, start_seq=seq)
        fpa_rows.extend(group_rows)
        seq += len(group_rows)
    return fpa_rows


def _fpa_judgement_rules_sheet_spec(manifest: dict[str, Any]) -> dict[str, Any]:
    sheets = manifest.get("sheets", {}) or {}
    spec = sheets.get("judgement_rules", {}) if isinstance(sheets, dict) else {}
    if isinstance(spec, str):
        spec = {"name": spec}
    if not isinstance(spec, dict):
        spec = {}
    header_row = _fpa_optional_manifest_int(spec.get("header_row"))
    data_start_row = _fpa_optional_manifest_int(spec.get("data_start_row"))
    if not data_start_row:
        data_start_row = header_row + 1 if header_row else 2
    return {
        "name": str(spec.get("name") or "附录1-FPA评估方法说明"),
        "header_row": header_row,
        "rule_header": str(spec.get("rule_header", spec.get("header", "")) or "").strip(),
        "data_start_row": data_start_row,
        "data_end_row": _fpa_optional_manifest_int(spec.get("data_end_row")),
        "max_rows": _fpa_optional_manifest_int(spec.get("max_rows")),
        "column": _fpa_judgement_rules_column(spec.get("column", spec.get("rule_column")), 3),
        "anchor": spec.get("anchor", {}) if isinstance(spec.get("anchor", {}), (dict, str)) else {},
    }


def _fpa_optional_manifest_int(value: object) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return 0
    return parsed if parsed > 0 else 0


def _fpa_judgement_rules_column(value: object, default: int) -> int:
    if value is None:
        return default
    try:
        if isinstance(value, str) and value.strip().isalpha():
            return column_index_from_string(value.strip())
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return parsed if parsed > 0 else default


def _fpa_judgement_rules_header_column(ws, spec: dict[str, Any]) -> int:
    header_row = int(spec.get("header_row") or 0)
    rule_header = str(spec.get("rule_header", "") or "").strip()
    if not header_row or not rule_header or header_row > ws.max_row:
        return int(spec["column"])
    for cell in ws[header_row]:
        if str(cell.value or "").strip() == rule_header:
            return int(cell.column)
    logger.warning(
        "fpa manifest judgement_rules 未在第 %s 行找到表头 %s，回退到列 %s",
        header_row,
        rule_header,
        spec["column"],
    )
    return int(spec["column"])


def _fpa_judgement_rules_anchor_start(ws, spec: dict[str, Any]) -> tuple[int, int]:
    anchor = spec.get("anchor", {}) or {}
    if not anchor:
        return int(spec["data_start_row"]), _fpa_judgement_rules_header_column(ws, spec)
    if isinstance(anchor, str):
        anchor = {"contains": anchor}
    if not isinstance(anchor, dict):
        return int(spec["data_start_row"]), _fpa_judgement_rules_header_column(ws, spec)

    offset_rows = _as_manifest_int(anchor.get("offset_rows"), 1)
    column = _fpa_judgement_rules_column(anchor.get("column"), int(spec["column"]))
    cell_ref = str(anchor.get("cell", "") or "").strip()
    contains = str(anchor.get("contains", "") or "").strip()

    if cell_ref:
        cell = ws[cell_ref]
        return cell.row + offset_rows, column
    if contains:
        for row in ws.iter_rows():
            for cell in row:
                if contains in str(cell.value or ""):
                    return cell.row + offset_rows, column
    return int(spec["data_start_row"]), int(spec["column"])


def _read_fpa_judgement_rules_from_template(template_path: str = "") -> list[str]:
    judgement_rules: list[str] = []
    if not template_path:
        return judgement_rules
    wb = None
    try:
        manifest, _, _ = load_template_manifest("fpa", template_path)
        sheet_spec = _fpa_judgement_rules_sheet_spec(manifest)
        wb = openpyxl.load_workbook(template_path, data_only=True)
        from ai_gen_reimbursement_docs.config_utils import _get_system_config_value
        appendix_sheet = _get_system_config_value('fpa_appendix_sheet', sheet_spec["name"])
        ws = wb[appendix_sheet]
        start_row, column = _fpa_judgement_rules_anchor_start(ws, sheet_spec)
        end_row = min(sheet_spec["data_end_row"], ws.max_row) if sheet_spec["data_end_row"] else ws.max_row
        if sheet_spec["max_rows"]:
            end_row = min(end_row, start_row + sheet_spec["max_rows"] - 1)
        for row_num in range(start_row, end_row + 1):
            val = ws.cell(row_num, column).value
            if val and str(val).strip():
                judgement_rules.append(str(val).strip())
        if judgement_rules:
            logger.debug("从模板附录读取判定原则 %d 条", len(judgement_rules))
    except Exception as e:
        logger.warning("从模板附录读取判定原则失败: %s", e)
    finally:
        if wb is not None:
            wb.close()
    return judgement_rules


def _read_fpa_judgement_rules(template_path: str = "") -> list[str]:
    from ai_gen_reimbursement_docs.config_utils import (
        load_fpa_judgement_rules_config,
        load_fpa_judgement_rules_source,
    )

    source = load_fpa_judgement_rules_source()
    if source == "config":
        return load_fpa_judgement_rules_config()
    return _read_fpa_judgement_rules_from_template(template_path)


def _fpa_type_from_classification_basis(basis: str) -> str:
    text = str(basis or "").strip()
    if not text:
        return ""
    upper_text = text.upper()
    for fpa_type in ("ILF", "EIF", "EI", "EQ", "EO"):
        if re.match(rf"^{fpa_type}(?:\b|[:：])", upper_text):
            return fpa_type
    for fpa_type, hints in BASIS_TYPE_HINTS.items():
        if any(hint in text for hint in hints):
            return fpa_type
    return ""


def _extract_json_obj(resp: str) -> dict[str, Any]:
    from ai_gen_reimbursement_docs.llm_client import strip_markdown_code_block

    clean = strip_markdown_code_block(resp or "").strip()
    try:
        data = _json.loads(clean)
    except _json.JSONDecodeError:
        start = clean.find("{")
        end = clean.rfind("}")
        if start < 0 or end <= start:
            raise
        data = _json.loads(clean[start:end + 1])
    if isinstance(data, list):
        data = {"rows": data}
    if not isinstance(data, dict):
        raise ValueError("AI 响应 JSON 根节点必须是对象或数组")
    return data


def _row_rule_hits(row: dict[str, object]) -> list[dict[str, object]]:
    hits = row.get(RULE_HITS_KEY)
    if not isinstance(hits, list):
        return []
    return [hit for hit in hits if isinstance(hit, dict)]


def _add_rule_hit(
    row: dict[str, object],
    *,
    hit_object: str,
    rule_id: str,
    rule_desc: str,
    suggested_type: str = "",
    adopted: bool = True,
    warnings: list[str] | None = None,
) -> None:
    hits = row.setdefault(RULE_HITS_KEY, [])
    if not isinstance(hits, list):
        hits = []
        row[RULE_HITS_KEY] = hits
    hits.append({
        "hit_object": hit_object,
        "rule_id": rule_id,
        "rule_desc": rule_desc,
        "suggested_type": suggested_type,
        "adopted": "是" if adopted else "否",
        "warnings": list(warnings or []),
    })


def _attach_profile_rule_hits(
    rows: list[dict[str, object]],
    *,
    profile: CustomRulesProfile,
    generation: str = "",
) -> None:
    """给规则生成行补充生成期命中信息，避免审核副本只靠落表后反推。"""
    for row in rows:
        if generation:
            row["生成方式"] = generation
        row_generation = str(row.get("生成方式", "") or generation or "fallback")
        existing_rule_ids = {
            str(hit.get("rule_id", "") or "")
            for hit in _row_rule_hits(row)
        }
        if any(rule_id and not rule_id.endswith(".fallback_classification_basis") for rule_id in existing_rule_ids):
            continue
        name = str(row.get("新增/修改功能点", "") or "")
        fpa_type = str(row.get("类型", "") or "")
        reason = str(row.get("类型理由", "") or "")
        rule_id = "profile.fallback"
        if row_generation == "rules_fallback":
            rule_id = "coverage.rules_fallback"
        elif profile.name in {"unified_ui", "multi_uis"} and "界面开发" in name:
            rule_id = f"{profile.name}.ui_merge"
        elif profile.name == "strict_fpa" and fpa_type == "EIF":
            rule_id = "strict_fpa.external_data_group"
        elif profile.name == "strict_fpa" and fpa_type == "ILF":
            rule_id = "strict_fpa.internal_data_group"
        elif profile.name == "strict_fpa":
            rule_id = f"strict_fpa.transaction.{fpa_type.lower()}"
        elif "关键词" in reason or fpa_type:
            rule_id = f"{profile.name}.keyword.{fpa_type.lower()}"
        _add_rule_hit(
            row,
            hit_object=str(row.get("源功能过程", "") or name),
            rule_id=rule_id,
            rule_desc=reason or "按当前 profile 兜底规则生成。",
            suggested_type=fpa_type,
            adopted=True,
            warnings=_warning_items(row.get("后处理警告", "")),
        )


def _trace_rule_hits_for_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    trace_hits: list[dict[str, object]] = []
    for row in rows:
        for hit in _row_rule_hits(row):
            trace_hits.append({
                "fpa_seq": row.get("序号", ""),
                "name": row.get("新增/修改功能点", ""),
                "generation": row.get("生成方式", ""),
                "hit_object": hit.get("hit_object", ""),
                "rule_id": hit.get("rule_id", ""),
                "rule_desc": hit.get("rule_desc", ""),
                "suggested_type": hit.get("suggested_type", ""),
                "adopted": hit.get("adopted", ""),
                "warnings": hit.get("warnings", []),
            })
    return trace_hits


def _renumber_rows(rows: list[dict[str, object]], start_seq: int) -> None:
    for offset, row in enumerate(rows):
        row["序号"] = start_seq + offset


def _normalize_ai_fpa_name_prefix(name: str, group: dict[str, object]) -> str:
    """Force AI row names to use the module path from source data."""
    clean_name = str(name or "").strip()
    if not clean_name:
        return clean_name

    prefix = _group_tag(group).strip()
    if not prefix:
        return clean_name
    if clean_name == prefix or clean_name.startswith(f"{prefix}-"):
        return clean_name
    l3 = str(group.get("l3", "") or "").strip()
    if clean_name.startswith(prefix):
        overlap_suffix = clean_name[len(prefix):].strip()
        if overlap_suffix:
            if l3 and prefix.endswith(l3) and not overlap_suffix.startswith(l3):
                overlap_suffix = f"{l3}{overlap_suffix}"
            return f"{prefix}-{overlap_suffix}"

    client_type = str(group.get("client_type", "") or "").strip()
    l1 = str(group.get("l1", "") or "").strip()
    l2 = str(group.get("l2", "") or "").strip()
    path_parts = [part for part in [l1, l2, l3] if part]
    suffix = clean_name

    if path_parts:
        path_pattern = "-".join(re.escape(part) for part in path_parts)
        path_match = re.match(rf"^.*?{path_pattern}-(?P<suffix>.+)$", clean_name)
        if path_match:
            suffix = path_match.group("suffix").strip()
    if suffix == clean_name and client_type:
        client_prefixes = [client_type, f"【{client_type}】"]
        for client_prefix in client_prefixes:
            if suffix == client_prefix:
                suffix = ""
                break
            if suffix.startswith(f"{client_prefix}-"):
                suffix = suffix[len(client_prefix) + 1:].strip()
                break
    if l1 and suffix:
        path_candidate = suffix
        bracket_match = re.match(r"^【[^】]+】(?P<rest>.+)$", path_candidate)
        if bracket_match:
            path_candidate = bracket_match.group("rest").strip()
        if path_candidate.startswith(f"{l1}-"):
            parts = [part.strip() for part in path_candidate.split("-") if part.strip()]
            if len(parts) >= 4:
                suffix = "-".join(parts[3:]).strip()
            elif len(parts) == 3:
                suffix = parts[2].strip()

    return f"{prefix}-{suffix}" if suffix else prefix


def _normalize_ai_fpa_name_connectors(name: str) -> str:
    clean_name = str(name or "").strip()
    if not clean_name:
        return clean_name
    return re.sub(r"_(?=(?:界面开发|接口开发)$)", "-", clean_name)


def _normalize_ai_fpa_rows_for_l3(
    *,
    group: dict[str, object],
    meta: dict[str, str],
    ai_rows: list[Any],
    judgement_rules: list[str],
    start_seq: int = 1,
    profile: CustomRulesProfile = FPA_PROFILE,
    strategy: str = "",
) -> tuple[list[dict[str, object]], list[str]]:
    warnings: list[str] = []
    subsystem = meta.get("子系统（模块）", "")
    asset = meta.get("资产标识", "")
    module_status = _module_change_status(group.get("processes", []) if isinstance(group.get("processes"), list) else [])

    ui_rows = [
        r for r in ai_rows
        if isinstance(r, dict) and "界面开发" in str(r.get("name", ""))
    ]
    if (
        profile.name != "ui_api_mapping"
        and len(ui_rows) > 1
        and any(not str(r.get("split_reason", "")).strip() for r in ui_rows)
    ):
        msg = f"{_group_tag(group)} AI 输出多条界面开发行但缺少 split_reason，已合并为三级模块级界面行"
        logger.warning(msg)
        warnings.append(msg)
        ai_rows = [
            profile.fallback_rows_for_l3(
                group,
                meta,
                start_seq=1,
                judgement_rules=judgement_rules,
            )[0],
            *[
                r for r in ai_rows
                if not (isinstance(r, dict) and "界面开发" in str(r.get("name", "")))
            ],
        ]

    normalized: list[dict[str, object]] = []
    multi_ui_names: set[str] = set()
    seq = start_seq
    type_judgement = build_fpa_type_judgement(group)
    for raw in ai_rows:
        if not isinstance(raw, dict):
            warnings.append(f"{_group_tag(group)} AI rows 中存在非对象项，已跳过")
            continue
        name = str(raw.get("name", "") or raw.get("新增/修改功能点", "") or "").strip()
        if not name:
            warnings.append(f"{_group_tag(group)} AI 行缺少 name，已跳过")
            continue
        row_warnings: list[str] = []
        row_hits: list[dict[str, object]] = []
        explanation = str(raw.get("explanation", "") or raw.get("计算依据说明", "") or "").strip()
        connector_normalized_name = _normalize_ai_fpa_name_connectors(name)
        if connector_normalized_name != name:
            warning = f"{_group_tag(group)} AI 行名称连接符已规范化: {name} -> {connector_normalized_name}"
            warnings.append(warning)
            row_warnings.append(warning)
            row_hits.append({
                "hit_object": name,
                "rule_id": "postprocess.ai_name_connector",
                "rule_desc": "FPA 行名称中界面开发/接口开发后缀前应使用短横线连接。",
                "suggested_type": "",
                "adopted": "是",
                "warnings": [warning],
            })
            name = connector_normalized_name
        normalized_name = profile.normalize_output_name(name, explanation)
        if normalized_name != name:
            warning = f"{_group_tag(group)} AI 行名称不符合 {profile.name} 口径，已规范化: {name} -> {normalized_name}"
            warnings.append(warning)
            row_warnings.append(warning)
            row_hits.append({
                "hit_object": name,
                "rule_id": f"{profile.name}.normalize_output_name",
                "rule_desc": f"AI 输出名称需符合 {profile.name} 口径。",
                "suggested_type": "",
                "adopted": "是",
                "warnings": [warning],
            })
            name = normalized_name
        if not explanation:
            explanation = f"{name}，具体为以下：\n1、基于该三级模块功能过程完成对应业务能力。"
            warning = f"{_group_tag(group)} AI 行缺少 explanation，已使用兜底说明: {name}"
            warnings.append(warning)
            row_warnings.append(warning)
            row_hits.append({
                "hit_object": name,
                "rule_id": "postprocess.explanation_fallback",
                "rule_desc": "AI 未提供计算依据说明时，使用兜底说明。",
                "suggested_type": "",
                "adopted": "是",
                "warnings": [warning],
            })

        ai_type = str(raw.get("type", "") or raw.get("类型", "") or "").strip().upper()
        fallback_type, fallback_reason = profile.infer_type(name, explanation)
        type_reason = str(raw.get("type_reason", "") or "").strip()
        if ai_type not in VALID_FPA_TYPES:
            if ai_type:
                warning = f"{name} AI 返回非法 type={ai_type}，已使用 {fallback_type} 兜底"
                warnings.append(warning)
                row_warnings.append(warning)
            fpa_type = fallback_type
            type_reason = type_reason or fallback_reason
            row_hits.append({
                "hit_object": name,
                "rule_id": "postprocess.invalid_ai_type",
                "rule_desc": fallback_reason or "AI type 非法时使用 profile 类型规则兜底。",
                "suggested_type": fallback_type,
                "adopted": "是",
                "warnings": row_warnings[-1:] if ai_type else [],
            })
        elif (
            fallback_type != ai_type
            and profile.has_obvious_conflict(name, explanation, ai_type)
            and not _ai_row_matches_type_judgement(raw, ai_type, type_judgement)
        ):
            rule_basis = fallback_reason or "规则认为 AI type 存在业务冲突。"
            conflict_detail = f"规则建议 type={fallback_type}；规则依据：{rule_basis}"
            if strategy in {"ai_first", "ai_only"}:
                warning = (
                    f"{name} AI type={ai_type} 与规则存在冲突（{conflict_detail}），"
                    f"AI 优先策略下保留 AI type={ai_type}"
                )
                warnings.append(warning)
                row_warnings.append(warning)
                fpa_type = ai_type
                type_reason = type_reason or "AI 根据功能点名称和业务说明判定。"
                row_hits.append({
                    "hit_object": name,
                    "rule_id": "postprocess.ai_first_type_conflict",
                    "rule_desc": f"{conflict_detail}；AI 优先策略保留合法 AI type。",
                    "suggested_type": fallback_type,
                    "adopted": "否",
                    "warnings": [warning],
                })
            else:
                warning = (
                    f"{name} AI type={ai_type} 与关键词规则明显冲突（{conflict_detail}），"
                    f"已使用规则建议 type={fallback_type} 兜底"
                )
                warnings.append(warning)
                row_warnings.append(warning)
                fpa_type = fallback_type
                type_reason = fallback_reason
                row_hits.append({
                    "hit_object": name,
                    "rule_id": "postprocess.keyword_type_conflict",
                    "rule_desc": f"{conflict_detail}；采用规则建议类型。",
                    "suggested_type": fallback_type,
                    "adopted": "是",
                    "warnings": [warning],
                })
        else:
            fpa_type = ai_type
            type_reason = type_reason or "AI 根据功能点名称和业务说明判定。"
            row_hits.append({
                "hit_object": name,
                "rule_id": "postprocess.ai_type_validation",
                "rule_desc": type_reason,
                "suggested_type": fpa_type,
                "adopted": "是",
                "warnings": [],
            })

        basis = ""
        idx = raw.get("classification_basis_index")
        parsed_idx: int | None = None
        if idx is not None:
            try:
                parsed_idx = int(idx)
            except (TypeError, ValueError):
                warning = f"{name} classification_basis_index 非数字: {idx}"
                warnings.append(warning)
                row_warnings.append(warning)
                row_hits.append({
                    "hit_object": name,
                    "rule_id": "postprocess.classification_basis_index",
                    "rule_desc": "classification_basis_index 必须是模板判定原则中的数字序号。",
                    "suggested_type": fpa_type,
                    "adopted": "是",
                    "warnings": [warning],
                })
        if parsed_idx is not None:
            if judgement_rules and 1 <= parsed_idx <= len(judgement_rules):
                basis = judgement_rules[parsed_idx - 1]
            else:
                warning = f"{name} classification_basis_index 越界: {parsed_idx}"
                warnings.append(warning)
                row_warnings.append(warning)
                row_hits.append({
                    "hit_object": name,
                    "rule_id": "postprocess.classification_basis_index",
                    "rule_desc": "classification_basis_index 必须落在模板判定原则范围内。",
                    "suggested_type": fpa_type,
                    "adopted": "是",
                    "warnings": [warning],
                })
        if not basis and raw.get("classification_basis"):
            val = str(raw.get("classification_basis", "")).strip()
            for rule in judgement_rules:
                if val and (val in rule or rule in val):
                    basis = rule
                    break
            if not basis:
                warning = f"{name} classification_basis 未匹配模板规则，已留空"
                warnings.append(warning)
                row_warnings.append(warning)
                row_hits.append({
                    "hit_object": name,
                    "rule_id": "postprocess.classification_basis_match",
                    "rule_desc": "classification_basis 必须能匹配模板判定原则。",
                    "suggested_type": fpa_type,
                    "adopted": "是",
                    "warnings": [warning],
                })

        review_warning = ""
        if not _ai_row_matches_type_judgement(raw, fpa_type, type_judgement):
            review_warning = profile.ai_data_group_review_warning(name, explanation, fpa_type)
        if review_warning:
            warnings.append(review_warning)
            row_warnings.append(review_warning)
            row_hits.append({
                "hit_object": name,
                "rule_id": "postprocess.ai_data_group_review",
                "rule_desc": "AI 识别出数据功能，但当前 strict_fpa 规则无法确认数据组边界，需人工复核。",
                "suggested_type": fpa_type,
                "adopted": "是",
                "warnings": [review_warning],
            })

        output_name = _normalize_ai_fpa_name_prefix(name, group)
        prefixed_name = output_name
        prefix_changed = prefixed_name != name

        basis_type = _fpa_type_from_classification_basis(basis)
        if basis_type and basis_type != fpa_type:
            warning = f"{output_name} 类型={fpa_type} 与计算依据归类指向的类型={basis_type} 不一致: {basis}"
            warnings.append(warning)
            row_warnings.append(warning)
            row_hits.append({
                "hit_object": output_name,
                "rule_id": "postprocess.classification_basis_type_conflict",
                "rule_desc": "计算依据归类应与最终 FPA 类型一致。",
                "suggested_type": basis_type,
                "adopted": "否",
                "warnings": [warning],
            })

        if _explanation_needs_full_source_path(
            group=group,
            name=output_name,
            explanation=explanation,
        ):
            normalized_explanation = _normalize_explanation_source_path(
                explanation=explanation,
                source_name=output_name,
            )
            if normalized_explanation != explanation:
                detail = f"{output_name} 计算依据说明来源场景已按完整功能点路径规范化"
                row_hits.append({
                    "hit_object": output_name,
                    "rule_id": "postprocess.explanation_source_path",
                    "rule_desc": "计算依据说明的来源场景优先使用当前 FPA 行完整功能点路径。",
                    "suggested_type": fpa_type,
                    "adopted": "是",
                    "warnings": [detail],
                })
                explanation = normalized_explanation

        normalized_table_detail_explanation = _normalize_explanation_table_count_detail(explanation)
        if normalized_table_detail_explanation != explanation:
            row_hits.append({
                "hit_object": output_name,
                "rule_id": "postprocess.explanation_table_count_detail",
                "rule_desc": "正式计算依据说明保留 FPA 类型和计量口径，不展开数据库表个数细节。",
                "suggested_type": fpa_type,
                "adopted": "是",
                "warnings": [
                    f"{output_name} 计算依据说明中的数据库表个数细节已规范化"
                ],
            })
            explanation = normalized_table_detail_explanation

        explanation_warnings = _explanation_quality_warnings(
            group=group,
            name=output_name,
            fpa_type=fpa_type,
            explanation=explanation,
            profile_name=profile.name,
        )
        if explanation_warnings:
            warnings.extend(explanation_warnings)
            row_warnings.extend(explanation_warnings)
            row_hits.append({
                "hit_object": name,
                "rule_id": "postprocess.explanation_quality",
                "rule_desc": "计算依据说明应符合结构化证据说明规则，并将缺失信息放入 check/debug。",
                "suggested_type": fpa_type,
                "adopted": "是",
                "warnings": explanation_warnings,
            })

        id_to_name = _process_id_to_name_for_group(group)
        valid_source_ids: list[str] = []
        unknown_source_ids: list[str] = []
        raw_source_ids = raw.get("source_process_ids", [])
        if isinstance(raw_source_ids, list):
            for item in raw_source_ids:
                source_id = str(item).strip()
                if not source_id:
                    continue
                if source_id in id_to_name:
                    valid_source_ids.append(source_id)
                else:
                    unknown_source_ids.append(source_id)
        elif str(raw_source_ids or "").strip():
            unknown_source_ids.append(str(raw_source_ids).strip())

        source_processes = raw.get("source_processes", [])
        raw_source_names = (
            [str(x).strip() for x in source_processes if str(x).strip()]
            if isinstance(source_processes, list)
            else [x.strip() for x in str(source_processes or "").split("、") if x.strip()]
        )
        if unknown_source_ids:
            warning = f"{_group_tag(group)} AI 返回未知 source_process_ids: {'、'.join(unknown_source_ids)}，已忽略"
            warnings.append(warning)
            row_warnings.append(warning)
            row_hits.append({
                "hit_object": name,
                "rule_id": "postprocess.unknown_source_process_ids",
                "rule_desc": "source_process_ids 必须来自当前三级模块的功能过程候选列表。",
                "suggested_type": fpa_type,
                "adopted": "是",
                "warnings": [warning],
            })
        if not valid_source_ids and raw_source_names:
            warning = f"{_group_tag(group)} AI 未返回合法 source_process_ids，已降级使用 source_processes 名称匹配"
            warnings.append(warning)
            row_warnings.append(warning)
            row_hits.append({
                "hit_object": name,
                "rule_id": "postprocess.missing_source_process_ids",
                "rule_desc": "AI 应返回 source_process_ids；缺失时仅以 source_processes 名称作为兜底覆盖依据。",
                "suggested_type": fpa_type,
                "adopted": "是",
                "warnings": [warning],
            })
        source_names_from_ids = [id_to_name[source_id] for source_id in valid_source_ids]
        source_text = "、".join(source_names_from_ids or raw_source_names)

        if len(valid_source_ids) == 1:
            suffixed_name = _normalize_ai_name_process_suffix(
                output_name,
                id_to_name[valid_source_ids[0]],
                fpa_type,
            )
            if suffixed_name != output_name:
                detail = f"{_group_tag(group)} AI 行名称末尾已按 source_process_id 规范化: {output_name} -> {suffixed_name}"
                row_hits.append({
                    "hit_object": output_name,
                    "rule_id": "postprocess.ai_name_process_suffix",
                    "rule_desc": "AI 行保留完整功能点结构，但末尾功能过程名优先使用 source_process_id 对应的源功能过程名称。",
                    "suggested_type": "",
                    "adopted": "是",
                    "warnings": [detail],
                })
                output_name = suffixed_name
        if prefix_changed:
            warning = f"{_group_tag(group)} AI 行名称前缀已按源功能清单规范化: {name} -> {prefixed_name}"
            warnings.append(warning)
            row_warnings.append(warning)
            row_hits.append({
                "hit_object": name,
                "rule_id": "postprocess.ai_name_prefix",
                "rule_desc": "新增/修改功能点前缀必须来自源功能清单的客户端类型、一级模块、二级模块、三级模块。",
                "suggested_type": "",
                "adopted": "是",
                "warnings": [warning],
            })

        if profile.name == "multi_uis" and "界面开发" in output_name:
            split_reason = str(raw.get("split_reason", "") or "").strip()
            if split_reason:
                reason_warning = f"{output_name} multi_uis 拆分理由: {split_reason}"
                row_warnings.append(reason_warning)
                row_hits.append({
                    "hit_object": output_name,
                    "rule_id": "multi_uis.split_reason",
                    "rule_desc": "multi_uis 多界面开发行的拆分理由记录到 check/review 元数据。",
                    "suggested_type": fpa_type,
                    "adopted": "是",
                    "warnings": [reason_warning],
                })
            if output_name in multi_ui_names:
                duplicate_warning = f"{output_name} multi_uis 存在同名多界面开发行，已保留并提示人工审阅。"
                warnings.append(duplicate_warning)
                row_warnings.append(duplicate_warning)
                row_hits.append({
                    "hit_object": output_name,
                    "rule_id": "multi_uis.duplicate_ui_name",
                    "rule_desc": "multi_uis 同名多界面开发行不自动合并，保留冲突行并进入 check/review。",
                    "suggested_type": fpa_type,
                    "adopted": "是",
                    "warnings": [duplicate_warning],
                })
            multi_ui_names.add(output_name)

        row = {
            "序号": seq,
            "子系统(模块)": subsystem,
            "资产标识": asset,
            "新增/修改功能点": output_name,
            "类型": fpa_type,
            "计算依据归类": basis,
            "计算依据说明": explanation,
            "变更状态": str(raw.get("change_status", "") or module_status),
            "调整值": _adjust_value_for_type(fpa_type, profile.name),
            "要素数量": int(raw.get("element_count", 1) or 1),
            "生成方式": "ai",
            "类型理由": type_reason,
            "源功能过程": source_text,
            "source_process_ids": valid_source_ids,
            "后处理警告": "；".join(row_warnings),
            "复杂度": raw.get("complexity", ""),
            "DET": raw.get("det_count", ""),
            "RET": raw.get("ret_count", ""),
            "FTR": raw.get("ftr_count", ""),
            "复杂度说明": raw.get("complexity_reason", ""),
        }
        adjustment_audit = calculate_fpa_adjustment_for_row(row, profile.name)
        row["调整值"] = adjustment_audit["adjustment_value"]
        row["复杂度"] = adjustment_audit["complexity"]
        row["DET"] = adjustment_audit["det_count"]
        row["RET"] = adjustment_audit["ret_count"]
        row["FTR"] = adjustment_audit["ftr_count"]
        row["复杂度说明"] = adjustment_audit["complexity_reason"]
        row["调整值计算方式"] = adjustment_audit["method"]
        for hit in row_hits:
            _add_rule_hit(
                row,
                hit_object=str(hit.get("hit_object", "")),
                rule_id=str(hit.get("rule_id", "")),
                rule_desc=str(hit.get("rule_desc", "")),
                suggested_type=str(hit.get("suggested_type", "")),
                adopted=str(hit.get("adopted", "是")) == "是",
                warnings=hit.get("warnings", []) if isinstance(hit.get("warnings"), list) else [],
            )
        normalized.append(row)
        seq += 1
    return normalized, warnings


def _fill_fallback_classification_basis(
    rows: list[dict[str, object]],
    judgement_rules: list[str],
    profile: CustomRulesProfile,
) -> None:
    if not judgement_rules:
        return
    for row in rows:
        generation = str(row.get("生成方式", "") or "")
        if generation not in {"fallback", "rules_fallback"}:
            continue
        if str(row.get("计算依据归类", "") or "").strip():
            continue
        basis = _basis_for_fpa_type(str(row.get("类型", "") or ""), judgement_rules)
        if basis:
            row["计算依据归类"] = basis
            _add_rule_hit(
                row,
                hit_object=str(row.get("新增/修改功能点", "") or ""),
                rule_id=f"{profile.name}.fallback_classification_basis",
                rule_desc="规则兜底行按 FPA 类型匹配判定原则，补齐计算依据归类。",
                suggested_type=str(row.get("类型", "") or ""),
                adopted=True,
            )


def _fill_strict_fpa_fallback_classification_basis(
    rows: list[dict[str, object]],
    judgement_rules: list[str],
    profile: CustomRulesProfile,
) -> None:
    _fill_fallback_classification_basis(rows, judgement_rules, profile)


def _process_names_for_group(group: dict[str, object]) -> list[str]:
    processes = group.get("processes", [])
    if not isinstance(processes, list):
        return []
    return [
        str(p.get("name", "") or "").strip()
        for p in processes
        if isinstance(p, dict) and str(p.get("name", "") or "").strip()
    ]


def _process_id_to_name_for_group(group: dict[str, object]) -> dict[str, str]:
    processes = group.get("processes", [])
    if not isinstance(processes, list):
        return {}
    result: dict[str, str] = {}
    for process in processes:
        if not isinstance(process, dict):
            continue
        process_id = str(process.get("process_id", "") or "").strip()
        process_name = str(
            process.get("process_name", "") or process.get("name", "") or ""
        ).strip()
        if process_id and process_name:
            result[process_id] = process_name
    return result


def _process_name_to_id_for_group(group: dict[str, object]) -> dict[str, str]:
    return {name: process_id for process_id, name in _process_id_to_name_for_group(group).items()}


def _source_process_ids_from_row(row: dict[str, object]) -> set[str]:
    source_ids = row.get("source_process_ids", [])
    if isinstance(source_ids, list):
        return {str(item).strip() for item in source_ids if str(item).strip()}
    if isinstance(source_ids, str):
        return {item.strip() for item in re.split(r"[、,，;；\s]+", source_ids) if item.strip()}
    return set()


def _source_process_names_from_row(row: dict[str, object]) -> set[str]:
    values: set[str] = set()
    for key in ("source_processes", "源功能过程"):
        raw = row.get(key, [])
        if isinstance(raw, list):
            values.update(str(item).strip() for item in raw if str(item).strip())
        elif isinstance(raw, str):
            values.update(item.strip() for item in re.split(r"[、,，;；\s]+", raw) if item.strip())
    return values


def _ai_row_matches_type_judgement(
    row: dict[str, object],
    ai_type: str,
    type_judgement: dict[str, object],
) -> bool:
    """Return true when a high-confidence agent judgement supports the AI row type."""
    if not ai_type:
        return False
    row_ids = _source_process_ids_from_row(row)
    row_names = _source_process_names_from_row(row)
    row_name = str(row.get("name", "") or row.get("新增/修改功能点", "") or "")
    row_text = f"{row_name} {row.get('explanation', '')} {row.get('计算依据说明', '')}"
    judgements = type_judgement.get("judgements", []) if isinstance(type_judgement, dict) else []
    if not isinstance(judgements, list):
        return False
    for judgement in judgements:
        if not isinstance(judgement, dict):
            continue
        if str(judgement.get("confidence", "") or "") != "high":
            continue
        judgement_ids = {
            str(item).strip()
            for item in judgement.get("source_process_ids", [])
            if str(item).strip()
        } if isinstance(judgement.get("source_process_ids", []), list) else set()
        judgement_names = {
            str(item).strip()
            for item in judgement.get("source_process_names", [])
            if str(item).strip()
        } if isinstance(judgement.get("source_process_names", []), list) else set()
        target = str(judgement.get("target_data_group", "") or "").strip()
        candidate_name = str(judgement.get("candidate_name", "") or "").strip()
        source_matches = bool(
            (row_ids and row_ids & judgement_ids)
            or (row_names and row_names & judgement_names)
            or (target and target in row_text)
            or (candidate_name and candidate_name in row_text)
        )
        suggested_type = str(judgement.get("suggested_type", "") or "").upper()
        judgement_kind = str(judgement.get("judgement_kind", "") or "")
        if suggested_type == "NONE" and judgement_kind == "ordinary_external_service" and ai_type != "EIF":
            if source_matches:
                return True
            continue
        if suggested_type != ai_type:
            continue
        if source_matches:
            return True
    return False


def _source_process_set(row: dict[str, object]) -> set[str]:
    raw = str(row.get("源功能过程", "") or "")
    return {item.strip() for item in raw.split("、") if item.strip()}


def _append_row_warning(row: dict[str, object], warning: str) -> None:
    clean_warning = str(warning or "").strip()
    if not clean_warning:
        return
    existing = _warning_items(row.get("后处理警告", ""))
    if clean_warning in existing:
        return
    row["后处理警告"] = "；".join([*existing, clean_warning])


def _apply_fpa_validation_issues(
    rows: list[dict[str, object]],
    issues: list[FpaValidationIssue],
) -> list[str]:
    warnings: list[str] = []
    for issue in issues:
        if issue.message in warnings:
            continue
        warnings.append(issue.message)
        if issue.row_index is None or not (0 <= issue.row_index < len(rows)):
            continue
        row = rows[issue.row_index]
        _append_row_warning(row, issue.message)
        _add_rule_hit(
            row,
            hit_object=str(row.get("新增/修改功能点", "") or ""),
            rule_id=issue.code,
            rule_desc="gen-fpa 输出稳定性 validator 命中项目口径检查。",
            suggested_type=str(row.get("类型", "") or ""),
            adopted=True,
            warnings=[issue.message],
        )
    return warnings


def _validate_fpa_rows_for_l3(
    *,
    group: dict[str, object],
    rows: list[dict[str, object]],
) -> tuple[list[FpaValidationIssue], list[str]]:
    issues = validate_fpa_rows(group=group, rows=rows)
    warnings = _apply_fpa_validation_issues(rows, issues)
    return issues, warnings


def _build_quality_review_for_l3(
    *,
    group: dict[str, object],
    rows: list[dict[str, object]],
    confirmed_decisions: object | None = None,
) -> dict[str, object]:
    return build_fpa_quality_review(
        group=group,
        rows=rows,
        confirmed_decisions=confirmed_decisions,
    )


def _build_agent_review_for_l3(
    *,
    group: dict[str, object],
    rows: list[dict[str, object]] | None = None,
    confirmed_decisions: object | None = None,
    profile: CustomRulesProfile = FPA_PROFILE,
) -> dict[str, object]:
    return build_fpa_agent_review(
        group=group,
        rows=rows,
        confirmed_decisions=confirmed_decisions,
        profile_name=profile.name,
        profile_kind=profile.agent_review_profile_kind(),
    )


def _covered_process_ids_for_row(row: dict[str, object], group: dict[str, object]) -> set[str]:
    """Prefer internal process IDs, falling back to exact source-process names."""
    valid_ids = set(_process_id_to_name_for_group(group))
    row_ids = _source_process_ids_from_row(row) & valid_ids
    if row_ids:
        return row_ids
    name_to_id = _process_name_to_id_for_group(group)
    return {
        name_to_id[name]
        for name in _source_process_set(row)
        if name in name_to_id
    }


def _is_similar_process_name(left: str, right: str) -> bool:
    clean_left = re.sub(r"\s+", "", str(left or ""))
    clean_right = re.sub(r"\s+", "", str(right or ""))
    if not clean_left or not clean_right:
        return False
    if clean_left == clean_right:
        return True
    return abs(len(clean_left) - len(clean_right)) <= 4 and SequenceMatcher(None, clean_left, clean_right).ratio() >= 0.75


def _normalize_ai_name_process_suffix(name: str, source_name: str, fpa_type: str) -> str:
    if fpa_type in {"ILF", "EIF"}:
        return name
    clean_name = str(name or "").strip()
    clean_source = str(source_name or "").strip()
    if not clean_name or not clean_source:
        return clean_name
    parts = clean_name.rsplit("-", 1)
    suffix = parts[-1].strip()
    if not suffix or suffix == clean_source or not _is_similar_process_name(suffix, clean_source):
        return clean_name
    return f"{parts[0]}-{clean_source}" if len(parts) == 2 else clean_source


def _supplement_ai_rows_with_rules(
    *,
    group: dict[str, object],
    meta: dict[str, str],
    ai_rows: list[dict[str, object]],
    profile: CustomRulesProfile,
    strategy: str,
    rule_set_config: FpaRuleSetConfig | None = None,
    judgement_rules: list[str] | None = None,
) -> tuple[list[dict[str, object]], list[str]]:
    """AI 优先策略下，用规则补齐 AI 没覆盖的功能过程，不覆盖 AI 已判定的类型。"""
    if strategy != "ai_first" or not ai_rows:
        return ai_rows, []
    coverage_rules = rule_set_config.coverage_rules if isinstance(rule_set_config, FpaRuleSetConfig) else None
    require_process_coverage = True if coverage_rules is None or coverage_rules.require_process_coverage is None else coverage_rules.require_process_coverage
    require_data_function = True if coverage_rules is None or coverage_rules.require_data_function is None else coverage_rules.require_data_function
    require_profile_exact_rows = profile.agent_review_profile_kind() in {"unified_ui", "ui_api_mapping"}
    if not require_process_coverage and not require_data_function and not require_profile_exact_rows:
        return ai_rows, []

    rule_rows = profile.fallback_rows_for_l3(
        group,
        meta,
        start_seq=1,
        judgement_rules=judgement_rules,
    )
    id_to_name = _process_id_to_name_for_group(group)
    expected_process_ids = set(id_to_name)
    expected_process_names = set(_process_names_for_group(group))
    covered_process_ids: set[str] = set()
    covered_process_names: set[str] = set()
    for row in ai_rows:
        covered_process_ids.update(_covered_process_ids_for_row(row, group))
        covered_process_names.update(_source_process_set(row))

    if expected_process_ids:
        missing_process_ids = expected_process_ids - covered_process_ids
        missing_processes = {id_to_name[process_id] for process_id in missing_process_ids}
    else:
        missing_processes = expected_process_names - covered_process_names
    data_function_types = {
        str(row.get("类型", "") or "").strip()
        for row in ai_rows
        if str(row.get("类型", "") or "").strip() in {"ILF", "EIF"}
    }
    supplemental: list[dict[str, object]] = []
    data_function_supplements = 0
    missing_process_supplements = 0
    required_row_supplements = 0
    required_type_normalizations = 0
    existing_by_name = {
        str(row.get("新增/修改功能点", "") or ""): row
        for row in ai_rows
        if str(row.get("新增/修改功能点", "") or "")
    }

    for row in rule_rows:
        row_type = str(row.get("类型", ""))
        row_sources = _source_process_set(row)
        row_name = str(row.get("新增/修改功能点", "") or "")
        existing = existing_by_name.get(row_name)
        include_profile_required_row = False
        if require_profile_exact_rows and row_name:
            if existing is None:
                include_profile_required_row = True
            elif profile.agent_review_profile_kind() == "ui_api_mapping" and str(existing.get("类型", "") or "") != row_type:
                old_type = str(existing.get("类型", "") or "")
                existing["类型"] = row_type
                existing["类型理由"] = str(row.get("类型理由", "") or "") or str(existing.get("类型理由", "") or "")
                warning = f"{row_name} AI 默认映射行类型 {old_type or '空'} 已按 {profile.name} contract 修正为 {row_type}。"
                old_warning = str(existing.get("后处理警告", "") or "")
                existing["后处理警告"] = f"{old_warning}；{warning}" if old_warning else warning
                adjustment_audit = calculate_fpa_adjustment_for_row(existing, profile.name)
                existing["调整值"] = adjustment_audit["adjustment_value"]
                existing["复杂度"] = adjustment_audit["complexity"]
                existing["DET"] = adjustment_audit["det_count"]
                existing["RET"] = adjustment_audit["ret_count"]
                existing["FTR"] = adjustment_audit["ftr_count"]
                existing["复杂度说明"] = adjustment_audit["complexity_reason"]
                existing["调整值计算方式"] = adjustment_audit["method"]
                _add_rule_hit(
                    existing,
                    hit_object=row_name,
                    rule_id=f"{profile.name}.contract_required_type",
                    rule_desc=f"{profile.name} contract 要求该默认映射行类型为 {row_type}。",
                    suggested_type=row_type,
                    adopted=True,
                    warnings=[warning],
                )
                required_type_normalizations += 1
                if row_type in {"ILF", "EIF"}:
                    data_function_types.add(row_type)
        include_data_row = require_data_function and row_type in {"ILF", "EIF"} and row_type not in data_function_types
        include_missing_process = require_process_coverage and bool(row_sources & missing_processes)
        if not include_data_row and not include_missing_process and not include_profile_required_row:
            continue

        copied = dict(row)
        copied["生成方式"] = "rules_fallback"
        reason = str(copied.get("类型理由", "") or "")
        if include_missing_process:
            missing_process_supplements += 1
        if include_data_row:
            data_function_supplements += 1
        if include_profile_required_row:
            required_row_supplements += 1
        copied["类型理由"] = reason or (
            "AI 未包含数据功能行，按规则集补齐。"
            if include_data_row and not include_missing_process
            else f"AI 未输出 {profile.name} contract 强制行，按规则集补齐。"
            if include_profile_required_row and not include_missing_process
            else "AI 未覆盖该功能过程，按规则集补齐。"
        )
        warning = (
            "AI 结果未包含数据功能行，已按规则集补齐；未覆盖 AI 已判定类型。"
            if include_data_row and not include_missing_process
            else f"AI 结果未输出 {profile.name} contract 强制行，已按规则集补齐。"
            if include_profile_required_row and not include_missing_process
            else "AI 结果未覆盖该功能过程，已按规则集补齐；未覆盖 AI 已判定类型。"
        )
        old_warning = str(copied.get("后处理警告", "") or "")
        copied["后处理警告"] = f"{old_warning}；{warning}" if old_warning else warning
        _attach_profile_rule_hits([copied], profile=profile, generation="rules_fallback")
        supplemental.append(copied)

    if not supplemental and not required_type_normalizations:
        return ai_rows, []

    combined = [*ai_rows, *supplemental]
    for seq, row in enumerate(combined, 1):
        row["序号"] = seq
    warning_parts: list[str] = []
    if missing_process_supplements:
        warning_parts.append(f"AI 结果未覆盖 {len(missing_processes)} 个功能过程")
    if data_function_supplements:
        warning_parts.append("AI 结果未包含数据功能行")
    if required_row_supplements:
        warning_parts.append(f"AI 结果未输出 {required_row_supplements} 条 {profile.name} contract 强制行")
    if required_type_normalizations:
        warning_parts.append(f"AI 结果有 {required_type_normalizations} 条 {profile.name} contract 默认行类型已修正")
    reason = "，".join(warning_parts) or "AI 结果需要规则补齐"
    action_parts: list[str] = []
    if supplemental:
        action_parts.append(f"已追加 {len(supplemental)} 条 rules_fallback 行")
    if required_type_normalizations:
        action_parts.append(f"已修正 {required_type_normalizations} 条 ai 行类型")
    warnings = [f"{_group_tag(group)} {reason}，{'；'.join(action_parts)}"]
    return combined, warnings


def _rules_first_ai_reasons(group: dict[str, object], rule_rows: list[dict[str, object]]) -> list[str]:
    """Return reasons why rules_first should ask AI to re-plan this module."""
    reasons: list[str] = []
    if not rule_rows:
        reasons.append("规则未生成 FPA 行")
        return reasons

    for index, row in enumerate(rule_rows, 1):
        name = str(row.get("新增/修改功能点", "") or "").strip()
        fpa_type = str(row.get("类型", "") or "").strip()
        if not name:
            reasons.append(f"规则行 {index} 新增/修改功能点为空")
        if fpa_type not in VALID_FPA_TYPES:
            reasons.append(f"规则行 {index} 类型无效: {fpa_type or '空'}")

    expected = set(_process_names_for_group(group))
    covered: set[str] = set()
    for row in rule_rows:
        covered.update(_source_process_set(row))
    missing = sorted(expected - covered)
    if missing:
        reasons.append(f"规则结果未覆盖功能过程: {'、'.join(missing)}")
    return reasons


def _build_fpa_audit_report(
    *,
    group: dict[str, object],
    module: dict[str, object],
    fpa_rows: list[dict[str, object]],
    warnings: list[str],
    profile: CustomRulesProfile,
    profile_version: str,
    strategy: str,
    rule_set: str,
    raw_source: str = "",
    raw_rows: list[object] | None = None,
    raw_warnings: list[str] | None = None,
    rule_hits: list[dict[str, object]] | None = None,
) -> FpaAuditReport:
    process_names = _process_names_for_group(group)
    id_to_name = _process_id_to_name_for_group(group)
    expected_ids = set(id_to_name)
    covered_ids: set[str] = set()
    covered_names: set[str] = set()
    generation_counts: dict[str, int] = {}
    for row in fpa_rows:
        generation = str(row.get("生成方式", "") or "unknown")
        generation_counts[generation] = generation_counts.get(generation, 0) + 1
        covered_ids.update(_covered_process_ids_for_row(row, group))
        covered_names.update(_source_process_set(row))

    if expected_ids:
        covered_processes = [name for process_id, name in id_to_name.items() if process_id in covered_ids]
        missing_processes = [name for process_id, name in id_to_name.items() if process_id not in covered_ids]
    else:
        covered_processes = [name for name in process_names if name in covered_names]
        missing_processes = [name for name in process_names if name not in covered_names]
    audit_warnings = list(warnings)
    if process_names and missing_processes:
        audit_warnings.append(f"仍有 {len(missing_processes)} 个功能过程未被 FPA 行覆盖")

    return FpaAuditReport(
        profile=profile.name,
        profile_version=profile_version,
        strategy=strategy,
        rule_set=rule_set,
        module=module,
        process_total=len(process_names),
        covered_processes=covered_processes,
        missing_processes=missing_processes,
        generation_counts=generation_counts,
        warnings=audit_warnings,
        raw_source=raw_source,
        raw_rows=list(raw_rows or []),
        raw_warnings=list(raw_warnings or []),
        rule_hits=list(rule_hits or []),
        agent_review=_build_agent_review_for_l3(group=group, rows=fpa_rows, profile=profile),
    )


def _module_payload_for_audit(index: int, group: dict[str, object]) -> dict[str, object]:
    processes = group.get("processes", [])
    process_count = len(processes) if isinstance(processes, list) else 0
    return {
        "index": index,
        "client_type": group.get("client_type", ""),
        "l1": group.get("l1", ""),
        "l2": group.get("l2", ""),
        "l3": group.get("l3", ""),
        "process_count": process_count,
    }


def _audit_rule_hits_for_rows(
    rows: list[dict[str, object]],
    traced_hits_by_seq: dict[str, list[dict[str, object]]],
    group: dict[str, object],
) -> list[dict[str, object]]:
    result: list[dict[str, object]] = []
    for row in rows:
        row_seq = str(row.get("序号", "") or "")
        traced_hits = traced_hits_by_seq.get(row_seq, [])
        if not traced_hits:
            rule_id, rule_desc = _rule_hit_for_audit(row)
            traced_hits = [{
                "hit_object": row.get("源功能过程", "") or group.get("l3", ""),
                "rule_id": rule_id,
                "rule_desc": rule_desc,
                "suggested_type": row.get("类型", ""),
                "adopted": "是",
                "warnings": _warning_items(row.get("后处理警告", "")),
            }]
        for hit in traced_hits:
            result.append({
                "fpa_seq": row.get("序号", ""),
                "function_point": row.get("新增/修改功能点", ""),
                "generation": row.get("生成方式", ""),
                "hit_object": hit.get("hit_object", "") or row.get("源功能过程", "") or group.get("l3", ""),
                "rule_id": hit.get("rule_id", ""),
                "rule_desc": hit.get("rule_desc", ""),
                "suggested_type": hit.get("suggested_type", "") or row.get("类型", ""),
                "adopted": hit.get("adopted", "") or "是",
                "warnings": hit.get("warnings", []) if isinstance(hit.get("warnings", []), list) else [],
            })
    return result


def _build_fpa_audit_reports_for_groups(
    *,
    groups: list[dict[str, object]],
    rows_by_module: dict[int, list[dict[str, object]]],
    warnings_by_module: dict[int, list[str]],
    profile: CustomRulesProfile,
    profile_version: str,
    strategy: str,
    rule_set: str,
    raw_audit_by_module: dict[int, dict[str, object]] | None = None,
    rule_hits_by_seq: dict[str, list[dict[str, object]]] | None = None,
) -> list[FpaAuditReport]:
    reports: list[FpaAuditReport] = []
    raw_audit_by_module = raw_audit_by_module or {}
    rule_hits_by_seq = rule_hits_by_seq or {}
    for idx, group in enumerate(groups, 1):
        rows = rows_by_module.get(idx, [])
        raw_audit = raw_audit_by_module.get(idx, {})
        raw_warnings = raw_audit.get("warnings", [])
        reports.append(_build_fpa_audit_report(
            group=group,
            module=_module_payload_for_audit(idx, group),
            fpa_rows=rows,
            warnings=warnings_by_module.get(idx, []),
            profile=profile,
            profile_version=profile_version,
            strategy=strategy,
            rule_set=rule_set,
            raw_source=str(raw_audit.get("source", "") or ""),
            raw_rows=raw_audit.get("raw_rows", []) if isinstance(raw_audit.get("raw_rows", []), list) else [],
            raw_warnings=raw_warnings if isinstance(raw_warnings, list) else _warning_items(raw_warnings),
            rule_hits=_audit_rule_hits_for_rows(rows, rule_hits_by_seq, group),
        ))
    return reports


def _rule_set_config_warnings(rule_set_config: object | None) -> list[str]:
    if isinstance(rule_set_config, FpaRuleSetConfig):
        return list(rule_set_config.config_warnings)
    return []


def _with_config_warnings(warnings: list[str], rule_set_config: object | None) -> list[str]:
    merged = list(_rule_set_config_warnings(rule_set_config))
    merged.extend(warnings)
    return merged


def _is_config_warning(warning: object) -> bool:
    return str(warning or "").strip().startswith(CONFIG_WARNING_PREFIX)


class FpaAiDebugError(ValueError):
    """FPA 预览 AI 调试信息异常。"""

    def __init__(self, message: str, debug: dict[str, object]):
        super().__init__(message)
        self.debug = debug


FPA_JSON_ONLY_RESPONSE_CONSTRAINT = """\
输出约束：
1. 不要输出 reasoning、分析过程、Markdown 或 JSON 外文本。
2. 只输出 JSON 对象。
3. 所有判断理由必须写入 rows[].type_reason、rows[].explanation、rows[].split_reason、rows[].complexity_reason。
4. 如需额外调试摘要，可输出 debug_summary，最多 5 条，每条不超过 40 字；没有必要时省略。
5. rows 必须完整，不得为了输出 debug_summary 牺牲 rows。"""


def _append_fpa_json_only_response_constraint(system_prompt: str) -> str:
    prompt = str(system_prompt or "").rstrip()
    if "不要输出 reasoning" in prompt and "debug_summary" in prompt:
        return prompt
    return f"{prompt}\n\n{FPA_JSON_ONLY_RESPONSE_CONSTRAINT}".strip()


def _build_fpa_ai_prompt_context(
    group: dict[str, object],
    judgement_rules: list[str],
    domain_context: dict[str, object] | None,
    profile: CustomRulesProfile = FPA_PROFILE,
) -> FpaPromptContext:
    from ai_gen_reimbursement_docs.config_utils import (
        load_fpa_core_rules_config,
        load_fpa_system_prompt_config,
        load_fpa_user_prompt_config,
    )

    core_rules_config = load_fpa_core_rules_config(profile.name)
    system_config = load_fpa_system_prompt_config(profile.name)
    user_config = load_fpa_user_prompt_config(profile.name)
    prompt = profile.build_prompt(group, judgement_rules, domain_context)
    return FpaPromptContext(
        system_prompt=_append_fpa_json_only_response_constraint(system_config.text),
        user_prompt=prompt,
        core_rules=core_rules_config.text,
        core_rules_source=core_rules_config.source_label,
        system_prompt_source=system_config.source_label,
        user_prompt_source=user_config.source_label,
    )


def _ai_plan_fpa_rows_for_l3_debug(
    group: dict[str, object],
    judgement_rules: list[str],
    domain_context: dict[str, object] | None,
    api_key: str,
    model: str,
    base_url: str,
    profile: CustomRulesProfile = FPA_PROFILE,
    confirmed_decisions: object | None = None,
) -> tuple[list[dict[str, object]], dict[str, object]]:
    """调用 AI 规划一个三级模块的 FPA 行，并返回 Web 预览调试信息。"""
    prompt_context = _build_fpa_ai_prompt_context(
        group, judgement_rules, domain_context, profile
    )
    debug: dict[str, object] = {
        "ai_called": True,
        "model": model,
        "system_prompt": prompt_context.system_prompt,
        "system_prompt_source": prompt_context.system_prompt_source,
        "user_prompt": prompt_context.user_prompt,
        "user_prompt_source": prompt_context.user_prompt_source,
        "core_rules_source": prompt_context.core_rules_source,
        "ai_prompt": f"[system]\n{prompt_context.system_prompt}\n\n[user]\n{prompt_context.user_prompt}",
        "raw_response": "",
        "thinking": "",
        "parsed_rows": [],
    }
    user_prompt = prompt_context.user_prompt
    confirmed_feedback = confirmation_feedback(confirmed_decisions)
    if confirmed_feedback:
        user_prompt = f"{user_prompt}\n\n{confirmed_feedback}"
        debug["user_prompt"] = user_prompt
        debug["ai_prompt"] = f"[system]\n{prompt_context.system_prompt}\n\n[user]\n{user_prompt}"
    resp, thinking = _call_llm(
        user_prompt,
        prompt_context.system_prompt,
        api_key,
        model,
        base_url,
        tag=f"fpa_l3_{group.get('l3', '')}",
        return_thinking=True,
    )
    debug["raw_response"] = resp
    debug["thinking"] = thinking
    if not resp:
        raise FpaAiDebugError("LLM 返回空响应", debug)
    try:
        data = _extract_json_obj(resp)
    except Exception as exc:
        raise FpaAiDebugError(str(exc), debug) from exc
    rows = data.get("rows")
    if not isinstance(rows, list):
        raise FpaAiDebugError("AI 响应缺少 rows 列表", debug)
    debug["parsed_rows"] = rows
    return rows, debug


def _ai_plan_fpa_rows_for_l3(
    group: dict[str, object],
    judgement_rules: list[str],
    domain_context: dict[str, object] | None,
    api_key: str,
    model: str,
    base_url: str,
    profile: CustomRulesProfile = FPA_PROFILE,
    validation_retry_feedback: str = "",
    confirmed_decisions: object | None = None,
) -> list[dict[str, object]]:
    """调用 AI 规划一个三级模块的 FPA 行，返回原始 AI rows。"""
    prompt_context = _build_fpa_ai_prompt_context(
        group, judgement_rules, domain_context, profile
    )
    user_prompt = prompt_context.user_prompt
    confirmed_feedback = confirmation_feedback(confirmed_decisions)
    if confirmed_feedback:
        user_prompt = f"{user_prompt}\n\n{confirmed_feedback}"
    if validation_retry_feedback:
        user_prompt = f"{user_prompt}\n\n{validation_retry_feedback}"
    resp = _call_llm(
        user_prompt,
        prompt_context.system_prompt,
        api_key,
        model,
        base_url,
        tag=f"fpa_l3_{group.get('l3', '')}",
    )
    if not resp:
        raise ValueError("LLM 返回空响应")
    data = _extract_json_obj(resp)
    rows = data.get("rows")
    if not isinstance(rows, list):
        raise ValueError("AI 响应缺少 rows 列表")
    return rows


def _build_domain_context(meta: dict[str, str]) -> dict[str, object]:
    from ai_gen_reimbursement_docs.config_utils import load_optional_fpa_domain_context

    keys = [
        "子系统（模块）",
        "资产标识",
        "新增/修改功能点前缀生成规则",
        "功能用户-接收者判定",
    ]
    context = {k: meta.get(k, "") for k in keys if meta.get(k)}
    configured_context = dict(load_optional_fpa_domain_context())
    configured_context.pop("project_description", None)
    context.update(configured_context)
    project_description = _build_project_description_from_work_order(meta)
    if project_description:
        context["project_description"] = project_description
    return context


def _first_meta_value(meta: dict[str, str], keys: tuple[str, ...]) -> str:
    for key in keys:
        value = str(meta.get(key, "") or "").strip()
        if value:
            return value
    return ""


def _build_project_description_from_work_order(meta: dict[str, str]) -> str:
    title = _first_meta_value(meta, ("工单标题", "1、工单需求-元数据录入.工单标题"))
    content = _first_meta_value(meta, ("工单内容", "1、工单需求-元数据录入.工单内容"))
    parts: list[str] = []
    if title:
        parts.append(f"工单标题：{title}")
    if content:
        parts.append(f"工单内容：{content}")
    description = "\n".join(parts).strip()
    if len(description) <= FPA_PROJECT_DESCRIPTION_MAX_CHARS:
        return description
    truncated = description[:FPA_PROJECT_DESCRIPTION_MAX_CHARS].rstrip()
    return f"{truncated}\n（工单内容已截断，完整内容以功能清单录入模板为准。）"


def _fpa_ai_cache_key(
    group: dict[str, object],
    judgement_rules: list[str],
    domain_context: dict[str, object],
    model: str,
    profile: CustomRulesProfile = FPA_PROFILE,
    strategy: str = "",
    rule_set: str = "",
    rule_set_config: object | None = None,
    core_rules: str = "",
    system_prompt: str = "",
    user_prompt: str = "",
) -> str:
    serializable_rule_set_config = (
        rule_set_config.raw if isinstance(rule_set_config, FpaRuleSetConfig) else rule_set_config or {}
    )
    payload = {
        "profile": profile.name,
        "profile_version": profile.version,
        "strategy": strategy,
        "rule_set": rule_set,
        "rule_set_config": serializable_rule_set_config,
        "core_rules": core_rules,
        "system_prompt": system_prompt,
        "user_prompt": user_prompt,
        "domain_context": domain_context,
        "group": group,
        "judgement_rules": judgement_rules,
        "model": model,
    }
    raw = _json.dumps(payload, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _load_fpa_ai_cache(cache_path: str) -> dict[str, object]:
    if not cache_path or not os.path.exists(cache_path):
        return {"version": 1, "entries": {}}
    try:
        with open(cache_path, encoding="utf-8") as f:
            data = _json.load(f)
        if not isinstance(data, dict):
            raise ValueError("cache root must be object")
        entries = data.get("entries")
        if not isinstance(entries, dict):
            data["entries"] = {}
        data["version"] = 1
        return data
    except Exception as exc:
        logger.warning("FPA AI 缓存读取失败，将忽略缓存: %s", exc)
        return {"version": 1, "entries": {}}


def _save_fpa_ai_cache(cache_path: str, cache: dict[str, object]) -> None:
    if not cache_path:
        return
    try:
        os.makedirs(os.path.dirname(cache_path) or ".", exist_ok=True)
        with open(cache_path, "w", encoding="utf-8") as f:
            _json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception as exc:
        logger.warning("FPA AI 缓存写入失败: %s", exc)


def _save_fpa_audit_trace(audit_trace_path: str, trace: dict[str, object]) -> None:
    if not audit_trace_path:
        return
    try:
        os.makedirs(os.path.dirname(audit_trace_path) or ".", exist_ok=True)
        with open(audit_trace_path, "w", encoding="utf-8") as f:
            _json.dump(trace, f, ensure_ascii=False, indent=2)
    except Exception as exc:
        logger.warning("FPA audit trace 写入失败: %s", exc)


def _load_fpa_audit_trace(audit_trace_path: str) -> dict[str, object]:
    if not audit_trace_path or not os.path.exists(audit_trace_path):
        return {}
    try:
        with open(audit_trace_path, encoding="utf-8") as f:
            data = _json.load(f)
        return data if isinstance(data, dict) else {}
    except Exception as exc:
        logger.warning("FPA audit trace 读取失败: %s", exc)
        return {}


def _as_float(value: object) -> float:
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except (ValueError, TypeError):
        return 0.0


def calculate_fpa_row_workload(row: dict[str, object]) -> float:
    """按代码化业务口径计算单行 FPA 工作量。"""
    return _as_float(row.get("调整值")) * _as_float(row.get("要素数量"))


def calculate_fpa_total(rows: list[dict[str, object]]) -> float:
    """按代码化业务口径计算 FPA 工作量总和。"""
    return sum(calculate_fpa_row_workload(row) for row in rows)


def _enrich_fpa_rows_with_adjustment(rows: list[dict[str, object]], profile_name: str = "") -> None:
    """Apply configured adjustment calculation and audit fields to rows in place."""
    for row in rows:
        adjustment_audit = calculate_fpa_adjustment_for_row(row, profile_name)
        row["调整值"] = adjustment_audit["adjustment_value"]
        row["复杂度"] = adjustment_audit["complexity"]
        row["DET"] = adjustment_audit["det_count"]
        row["RET"] = adjustment_audit["ret_count"]
        row["FTR"] = adjustment_audit["ftr_count"]
        row["复杂度说明"] = adjustment_audit["complexity_reason"]
        row["调整值计算方式"] = adjustment_audit["method"]


def calculate_fpa_excel_formula_projection(xlsx_path: str) -> float:
    """读取 Excel 公式输入列，确定性投影 L 列工作量总和。"""
    from ai_gen_reimbursement_docs.config_utils import _get_system_config_value

    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    try:
        sheet_name = _get_system_config_value("fpa_sheet", "FPA功能点估算")
        ws = wb[sheet_name]
        last_data_row = ws.max_row
        expected_total_formula = f"=SUM(L3:L{last_data_row})"
        if ws.cell(1, FPA_COL_FORMULA_WORKLOAD).value != expected_total_formula:
            raise ValueError(f"FPA Excel 汇总公式无效，应为 {expected_total_formula}")

        total = 0.0
        for row_num in range(3, last_data_row + 1):
            expected_row_formula = f"=J{row_num}*K{row_num}"
            if ws.cell(row_num, FPA_COL_FORMULA_WORKLOAD).value != expected_row_formula:
                raise ValueError(f"FPA Excel 第 {row_num} 行工作量公式无效，应为 {expected_row_formula}")
            total += _as_float(ws.cell(row_num, FPA_COL_ADJUST).value) * _as_float(
                ws.cell(row_num, FPA_COL_ELEMENTS).value
            )
        return total
    finally:
        wb.close()


def _recalculate_with_excel_com(xlsx_path: str) -> tuple[bool, str]:
    """Use local Excel COM to recalculate and save formula caches."""
    if os.name != "nt":
        return False, "Excel COM 仅支持 Windows"
    try:
        import win32com.client  # type: ignore
    except Exception as exc:
        return False, f"Excel COM 不可用: {exc}"

    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        workbook = excel.Workbooks.Open(os.path.abspath(xlsx_path))
        excel.CalculateFullRebuild()
        workbook.Save()
        return True, "Excel COM"
    except Exception as exc:
        return False, f"Excel COM 复算失败: {exc}"
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=True)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass


def _recalculate_with_libreoffice(xlsx_path: str) -> tuple[bool, str]:
    """Use LibreOffice/soffice headless conversion to recalculate formula caches."""
    executable = shutil.which("soffice") or shutil.which("libreoffice")
    if not executable:
        return False, "未找到 LibreOffice/soffice"
    with tempfile.TemporaryDirectory(prefix="ard-fpa-recalc-") as tmp_dir:
        completed = subprocess.run(
            [
                executable,
                "--headless",
                "--convert-to",
                "xlsx",
                "--outdir",
                tmp_dir,
                os.path.abspath(xlsx_path),
            ],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=120,
            check=False,
        )
        if completed.returncode != 0:
            detail = (completed.stderr or completed.stdout or "").strip()
            return False, f"LibreOffice 复算失败: {detail or completed.returncode}"
        recalculated = os.path.join(tmp_dir, os.path.basename(xlsx_path))
        if not os.path.exists(recalculated):
            return False, "LibreOffice 未生成复算后的 xlsx"
        shutil.copy2(recalculated, xlsx_path)
    return True, "LibreOffice"


def _read_fpa_excel_cached_total(xlsx_path: str) -> float | None:
    from ai_gen_reimbursement_docs.config_utils import _get_system_config_value

    sheet_name = _get_system_config_value('fpa_sheet', 'FPA功能点估算')
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    try:
        ws = wb[sheet_name]
        value = ws.cell(1, FPA_COL_FORMULA_WORKLOAD).value
    finally:
        wb.close()
    try:
        return None if value in (None, "") else float(value)
    except (TypeError, ValueError):
        return None


def validate_fpa_excel_recalculation(
    xlsx_path: str,
    expected_total: float,
    tolerance: float = 0.01,
) -> list[str]:
    """Try to recalculate FPA Excel formulas and return non-blocking warnings."""
    attempts: list[str] = []
    for recalculator in (_recalculate_with_excel_com, _recalculate_with_libreoffice):
        ok, message = recalculator(xlsx_path)
        if not ok:
            attempts.append(message)
            continue

        cached_total = _read_fpa_excel_cached_total(xlsx_path)
        if cached_total is None:
            return [f"FPA Excel 公式复算已执行（{message}），但无法读取 L1 缓存总工作量"]
        if abs(cached_total - expected_total) > tolerance:
            return [
                "FPA Excel 公式复算结果与代码汇总不一致: "
                f"Excel={cached_total}, 代码汇总={expected_total}"
            ]
        return []

    return [f"未执行 FPA Excel 公式复算校验: {'；'.join(attempts)}"]


def write_fpa_summary_md(summary_md_path: str, total: float) -> None:
    os.makedirs(os.path.dirname(summary_md_path) or '.', exist_ok=True)
    with open(summary_md_path, 'w', encoding='utf-8') as f:
        f.write("# FPA 工作量\n\n")
        f.write(f"FPA工作量（人/天）: {total}\n")


def _wait_for_batch_fpa_confirmation(
    *,
    group: dict[str, object],
    group_index: int,
    group_total: int,
    confirmation_mode: str,
    questions: list[dict[str, object]],
) -> dict[str, object]:
    payload = {
        "confirmation_mode": confirmation_mode,
        "module": {
            "index": group_index,
            "total": group_total,
            "client_type": group.get("client_type", ""),
            "l1": group.get("l1", ""),
            "l2": group.get("l2", ""),
            "l3": group.get("l3", ""),
            "process_count": len(group.get("processes", []) or []),
        },
        "confirmation_questions": questions,
    }
    decisions = current_callbacks().wait_for_fpa_confirmation(payload)
    if not isinstance(decisions, dict):
        return {}
    return decisions


def _merge_confirmed_decisions(
    base: object | None,
    extra: object | None,
) -> dict[str, object]:
    merged: dict[str, object] = dict(normalize_confirmed_decisions(base or {}))
    merged.update(normalize_confirmed_decisions(extra or {}))
    return merged


def _plan_fpa_rows_with_ai(
    rows: list[dict[str, str]],
    meta: dict[str, str],
    judgement_rules: list[str],
    api_key: str,
    model: str,
    base_url: str,
    cache_path: str = "",
    profile: CustomRulesProfile = FPA_PROFILE,
    strategy: str = "",
    rule_set: str = "",
    rule_set_config: object | None = None,
    audit_trace_path: str = "",
    confirmed_decisions: object | None = None,
    fpa_confirmation_mode: str = "auto",
) -> list[dict[str, object]]:
    execution = resolve_fpa_execution_config(profile.name, strategy, rule_set)
    profile = execution.profile
    strategy = execution.strategy
    rule_set = execution.rule_set
    effective_rule_set_config = rule_set_config if isinstance(rule_set_config, FpaRuleSetConfig) else execution.rule_set_config
    rule_set_token = set_current_fpa_rule_set_config(effective_rule_set_config)
    try:
        return _plan_fpa_rows_with_execution(
            rows,
            meta,
            judgement_rules,
            api_key,
            model,
            base_url,
            cache_path=cache_path,
            profile=profile,
            strategy=strategy,
            rule_set=rule_set,
            rule_set_config=effective_rule_set_config,
            audit_trace_path=audit_trace_path,
            confirmed_decisions=confirmed_decisions,
            fpa_confirmation_mode=fpa_confirmation_mode,
        )
    finally:
        reset_current_fpa_rule_set_config(rule_set_token)


def _plan_fpa_rows_with_execution(
    rows: list[dict[str, str]],
    meta: dict[str, str],
    judgement_rules: list[str],
    api_key: str,
    model: str,
    base_url: str,
    cache_path: str = "",
    profile: CustomRulesProfile = FPA_PROFILE,
    strategy: str = "",
    rule_set: str = "",
    rule_set_config: object | None = None,
    audit_trace_path: str = "",
    confirmed_decisions: object | None = None,
    fpa_confirmation_mode: str = "auto",
) -> list[dict[str, object]]:
    groups = _group_rows_by_l3(rows)
    config_warnings = _rule_set_config_warnings(rule_set_config)
    confirmation_mode = normalize_confirmation_mode(fpa_confirmation_mode or "auto")
    confirmed_decisions = normalize_confirmed_decisions(confirmed_decisions or {})
    if strategy == "rules_only":
        logger.info("FPA strategy=%s，使用规则集 %s 生成 FPA", strategy, rule_set)
        all_rows: list[dict[str, object]] = []
        audit_modules: list[dict[str, object]] = []
        seq = 1
        for group in groups:
            group_rows = profile.fallback_rows_for_l3(
                group,
                meta,
                start_seq=seq,
                judgement_rules=judgement_rules,
            )
            _attach_profile_rule_hits(group_rows, profile=profile, generation="fallback")
            _fill_fallback_classification_basis(group_rows, judgement_rules, profile)
            quality_review = _build_quality_review_for_l3(
                group=group,
                rows=group_rows,
                confirmed_decisions=confirmed_decisions,
            )
            agent_review = _build_agent_review_for_l3(
                group=group,
                rows=group_rows,
                confirmed_decisions=confirmed_decisions,
                profile=profile,
            )
            all_rows.extend(group_rows)
            seq += len(group_rows)
            audit_modules.append({
                "module": _group_tag(group),
                "l3": group.get("l3", ""),
                "source": "rules",
                "raw_rows": [],
                "warnings": [*config_warnings, "仅规则策略未调用 AI"],
                "rule_hits": _trace_rule_hits_for_rows(group_rows),
                "agent_review": agent_review,
                "quality_review": quality_review,
            })
        audit_trace = {
            "version": 1,
            "profile": profile.name,
            "strategy": strategy,
            "rule_set": rule_set,
            "modules": audit_modules,
        }
        audit_trace["stability_report"] = build_fpa_stability_report(audit_trace)
        _save_fpa_audit_trace(audit_trace_path, audit_trace)
        return all_rows
    if strategy != "rules_first" and not api_key:
        raise ValueError(f"FPA strategy={strategy} 需要 API Key，当前未配置")

    from ai_gen_reimbursement_docs.config_utils import FpaPromptConfigError, load_flow_max_ai, load_gen_fpa_ai_limit

    max_ai = load_flow_max_ai("gen_fpa")
    module_limit = load_gen_fpa_ai_limit()
    all_rows: list[dict[str, object]] = []
    seq = 1
    attempted = success = parse_failed = empty_response = generated = fallback = skipped = warning_count = cache_hits = 0
    domain_context = _build_domain_context(meta)
    cache = _load_fpa_ai_cache(cache_path) if cache_path else {"version": 1, "entries": {}}
    cache_entries = cache.get("entries")
    if not isinstance(cache_entries, dict):
        cache_entries = {}
        cache["entries"] = cache_entries
    audit_modules: list[dict[str, object]] = []

    for idx, group in enumerate(groups, 1):
        rules_first_rows: list[dict[str, object]] = []
        rules_first_reasons: list[str] = []
        if strategy == "rules_first":
            rules_first_rows = profile.fallback_rows_for_l3(
                group,
                meta,
                start_seq=seq,
                judgement_rules=judgement_rules,
            )
            _attach_profile_rule_hits(rules_first_rows, profile=profile, generation="fallback")
            _fill_fallback_classification_basis(rules_first_rows, judgement_rules, profile)
            rules_first_reasons = _rules_first_ai_reasons(group, rules_first_rows)
            if not rules_first_reasons:
                quality_review = _build_quality_review_for_l3(
                    group=group,
                    rows=rules_first_rows,
                    confirmed_decisions=confirmed_decisions,
                )
                agent_review = _build_agent_review_for_l3(
                    group=group,
                    rows=rules_first_rows,
                    confirmed_decisions=confirmed_decisions,
                    profile=profile,
                )
                logger.info("  FPA rules_first 使用规则结果 [%d/%d] %s", idx, len(groups), _group_tag(group))
                all_rows.extend(rules_first_rows)
                seq += len(rules_first_rows)
                audit_modules.append({
                    "module": _group_tag(group),
                    "l3": group.get("l3", ""),
                    "source": "rules",
                    "raw_rows": [],
                    "warnings": [*config_warnings, "规则优先策略未调用 AI：规则结果覆盖完整且基础字段有效"],
                    "rule_hits": _trace_rule_hits_for_rows(rules_first_rows),
                    "agent_review": agent_review,
                    "quality_review": quality_review,
                })
                continue
            if not api_key:
                warning = "规则结果需要 AI 复核但未配置 API Key，已保留规则生成结果: " + "；".join(rules_first_reasons)
                quality_review = _build_quality_review_for_l3(
                    group=group,
                    rows=rules_first_rows,
                    confirmed_decisions=confirmed_decisions,
                )
                agent_review = _build_agent_review_for_l3(
                    group=group,
                    rows=rules_first_rows,
                    confirmed_decisions=confirmed_decisions,
                    profile=profile,
                )
                logger.warning("FPA rules_first 未调用 AI [%s]: %s", _group_tag(group), warning)
                all_rows.extend(rules_first_rows)
                seq += len(rules_first_rows)
                audit_modules.append({
                    "module": _group_tag(group),
                    "l3": group.get("l3", ""),
                    "source": "rules",
                    "raw_rows": [],
                    "warnings": [*config_warnings, warning],
                    "rule_hits": _trace_rule_hits_for_rows(rules_first_rows),
                    "agent_review": agent_review,
                    "quality_review": quality_review,
                })
                continue

        limited = (max_ai > 0 and idx > max_ai) or (module_limit > 0 and idx > module_limit)
        if limited:
            if strategy == "ai_only":
                raise ValueError(f"FPA strategy=ai_only 但三级模块 {_group_tag(group)} 被 AI 限制跳过")
            skipped += 1
            group_rows = rules_first_rows or profile.fallback_rows_for_l3(
                group,
                meta,
                start_seq=seq,
                judgement_rules=judgement_rules,
            )
            generation = "fallback" if strategy == "rules_first" else "rules_fallback"
            _attach_profile_rule_hits(group_rows, profile=profile, generation=generation)
            _fill_fallback_classification_basis(group_rows, judgement_rules, profile)
            fallback += len(group_rows)
            all_rows.extend(group_rows)
            seq += len(group_rows)
            warning = "模块超过 AI 调用限制，已使用规则生成"
            if rules_first_reasons:
                warning = f"规则结果需要 AI 复核但{warning}: {'；'.join(rules_first_reasons)}"
            quality_review = _build_quality_review_for_l3(
                group=group,
                rows=group_rows,
                confirmed_decisions=confirmed_decisions,
            )
            agent_review = _build_agent_review_for_l3(
                group=group,
                rows=group_rows,
                confirmed_decisions=confirmed_decisions,
                profile=profile,
            )
            audit_modules.append({
                "module": _group_tag(group),
                "l3": group.get("l3", ""),
                "source": "rules" if strategy == "rules_first" else "rules_fallback",
                "raw_rows": [],
                "warnings": [*config_warnings, warning],
                "rule_hits": _trace_rule_hits_for_rows(group_rows),
                "agent_review": agent_review,
                "quality_review": quality_review,
            })
            continue

        prompt_context = _build_fpa_ai_prompt_context(
            group, judgement_rules, domain_context, profile
        )
        cache_key = _fpa_ai_cache_key(
            group, judgement_rules, domain_context, model,
            profile=profile,
            strategy=strategy,
            rule_set=rule_set,
            rule_set_config=rule_set_config.raw if isinstance(rule_set_config, FpaRuleSetConfig) else rule_set_config or {},
            core_rules=prompt_context.core_rules,
            system_prompt=prompt_context.system_prompt,
            user_prompt=prompt_context.user_prompt,
        )
        cached = cache_entries.get(cache_key) if cache_path else None
        if isinstance(cached, dict) and isinstance(cached.get("rows"), list):
            try:
                raw_rows = cached["rows"]
                group_rows, warnings = _normalize_ai_fpa_rows_for_l3(
                    group=group,
                    meta=meta,
                    ai_rows=raw_rows,
                    judgement_rules=judgement_rules,
                    start_seq=seq,
                    profile=profile,
                    strategy=strategy,
                )
                if rules_first_reasons:
                    warnings.insert(0, "规则结果触发 AI 复核: " + "；".join(rules_first_reasons))
                group_rows, supplement_warnings = _supplement_ai_rows_with_rules(
                    group=group,
                    meta=meta,
                    ai_rows=group_rows,
                    profile=profile,
                    strategy=strategy,
                    rule_set_config=rule_set_config,
                    judgement_rules=judgement_rules,
                )
                _renumber_rows(group_rows, seq)
                warnings.extend(supplement_warnings)
                _fill_fallback_classification_basis(group_rows, judgement_rules, profile)
                validation_issues, validation_warnings = _validate_fpa_rows_for_l3(
                    group=group,
                    rows=group_rows,
                )
                warnings.extend(validation_warnings)
                if retryable_validation_issues(validation_issues):
                    logger.warning(
                        "FPA AI 缓存未通过稳定性校验 [%s]，将重新调用 AI: %s",
                        _group_tag(group),
                        "；".join(issue.message for issue in retryable_validation_issues(validation_issues)),
                    )
                    group_rows = []
                if group_rows:
                    quality_review = _build_quality_review_for_l3(
                        group=group,
                        rows=group_rows,
                        confirmed_decisions=confirmed_decisions,
                    )
                    agent_review = _build_agent_review_for_l3(
                        group=group,
                        rows=group_rows,
                        confirmed_decisions=confirmed_decisions,
                        profile=profile,
                    )
                    cache_hits += 1
                    success += 1
                    generated += len(group_rows)
                    warning_count += len(warnings)
                    logger.info("  FPA AI 缓存命中 [%d/%d] %s", idx, len(groups), _group_tag(group))
                    all_rows.extend(group_rows)
                    seq += len(group_rows)
                    audit_modules.append({
                        "module": _group_tag(group),
                        "l3": group.get("l3", ""),
                        "source": "ai_cache",
                        "raw_rows": raw_rows,
                        "warnings": _with_config_warnings(warnings, rule_set_config),
                        "rule_hits": _trace_rule_hits_for_rows(group_rows),
                        "agent_review": agent_review,
                        "quality_review": quality_review,
                    })
                    continue
            except Exception as exc:
                logger.warning("FPA AI 缓存内容无效 [%s]: %s，将重新调用 AI", _group_tag(group), exc)

        attempted += 1
        retry_trigger_source = ""
        try:
            logger.info("  FPA AI 规划三级模块 [%d/%d] %s", idx, len(groups), _group_tag(group))
            raw_rows = _ai_plan_fpa_rows_for_l3(
                group,
                judgement_rules,
                domain_context,
                api_key,
                model,
                base_url,
                profile=profile,
                confirmed_decisions=confirmed_decisions,
            )
            group_rows, warnings = _normalize_ai_fpa_rows_for_l3(
                group=group,
                meta=meta,
                ai_rows=raw_rows,
                judgement_rules=judgement_rules,
                start_seq=seq,
                profile=profile,
                strategy=strategy,
            )
            if rules_first_reasons:
                warnings.insert(0, "规则结果触发 AI 复核: " + "；".join(rules_first_reasons))
            group_rows, supplement_warnings = _supplement_ai_rows_with_rules(
                group=group,
                meta=meta,
                ai_rows=group_rows,
                profile=profile,
                strategy=strategy,
                rule_set_config=rule_set_config,
                judgement_rules=judgement_rules,
            )
            _renumber_rows(group_rows, seq)
            warnings.extend(supplement_warnings)
            _fill_fallback_classification_basis(group_rows, judgement_rules, profile)
            validation_issues, validation_warnings = _validate_fpa_rows_for_l3(
                group=group,
                rows=group_rows,
            )
            warnings.extend(validation_warnings)
            initial_quality_review = _build_quality_review_for_l3(
                group=group,
                rows=group_rows,
                confirmed_decisions=confirmed_decisions,
            )
            retry_feedback = validation_feedback(validation_issues)
            if retry_feedback:
                retry_trigger_source = "validator"
            else:
                retry_feedback = quality_feedback(initial_quality_review)
                if retry_feedback:
                    retry_trigger_source = "quality_review"
            if retry_feedback and strategy == "ai_first":
                logger.warning("FPA AI 输出触发稳定性重试 [%s]", _group_tag(group))
                retry_notice = f"{_group_tag(group)} AI 输出稳定性校验触发一次重试"
                retry_group_rows: list[dict[str, object]] = []
                retry_warnings: list[str] = []
                retry_issues: list[FpaValidationIssue] = []
                retry_quality_review: dict[str, object] = {}
                try:
                    retry_raw_rows = _ai_plan_fpa_rows_for_l3(
                        group,
                        judgement_rules,
                        domain_context,
                        api_key,
                        model,
                        base_url,
                        profile=profile,
                        validation_retry_feedback=retry_feedback,
                        confirmed_decisions=confirmed_decisions,
                    )
                    retry_group_rows, retry_warnings = _normalize_ai_fpa_rows_for_l3(
                        group=group,
                        meta=meta,
                        ai_rows=retry_raw_rows,
                        judgement_rules=judgement_rules,
                        start_seq=seq,
                        profile=profile,
                        strategy=strategy,
                    )
                    if rules_first_reasons:
                        retry_warnings.insert(0, "规则结果触发 AI 复核: " + "；".join(rules_first_reasons))
                    retry_group_rows, retry_supplement_warnings = _supplement_ai_rows_with_rules(
                        group=group,
                        meta=meta,
                        ai_rows=retry_group_rows,
                        profile=profile,
                        strategy=strategy,
                        rule_set_config=rule_set_config,
                        judgement_rules=judgement_rules,
                    )
                    _renumber_rows(retry_group_rows, seq)
                    retry_warnings.extend(retry_supplement_warnings)
                    _fill_fallback_classification_basis(retry_group_rows, judgement_rules, profile)
                    retry_issues, retry_validation_warnings = _validate_fpa_rows_for_l3(
                        group=group,
                        rows=retry_group_rows,
                    )
                    retry_warnings.extend(retry_validation_warnings)
                    retry_quality_review = _build_quality_review_for_l3(
                        group=group,
                        rows=retry_group_rows,
                        confirmed_decisions=confirmed_decisions,
                    )
                except Exception as retry_exc:
                    retry_warnings.append(f"{retry_notice}，但重试调用或解析失败，已保留首次可解析 AI 输出: {retry_exc}")
                if retry_group_rows:
                    raw_rows = retry_raw_rows
                    group_rows = retry_group_rows
                    validation_issues = retry_issues
                    warnings = [retry_notice, *retry_warnings]
                    if retryable_validation_issues(retry_issues):
                        warnings.append(
                            f"{_group_tag(group)} AI 重试后仍存在稳定性 warning: "
                            + "；".join(issue.message for issue in retryable_validation_issues(retry_issues))
                        )
                    retry_quality_issues = [
                        issue for issue in retryable_quality_issues(retry_quality_review)
                        if not str(issue.get("code", "") or "").startswith("validator.")
                    ]
                    if retry_quality_issues:
                        warnings.append(
                            f"{_group_tag(group)} AI 重试后仍存在质量审核 warning: "
                            + "；".join(str(issue.get("message", "") or "") for issue in retry_quality_issues)
                        )
                else:
                    warnings.extend(retry_warnings)
                    warnings.append(f"{retry_notice}，但重试未生成有效 FPA 行，保留首次结果")
            if not group_rows:
                raise ValueError("AI 规划未生成有效 FPA 行")
            quality_review = _build_quality_review_for_l3(
                group=group,
                rows=group_rows,
                confirmed_decisions=confirmed_decisions,
            )
            agent_review = _build_agent_review_for_l3(
                group=group,
                rows=group_rows,
                confirmed_decisions=confirmed_decisions,
                profile=profile,
            )
            confirmation_questions = build_fpa_confirmation_questions(
                group=group,
                issues=validation_issues,
                mode=confirmation_mode,
                confirmed_decisions=confirmed_decisions,
            )
            if confirmation_questions:
                logger.info(
                    "FPA 批量生成等待计量口径确认 [%d/%d] %s",
                    idx,
                    len(groups),
                    _group_tag(group),
                )
                user_decisions = _wait_for_batch_fpa_confirmation(
                    group=group,
                    group_index=idx,
                    group_total=len(groups),
                    confirmation_mode=confirmation_mode,
                    questions=confirmation_questions,
                )
                if user_decisions:
                    confirmed_decisions = _merge_confirmed_decisions(
                        confirmed_decisions,
                        user_decisions,
                    )
                    raw_rows = _ai_plan_fpa_rows_for_l3(
                        group,
                        judgement_rules,
                        domain_context,
                        api_key,
                        model,
                        base_url,
                        profile=profile,
                        confirmed_decisions=confirmed_decisions,
                    )
                    group_rows, confirmation_warnings = _normalize_ai_fpa_rows_for_l3(
                        group=group,
                        meta=meta,
                        ai_rows=raw_rows,
                        judgement_rules=judgement_rules,
                        start_seq=seq,
                        profile=profile,
                        strategy=strategy,
                    )
                    if rules_first_reasons:
                        confirmation_warnings.insert(0, "规则结果触发 AI 复核: " + "；".join(rules_first_reasons))
                    group_rows, confirmation_supplement_warnings = _supplement_ai_rows_with_rules(
                        group=group,
                        meta=meta,
                        ai_rows=group_rows,
                        profile=profile,
                        strategy=strategy,
                        rule_set_config=rule_set_config,
                        judgement_rules=judgement_rules,
                    )
                    _renumber_rows(group_rows, seq)
                    confirmation_warnings.extend(confirmation_supplement_warnings)
                    _fill_fallback_classification_basis(group_rows, judgement_rules, profile)
                    validation_issues, confirmation_validation_warnings = _validate_fpa_rows_for_l3(
                        group=group,
                        rows=group_rows,
                    )
                    confirmation_warnings.extend(confirmation_validation_warnings)
                    warnings = [
                        f"{_group_tag(group)} 已应用用户确认的 FPA 计量口径并重新生成",
                        *confirmation_warnings,
                    ]
                    if not group_rows:
                        raise ValueError("用户确认后 AI 规划未生成有效 FPA 行")
                    quality_review = _build_quality_review_for_l3(
                        group=group,
                        rows=group_rows,
                        confirmed_decisions=confirmed_decisions,
                    )
                    agent_review = _build_agent_review_for_l3(
                        group=group,
                        rows=group_rows,
                        confirmed_decisions=confirmed_decisions,
                        profile=profile,
                    )
                else:
                    warnings.append(f"{_group_tag(group)} 未收到 FPA 计量口径确认，保留当前生成结果")
            warning_count += len(warnings)
            success += 1
            generated += len(group_rows)
            if cache_path:
                cache_entries[cache_key] = {
                    "created_at": datetime.now().isoformat(timespec="seconds"),
                    "model": model,
                    "profile": profile.name,
                    "profile_version": profile.version,
                    "strategy": strategy,
                    "rule_set": rule_set,
                    "rows": raw_rows,
                }
            audit_modules.append({
                "module": _group_tag(group),
                "l3": group.get("l3", ""),
                "source": "ai",
                "raw_rows": raw_rows,
                "warnings": _with_config_warnings(warnings, rule_set_config),
                "retry_trigger_source": retry_trigger_source,
                "rule_hits": _trace_rule_hits_for_rows(group_rows),
                "agent_review": agent_review,
                "quality_review": quality_review,
            })
        except FpaPromptConfigError:
            raise
        except Exception as exc:
            if strategy == "ai_only":
                raise
            msg = str(exc)
            if "空响应" in msg:
                empty_response += 1
            else:
                parse_failed += 1
            logger.warning("FPA AI 响应解析失败 [%s]: %s", _group_tag(group), exc)
            group_rows = rules_first_rows or profile.fallback_rows_for_l3(
                group,
                meta,
                start_seq=seq,
                judgement_rules=judgement_rules,
            )
            _attach_profile_rule_hits(group_rows, profile=profile, generation="fallback")
            _fill_fallback_classification_basis(group_rows, judgement_rules, profile)
            fallback += len(group_rows)
            fallback_warning = f"AI 调用或解析失败: {exc}"
            if rules_first_reasons:
                fallback_warning = (
                    "规则结果触发 AI 复核，但 AI 调用或解析失败，已保留规则生成结果: "
                    + "；".join(rules_first_reasons)
                    + f"；AI错误: {exc}"
                )
            quality_review = _build_quality_review_for_l3(
                group=group,
                rows=group_rows,
                confirmed_decisions=confirmed_decisions,
            )
            agent_review = _build_agent_review_for_l3(
                group=group,
                rows=group_rows,
                confirmed_decisions=confirmed_decisions,
                profile=profile,
            )
            audit_modules.append({
                "module": _group_tag(group),
                "l3": group.get("l3", ""),
                "source": "rules_fallback",
                "raw_rows": locals().get("raw_rows", []),
                "warnings": [*config_warnings, fallback_warning],
                "rule_hits": _trace_rule_hits_for_rows(group_rows),
                "agent_review": agent_review,
                "quality_review": quality_review,
            })

        all_rows.extend(group_rows)
        seq += len(group_rows)

    msg = (
        "FPA AI 规划完成: 尝试 %d 个三级模块，成功 %d 个，空响应 %d 个，解析失败 %d 个，"
        "AI 生成 %d 行，兜底生成 %d 行，配置跳过 %d 个三级模块，缓存命中 %d 个，后处理 warning %d 个"
    ) % (attempted, success, empty_response, parse_failed, generated, fallback, skipped, cache_hits, warning_count)
    if attempted and success == 0:
        logger.warning(msg)
    else:
        logger.info(msg)
    if cache_path:
        _save_fpa_ai_cache(cache_path, cache)
    audit_trace = {
        "version": 1,
        "profile": profile.name,
        "strategy": strategy,
        "rule_set": rule_set,
        "modules": audit_modules,
    }
    audit_trace["stability_report"] = build_fpa_stability_report(audit_trace)
    _save_fpa_audit_trace(audit_trace_path, audit_trace)
    return all_rows


def _write_fpa_rows_md(
    fpa_rows: list[dict[str, object]],
    output_md_path: str,
    ai_filled: bool = False,
    execution_meta: dict[str, str] | None = None,
) -> float:
    _enrich_fpa_rows_with_adjustment(fpa_rows, (execution_meta or {}).get("profile", ""))
    with open(output_md_path, 'w', encoding='utf-8') as f:
        f.write("# FPA 工作量评估\n\n")
        f.write(f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        if ai_filled:
            f.write(f"**AI 规划**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        if execution_meta:
            for key in ("profile", "strategy", "rule_set"):
                val = execution_meta.get(key, "")
                if val:
                    f.write(f"**{key}**: {val}\n")
        f.write("\n")
        f.write("| 序号 | 子系统(模块) | 资产标识 | 新增/修改功能点 | 类型 | 计算依据归类 | 计算依据说明 | 变更状态 | 调整值 | 要素数量 | 生成方式 | 类型理由 | 源功能过程 | 后处理警告 | 复杂度 | DET | RET | FTR | 复杂度说明 | 调整值计算方式 |\n")
        f.write("|------|-------------|---------|----------------|------|-------------|-------------|---------|-------|---------|---------|---------|-----------|-----------|--------|-----|-----|-----|------------|----------------|\n")
        for row in fpa_rows:
            vals = [
                str(row.get("序号", "")),
                str(row.get("子系统(模块)", "")),
                str(row.get("资产标识", "")),
                str(row.get("新增/修改功能点", "")).replace('|', '\\|'),
                str(row.get("类型", "")),
                str(row.get("计算依据归类", "")),
                str(row.get("计算依据说明", "")).replace("|", chr(92) + "|").replace(chr(10), " "),
                str(row.get("变更状态", "")),
                str(row.get("调整值", "")),
                str(row.get("要素数量", "")),
                str(row.get("生成方式", "")),
                str(row.get("类型理由", "")).replace('|', '\\|'),
                str(row.get("源功能过程", "")).replace('|', '\\|'),
                str(row.get("后处理警告", "")).replace('|', '\\|'),
                str(row.get("复杂度", "")),
                str(row.get("DET", "")),
                str(row.get("RET", "")),
                str(row.get("FTR", "")),
                str(row.get("复杂度说明", "")).replace("|", chr(92) + "|").replace(chr(10), " "),
                str(row.get("调整值计算方式", "")),
            ]
            f.write("| " + " | ".join(vals) + " |\n")
    return calculate_fpa_total(fpa_rows)


def _read_fpa_rows_md_for_audit(fpa_md_path: str) -> tuple[dict[str, str], list[dict[str, object]]]:
    execution_meta: dict[str, str] = {}
    fpa_rows: list[dict[str, object]] = []
    with open(fpa_md_path, encoding="utf-8") as f:
        in_table = False
        for line in f:
            line = line.rstrip()
            meta_match = re.match(r"^\*\*(profile|strategy|rule_set)\*\*:\s*(.*)$", line)
            if meta_match:
                key, value = meta_match.group(1), meta_match.group(2)
                if key in {"profile", "strategy", "rule_set"}:
                    execution_meta[key] = value.strip()
            if "| 序号 | 子系统" in line:
                in_table = True
                continue
            if "|------|" in line and in_table:
                continue
            if in_table:
                cells = parse_md_table_row(line, min_cols=14)
                if cells is not None:
                    fpa_rows.append({
                        "序号": cells[0],
                        "子系统(模块)": cells[1],
                        "资产标识": cells[2],
                        "新增/修改功能点": cells[3],
                        "类型": cells[4],
                        "计算依据归类": cells[5],
                        "计算依据说明": cells[6],
                        "变更状态": cells[7],
                        "调整值": cells[8],
                        "要素数量": cells[9],
                        "生成方式": cells[10],
                        "类型理由": cells[11],
                        "源功能过程": cells[12],
                        "后处理警告": cells[13],
                        "复杂度": cells[14] if len(cells) > 14 else "",
                        "DET": cells[15] if len(cells) > 15 else "",
                        "RET": cells[16] if len(cells) > 16 else "",
                        "FTR": cells[17] if len(cells) > 17 else "",
                        "复杂度说明": cells[18] if len(cells) > 18 else "",
                        "调整值计算方式": cells[19] if len(cells) > 19 else "",
                    })
    return execution_meta, fpa_rows


def _warning_items(text: object) -> list[str]:
    raw = str(text or "").strip()
    if not raw:
        return []
    return [item.strip() for item in re.split(r"[；;]", raw) if item.strip()]


def _group_rows_for_audit(
    fpa_rows: list[dict[str, object]],
    groups: list[dict[str, object]],
) -> dict[int, list[dict[str, object]]]:
    result: dict[int, list[dict[str, object]]] = {idx: [] for idx in range(1, len(groups) + 1)}
    remaining: list[dict[str, object]] = []
    process_names_by_idx = {
        idx: set(_process_names_for_group(group))
        for idx, group in enumerate(groups, 1)
    }

    def module_markers(group: dict[str, object]) -> list[str]:
        client = str(group.get("client_type", "") or "").strip()
        l1 = str(group.get("l1", "") or "").strip()
        l2 = str(group.get("l2", "") or "").strip()
        l3 = str(group.get("l3", "") or "").strip()
        markers = [
            f"【{client}】{l1}-{l2}-{l3}" if client and l1 and l2 and l3 else "",
            f"{l1}-{l2}-{l3}" if l1 and l2 and l3 else "",
            f"{l2}-{l3}" if l2 and l3 else "",
        ]
        return [marker for marker in markers if marker]

    def choose_by_module_marker(row: dict[str, object], candidate_idxs: list[int]) -> int | None:
        point_name = str(row.get("新增/修改功能点", "") or "")
        explanation = str(row.get("计算依据说明", "") or "")
        haystack = f"{point_name}\n{explanation}"
        matches: list[tuple[int, int]] = []
        for idx in candidate_idxs:
            marker_len = max(
                (len(marker) for marker in module_markers(groups[idx - 1]) if marker in haystack),
                default=0,
            )
            if marker_len:
                matches.append((marker_len, idx))
        if not matches:
            return None
        matches.sort(reverse=True)
        return matches[0][1]

    for row in fpa_rows:
        sources = _source_process_set(row)
        source_matches = [
            idx
            for idx, process_names in process_names_by_idx.items()
            if process_names and sources & process_names
        ]
        if len(source_matches) == 1:
            result[source_matches[0]].append(row)
            continue
        if len(source_matches) > 1:
            idx = choose_by_module_marker(row, source_matches) or source_matches[0]
            result[idx].append(row)
            continue

        marker_matches = [
            idx
            for idx in range(1, len(groups) + 1)
            if choose_by_module_marker(row, [idx]) is not None
        ]
        if marker_matches:
            idx = choose_by_module_marker(row, marker_matches) or marker_matches[0]
            result[idx].append(row)
        else:
            remaining.append(row)

    if remaining and groups:
        result.setdefault(len(groups), []).extend(remaining)
    return result


def _rule_hit_for_audit(row: dict[str, object]) -> tuple[str, str]:
    """根据当前已落表的审计字段还原规则/校验命中说明。"""
    generation = str(row.get("生成方式", "") or "").strip()
    name = str(row.get("新增/修改功能点", "") or "")
    fpa_type = str(row.get("类型", "") or "")
    reason = str(row.get("类型理由", "") or "").strip()
    warning = str(row.get("后处理警告", "") or "").strip()
    text = f"{name} {reason} {warning}"

    if generation == "rules_fallback":
        return "coverage.rules_fallback", reason or "AI 未覆盖功能过程时，按规则集补齐。"
    if "界面开发" in name:
        return "unified_ui.ui_merge", reason or "同一三级模块界面能力默认合并为一条。"
    if "外部系统维护" in text or "外部应用维护" in text or "引用外部" in text:
        return "strict_fpa.external_data_group", reason or "命中外部维护数据组规则。"
    if "本系统维护" in text or "逻辑数据组" in text:
        return "strict_fpa.internal_data_group", reason or "命中内部逻辑数据组规则。"
    if "事务功能" in text:
        return f"strict_fpa.transaction.{fpa_type.lower()}", reason or "命中事务功能类型规则。"
    if "关键词命中" in text:
        return f"unified_ui.keyword.{fpa_type.lower()}", reason or "命中关键词类型兜底规则。"
    if generation in {"ai", "ai_cache"}:
        return "postprocess.ai_type_validation", reason or "AI 输出经类型合法性和业务冲突规则校验后采用。"
    return "profile.fallback", reason or "按当前 profile 兜底规则生成。"


def _rule_hits_from_audit_trace(audit_trace: dict[str, object]) -> dict[str, list[dict[str, object]]]:
    by_seq: dict[str, list[dict[str, object]]] = {}
    modules = audit_trace.get("modules", [])
    if not isinstance(modules, list):
        return by_seq
    for module in modules:
        if not isinstance(module, dict):
            continue
        hits = module.get("rule_hits", [])
        if not isinstance(hits, list):
            continue
        for hit in hits:
            if not isinstance(hit, dict):
                continue
            seq = str(hit.get("fpa_seq", "") or "").strip()
            if seq:
                by_seq.setdefault(seq, []).append(hit)
    return by_seq


def _warning_source_for_row(
    rule_hits_by_seq: dict[str, list[dict[str, object]]],
    row_seq: object,
    warning: str,
) -> tuple[str, str]:
    for hit in rule_hits_by_seq.get(str(row_seq or ""), []):
        hit_warnings = hit.get("warnings", [])
        if isinstance(hit_warnings, list) and any(str(item).strip() == warning for item in hit_warnings):
            return str(hit.get("rule_id", "") or ""), str(hit.get("rule_desc", "") or "")
    return "", ""


FPA_CHECK_DEFAULT_COLUMNS: dict[str, list[str]] = {
    "FPA结果": [
        "序号", "子系统(模块)", "资产标识", "新增/修改功能点", "类型", "计算依据归类",
        "计算依据说明", "变更状态", "调整值", "要素数量", "生成方式", "类型理由",
        "源功能过程", "后处理警告", "复杂度", "DET", "RET", "FTR", "复杂度说明",
        "调整值计算方式", "profile", "strategy", "rule_set",
    ],
    "覆盖审核": [
        "模块序号", "客户端类型", "一级模块", "二级模块", "三级模块",
        "功能过程总数", "已覆盖数", "未覆盖数", "已覆盖功能过程", "未覆盖功能过程",
        "生成方式统计", "Warnings",
    ],
    "Warnings": ["级别", "FPA行序号", "模块序号", "对象", "Warning", "来源规则ID", "来源说明"],
    "规则命中详情": [
        "模块序号", "客户端类型", "一级模块", "二级模块", "三级模块",
        "FPA行序号", "功能点名称", "生成方式", "rule_set",
        "命中对象", "规则ID", "规则说明", "建议类型", "是否采用", "Warnings",
    ],
    "AI原始返回": ["模块", "三级模块", "来源", "Warnings", "AI原始Rows JSON"],
    "稳定性报告": [
        "范围", "模块序号", "模块", "三级模块", "来源", "Warning数", "Quality Issue数",
        "可重试Quality Issue数", "确认数", "重试次数", "来源统计", "Issue Codes",
    ],
}


def _fpa_check_columns() -> dict[str, list[str]]:
    from ai_gen_reimbursement_docs.config_utils import load_fpa_check_columns

    configured = load_fpa_check_columns()
    columns: dict[str, list[str]] = {}
    for sheet_name, defaults in FPA_CHECK_DEFAULT_COLUMNS.items():
        selected = configured.get(sheet_name, [])
        if selected:
            filtered = [column for column in selected if column in defaults]
            columns[sheet_name] = filtered or list(defaults)
        else:
            columns[sheet_name] = list(defaults)
    return columns


def _append_audit_row(ws, columns: list[str], values: dict[str, object]) -> None:
    ws.append([values.get(column, "") for column in columns])


def _header_index(ws) -> dict[str, int]:
    return {
        str(cell.value): index
        for index, cell in enumerate(ws[1])
        if cell.value is not None
    }


def generate_fpa_check_xlsx_from_md(
    fpa_md_path: str,
    tree_md_path: str,
    output_path: str,
    audit_trace_path: str = "",
) -> str:
    """生成 FPA 审核副本，不影响正式交付 Excel。"""
    logger.info("第1.5步：生成 FPA 审核副本...")
    execution_meta, fpa_rows = _read_fpa_rows_md_for_audit(fpa_md_path)
    tree_rows = parse_module_tree_md(tree_md_path)
    groups = _group_rows_by_l3(tree_rows)
    rows_by_module = _group_rows_for_audit(fpa_rows, groups)
    audit_trace = _load_fpa_audit_trace(audit_trace_path)
    stability_report = audit_trace.get("stability_report", {})
    if not isinstance(stability_report, dict):
        stability_report = {}
    if not stability_report:
        stability_report = build_fpa_stability_report(audit_trace)
    rule_hits_by_seq = _rule_hits_from_audit_trace(audit_trace)
    audit_modules = audit_trace.get("modules", [])
    if not isinstance(audit_modules, list):
        audit_modules = []
    config_warnings_by_module_idx: dict[int, list[str]] = {}
    trace_warnings_by_module_idx: dict[int, list[str]] = {}
    for module_idx, item in enumerate(audit_modules, 1):
        if not isinstance(item, dict):
            continue
        item_warnings = item.get("warnings", [])
        if not isinstance(item_warnings, list):
            continue
        config_warnings = [str(w).strip() for w in item_warnings if _is_config_warning(w)]
        if config_warnings:
            config_warnings_by_module_idx[module_idx] = config_warnings
        trace_warnings = [str(w).strip() for w in item_warnings if str(w).strip() and not _is_config_warning(w)]
        if trace_warnings:
            trace_warnings_by_module_idx[module_idx] = trace_warnings
    raw_audit_by_module_idx: dict[int, dict[str, object]] = {
        module_idx: item
        for module_idx, item in enumerate(audit_modules, 1)
        if isinstance(item, dict)
    }
    warnings_by_module_idx: dict[int, list[str]] = {}
    for idx in range(1, len(groups) + 1):
        module_warnings: list[str] = []
        for row in rows_by_module.get(idx, []):
            module_warnings.extend(_warning_items(row.get("后处理警告", "")))
        module_warnings.extend(config_warnings_by_module_idx.get(idx, []))
        for warning in trace_warnings_by_module_idx.get(idx, []):
            if warning not in module_warnings:
                module_warnings.append(warning)
        warnings_by_module_idx[idx] = module_warnings
    profile = get_fpa_profile(execution_meta.get("profile", "") or "unified_ui")
    audit_reports = _build_fpa_audit_reports_for_groups(
        groups=groups,
        rows_by_module=rows_by_module,
        warnings_by_module=warnings_by_module_idx,
        profile=profile,
        profile_version=profile.version,
        strategy=execution_meta.get("strategy", ""),
        rule_set=execution_meta.get("rule_set", ""),
        raw_audit_by_module=raw_audit_by_module_idx,
        rule_hits_by_seq=rule_hits_by_seq,
    )
    sheet_columns = _fpa_check_columns()

    wb = openpyxl.Workbook()
    ws_result = wb.active
    ws_result.title = "FPA结果"
    ws_result.append(sheet_columns["FPA结果"])
    warning_rows: list[dict[str, object]] = []
    for row in fpa_rows:
        _append_audit_row(ws_result, sheet_columns["FPA结果"], {
            "序号": row.get("序号", ""),
            "子系统(模块)": row.get("子系统(模块)", ""),
            "资产标识": row.get("资产标识", ""),
            "新增/修改功能点": row.get("新增/修改功能点", ""),
            "类型": row.get("类型", ""),
            "计算依据归类": row.get("计算依据归类", ""),
            "计算依据说明": row.get("计算依据说明", ""),
            "变更状态": row.get("变更状态", ""),
            "调整值": row.get("调整值", ""),
            "要素数量": row.get("要素数量", ""),
            "生成方式": row.get("生成方式", ""),
            "类型理由": row.get("类型理由", ""),
            "源功能过程": row.get("源功能过程", ""),
            "后处理警告": row.get("后处理警告", ""),
            "复杂度": row.get("复杂度", ""),
            "DET": row.get("DET", ""),
            "RET": row.get("RET", ""),
            "FTR": row.get("FTR", ""),
            "复杂度说明": row.get("复杂度说明", ""),
            "调整值计算方式": row.get("调整值计算方式", ""),
            "profile": execution_meta.get("profile", ""),
            "strategy": execution_meta.get("strategy", ""),
            "rule_set": execution_meta.get("rule_set", ""),
        })
        for warning in _warning_items(row.get("后处理警告", "")):
            source_rule_id, source_rule_desc = _warning_source_for_row(
                rule_hits_by_seq,
                row.get("序号", ""),
                warning,
            )
            warning_rows.append({
                "级别": "row",
                "FPA行序号": row.get("序号", ""),
                "模块序号": "",
                "对象": row.get("新增/修改功能点", ""),
                "Warning": warning,
                "来源规则ID": source_rule_id,
                "来源说明": source_rule_desc,
            })

    ws_coverage = wb.create_sheet("覆盖审核")
    ws_coverage.append(sheet_columns["覆盖审核"])
    for idx, audit_report in enumerate(audit_reports, 1):
        group = groups[idx - 1]
        module_payload = audit_report.module
        module_warnings = list(warnings_by_module_idx.get(idx, []))
        coverage_warnings: list[str] = []
        if audit_report.missing_processes:
            coverage_warnings.append(f"未覆盖功能过程: {'、'.join(audit_report.missing_processes)}")
            module_warnings.extend(coverage_warnings)
        _append_audit_row(ws_coverage, sheet_columns["覆盖审核"], {
            "模块序号": idx,
            "客户端类型": module_payload.get("client_type", ""),
            "一级模块": module_payload.get("l1", ""),
            "二级模块": module_payload.get("l2", ""),
            "三级模块": module_payload.get("l3", ""),
            "功能过程总数": audit_report.process_total,
            "已覆盖数": len(audit_report.covered_processes),
            "未覆盖数": len(audit_report.missing_processes),
            "已覆盖功能过程": "、".join(audit_report.covered_processes),
            "未覆盖功能过程": "、".join(audit_report.missing_processes),
            "生成方式统计": _json.dumps(audit_report.generation_counts, ensure_ascii=False, sort_keys=True),
            "Warnings": "；".join(module_warnings),
        })
        for warning in coverage_warnings:
            warning_rows.append({
                "级别": "module",
                "FPA行序号": "",
                "模块序号": idx,
                "对象": group.get("l3", ""),
                "Warning": warning,
                "来源规则ID": "coverage.missing_process",
                "来源说明": "功能过程未被任何 FPA 行源功能过程覆盖。",
            })
        for warning in config_warnings_by_module_idx.get(idx, []):
            warning_rows.append({
                "级别": "module",
                "FPA行序号": "",
                "模块序号": idx,
                "对象": group.get("l3", ""),
                "Warning": warning,
                "来源规则ID": "config.external_data_rules.external_service",
                "来源说明": "rule_set.external_data_rules 将普通外部服务配置为外部数据组，配置加载不中断但需要人工复核。",
            })
        for warning in trace_warnings_by_module_idx.get(idx, []):
            if any(item.get("Warning") == warning for item in warning_rows):
                continue
            warning_rows.append({
                "级别": "module",
                "FPA行序号": "",
                "模块序号": idx,
                "对象": group.get("l3", ""),
                "Warning": warning,
                "来源规则ID": "audit.module_warning",
                "来源说明": "生成期模块级审核 warning。",
            })

    ws_warnings = wb.create_sheet("Warnings")
    ws_warnings.append(sheet_columns["Warnings"])
    for item in warning_rows:
        _append_audit_row(ws_warnings, sheet_columns["Warnings"], item)

    ws_rule_hits = wb.create_sheet("规则命中详情")
    ws_rule_hits.append(sheet_columns["规则命中详情"])
    for idx, audit_report in enumerate(audit_reports, 1):
        module_payload = audit_report.module
        for hit in audit_report.rule_hits:
            warnings = "；".join(
                str(x) for x in hit.get("warnings", [])
                if isinstance(hit.get("warnings", []), list) and str(x).strip()
            )
            _append_audit_row(ws_rule_hits, sheet_columns["规则命中详情"], {
                "模块序号": idx,
                "客户端类型": module_payload.get("client_type", ""),
                "一级模块": module_payload.get("l1", ""),
                "二级模块": module_payload.get("l2", ""),
                "三级模块": module_payload.get("l3", ""),
                "FPA行序号": hit.get("fpa_seq", ""),
                "功能点名称": hit.get("function_point", ""),
                "生成方式": hit.get("generation", ""),
                "rule_set": audit_report.rule_set,
                "命中对象": hit.get("hit_object", ""),
                "规则ID": hit.get("rule_id", ""),
                "规则说明": hit.get("rule_desc", ""),
                "建议类型": hit.get("suggested_type", ""),
                "是否采用": hit.get("adopted", ""),
                "Warnings": warnings,
            })

    ws_raw = wb.create_sheet("AI原始返回")
    ws_raw.append(sheet_columns["AI原始返回"])
    for audit_report in audit_reports:
        if not audit_report.raw_source and not audit_report.raw_rows and not audit_report.raw_warnings:
            continue
        _append_audit_row(ws_raw, sheet_columns["AI原始返回"], {
            "模块": _group_tag(audit_report.module),
            "三级模块": audit_report.module.get("l3", ""),
            "来源": audit_report.raw_source,
            "Warnings": "；".join(str(x) for x in audit_report.raw_warnings if str(x).strip()),
            "AI原始Rows JSON": _json.dumps(audit_report.raw_rows, ensure_ascii=False, indent=2),
        })

    ws_stability = wb.create_sheet("稳定性报告")
    ws_stability.append(sheet_columns["稳定性报告"])
    stability_summary = stability_report.get("summary", {}) if isinstance(stability_report, dict) else {}
    if not isinstance(stability_summary, dict):
        stability_summary = {}
    _append_audit_row(ws_stability, sheet_columns["稳定性报告"], {
        "范围": "summary",
        "Warning数": stability_summary.get("warning_count", 0),
        "Quality Issue数": stability_summary.get("quality_issue_count", 0),
        "可重试Quality Issue数": stability_summary.get("retryable_quality_issue_count", 0),
        "确认数": stability_summary.get("confirmed_decision_count", 0),
        "重试次数": stability_summary.get("retry_count", 0),
        "来源统计": _json.dumps(stability_summary.get("source_counts", {}), ensure_ascii=False, sort_keys=True),
        "Issue Codes": _json.dumps(stability_summary.get("issue_code_counts", {}), ensure_ascii=False, sort_keys=True),
    })
    stability_modules = stability_report.get("modules", []) if isinstance(stability_report, dict) else []
    if not isinstance(stability_modules, list):
        stability_modules = []
    for module in stability_modules:
        if not isinstance(module, dict):
            continue
        _append_audit_row(ws_stability, sheet_columns["稳定性报告"], {
            "范围": "module",
            "模块序号": module.get("module_index", ""),
            "模块": module.get("module", ""),
            "三级模块": module.get("l3", ""),
            "来源": module.get("source", ""),
            "Warning数": module.get("warning_count", 0),
            "Quality Issue数": module.get("quality_issue_count", 0),
            "可重试Quality Issue数": module.get("retryable_quality_issue_count", 0),
            "确认数": module.get("confirmed_decision_count", 0),
            "重试次数": module.get("retry_count", 0),
            "Issue Codes": _json.dumps(module.get("issue_code_counts", {}), ensure_ascii=False, sort_keys=True),
        })

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    missing_fill = PatternFill("solid", fgColor="FCE4D6")

    for ws in (ws_result, ws_coverage, ws_warnings, ws_rule_hits, ws_raw, ws_stability):
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        for column_cells in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, 10), 48)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="center")
        headers = _header_index(ws)
        if ws.title == "FPA结果":
            warning_idx = headers.get("后处理警告")
            generation_idx = headers.get("生成方式")
            for row in ws.iter_rows(min_row=2):
                warning_value = row[warning_idx].value if warning_idx is not None else ""
                generation_value = row[generation_idx].value if generation_idx is not None else ""
                if str(warning_value or "").strip():
                    for cell in row:
                        cell.fill = warning_fill
                if str(generation_value or "") == "rules_fallback":
                    for cell in row:
                        cell.fill = missing_fill
        if ws.title == "覆盖审核":
            missing_idx = headers.get("未覆盖数")
            for row in ws.iter_rows(min_row=2):
                missing_count = row[missing_idx].value if missing_idx is not None else 0
                try:
                    has_missing = int(missing_count or 0) > 0
                except (TypeError, ValueError):
                    has_missing = False
                if has_missing:
                    for cell in row:
                        cell.fill = missing_fill
        if ws.title == "规则命中详情":
            generation_idx = headers.get("生成方式")
            warning_idx = headers.get("Warnings")
            for row in ws.iter_rows(min_row=2):
                generation_value = row[generation_idx].value if generation_idx is not None else ""
                warning_value = row[warning_idx].value if warning_idx is not None else ""
                if str(warning_value or "").strip():
                    for cell in row:
                        cell.fill = warning_fill
                if str(generation_value or "") == "rules_fallback":
                    for cell in row:
                        cell.fill = missing_fill

    os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
    wb.save(output_path)
    logger.info("FPA 审核副本已生成: %s", output_path)
    return output_path


def init_fpa_template_md(
    tree_md_path: str,
    meta_md_path: str,
    output_md_path: str,
    summary_md_path: str = "",
    profile_name: str = "",
    rule_set: str = "",
) -> str:
    """生成 FPA 模板 MD（规则骨架，F/G 列留空待 AI 填充）。

    Args:
        summary_md_path: 非空时同步写入 gen-fpa-FPA工作量-总和.md（调整值×要素数量 的求和）
    """
    logger.info("第1.1步：生成 FPA 模板 MD...")
    meta = parse_meta_md(meta_md_path)
    rows = parse_module_tree_md(tree_md_path)
    execution = resolve_fpa_execution_config(profile_name, rule_set=rule_set)
    profile = execution.profile
    rule_set_token = set_current_fpa_rule_set_config(execution.rule_set_config)
    try:
        fpa_rows = _build_fpa_rule_rows(rows, meta, profile=profile)
    finally:
        reset_current_fpa_rule_set_config(rule_set_token)

    total = _write_fpa_rows_md(
        fpa_rows,
        output_md_path,
        execution_meta={
            "profile": execution.profile.name,
            "strategy": execution.strategy,
            "rule_set": execution.rule_set,
        },
    )

    logger.info(f"FPA 模板 MD 已生成: {output_md_path} ({len(fpa_rows)} 行)")

    if summary_md_path:
        write_fpa_summary_md(summary_md_path, total)
        logger.info(f"第1.2步：FPA工作量已写入: {summary_md_path} ({total})")

    return output_md_path


def plan_fpa_md_from_tree(
    tree_md_path: str,
    meta_md_path: str,
    output_md_path: str,
    template_path: str = "",
    api_key: str = "",
    model: str = "",
    base_url: str = "",
    summary_md_path: str = "",
    profile_name: str = "",
    strategy: str = "",
    rule_set: str = "",
    audit_trace_path: str = "",
    confirmed_decisions: object | None = None,
    fpa_confirmation_mode: str = "auto",
) -> str:
    """以三级模块为单位 AI 规划 FPA 行，并写入 MD。"""
    logger.info("第1.3步：AI 规划 FPA 数据...")
    meta = parse_meta_md(meta_md_path)
    rows = parse_module_tree_md(tree_md_path)
    profile = get_fpa_profile(profile_name)
    execution = resolve_fpa_execution_config(profile.name, strategy, rule_set)
    profile = execution.profile
    judgement_rules = _read_fpa_judgement_rules(template_path)
    if not judgement_rules:
        logger.warning("未配置「计算依据归类判定原则」，AI 输出的归类将无法按模板 index 映射")
    cache_path = os.path.join(os.path.dirname(output_md_path) or ".", "fpa_ai_cache.json") if api_key else ""
    rule_set_token = set_current_fpa_rule_set_config(execution.rule_set_config)
    try:
        fpa_rows = _plan_fpa_rows_with_ai(
            rows,
            meta,
            judgement_rules,
            api_key,
            model,
            base_url,
            cache_path=cache_path,
            profile=profile,
            strategy=execution.strategy,
            rule_set=execution.rule_set,
            rule_set_config=execution.rule_set_config,
            audit_trace_path=audit_trace_path,
            confirmed_decisions=confirmed_decisions,
            fpa_confirmation_mode=fpa_confirmation_mode,
        )
    finally:
        reset_current_fpa_rule_set_config(rule_set_token)
    total = _write_fpa_rows_md(
        fpa_rows,
        output_md_path,
        ai_filled=any(str(row.get("生成方式", "")) in {"ai", "ai_cache"} for row in fpa_rows),
        execution_meta={
            "profile": execution.profile.name,
            "strategy": execution.strategy,
            "rule_set": execution.rule_set,
        },
    )
    if summary_md_path:
        write_fpa_summary_md(summary_md_path, total)
        logger.info("第1.2步：FPA工作量已更新: %s (%s)", summary_md_path, total)
    logger.info("FPA 规划 MD 已生成: %s (%d 行)", output_md_path, len(fpa_rows))
    return output_md_path


def _prepare_fpa_preview_md_dir(
    *,
    file_path: str,
    work_dir: str,
    keep_preview_files: bool,
    use_preview_cache: bool,
    temp_prefix: str,
    required_files: list[str],
) -> tuple[str, tempfile.TemporaryDirectory[str] | None, bool]:
    temp_ctx: tempfile.TemporaryDirectory[str] | None = None
    if work_dir:
        md_dir = os.path.join(work_dir, "fpa-preview-md")
        os.makedirs(md_dir, exist_ok=True)
    elif keep_preview_files:
        md_dir = os.path.join(os.path.dirname(os.path.abspath(file_path)), ".fpa-preview", "fpa-preview-md")
        os.makedirs(md_dir, exist_ok=True)
    else:
        temp_ctx = tempfile.TemporaryDirectory(prefix=temp_prefix)
        md_dir = temp_ctx.name

    cache_ready = all(os.path.exists(os.path.join(md_dir, name)) for name in required_files)
    if use_preview_cache and cache_ready:
        return md_dir, temp_ctx, True

    from ai_gen_reimbursement_docs.excel_source import generate_md_files

    generate_md_files(file_path, md_dir)
    return md_dir, temp_ctx, False


def preview_fpa_modules(
    *,
    file_path: str,
    work_dir: str = "",
    use_preview_cache: bool = False,
    keep_preview_files: bool = False,
) -> dict[str, object]:
    """解析功能清单并返回可预览的三级模块列表，不调用 AI，不生成正式交付物。"""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"功能清单输入文件不存在: {file_path}")
    if not work_dir and not keep_preview_files and not use_preview_cache:
        base_data = read_base_data_from_excel(file_path)
        rows = base_data["tree_rows"]
        groups = _group_rows_by_l3(rows if isinstance(rows, list) else [])
        modules: list[dict[str, object]] = []
        for idx, group in enumerate(groups, 1):
            processes = group.get("processes", [])
            process_count = len(processes) if isinstance(processes, list) else 0
            path_parts = [
                str(group.get("client_type", "") or "").strip(),
                str(group.get("l1", "") or "").strip(),
                str(group.get("l2", "") or "").strip(),
                str(group.get("l3", "") or "").strip(),
            ]
            label_path = " / ".join([part for part in path_parts if part])
            modules.append({
                "index": idx,
                "client_type": group.get("client_type", ""),
                "l1": group.get("l1", ""),
                "l2": group.get("l2", ""),
                "l3": group.get("l3", ""),
                "l3_desc": group.get("l3_desc", ""),
                "process_count": process_count,
                "label": f"{idx}. {label_path}" if label_path else str(idx),
            })
        return {
            "modules": modules,
            "warnings": [] if modules else ["未解析到三级模块"],
            "preview_md_dir": "",
            "preview_cache_used": False,
        }
    md_dir, temp_ctx, cache_used = _prepare_fpa_preview_md_dir(
        file_path=file_path,
        work_dir=work_dir,
        keep_preview_files=keep_preview_files,
        use_preview_cache=use_preview_cache,
        temp_prefix="ard-fpa-preview-modules-",
        required_files=["0.1.gen-basedata-功能清单-模块树.md"],
    )
    try:
        tree_md = os.path.join(md_dir, "0.1.gen-basedata-功能清单-模块树.md")
        rows = parse_module_tree_md(tree_md)
        modules: list[dict[str, object]] = []
        for idx, group in enumerate(_group_rows_by_l3(rows), 1):
            processes = group.get("processes", [])
            process_count = len(processes) if isinstance(processes, list) else 0
            path_parts = [
                str(group.get("client_type", "") or "").strip(),
                str(group.get("l1", "") or "").strip(),
                str(group.get("l2", "") or "").strip(),
                str(group.get("l3", "") or "").strip(),
            ]
            label_path = " / ".join([part for part in path_parts if part])
            modules.append({
                "index": idx,
                "client_type": group.get("client_type", ""),
                "l1": group.get("l1", ""),
                "l2": group.get("l2", ""),
                "l3": group.get("l3", ""),
                "l3_desc": group.get("l3_desc", ""),
                "process_count": process_count,
                "label": f"{idx}. {label_path}" if label_path else str(idx),
            })
        return {
            "modules": modules,
            "warnings": [] if modules else ["未解析到三级模块"],
            "preview_md_dir": md_dir,
            "preview_cache_used": cache_used,
        }
    finally:
        if temp_ctx is not None:
            temp_ctx.cleanup()


def preview_fpa_module(
    *,
    file_path: str,
    module_name: str = "",
    module_index: int | None = None,
    api_key: str = "",
    model: str = "",
    base_url: str = "",
    template_path: str = "",
    work_dir: str = "",
    profile_name: str = "",
    strategy: str = "",
    rule_set: str = "",
    fpa_confirmation_mode: str = "auto",
    confirmed_decisions: object | None = None,
    use_preview_cache: bool = False,
    keep_preview_files: bool = False,
) -> dict[str, object]:
    """预览单个三级模块的 FPA 规划结果，不生成正式 Excel。"""
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"功能清单输入文件不存在: {file_path}")
    execution = resolve_fpa_execution_config(profile_name, strategy, rule_set)
    profile = execution.profile
    confirmation_mode = normalize_confirmation_mode(fpa_confirmation_mode or "auto")
    normalized_decisions = normalize_confirmed_decisions(confirmed_decisions or {})
    temp_ctx = None
    cache_used = False
    md_dir = ""
    try:
        if not work_dir and not keep_preview_files and not use_preview_cache:
            base_data = read_base_data_from_excel(file_path)
            rows = base_data["tree_rows"] if isinstance(base_data["tree_rows"], list) else []
            meta = base_data["meta"] if isinstance(base_data["meta"], dict) else {}
        else:
            md_dir, temp_ctx, cache_used = _prepare_fpa_preview_md_dir(
                file_path=file_path,
                work_dir=work_dir,
                keep_preview_files=keep_preview_files,
                use_preview_cache=use_preview_cache,
                temp_prefix="ard-fpa-preview-",
                required_files=[
                    "0.1.gen-basedata-功能清单-模块树.md",
                    "0.2.gen-basedata-录入文档元数据-模板.md",
                ],
            )
            tree_md = os.path.join(md_dir, "0.1.gen-basedata-功能清单-模块树.md")
            meta_md = os.path.join(md_dir, "0.2.gen-basedata-录入文档元数据-模板.md")
            rows = parse_module_tree_md(tree_md)
            meta = parse_meta_md(meta_md) if os.path.exists(meta_md) else {}
        groups = _group_rows_by_l3(rows)

        matches: list[tuple[int, dict[str, object]]] = []
        for idx, group in enumerate(groups, 1):
            if module_index is not None and idx == module_index:
                matches.append((idx, group))
            elif module_index is None and module_name and str(group.get("l3", "")) == module_name:
                matches.append((idx, group))
        if module_index is None and not module_name:
            raise ValueError("请指定 module_name 或 module_index")
        if not matches:
            raise ValueError(f"未找到三级模块: {module_name or module_index}")
        if module_index is None and len(matches) > 1:
            indexes = [idx for idx, _ in matches]
            raise ValueError(f"存在多个同名三级模块「{module_name}」，请用 module_index 指定: {indexes}")

        idx, group = matches[0]
        judgement_rules = _read_fpa_judgement_rules(template_path)
        warnings: list[str] = []
        confirmation_issues: list[FpaValidationIssue] = []
        used_ai = execution.strategy in {"ai_first", "ai_only"}
        debug: dict[str, object] = {
            "ai_called": False,
            "reason": execution.strategy,
            "model": model,
            "system_prompt": "",
            "system_prompt_source": "未配置",
            "user_prompt": "",
            "user_prompt_source": "未配置",
            "core_rules_source": "未配置",
            "ai_prompt": "",
            "raw_response": "",
            "thinking": "",
            "parsed_rows": [],
            "final_rows": [],
            "agent_review": {},
            "quality_review": {},
        }
        rule_set_token = set_current_fpa_rule_set_config(execution.rule_set_config)
        try:
            if execution.strategy in {"rules_first", "rules_only"}:
                used_ai = False
                debug["reason"] = execution.strategy
                fpa_rows = profile.fallback_rows_for_l3(
                    group,
                    meta,
                    start_seq=1,
                    judgement_rules=judgement_rules,
                )
                _attach_profile_rule_hits(fpa_rows, profile=profile, generation="fallback")
                _fill_fallback_classification_basis(fpa_rows, judgement_rules, profile)
                if execution.strategy == "rules_first":
                    rules_first_reasons = _rules_first_ai_reasons(group, fpa_rows)
                    if rules_first_reasons and api_key:
                        debug["reason"] = "rules_first_needs_ai"
                        raw_rows, debug = _ai_plan_fpa_rows_for_l3_debug(
                            group, judgement_rules, _build_domain_context(meta), api_key, model, base_url,
                            profile=profile,
                            confirmed_decisions=normalized_decisions,
                        )
                        debug["reason"] = "rules_first_needs_ai"
                        fpa_rows, warnings = _normalize_ai_fpa_rows_for_l3(
                            group=group,
                            meta=meta,
                            ai_rows=raw_rows,
                            judgement_rules=judgement_rules,
                            start_seq=1,
                            profile=profile,
                            strategy=execution.strategy,
                        )
                        fpa_rows, supplement_warnings = _supplement_ai_rows_with_rules(
                            group=group,
                            meta=meta,
                            ai_rows=fpa_rows,
                            profile=profile,
                            strategy=execution.strategy,
                            rule_set_config=execution.rule_set_config,
                            judgement_rules=judgement_rules,
                        )
                        warnings.extend(supplement_warnings)
                        _fill_fallback_classification_basis(fpa_rows, judgement_rules, profile)
                        validation_issues, validation_warnings = _validate_fpa_rows_for_l3(
                            group=group,
                            rows=fpa_rows,
                        )
                        confirmation_issues.extend(validation_issues)
                        warnings.extend(validation_warnings)
                        warnings.insert(0, "规则结果触发 AI 复核: " + "；".join(rules_first_reasons))
                        if not fpa_rows:
                            raise ValueError("AI 规划未生成有效 FPA 行")
                        used_ai = True
                    elif rules_first_reasons:
                        warning = "规则结果需要 AI 复核但未配置 API Key，已保留规则生成结果: " + "；".join(rules_first_reasons)
                        warnings.append(warning)
                        debug["reason"] = "rules_first_needs_ai_missing_api_key"
                    else:
                        debug["reason"] = "rules_first_rules_ok"
            elif not api_key:
                debug["reason"] = "missing_api_key"
                raise ValueError(f"FPA strategy={execution.strategy} 需要 API Key，当前未配置")
            else:
                raw_rows, debug = _ai_plan_fpa_rows_for_l3_debug(
                    group, judgement_rules, _build_domain_context(meta), api_key, model, base_url,
                    profile=profile,
                    confirmed_decisions=normalized_decisions,
                )
                fpa_rows, warnings = _normalize_ai_fpa_rows_for_l3(
                    group=group,
                    meta=meta,
                    ai_rows=raw_rows,
                    judgement_rules=judgement_rules,
                    start_seq=1,
                    profile=profile,
                    strategy=execution.strategy,
                )
                fpa_rows, supplement_warnings = _supplement_ai_rows_with_rules(
                    group=group,
                    meta=meta,
                    ai_rows=fpa_rows,
                    profile=profile,
                    strategy=execution.strategy,
                    rule_set_config=execution.rule_set_config,
                    judgement_rules=judgement_rules,
                )
                warnings.extend(supplement_warnings)
                _fill_fallback_classification_basis(fpa_rows, judgement_rules, profile)
                validation_issues, validation_warnings = _validate_fpa_rows_for_l3(
                    group=group,
                    rows=fpa_rows,
                )
                confirmation_issues.extend(validation_issues)
                warnings.extend(validation_warnings)
                if not fpa_rows:
                    raise ValueError("AI 规划未生成有效 FPA 行")
        except Exception as exc:
            from ai_gen_reimbursement_docs.config_utils import FpaPromptConfigError

            if isinstance(exc, FpaAiDebugError):
                debug = exc.debug
            debug["error"] = str(exc)
            if isinstance(exc, FpaPromptConfigError):
                raise
            if execution.strategy == "ai_only" or not api_key:
                raise
            used_ai = False
            debug["reason"] = "ai_failed_fallback"
            warnings.append(f"AI 调用或解析失败，已使用兜底生成: {exc}")
            logger.warning("FPA 预览 AI 失败 [%s]: %s", _group_tag(group), exc)
            fpa_rows = profile.fallback_rows_for_l3(
                group,
                meta,
                start_seq=1,
                judgement_rules=judgement_rules,
            )
            _attach_profile_rule_hits(fpa_rows, profile=profile, generation="fallback")
            _fill_fallback_classification_basis(fpa_rows, judgement_rules, profile)
        finally:
            reset_current_fpa_rule_set_config(rule_set_token)
        warnings = _with_config_warnings(warnings, execution.rule_set_config)

        def _row_to_preview(row: dict[str, object]) -> dict[str, object]:
            basis = str(row.get("计算依据归类", ""))
            basis_index = judgement_rules.index(basis) + 1 if basis in judgement_rules else None
            return {
                "name": row.get("新增/修改功能点", ""),
                "type": row.get("类型", ""),
                "type_reason": row.get("类型理由", ""),
                "classification_basis": basis,
                "classification_basis_index": basis_index,
                "explanation": _format_fpa_explanation(str(row.get("计算依据说明", "") or "")),
                "source_processes": [
                    x for x in str(row.get("源功能过程", "")).split("、") if x
                ],
                "source_process_ids": sorted(_source_process_ids_from_row(row)),
                "generation": row.get("生成方式", "ai" if used_ai else "fallback"),
            }

        preview_rows = [_row_to_preview(row) for row in fpa_rows]
        debug["final_rows"] = preview_rows
        quality_review = _build_quality_review_for_l3(
            group=group,
            rows=fpa_rows,
            confirmed_decisions=normalized_decisions,
        )
        agent_review = _build_agent_review_for_l3(
            group=group,
            rows=fpa_rows,
            confirmed_decisions=normalized_decisions,
            profile=profile,
        )
        debug["agent_review"] = agent_review
        debug["quality_review"] = quality_review
        confirmation_questions = build_fpa_confirmation_questions(
            group=group,
            issues=confirmation_issues,
            mode=confirmation_mode,
            confirmed_decisions=normalized_decisions,
        )
        raw_source = "ai" if debug.get("ai_called") and used_ai else "rules"
        if debug.get("reason") == "ai_failed_fallback":
            raw_source = "rules_fallback"
        raw_rows = debug.get("parsed_rows", [])
        audit = _build_fpa_audit_reports_for_groups(
            groups=[group],
            rows_by_module={1: fpa_rows},
            warnings_by_module={1: warnings},
            profile=profile,
            profile_version=profile.version,
            strategy=execution.strategy,
            rule_set=execution.rule_set,
            raw_audit_by_module={
                1: {
                    "source": raw_source,
                    "raw_rows": raw_rows if isinstance(raw_rows, list) else [],
                    "warnings": warnings,
                }
            },
        )[0]
        audit.module["index"] = idx
        module_payload = dict(audit.module)
        return {
            "module": module_payload,
            "rows": preview_rows,
            "warnings": warnings,
            "status": "needs_confirmation" if confirmation_questions else "ok",
            "confirmation_mode": confirmation_mode,
            "confirmation_questions": confirmation_questions,
            "confirmed_decision_count": confirmed_decision_count(normalized_decisions),
            "used_ai": used_ai,
            "profile": profile.name,
            "profile_version": profile.version,
            "strategy": execution.strategy,
            "rule_set": execution.rule_set,
            "audit": audit.to_dict(),
            "preview_md_dir": md_dir,
            "preview_cache_used": cache_used,
            "debug": debug,
        }
    finally:
        if temp_ctx is not None:
            temp_ctx.cleanup()


def _format_fpa_explanation(text: str) -> str:
    """格式化 FPA 计算依据说明：加换行排版，不改变原文内容。"""
    NL = chr(10)
    text = text.lstrip("：: ")
    text = text.replace("具体如下", "具体如下" + NL + NL)
    text = text.replace("事件流：", NL + "事件流：")
    text = text.replace("触发事件：", NL + "触发事件：")
    text = text.replace("；", "；" + NL)
    text = text.replace("业务规则", NL + "业务规则")
    text = text.replace("业务数据", NL + "业务数据")
    text = text.replace("涉及表", NL + "涉及表")
    text = text.replace("涉及服务", NL + "涉及服务")
    text = text.replace("；涉及接口", "；" + NL + "涉及接口")
    text = re.sub(re.compile(r"(?<=\S)\s+(?=\d+\.)"), NL, text)
    text = re.sub(re.compile(r"^[	 ]+", re.MULTILINE), "", text)
    text = re.sub(re.compile(r'\n[：:;；]\s*\n'), '\n', text)
    text = re.sub(re.compile(NL + "{3,}"), NL + NL, text)
    return text


def _as_manifest_int(value: object, default: int) -> int:
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return default
    return parsed if parsed > 0 else default


def _fpa_result_sheet_spec(manifest: dict[str, Any]) -> dict[str, Any]:
    sheets = manifest.get("sheets", {}) or {}
    spec = sheets.get("result", {}) if isinstance(sheets, dict) else {}
    if not isinstance(spec, dict):
        spec = {}
    return {
        "name": str(spec.get("name") or "FPA功能点估算"),
        "header_row": _as_manifest_int(spec.get("header_row"), 2),
        "data_start_row": _as_manifest_int(spec.get("data_start_row"), 3),
        "style_source_row": _as_manifest_int(spec.get("style_source_row"), 3),
        "columns": spec.get("columns", {}) if isinstance(spec.get("columns", {}), dict) else {},
        "named_cells": spec.get("named_cells", {}) if isinstance(spec.get("named_cells", {}), dict) else {},
    }


def _fpa_header_map(ws, header_row: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col_idx in range(1, ws.max_column + 1):
        value = ws.cell(header_row, col_idx).value
        if value is None:
            continue
        header = str(value).strip()
        if header and header not in headers:
            headers[header] = col_idx
    return headers


def _fpa_manifest_header(sheet_spec: dict[str, Any], key: str) -> str:
    columns = sheet_spec.get("columns", {}) or {}
    if not isinstance(columns, dict):
        return ""
    spec = columns.get(key, {})
    if isinstance(spec, str):
        return spec.strip()
    if isinstance(spec, dict):
        return str(spec.get("header", "") or "").strip()
    return ""


def _fpa_column_by_header(
    headers: dict[str, int],
    sheet_spec: dict[str, Any],
    key: str,
    candidates: tuple[str, ...],
    fallback_col: int,
) -> int:
    manifest_header = _fpa_manifest_header(sheet_spec, key)
    search = (manifest_header,) + candidates if manifest_header else candidates
    for header in search:
        if header and header in headers:
            return headers[header]
    return fallback_col


def _fpa_manifest_named_cell(sheet_spec: dict[str, Any], key: str) -> str:
    named_cells = sheet_spec.get("named_cells", {}) or {}
    if not isinstance(named_cells, dict):
        return ""
    spec = named_cells.get(key, "")
    if isinstance(spec, str):
        return spec.strip()
    if isinstance(spec, dict):
        return str(spec.get("name", "") or "").strip()
    return ""


def _fpa_named_cell_target(wb, name: str, *, expected_sheet: str) -> tuple[str, int, int] | None:
    if not name:
        return None
    defined_name = wb.defined_names.get(name)
    if defined_name is None:
        logger.warning("fpa manifest named_cells 指向的命名单元格不存在: %s", name)
        return None
    try:
        destinations = list(defined_name.destinations)
    except Exception as exc:
        logger.warning("fpa manifest named_cells 无法解析命名单元格 %s: %s", name, exc)
        return None
    if len(destinations) != 1:
        logger.warning("fpa manifest named_cells 仅支持单一目标命名单元格: %s", name)
        return None
    sheet_name, coord = destinations[0]
    if sheet_name != expected_sheet:
        logger.warning(
            "fpa manifest named_cells 命名单元格 %s 指向 sheet %s，期望 %s",
            name,
            sheet_name,
            expected_sheet,
        )
        return None
    try:
        min_col, min_row, max_col, max_row = openpyxl.utils.cell.range_boundaries(str(coord))
    except ValueError:
        logger.warning("fpa manifest named_cells 命名单元格 %s 坐标无效: %s", name, coord)
        return None
    if min_col != max_col or min_row != max_row:
        logger.warning("fpa manifest named_cells 仅支持单个单元格目标: %s -> %s", name, coord)
        return None
    return sheet_name, min_row, min_col


def _fpa_named_cell_row(wb, ws, sheet_spec: dict[str, Any], key: str, fallback: int) -> int:
    name = _fpa_manifest_named_cell(sheet_spec, key)
    target = _fpa_named_cell_target(wb, name, expected_sheet=ws.title) if name else None
    return target[1] if target else fallback


def _fpa_named_cell_address(wb, ws, sheet_spec: dict[str, Any], key: str) -> tuple[int, int] | None:
    name = _fpa_manifest_named_cell(sheet_spec, key)
    target = _fpa_named_cell_target(wb, name, expected_sheet=ws.title) if name else None
    if target is None:
        return None
    _sheet_name, row, col = target
    return row, col


def _replace_fpa_formula_refs(formula: str, *, row: int, columns: dict[str, int]) -> str:
    replacements = {
        "E3": f"{openpyxl.utils.get_column_letter(columns['type'])}{row}",
        "H3": f"{openpyxl.utils.get_column_letter(columns['status'])}{row}",
        "I3": f"{openpyxl.utils.get_column_letter(columns['formula_base'])}{row}",
        "J3": f"{openpyxl.utils.get_column_letter(columns['adjust'])}{row}",
        "K3": f"{openpyxl.utils.get_column_letter(columns['elements'])}{row}",
        "J{row}": f"{openpyxl.utils.get_column_letter(columns['adjust'])}{row}",
        "K{row}": f"{openpyxl.utils.get_column_letter(columns['elements'])}{row}",
    }
    result = formula
    for old, new in replacements.items():
        result = result.replace(old, new)
    return result


def generate_fpa_xlsx_from_md(
    fpa_md_path: str,
    meta_md_path: str,
    template_path: str,
    output_path: str,
) -> str:
    """从已填充的 FPA MD 生成 FPA工作量评估.xlsx。"""
    logger.info("第1.4步：从 FPA MD 生成 Excel...")

    meta = parse_meta_md(meta_md_path)
    base_formula = meta.get("基准值公式", "")
    workload_formula = meta.get("FPA工作量公式", "J{row}*K{row}")

    fpa_rows = []
    with open(fpa_md_path, encoding='utf-8') as f:
        in_table = False
        for line in f:
            line = line.rstrip()
            if "| 序号 | 子系统" in line:
                in_table = True
                continue
            if "|------|" in line and in_table:
                continue
            if in_table:
                cells = parse_md_table_row(line, min_cols=14)
                if cells is not None:
                    fpa_rows.append({
                        "序号": cells[0],
                        "子系统(模块)": cells[1],
                        "资产标识": cells[2],
                        "新增/修改功能点": cells[3],
                        "类型": cells[4],
                        "计算依据归类": cells[5],
                        "计算依据说明": cells[6],
                        "变更状态": cells[7],
                        "调整值": cells[8],
                        "要素数量": cells[9],
                    })

    manifest, _, _ = load_template_manifest("fpa", template_path)
    sheet_spec = _fpa_result_sheet_spec(manifest)
    wb = safe_load_workbook(template_path, 'FPA工作量评估')
    from ai_gen_reimbursement_docs.config_utils import _get_system_config_value
    _fpa_sheet = _get_system_config_value('fpa_sheet', sheet_spec["name"])
    ws = wb[_fpa_sheet if _fpa_sheet in wb.sheetnames else sheet_spec["name"]]
    header_row = sheet_spec["header_row"]
    data_start_row = _fpa_named_cell_row(wb, ws, sheet_spec, "data_start", sheet_spec["data_start_row"])
    style_source_row = sheet_spec["style_source_row"]
    headers = _fpa_header_map(ws, header_row)
    fpa_columns = {
        "seq": _fpa_column_by_header(headers, sheet_spec, "seq", ("序号",), FPA_COL_SEQ),
        "subsystem": _fpa_column_by_header(headers, sheet_spec, "subsystem", ("子系统(模块)", "子系统"), FPA_COL_SUBSYSTEM),
        "asset": _fpa_column_by_header(headers, sheet_spec, "asset", ("资产标识",), FPA_COL_ASSET),
        "function_point_name": _fpa_column_by_header(headers, sheet_spec, "function_point_name", ("新增/修改功能点",), FPA_COL_FUNC_POINT),
        "type": _fpa_column_by_header(headers, sheet_spec, "type", ("类型",), FPA_COL_TYPE),
        "classification_basis": _fpa_column_by_header(headers, sheet_spec, "classification_basis", ("计算依据归类",), FPA_COL_CLASSIFICATION),
        "explanation": _fpa_column_by_header(
            headers,
            sheet_spec,
            "explanation",
            ("计算依据说明", "计算依据说明，记录关键信息如：事件流、业务规则、业务数据、非功能性规约、表、服务、接口等内容"),
            FPA_COL_EXPLANATION,
        ),
        "status": _fpa_column_by_header(headers, sheet_spec, "status", ("变更状态",), FPA_COL_STATUS),
        "formula_base": _fpa_column_by_header(headers, sheet_spec, "formula_base", ("基准值",), FPA_COL_FORMULA_BASE),
        "adjust": _fpa_column_by_header(headers, sheet_spec, "adjust", ("调整值",), FPA_COL_ADJUST),
        "elements": _fpa_column_by_header(headers, sheet_spec, "elements", ("要素数量",), FPA_COL_ELEMENTS),
        "formula_workload": _fpa_column_by_header(headers, sheet_spec, "formula_workload", ("FPA工作量",), FPA_COL_FORMULA_WORKLOAD),
    }
    fpa_key_columns = {
        "序号": fpa_columns["seq"],
        "子系统(模块)": fpa_columns["subsystem"],
        "资产标识": fpa_columns["asset"],
        "新增/修改功能点": fpa_columns["function_point_name"],
        "类型": fpa_columns["type"],
        "计算依据归类": fpa_columns["classification_basis"],
        "计算依据说明": fpa_columns["explanation"],
        "变更状态": fpa_columns["status"],
        "调整值": fpa_columns["adjust"],
        "要素数量": fpa_columns["elements"],
    }
    total_cols = max(FPA_TOTAL_COLS, ws.max_column, *fpa_columns.values())

    tmpl_format = {}
    for col_idx in range(1, total_cols + 1):
        c = ws.cell(style_source_row, col_idx)
        tmpl_format[col_idx] = {
            'font': copy(c.font) if c.font else None,
            'fill': copy(c.fill) if c.fill else None,
            'border': copy(c.border) if c.border else None,
            'number_format': c.number_format,
            'alignment': copy(c.alignment) if c.alignment else None,
        }
    for col_idx in (fpa_columns["formula_base"], fpa_columns["formula_workload"]):
        c = ws.cell(header_row, col_idx)
        if c.fill:
            tmpl_format[col_idx]['fill'] = copy(c.fill)

    if ws.max_row >= data_start_row:
        ws.delete_rows(data_start_row, ws.max_row - data_start_row + 1)

    for i, fpa_row in enumerate(fpa_rows):
        excel_row = i + data_start_row
        for key, col_idx in fpa_key_columns.items():
            val = fpa_row.get(key, "")
            cell = ws.cell(excel_row, col_idx)
            if col_idx in (fpa_columns["seq"], fpa_columns["adjust"], fpa_columns["elements"]):
                try:
                    cell.value = int(val)
                except (ValueError, TypeError):
                    cell.value = val
            elif col_idx == fpa_columns["explanation"]:
                cell.value = _format_fpa_explanation(val)
            else:
                cell.value = val
        if base_formula:
            formula = _replace_fpa_formula_refs(base_formula, row=excel_row, columns=fpa_columns)
            ws.cell(excel_row, fpa_columns["formula_base"]).value = f"={formula}" if not formula.startswith('=') else formula
        if workload_formula:
            formula = _replace_fpa_formula_refs(workload_formula, row=excel_row, columns=fpa_columns)
            ws.cell(excel_row, fpa_columns["formula_workload"]).value = f"={formula}" if not formula.startswith('=') else formula
        ws.cell(excel_row, FPA_TOTAL_COLS - 1, "")
        ws.cell(excel_row, FPA_TOTAL_COLS, "")

        for col_idx in range(1, total_cols + 1):
            c = ws.cell(excel_row, col_idx)
            fmt = tmpl_format.get(col_idx, {})
            if fmt.get('font'):
                c.font = fmt['font']
            if fmt.get('border'):
                c.border = fmt['border']
            if fmt.get('number_format'):
                c.number_format = fmt['number_format']
            if col_idx in (fpa_columns["formula_base"], fpa_columns["formula_workload"]) and fmt.get('fill'):
                c.fill = fmt['fill']
            if col_idx in (fpa_columns["function_point_name"], fpa_columns["explanation"]):
                orig_align = fmt.get('alignment')
                h = 'left' if col_idx == fpa_columns["explanation"] else (orig_align.horizontal or 'center')
                if orig_align:
                    c.alignment = Alignment(
                        wrap_text=True,
                        vertical='center',
                        horizontal=h,
                    )
                else:
                    c.alignment = Alignment(wrap_text=True, vertical='center', horizontal=h)
            else:
                if fmt.get('alignment'):
                    c.alignment = fmt['alignment']

    last_data_row = data_start_row + len(fpa_rows) - 1
    summary_total = _fpa_named_cell_address(wb, ws, sheet_spec, "summary_total")
    summary_columns = [
        fpa_columns["formula_base"],
        fpa_columns["adjust"],
        fpa_columns["elements"],
        fpa_columns["formula_workload"],
        FPA_TOTAL_COLS - 1,
    ]
    if summary_total is not None:
        summary_row, summary_col = summary_total
        ws.cell(summary_row, summary_col).value = (
            f"=SUM({openpyxl.utils.get_column_letter(fpa_columns['formula_workload'])}"
            f"{data_start_row}:{openpyxl.utils.get_column_letter(fpa_columns['formula_workload'])}{last_data_row})"
        )
        summary_columns = [
            col_idx
            for col_idx in summary_columns
            if col_idx not in {summary_col, fpa_columns["formula_workload"]}
        ]
    for col_idx in summary_columns:
        cell = ws.cell(1, col_idx)
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        cell.value = f"=SUM({col_letter}{data_start_row}:{col_letter}{last_data_row})"
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    try:
        wb.save(output_path)
    except PermissionError:
        temp_path = output_path.rsplit('.', 1)[0] + '_TEMP.xlsx'
        wb.save(temp_path)
        logger.warning(
            "文件被占用，已保存到临时文件: %s\n"
            "关闭 Excel/WPS 后，将 _TEMP 文件重命名替换原文件即可", temp_path
        )
        return temp_path
    logger.info(f"FPA工作量评估已生成: {output_path} ({len(fpa_rows)} 行)")

    return output_path
