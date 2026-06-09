"""生成 项目需求清单.xlsx"""

import copy as _cpy
import logging
import os
from typing import Any

from openpyxl.utils.cell import range_boundaries
from openpyxl.styles import Alignment

from ai_gen_reimbursement_docs.constants import (
    REQ_COL_SEQ, REQ_COL_PROJECT, REQ_COL_SUBSYSTEM, REQ_COL_L1, REQ_COL_L2,
    REQ_COL_L3, REQ_COL_PROC_TYPE, REQ_COL_WORKLOAD, REQ_COL_CFP, REQ_TOTAL_COLS,
    REQ_COL_KEY_MAP,
)
from ai_gen_reimbursement_docs.excel_source import (
    parse_module_tree_md, safe_load_workbook,
)
from ai_gen_reimbursement_docs.template_manifest import load_template_manifest

logger = logging.getLogger('ai_gen_reimbursement_docs.gen_list')


def _as_int(value: Any, default: int) -> int:
    try:
        return int(value)
    except (TypeError, ValueError):
        return default


def _sheet_spec(
    manifest: dict[str, Any],
    key: str,
    *,
    default_name: str,
    default_header_row: int = 2,
    default_data_start_row: int = 3,
    default_style_source_row: int = 3,
) -> dict[str, Any]:
    sheets = manifest.get("sheets", {}) or {}
    spec = sheets.get(key, {}) if isinstance(sheets, dict) else {}
    if isinstance(spec, str):
        spec = {"name": spec}
    if not isinstance(spec, dict):
        spec = {}
    return {
        **spec,
        "name": str(spec.get("name") or default_name),
        "header_row": _as_int(spec.get("header_row"), default_header_row),
        "data_start_row": _as_int(spec.get("data_start_row"), default_data_start_row),
        "style_source_row": _as_int(spec.get("style_source_row"), default_style_source_row),
    }


def _cell_text(value: Any) -> str:
    return str(value or "").strip()


def _header_map(ws, header_row: int) -> dict[str, int]:
    headers: dict[str, int] = {}
    if header_row < 1 or header_row > ws.max_row:
        return headers
    for col_idx in range(1, ws.max_column + 1):
        text = _cell_text(ws.cell(header_row, col_idx).value)
        if text and text not in headers:
            headers[text] = col_idx
    return headers


def _manifest_header(sheet_spec: dict[str, Any], key: str) -> str:
    columns = sheet_spec.get("columns", {}) or {}
    if not isinstance(columns, dict):
        return ""
    spec = columns.get(key, {})
    if isinstance(spec, str):
        return spec
    if isinstance(spec, dict):
        return str(spec.get("header", "") or "").strip()
    return ""


def _manifest_named_cell(sheet_spec: dict[str, Any], key: str) -> str:
    named_cells = sheet_spec.get("named_cells", {}) or {}
    if not isinstance(named_cells, dict):
        return ""
    spec = named_cells.get(key, "")
    if isinstance(spec, str):
        return spec.strip()
    if isinstance(spec, dict):
        return str(spec.get("name", "") or "").strip()
    return ""


def _named_cell_target(wb, name: str, *, expected_sheet: str) -> tuple[str, int, int] | None:
    if not name:
        return None
    defined_name = wb.defined_names.get(name)
    if defined_name is None:
        logger.warning("list manifest named_cells 指向的命名单元格不存在: %s", name)
        return None
    try:
        destinations = list(defined_name.destinations)
    except Exception as exc:
        logger.warning("list manifest named_cells 无法解析命名单元格 %s: %s", name, exc)
        return None
    if len(destinations) != 1:
        logger.warning("list manifest named_cells 仅支持单一目标命名单元格: %s", name)
        return None
    sheet_name, coord = destinations[0]
    if sheet_name != expected_sheet:
        logger.warning(
            "list manifest named_cells 命名单元格 %s 指向 sheet %s，期望 %s",
            name,
            sheet_name,
            expected_sheet,
        )
        return None
    try:
        min_col, min_row, max_col, max_row = range_boundaries(str(coord))
    except ValueError:
        logger.warning("list manifest named_cells 命名单元格 %s 坐标无效: %s", name, coord)
        return None
    if min_col != max_col or min_row != max_row:
        logger.warning("list manifest named_cells 仅支持单个单元格目标: %s -> %s", name, coord)
        return None
    return sheet_name, min_row, min_col


def _write_named_cell_if_configured(
    wb,
    ws,
    sheet_spec: dict[str, Any],
    key: str,
    value: Any,
) -> bool:
    name = _manifest_named_cell(sheet_spec, key)
    if not name:
        return False
    target = _named_cell_target(wb, name, expected_sheet=ws.title)
    if target is None:
        return False
    _sheet_name, row, col = target
    ws.cell(row, col, value)
    return True


def _column_by_header(
    headers: dict[str, int],
    sheet_spec: dict[str, Any],
    key: str,
    aliases: tuple[str, ...],
    fallback: int,
) -> int:
    candidates = []
    manifest_header = _manifest_header(sheet_spec, key)
    if manifest_header:
        candidates.append(manifest_header)
    candidates.extend(aliases)
    for header in candidates:
        if header in headers:
            return headers[header]
    return fallback


def parse_meta_md(meta_md_path: str) -> dict[str, str]:
    """解析文档元数据.md 为扁平字典。支持跨多行的表格值。"""
    from ai_gen_reimbursement_docs.gen_spec import parse_meta_md
    return parse_meta_md(meta_md_path)


def generate_list_xlsx_from_md(
    meta_md_path: str,
    tree_md_path: str,
    template_path: str,
    output_path: str,
    cfp_total: float = 0,
    fpa_reduced: float = 0,
) -> str:
    """生成项目需求清单.xlsx。

    cfp_total: 送审功能点 = gen-fpa-FPA工作量-总和.md 的原始值
    fpa_reduced: 送审工作量 = FPA 核减后的工作量
    """
    logger.info("第4.1步：开始生成项目需求清单.xlsx...")

    meta = parse_meta_md(meta_md_path)
    rows = parse_module_tree_md(tree_md_path)
    manifest, _, _ = load_template_manifest("list", template_path)

    wb = safe_load_workbook(template_path, '项目需求清单')

    # ====== Sheet 1: 项目信息概览 ======
    project_spec = _sheet_spec(manifest, "project_info", default_name="项目信息概览")
    ws1 = wb[project_spec["name"]]
    project_header_row = project_spec["header_row"]
    project_data_row = project_spec["data_start_row"]
    project_headers = _header_map(ws1, project_header_row)

    title = meta.get("项目信息概览-标题", "")
    ws1.cell(1, 1, title)

    project_field_map = {
        "project_name": ("项目名称", "项目信息概览-项目名称", 2),
        "subsystem": ("子系统名称", "项目信息概览-子系统名称", 3),
        "project_type": ("项目类型", "项目信息概览-项目类型", 4),
        "domain": ("所属域", "项目信息概览-所属域", 5),
        "system": ("所属系统", "项目信息概览-所属系统", 6),
        "department": ("需求部门", "项目信息概览-需求部门", 7),
        "owner": ("需求负责人", "项目信息概览-需求负责人", 8),
        "owner_contact": ("需求负责人联系方式", "项目信息概览-需求负责人联系方式", 9),
    }
    for key, (header, meta_key, fallback_col) in project_field_map.items():
        value = meta.get(meta_key, "")
        if _write_named_cell_if_configured(wb, ws1, project_spec, key, value):
            continue
        col_idx = _column_by_header(project_headers, project_spec, key, (header,), fallback_col)
        if header in project_headers or _manifest_header(project_spec, key) or not project_headers:
            ws1.cell(project_data_row, col_idx, value)

    if not _write_named_cell_if_configured(wb, ws1, project_spec, "workload", fpa_reduced):
        workload_col = _column_by_header(project_headers, project_spec, "workload", ("送审工作量",), 10)
        ws1.cell(project_data_row, workload_col, fpa_reduced)
    if not _write_named_cell_if_configured(wb, ws1, project_spec, "cfp", cfp_total):
        cfp_col = _column_by_header(project_headers, project_spec, "cfp", ("送审功能点",), 11)
        ws1.cell(project_data_row, cfp_col, cfp_total)

    # ====== Sheet 2: 功能清单 ======
    function_spec = _sheet_spec(manifest, "function_list", default_name="功能清单")
    ws2 = wb[function_spec["name"]]
    function_header_row = function_spec["header_row"]
    function_data_start_row = function_spec["data_start_row"]
    function_style_source_row = function_spec["style_source_row"]
    function_headers = _header_map(ws2, function_header_row)

    req_cols = {
        "序号": _column_by_header(function_headers, function_spec, "seq", ("序号",), REQ_COL_SEQ),
        "项目名称": _column_by_header(function_headers, function_spec, "project_name", ("项目名称",), REQ_COL_PROJECT),
        "子系统": _column_by_header(function_headers, function_spec, "subsystem", ("子系统",), REQ_COL_SUBSYSTEM),
        "一级模块": _column_by_header(function_headers, function_spec, "module_l1", ("一级功能模块名称", "一级模块"), REQ_COL_L1),
        "二级模块": _column_by_header(function_headers, function_spec, "module_l2", ("二级功能模块名称", "二级模块"), REQ_COL_L2),
        "三级模块": _column_by_header(function_headers, function_spec, "module_l3", ("三级功能模块名称", "三级模块"), REQ_COL_L3),
        "功能过程类型": _column_by_header(function_headers, function_spec, "type", ("类型", "功能过程类型"), REQ_COL_PROC_TYPE),
        "送审工作量": _column_by_header(function_headers, function_spec, "workload", ("送审工作量",), REQ_COL_WORKLOAD),
        "送审功能点": _column_by_header(function_headers, function_spec, "cfp", ("送审功能点",), REQ_COL_CFP),
    }
    total_cols = max(REQ_TOTAL_COLS, ws2.max_column, *req_cols.values())

    fl_title = meta.get("功能清单-标题", "")
    ws2.cell(1, 1, fl_title)

    _center = Alignment(horizontal='center', vertical='center')
    _center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)

    _tmpl_border = _cpy.copy(ws2.cell(function_style_source_row, 1).border)

    for merge_range in list(ws2.merged_cells.ranges):
        ws2.unmerge_cells(str(merge_range))

    if ws2.max_row >= function_data_start_row:
        ws2.delete_rows(function_data_start_row, ws2.max_row - function_data_start_row + 1)

    project_name = meta.get("功能清单-项目名称", "")
    subsystem = meta.get("功能清单-子系统", "")

    data_rows_data = []
    seen_modules = set()
    seq = 0

    for r in rows:
        key = (r["一级模块"], r["二级模块"], r["三级模块"])
        if key not in seen_modules:
            seen_modules.add(key)
            seq += 1
            row_idx = function_data_start_row + seq - 1
            _req_data = {
                "序号": seq, "项目名称": project_name, "子系统": subsystem,
                "一级模块": r["一级模块"], "二级模块": r["二级模块"],
                "三级模块": r["三级模块"], "功能过程类型": r["功能过程类型"],
            }
            for col_idx in range(1, total_cols + 1):
                c = ws2.cell(row_idx, col_idx)
                c.alignment = _center
                c.border = _tmpl_border
            for key_name in REQ_COL_KEY_MAP.values():
                ws2.cell(row_idx, req_cols[key_name], _req_data.get(key_name, ""))
            ws2.cell(row_idx, req_cols["项目名称"]).alignment = _center_wrap

            ws2.cell(row_idx, req_cols["送审工作量"], fpa_reduced)
            ws2.cell(row_idx, req_cols["送审功能点"], cfp_total)
            
            data_rows_data.append({
                "row": row_idx,
                "project_name": project_name,
                "subsystem": subsystem,
                "module_l1": r["一级模块"],
                "module_l2": r["二级模块"],
            })

    if seq > 0:
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
        ws2.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    merge_cols = {
        req_cols["项目名称"]: "project_name",
        req_cols["子系统"]: "subsystem",
        req_cols["一级模块"]: "module_l1",
        req_cols["二级模块"]: "module_l2",
    }
    for col_idx, val_key in merge_cols.items():
        i = 0
        while i < len(data_rows_data):
            curr_val = data_rows_data[i][val_key]
            j = i
            while j < len(data_rows_data) and data_rows_data[j][val_key] == curr_val:
                j += 1
            count = j - i
            if count > 1:
                ws2.merge_cells(
                    start_row=data_rows_data[i]["row"],
                    start_column=col_idx,
                    end_row=data_rows_data[j - 1]["row"],
                    end_column=col_idx
                )
                ws2.cell(data_rows_data[i]["row"], col_idx).border = _tmpl_border
            i = j

    if len(data_rows_data) > 1:
        for _col in (req_cols["送审工作量"], req_cols["送审功能点"]):
            ws2.merge_cells(
                start_row=data_rows_data[0]["row"],
                start_column=_col,
                end_row=data_rows_data[-1]["row"],
                end_column=_col
            )
            _top_cell = ws2.cell(data_rows_data[0]["row"], _col)
            _top_cell.border = _tmpl_border
            _top_cell.alignment = _center

        for _col in (req_cols["送审工作量"], req_cols["送审功能点"]):
            for _r in range(data_rows_data[0]["row"], data_rows_data[-1]["row"] + 1):
                ws2.cell(_r, _col).border = _tmpl_border

    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    try:
        wb.save(output_path)
    except PermissionError:
        temp_path = output_path.rsplit('.', 1)[0] + '_TEMP.xlsx'
        wb.save(temp_path)
        logger.warning(
            "文件被占用，已保存到临时文件: %s\n"
            "关闭 Excel/WPS 后，将 _TEMP 文件重命名替换原文件即可", temp_path
        )
        return temp_path
    logger.info(f"项目需求清单已生成: {output_path} ({len(seen_modules)} 模块)")

    return output_path
