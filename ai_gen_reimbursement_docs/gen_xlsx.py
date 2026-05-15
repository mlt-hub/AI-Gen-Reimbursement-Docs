"""生成 FPA工作量评估.xlsx 和 项目需求清单.xlsx"""

import logging
import os
import re
from copy import copy
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Font, Border

from ai_gen_reimbursement_docs.constants import (
    FPA_COL_SEQ, FPA_COL_SUBSYSTEM, FPA_COL_ASSET, FPA_COL_FUNC_POINT,
    FPA_COL_TYPE, FPA_COL_CLASSIFICATION, FPA_COL_EXPLANATION, FPA_COL_STATUS,
    FPA_COL_FORMULA_BASE, FPA_COL_ADJUST, FPA_COL_ELEMENTS, FPA_COL_FORMULA_WORKLOAD,
    FPA_TOTAL_COLS, FPA_COL_KEY_MAP,
    REQ_COL_SEQ, REQ_COL_PROJECT, REQ_COL_SUBSYSTEM, REQ_COL_L1, REQ_COL_L2,
    REQ_COL_L3, REQ_COL_PROC_TYPE, REQ_COL_WORKLOAD, REQ_COL_CFP, REQ_TOTAL_COLS,
    REQ_COL_KEY_MAP,
)
from ai_gen_reimbursement_docs.excel_source import replace_placeholders, strip_ai_marker
from ai_gen_reimbursement_docs.md_table import parse_md_table_row

logger = logging.getLogger('ai_gen_reimbursement_docs.gen_xlsx')


def _safe_load_workbook(path: str, label: str) -> openpyxl.Workbook:
    """安全加载 xlsx，失败时抛出可读的错误而非 InvalidFileException。"""
    try:
        return openpyxl.load_workbook(path)
    except FileNotFoundError:
        raise FileNotFoundError(f"「{label}」模板文件不存在: {path}")
    except Exception as e:
        raise ValueError(
            f"「{label}」模板无法打开，请检查文件是否为有效的 .xlsx 格式: {path}\n"
            f"内部错误: {e}"
        ) from e


# ============================================================
#  公用
# ============================================================

def _load_meta_md(meta_md_path: str) -> dict[str, str]:
    """解析文档元数据.md 为扁平字典。支持跨多行的表格值。"""
    from ai_gen_reimbursement_docs.gen_spec import _parse_meta_md
    return _parse_meta_md(meta_md_path)


def _load_module_rows(tree_md_path: str) -> list[dict[str, str]]:
    """解析功能清单模块树.md 为行字典列表。"""
    rows = []
    with open(tree_md_path, encoding='utf-8') as f:
        in_table = False
        for line in f:
            line = line.rstrip()
            if "| 入口 | 一级模块" in line:
                in_table = True
                continue
            if "|------" in line and in_table:
                continue
            if in_table:
                cells = parse_md_table_row(line, min_cols=9)
                if cells is not None:
                    rows.append({
                        "入口": cells[0],
                        "一级模块": cells[1],
                        "二级模块": cells[2],
                        "三级模块": cells[3],
                        "客户端类型": cells[4],
                        "三级模块整体功能描述": cells[5],
                        "功能过程": cells[6],
                        "功能过程类型": cells[7],
                        "功能过程描述": cells[8],
                    })
    return rows


def _receiver_from_client_type(client_type: str, rules_text: str) -> str:
    """根据客户端类型判定接收者。"""
    # 默认值
    default = "操作员"
    if not rules_text:
        logger.warning(
            "Excel 模板未配置「功能用户-接收者判定」，接收者将使用默认值，请在模板 Sheet 6 中补充"
        )
        return default

    # 解析规则文本：后台：后台管理员\n前台：普通用户\n...
    for line in rules_text.split('\n'):
        line = line.strip()
        if '：' in line:
            keyword, receiver = line.split('：', 1)
            if keyword in client_type:
                return receiver.strip()
    return default


# ============================================================
#  FPA工作量评估.xlsx
# ============================================================

def _call_llm(prompt: str, system_prompt: str, api_key: str, model: str,
              base_url: str, tag: str = "") -> str:
    """调用 LLM（委托至 llm_client 公共模块）。"""
    from ai_gen_reimbursement_docs.llm_client import call_llm

    try:
        return call_llm(
            prompt=prompt, system=system_prompt,
            api_key=api_key, model=model, base_url=base_url, tag=tag,
        )
    except Exception as e:
        logger.warning("LLM 调用失败 [%s]: %s", tag, e)
        return ""


def _build_fpa_rule_rows(rows: list[dict[str, str]], meta: dict[str, str]) -> list[dict[str, object]]:
    """从功能清单行构建 FPA 模板行（规则骨架）。"""
    prefix_rule = meta.get("新增/修改功能点前缀生成规则",
                           "【客户端类型】一级模块-二级模块-三级模块-功能过程")
    receiver_rules = meta.get("功能用户-接收者判定", "")
    subsystem = meta.get("子系统（模块）", "")
    asset = meta.get("资产标识", "")

    fpa_rows = []
    seq = 0

    for r in rows:
        client_type = r["客户端类型"]
        l1 = r["一级模块"]
        l2 = r["二级模块"]
        l3 = r["三级模块"]
        proc = r["功能过程"]
        proc_desc = r["功能过程描述"]
        proc_type = r["功能过程类型"]

        # 接收者
        receiver = _receiver_from_client_type(client_type, receiver_rules)

        # 功能点前缀（保留 【】 括号）
        fp_prefix = prefix_rule \
            .replace("【客户端类型】", f"【{client_type}】") \
            .replace("一级模块", l1) \
            .replace("二级模块", l2) \
            .replace("三级模块", l3) \
            .replace("功能过程", proc)

        # 界面行
        seq += 1
        fpa_rows.append({
            "序号": seq,
            "子系统(模块)": subsystem,
            "资产标识": asset,
            "新增/修改功能点": f"{fp_prefix}-界面开发",
            "类型": "EI",
            "计算依据归类": "",
            "计算依据说明": f"【{client_type}】{l1}-{l2}-{l3}-{proc}-界面开发，具体如下：\n1、",
            "变更状态": proc_type,
            "基准值": "",
            "调整值": 2,
            "要素数量": 1,
            "FPA工作量": "",
            "核减后工作量": "",
            "备注说明": "",
        })

        # 接口行
        seq += 1
        fpa_rows.append({
            "序号": seq,
            "子系统(模块)": subsystem,
            "资产标识": asset,
            "新增/修改功能点": f"{fp_prefix}-接口开发",
            "类型": "ILF",
            "计算依据归类": "",
            "计算依据说明": f"【{client_type}】{l1}-{l2}-{l3}-{proc}-接口开发，具体如下：\n1、",
            "变更状态": proc_type,
            "基准值": "",
            "调整值": 1,
            "要素数量": 1,
            "FPA工作量": "",
            "核减后工作量": "",
            "备注说明": "",
        })

    return fpa_rows


def _ai_fill_fpa(
    fpa_rows: list[dict[str, object]],
    judgement_rules: list[str],
    api_key: str,
    model: str,
    base_url: str,
) -> list[dict[str, object]]:
    """AI 填充 FPA 行的 F/G 列。"""
    if not api_key:
        logger.warning("未设置 API Key，跳过 AI 填充 FPA")
        return fpa_rows

    from ai_gen_reimbursement_docs.config_utils import load_ai_system_prompt
    system_prompt = load_ai_system_prompt("fpa_eval") or (
        "你是一个 FPA 功能点评估助手。根据功能过程描述和类型，"
        "从判定原则中选择最匹配的一项，并展开计算依据说明。"
        "直接输出结果，不要输出其他内容。"
    )

    total = len(fpa_rows)
    from ai_gen_reimbursement_docs.config_utils import load_flow_max_ai, load_gen_fpa_ai_limit
    _max_ai = load_flow_max_ai("gen_fpa")
    _proc_limit = load_gen_fpa_ai_limit()
    _seen_fpa_procs: list[str] = []  # 用 list 保持首次出现顺序，避免 set 无序
    _allowed_fpa_procs: set[str] | None = None
    if _proc_limit > 0:
        for _r in fpa_rows:
            _base = _r["新增/修改功能点"].rsplit('-', 1)[0]
            if _base not in _seen_fpa_procs:
                _seen_fpa_procs.append(_base)
        _allowed_fpa_procs = set(_seen_fpa_procs[:_proc_limit])
        _seen_fpa_procs.clear()
    _skip_ai_limit = 0
    _skip_proc_limit = 0
    _filled_count = 0
    for idx, row in enumerate(fpa_rows, 1):
        _base_name = row["新增/修改功能点"].rsplit('-', 1)[0]
        if _max_ai > 0 and idx > _max_ai:
            logger.info(f"  [{idx}/{total}] 跳过（超过 AI 限制 {_max_ai}）")
            _skip_ai_limit += 1
            continue
        if _allowed_fpa_procs is not None and _base_name not in _allowed_fpa_procs:
            logger.info(f"  [{idx}/{total}] 跳过（超过功能过程限制 {_proc_limit}）")
            _skip_proc_limit += 1
            continue
        if not row["计算依据说明"]:
            continue

        # 判定原则列表（已由调用方准备好）
        _rules_list = judgement_rules
        _numbered_rules = "\n".join(f"{i}) {r}" for i, r in enumerate(_rules_list, 1))

        row_tag = f"fpa_{row['类型']}_{row['新增/修改功能点']}"
        prompt = (
            f"功能过程描述：{row['计算依据说明']}\n\n"
            f"判定原则列表（请返回最匹配的序号，序号从1开始）：\n{_numbered_rules}\n\n"
            f"请直接输出JSON，不要输出其他内容：\n"
            f'{{"type":"EI/EO/EQ/ILF/EIF","classification_basis_index":1,"explanation":"<展开的说明，包含触发事件、事件流、业务规则、业务数据、涉及表/文件/接口>"}}'
        )

        _filled_count += 1
        logger.info(f"  FPA AI 填充 [{idx}/{total}] {row['新增/修改功能点'][:40]}...")
        resp = _call_llm(prompt, system_prompt, api_key, model, base_url, tag=row_tag)
        import json as _json
        if resp:
            try:
                from ai_gen_reimbursement_docs.llm_client import strip_markdown_code_block
                _clean = strip_markdown_code_block(resp)
                _data = _json.loads(_clean)
                if isinstance(_data, list):
                    _data = _data[0]
                if _data.get("type"):
                    row["类型"] = _data["type"].strip()
                # 优先按序号索引（兼容字符串序号），兜底按文本模糊匹配
                _basis = None
                _idx = _data.get("classification_basis_index")
                if _idx is not None:
                    try:
                        _idx = int(_idx)
                    except (ValueError, TypeError):
                        pass
                if isinstance(_idx, int) and _rules_list and 1 <= _idx <= len(_rules_list):
                    _basis = _rules_list[_idx - 1]
                elif _data.get("classification_basis"):
                    _val = _data["classification_basis"].strip()
                    for _rule in _rules_list:
                        if _rule and (_val in _rule or _rule in _val):
                            _basis = _rule
                            break
                    if _basis is None:
                        _basis = _val
                if _basis is not None:
                    row["计算依据归类"] = _basis
                else:
                    logger.warning(f"AI 响应中未找到归类依据，序号={_idx}，原始响应片段={resp[:200]}")
                if _data.get("explanation"):
                    exp = _data["explanation"].strip()
                    exp = exp.replace("具体如下", "具体如下" + chr(10))
                    exp = exp.replace("；", "；" + chr(10))
                    exp = exp.replace("事件流：", "事件流：" + chr(10))
                    row["计算依据说明"] = exp
            except Exception:
                pass

    # 汇总：全部跳过时提示配置限制
    if _filled_count == 0 and total > 0:
        _reasons = []
        if _max_ai > 0 and _skip_ai_limit > 0:
            _reasons.append(f"max_ai_l3_modules={_max_ai}（跳过 {_skip_ai_limit} 行）")
        if _proc_limit > 0 and _skip_proc_limit > 0:
            _reasons.append(f"gen_fpa_ai_limit={_proc_limit}（跳过 {_skip_proc_limit} 行）")
        if _reasons:
            logger.warning(
                "⚠ FPA AI 填充全部跳过（共 %d 行），请检查配置限制：%s。"
                "如需 AI 填充，请在 ~/.ai-gen-reimbursement-docs/system_config.yaml 中将对应值设为 0",
                total, "、".join(_reasons),
            )

    return fpa_rows


def init_fpa_template_md(
    tree_md_path: str,
    meta_md_path: str,
    output_md_path: str,
    summary_md_path: str = "",
) -> str:
    """生成 FPA 模板 MD（规则骨架，F/G 列留空待 AI 填充）。

    Args:
        summary_md_path: 非空时同步写入 gen-fpa-FPA工作量-总和.md（调整值×要素数量 的求和）
    """
    logger.info("生成 FPA 模板 MD...")
    meta = _load_meta_md(meta_md_path)
    rows = _load_module_rows(tree_md_path)
    fpa_rows = _build_fpa_rule_rows(rows, meta)

    total = 0.0
    with open(output_md_path, 'w', encoding='utf-8') as f:
        f.write("# FPA 工作量评估\n\n")
        f.write(f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        f.write("| 序号 | 子系统(模块) | 资产标识 | 新增/修改功能点 | 类型 | 计算依据归类 | 计算依据说明 | 变更状态 | 调整值 | 要素数量 |\n")
        f.write("|------|-------------|---------|----------------|------|-------------|-------------|---------|-------|---------|\n")
        for row in fpa_rows:
            vals = [
                str(row["序号"]),
                row["子系统(模块)"],
                row["资产标识"],
                row["新增/修改功能点"].replace('|', '\\|'),
                row["类型"],
                row["计算依据归类"],
                row["计算依据说明"].replace("|", chr(92) + "|").replace(chr(10), " "),
                row["变更状态"],
                str(row["调整值"]),
                str(row["要素数量"]),
            ]
            f.write("| " + " | ".join(vals) + " |\n")
            try:
                total += float(row["调整值"]) * float(row["要素数量"])
            except (ValueError, TypeError):
                pass

    logger.info(f"FPA 模板 MD 已生成: {output_md_path} ({len(fpa_rows)} 行)")

    if summary_md_path:
        os.makedirs(os.path.dirname(summary_md_path) or '.', exist_ok=True)
        with open(summary_md_path, 'w', encoding='utf-8') as f:
            f.write("# FPA 工作量\n\n")
            f.write(f"FPA工作量（人/天）: {total}\n")
        logger.info(f"FPA 工作量已写入: {summary_md_path} ({total})")

    return output_md_path


def ai_fill_fpa_md(
    fpa_md_path: str,
    meta_md_path: str,
    template_path: str = "",
    api_key: str = "",
    model: str = "",
    base_url: str = "",
) -> str:
    """读取 FPA 模板 MD，AI 填充 F/G 列，写回 MD。"""
    logger.info("AI 填充 FPA 数据...")
    logger.info(f"AI 模型: {model}  端点: {base_url or '默认'}  API Key: {'已设置' if api_key else '未设置'}")

    meta = _load_meta_md(meta_md_path)
    # 优先从模板附录 sheet 读取判定原则，兜底用元数据
    judgement_rules: list[str] = []
    if template_path:
        try:
            wb = openpyxl.load_workbook(template_path)
            ws = wb['附录1-FPA评估方法说明']
            for row_num in range(2, 15):
                val = ws.cell(row_num, 3).value
                if val and str(val).strip():
                    judgement_rules.append(str(val).strip())
            wb.close()
            if judgement_rules:
                logger.info(f"从模板附录读取判定原则 {len(judgement_rules)} 条")
        except Exception as e:
            logger.warning(f"从模板附录读取判定原则失败: {e}")
    if not judgement_rules:
        meta_rules = meta.get("计算依据归类判定原则", "")
        if meta_rules:
            judgement_rules = [r.strip() for r in meta_rules.split('\n') if r.strip()]
            logger.info(f"从元数据读取判定原则 {len(judgement_rules)} 条")
    if not judgement_rules:
        logger.warning("未配置「计算依据归类判定原则」，AI 输出的归类将原样保留")

    # 解析 MD 表格
    fpa_rows = []
    with open(fpa_md_path, encoding='utf-8') as f:
        in_table = False
        for line in f:
            line = line.rstrip()
            if "| 序号 | 子系统" in line:
                in_table = True
                continue
            if "|------|" in line and in_table:
                continue
            if in_table:
                cells = parse_md_table_row(line, min_cols=10)
                if cells is not None:
                    fpa_rows.append({
                        "序号": cells[0],
                        "子系统(模块)": cells[1],
                        "资产标识": cells[2],
                        "新增/修改功能点": cells[3],
                        "类型": cells[4],
                        "计算依据归类": cells[5],
                        "计算依据说明": cells[6],
                        "变更状态": cells[7],
                        "调整值": cells[8],
                        "要素数量": cells[9],
                    })

    # AI 填充
    fpa_rows = _ai_fill_fpa(fpa_rows, judgement_rules, api_key, model, base_url)

    # 写回 MD
    with open(fpa_md_path, 'w', encoding='utf-8') as f:
        f.write("# FPA 工作量评估\n\n")
        f.write(f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write(f"**AI 填充**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
        f.write("| 序号 | 子系统(模块) | 资产标识 | 新增/修改功能点 | 类型 | 计算依据归类 | 计算依据说明 | 变更状态 | 调整值 | 要素数量 |\n")
        f.write("|------|-------------|---------|----------------|------|-------------|-------------|---------|-------|---------|\n")
        for row in fpa_rows:
            vals = [
                str(row["序号"]),
                row["子系统(模块)"],
                row["资产标识"],
                row["新增/修改功能点"].replace('|', '\\|'),
                row["类型"],
                row["计算依据归类"],
                row["计算依据说明"].replace('|', '\\|').replace(chr(10), ' '),
                row["变更状态"],
                str(row["调整值"]),
                str(row["要素数量"]),
            ]
            f.write("| " + " | ".join(vals) + " |\n")

    logger.info(f"FPA MD 已填充: {fpa_md_path}")
    return fpa_md_path



def _format_fpa_explanation(text: str) -> str:
    """格式化 FPA 计算依据说明：加换行排版，不改变原文内容。"""
    import re as _re
    NL = chr(10)
    # 去掉开头多余的冒号/空格
    text = text.lstrip("：: ")
    # 具体如下后空一行；触发事件、事件流前换行但冒号后不换行
    text = text.replace("具体如下", "具体如下" + NL + NL)
    text = text.replace("事件流：", NL + "事件流：")
    text = text.replace("触发事件：", NL + "触发事件：")
    # 分号换行
    text = text.replace("；", "；" + NL)
    text = text.replace("业务规则", NL + "业务规则")
    text = text.replace("业务数据", NL + "业务数据")
    text = text.replace("涉及表", NL + "涉及表")
    text = text.replace("涉及服务", NL + "涉及服务")
    text = text.replace("；涉及接口", "；" + NL + "涉及接口")
    text = _re.sub(_re.compile(r"(?<=\S)\s+(?=\d+\.)"), NL, text)
    text = _re.sub(_re.compile(r"^[	 ]+", _re.MULTILINE), "", text)
    # 去掉只有分号或冒号的行
    text = _re.sub(_re.compile(r'\n[：:;；]\s*\n'), '\n', text)
    text = _re.sub(_re.compile(NL + "{3,}"), NL + NL, text)
    return text

def generate_fpa_xlsx_from_md(
    fpa_md_path: str,
    meta_md_path: str,
    template_path: str,
    output_path: str,
) -> str:
    """从已填充的 FPA MD 生成 FPA工作量评估.xlsx。"""
    logger.info("从 FPA MD 生成 Excel...")

    meta = _load_meta_md(meta_md_path)
    base_formula = meta.get("基准值公式", "")
    workload_formula = meta.get("FPA工作量公式", "J{row}*K{row}")

    # 从 MD 解析 FPA 行
    fpa_rows = []
    with open(fpa_md_path, encoding='utf-8') as f:
        in_table = False
        for line in f:
            line = line.rstrip()
            if "| 序号 | 子系统" in line:
                in_table = True
                continue
            if "|------|" in line and in_table:
                continue
            if in_table:
                cells = parse_md_table_row(line, min_cols=10)
                if cells is not None:
                    fpa_rows.append({
                        "序号": cells[0],
                        "子系统(模块)": cells[1],
                        "资产标识": cells[2],
                        "新增/修改功能点": cells[3],
                        "类型": cells[4],
                        "计算依据归类": cells[5],
                        "计算依据说明": cells[6],
                        "变更状态": cells[7],
                        "调整值": cells[8],
                        "要素数量": cells[9],
                    })

    # 填充模板
    wb = _safe_load_workbook(template_path, 'FPA工作量评估')
    ws = wb['FPA功能点估算']

        # 保存模板第3行的格式作为参照（I 和 L 列沿用第2行标题样式）
    tmpl_format = {}
    for col_idx in range(1, FPA_TOTAL_COLS):
        c = ws.cell(3, col_idx)
        tmpl_format[col_idx] = {
            'font': c.font.copy() if c.font else None,
            'fill': c.fill.copy() if c.fill else None,
            'border': c.border.copy() if c.border else None,
            'number_format': c.number_format,
            'alignment': c.alignment.copy() if c.alignment else None,
        }
    # I/L 列：沿用第2行标题单元格的背景色
    for col_idx in (FPA_COL_FORMULA_BASE, FPA_COL_FORMULA_WORKLOAD):
        c = ws.cell(2, col_idx)
        if c.fill:
            tmpl_format[col_idx]['fill'] = c.fill.copy()

    # 删除旧数据行（从第3行起），彻底清除旧数据及格式
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)

    # 写入数据
    for i, fpa_row in enumerate(fpa_rows):
        excel_row = i + 3
        # 通用列：从 dict 取值
        for col_idx, key in FPA_COL_KEY_MAP.items():
            val = fpa_row.get(key, "")
            cell = ws.cell(excel_row, col_idx)
            # int 转换列
            if col_idx in (FPA_COL_SEQ, FPA_COL_ADJUST, FPA_COL_ELEMENTS):
                try:
                    cell.value = int(val)
                except (ValueError, TypeError):
                    cell.value = val
            elif col_idx == FPA_COL_EXPLANATION:
                cell.value = _format_fpa_explanation(val)
            else:
                cell.value = val
        # 公式列
        if base_formula:
            formula = base_formula.replace("E3", f"E{excel_row}")                 .replace("H3", f"H{excel_row}").replace("I3", f"I{excel_row}")                 .replace("J3", f"J{excel_row}").replace("K3", f"K{excel_row}")
            ws.cell(excel_row, FPA_COL_FORMULA_BASE).value = f"={formula}" if not formula.startswith('=') else formula
        if workload_formula:
            formula = workload_formula.replace("J{row}", f"J{excel_row}")                 .replace("K{row}", f"K{excel_row}")                 .replace("J3", f"J{excel_row}").replace("K3", f"K{excel_row}")
            ws.cell(excel_row, FPA_COL_FORMULA_WORKLOAD).value = f"={formula}" if not formula.startswith('=') else formula
        ws.cell(excel_row, FPA_TOTAL_COLS - 1, "")
        ws.cell(excel_row, FPA_TOTAL_COLS, "")

        # 从模板第3行复制格式（跳过 fill，避免空单元格带底色；I/L 列除外）
        for col_idx in range(1, FPA_TOTAL_COLS):
            c = ws.cell(excel_row, col_idx)
            fmt = tmpl_format.get(col_idx, {})
            if fmt.get('font'):
                c.font = fmt['font']
            if fmt.get('border'):
                c.border = fmt['border']
            if fmt.get('number_format'):
                c.number_format = fmt['number_format']
            # I 列(9) 和 L 列(12)：应用第2行标题的背景色
            if col_idx in (9, 12) and fmt.get('fill'):
                c.fill = fmt['fill']
            if col_idx in (FPA_COL_FUNC_POINT, FPA_COL_EXPLANATION):
                orig_align = fmt.get('alignment')
                h = 'left' if col_idx == 7 else (orig_align.horizontal or 'center')
                if orig_align:
                    c.alignment = Alignment(
                        wrap_text=True,
                        vertical='center',
                        horizontal=h,
                    )
                else:
                    c.alignment = Alignment(wrap_text=True, vertical='center', horizontal=h)
            else:
                if fmt.get('alignment'):
                    c.alignment = fmt['alignment']

    # 更新第1行合计公式
    last_data_row = len(fpa_rows) + 2
    for col_idx in [FPA_COL_FORMULA_BASE, FPA_COL_ADJUST, FPA_COL_ELEMENTS, FPA_COL_FORMULA_WORKLOAD, FPA_TOTAL_COLS - 1]:
        cell = ws.cell(1, col_idx)
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        cell.value = f"=SUM({col_letter}3:{col_letter}{last_data_row})"
    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    try:
        wb.save(output_path)
    except PermissionError:
        logger.error(
            "无法写入 %s —— 文件可能被 Excel/WPS 占用，请关闭后重试", output_path
        )
        raise
    logger.info(f"FPA工作量评估已生成: {output_path} ({len(fpa_rows)} 行)")

    return output_path


# ============================================================
#  项目需求清单.xlsx
# ============================================================

def generate_list_xlsx_from_md(
    meta_md_path: str,
    tree_md_path: str,
    template_path: str,
    output_path: str,
    cfp_total: float = 0,
    fpa_reduced: float = 0,
) -> str:
    """生成项目需求清单.xlsx。

    cfp_total: 送审功能点 = gen-fpa-FPA工作量-总和.md 的原始值
    fpa_reduced: 送审工作量 = FPA 核减后的工作量
    """
    logger.info("开始生成项目需求清单.xlsx...")

    meta = _load_meta_md(meta_md_path)
    rows = _load_module_rows(tree_md_path)

    wb = _safe_load_workbook(template_path, '项目需求清单')

    # ====== Sheet 1: 项目信息概览 ======
    ws1 = wb['项目信息概览']

    # 替换标题
    title = meta.get("项目信息概览-标题", "")
    ws1.cell(1, 1, title)

    # 替换数据行（第3行，第2行为表头）
    ws1.cell(3, 2, meta.get("项目信息概览-项目名称", ""))
    ws1.cell(3, 3, meta.get("项目信息概览-子系统名称", ""))
    ws1.cell(3, 4, meta.get("项目信息概览-项目类型", ""))
    ws1.cell(3, 5, meta.get("项目信息概览-所属域", ""))
    ws1.cell(3, 6, meta.get("项目信息概览-所属系统", ""))
    ws1.cell(3, 7, meta.get("项目信息概览-需求部门", ""))
    ws1.cell(3, 8, meta.get("项目信息概览-需求负责人", ""))
    ws1.cell(3, 9, meta.get("项目信息概览-需求负责人联系方式", ""))

    # 送审工作量 = FPA 核减后的工作量
    if fpa_reduced > 0:
        ws1.cell(3, 10, fpa_reduced)

    # 送审功能点 = CFP 总量
    if cfp_total > 0:
        ws1.cell(3, 11, cfp_total)

    # ====== Sheet 2: 功能清单 ======
    ws2 = wb['功能清单']

    # 替换标题
    fl_title = meta.get("功能清单-标题", "")
    ws2.cell(1, 1, fl_title)

    # 取消所有合并单元格，避免逐行写入时遇到 MergedCell
    for merge_range in list(ws2.merged_cells.ranges):
        ws2.unmerge_cells(str(merge_range))

    # 删除旧数据行（从第3行起），彻底清除旧数据及格式
    if ws2.max_row >= 3:
        ws2.delete_rows(3, ws2.max_row - 2)

    project_name = meta.get("功能清单-项目名称", "")
    subsystem = meta.get("功能清单-子系统", "")

    # 按三级模块去重写入（先存为列表，之后合并单元格）
    data_rows_data = []
    seen_modules = set()
    seq = 0

    _center = Alignment(horizontal='center', vertical='center')
    _center_wrap = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 保存模板第2行（表头）的边框样式
    import copy as _cpy
    _tmpl_border = _cpy.copy(ws2.cell(2, 1).border)

    for r in rows:
        key = (r["一级模块"], r["二级模块"], r["三级模块"])
        if key not in seen_modules:
            seen_modules.add(key)
            seq += 1
            row_idx = seq + 2
            # 构建行数据 dict，统一通过 REQ_COL_KEY_MAP 写入
            _req_data = {
                "序号": seq, "项目名称": project_name, "子系统": subsystem,
                "一级模块": r["一级模块"], "二级模块": r["二级模块"],
                "三级模块": r["三级模块"], "功能过程类型": r["功能过程类型"],
            }
            for col_idx in range(1, REQ_TOTAL_COLS):
                c = ws2.cell(row_idx, col_idx)
                c.alignment = _center
                c.border = _tmpl_border
            for col_idx, key in REQ_COL_KEY_MAP.items():
                ws2.cell(row_idx, col_idx, _req_data.get(key, ""))
            ws2.cell(row_idx, REQ_COL_PROJECT).alignment = _center_wrap
            if fpa_reduced > 0:
                ws2.cell(row_idx, REQ_COL_WORKLOAD, fpa_reduced)
            if cfp_total > 0:
                ws2.cell(row_idx, REQ_COL_CFP, cfp_total)
            data_rows_data.append({
                "row": row_idx,
                "project_name": project_name,
                "subsystem": subsystem,
                "module_l1": r["一级模块"],
                "module_l2": r["二级模块"],
            })

    # 合并行1标题居中（A~I列）
    if seq > 0:
        ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=REQ_TOTAL_COLS)
        ws2.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 合并相同值的列：B(项目名称), C(子系统), D(一级模块), E(二级模块)
    for col_idx in [REQ_COL_PROJECT, REQ_COL_SUBSYSTEM, REQ_COL_L1, REQ_COL_L2]:
        i = 0
        while i < len(data_rows_data):
            val_key = {REQ_COL_PROJECT: "project_name", REQ_COL_SUBSYSTEM: "subsystem",
                       REQ_COL_L1: "module_l1", REQ_COL_L2: "module_l2"}[col_idx]
            curr_val = data_rows_data[i][val_key]
            j = i
            while j < len(data_rows_data) and data_rows_data[j][val_key] == curr_val:
                j += 1
            count = j - i
            if count > 1:
                ws2.merge_cells(
                    start_row=data_rows_data[i]["row"],
                    start_column=col_idx,
                    end_row=data_rows_data[j - 1]["row"],
                    end_column=col_idx
                )
                ws2.cell(data_rows_data[i]["row"], col_idx).border = _tmpl_border
            i = j

    # 合并送审工作量和送审功能点列 — 所有行相同值
    if len(data_rows_data) > 1:
        for _col in (REQ_COL_WORKLOAD, REQ_COL_CFP):
            ws2.merge_cells(
                start_row=data_rows_data[0]["row"],
                start_column=_col,
                end_row=data_rows_data[-1]["row"],
                end_column=_col
            )
            _top_cell = ws2.cell(data_rows_data[0]["row"], _col)
            _top_cell.border = _tmpl_border
            _top_cell.alignment = _center

        # 合并区域补全各单元格边框（合并后非左上角单元格边框丢失）
        for _col in (REQ_COL_WORKLOAD, REQ_COL_CFP):
            for _r in range(data_rows_data[0]["row"], data_rows_data[-1]["row"] + 1):
                ws2.cell(_r, _col).border = _tmpl_border

    os.makedirs(os.path.dirname(output_path) or '.', exist_ok=True)
    try:
        wb.save(output_path)
    except PermissionError:
        logger.error(
            "无法写入 %s —— 文件可能被 Excel/WPS 占用，请关闭后重试", output_path
        )
        raise
    logger.info(f"项目需求清单已生成: {output_path} ({len(seen_modules)} 模块)")

    return output_path
