"""AI生成项目报账文档 - CLI入口"""

import argparse
import json
import logging
import os
from pathlib import Path as _Path
import re
import shutil
import sys
import uuid

from ai_gen_reimbursement_docs.config_utils import load_api_key, load_base_url, load_model_name
from ai_gen_reimbursement_docs.run_history import (
    get_run,
    list_runs,
    now_iso,
    upsert_run,
    user_history_path,
)

logger = logging.getLogger('ai_gen_reimbursement_docs')


def _new_cli_run_id() -> str:
    from datetime import datetime

    return f"{datetime.now():%Y%m%d_%H%M%S}_{uuid.uuid4().hex[:6]}"


def _summary_files_from_result(result) -> list[tuple[str, str]]:
    return [
        ("FPA 工作量评估", getattr(result, "fpa_xlsx", "")),
        ("项目功能点拆分表", getattr(result, "cosmic_xlsx", "")),
        ("项目需求清单", getattr(result, "require_xlsx", "")),
        ("项目需求说明书", getattr(result, "spec_docx", "")),
    ]


def _done_files_from_result(result) -> list[dict]:
    files = []
    for label, path in _summary_files_from_result(result):
        if path and os.path.exists(path):
            files.append(
                {
                    "label": label,
                    "name": os.path.basename(path),
                    "path": path,
                    "size_kb": round(os.path.getsize(path) / 1024),
                    "is_temp": "_TEMP" in os.path.basename(path),
                }
            )
    return files


def _infer_output_dir(result, fallback: str = "") -> str:
    for _, path in _summary_files_from_result(result):
        if path:
            return os.path.dirname(os.path.abspath(path))
    return fallback


def _record_cli_history(
    *,
    run_id: str,
    task_mode: str,
    run_state: str,
    input_path: str = "",
    output_dir: str = "",
    done_files: list[dict] | None = None,
    error: str = "",
    created_at: str | None = None,
    started_at: str | None = None,
) -> None:
    now = now_iso()
    try:
        upsert_run(
            {
                "run_id": run_id,
                "source": "cli",
                "session_id": "",
                "mode": "local",
                "task_mode": task_mode,
                "run_state": run_state,
                "input_name": os.path.basename(input_path) if input_path else "",
                "input_path": input_path,
                "output_dir": output_dir,
                "artifact_kind": "local_dir",
                "done_files": done_files or [],
                "error": error,
                "created_at": created_at or now,
                "started_at": started_at or now,
                "finished_at": now if run_state in {"done", "error", "cancelled"} else "",
                "updated_at": now,
            },
            user_history_path(),
        )
    except Exception as exc:
        logger.warning("运行历史写入失败: %s", exc)


def _print_history(*, limit: int, as_json: bool) -> None:
    try:
        items = list_runs(user_history_path(), limit=limit, offset=0)
    except Exception as exc:
        logger.error("运行历史读取失败: %s", exc)
        return
    if as_json:
        print(json.dumps(items, ensure_ascii=False, indent=2))
        return
    if not items:
        print("暂无运行历史")
        return
    for item in items:
        created_at = item.get("created_at", "")
        state = item.get("run_state", "")
        task_mode = item.get("task_mode", "")
        input_name = item.get("input_name", "")
        output_dir = item.get("output_dir", "")
        run_id = item.get("run_id", "")
        availability = "目录可用" if item.get("open_folder_available") else "目录不存在"
        print(f"{created_at}  {state:9}  {task_mode:12}  {input_name}")
        print(f"  run_id: {run_id}")
        print(f"  输出目录: {output_dir or '-'} ({availability})")


def _get_version() -> str:
    """Read version from pyproject.toml."""
    import tomllib
    try:
        root = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
        toml_path = os.path.join(root, 'pyproject.toml')
        with open(toml_path, 'rb') as f:
            return tomllib.load(f)['project']['version']
    except Exception:
        return "unknown"


def _start_web_ui(root: str, port: int = 0) -> None:
    """启动 Web UI 服务器并打开浏览器。port 为 0 时从配置文件读取。"""
    try:
        import uvicorn
        import webbrowser
    except ImportError:
        print("Web UI 需要安装 uvicorn: pip install uvicorn[standard]")
        return

    from ai_gen_reimbursement_docs.config_utils import load_web_port
    host = "127.0.0.1"
    if port <= 0:
        port = load_web_port()

    import socket
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        sock.bind((host, port))
        sock.close()
    except OSError:
        sock.close()
        print(f"端口 {port} 被占用或无权访问，启动失败")
        print("解决方法：")
        print("  1. 换个端口：ard --web --port 9090")
        print("  2. 修改配置：~/.ai-gen-reimbursement-docs/system_config.yaml → web_port")
        print("  3. 关闭占用端口的进程后重试")
        return

    webbrowser.open(f"http://{host}:{port}")
    print(f"Web UI 已启动: http://{host}:{port}")
    # exe 模式用 exe 所在目录（web_app/ 外挂在该目录），源码模式用项目根
    if getattr(sys, 'frozen', False):
        app_dir = os.path.dirname(sys.executable)
    else:
        app_dir = root
    uvicorn.run("web_app.server:app", host=host, port=port,
                app_dir=app_dir, log_level="info",
                timeout_graceful_shutdown=2)


def _auto_detect_and_run(api_key: str, model: str, base_url: str,
                         search_dir: str = "") -> None:
    """零参数模式：在指定目录（或当前目录）搜索功能清单 xlsx，唯一匹配则自动全流程执行。"""
    import glob

    from ai_gen_reimbursement_docs.excel_source import is_valid_input_xlsx
    from ai_gen_reimbursement_docs.pipeline import _try_read_project_name

    _base = search_dir if search_dir else "."
    _label = _base if search_dir else "当前目录"
    xlsx_files = glob.glob(os.path.join(_base, "*.xlsx"))
    if not xlsx_files:
        print(f"{_label}未找到任何 .xlsx 文件")
        print("使用方式: ard --from-excel <功能清单路径> --gen-all")
        return

    valid = [f for f in xlsx_files if is_valid_input_xlsx(f)]

    if len(valid) == 0:
        print(f"{_label}找到 {len(xlsx_files)} 个 .xlsx 文件，但都不符合功能清单录入文档规范")
        print(f"文件列表: {', '.join(xlsx_files)}")
        print("使用方式: ard --from-excel <功能清单路径> --gen-all")
        return

    if len(valid) > 1:
        print(f"{_label}找到 {len(valid)} 个符合规范的功能清单文件，请指定其中一个:")
        for f in valid:
            print(f"  ard --from-excel \"{f}\" --gen-all")
        return

    excel_path = valid[0]
    print(f"检测到功能清单: {excel_path}")

    project_name = _try_read_project_name(excel_path)
    if project_name:
        print(f"项目名称: {project_name}")

    print("自动执行全流程...")
    _run_pipeline_with_args(excel_path, api_key, model, base_url, project_name)


def _run_pipeline_with_args(
    excel_path: str,
    api_key: str = "",
    model: str = "",
    base_url: str = "",
    project_name: str = "",
) -> None:
    """执行管道并输出摘要（零参数模式复用）。"""
    from ai_gen_reimbursement_docs.cli.logging import build_cli_callbacks, setup_logging
    from ai_gen_reimbursement_docs.cli.notify import play_notify_sound
    from ai_gen_reimbursement_docs.exceptions import CosmicToolError
    from ai_gen_reimbursement_docs.pipeline import run_pipeline_simple

    run_id = _new_cli_run_id()
    started_at = now_iso()
    _record_cli_history(
        run_id=run_id,
        task_mode="gen-all",
        run_state="running",
        input_path=excel_path,
        created_at=started_at,
        started_at=started_at,
    )
    try:
        result = run_pipeline_simple(
            mode='gen-all',
            file_path=excel_path,
            api_key=api_key,
            model=model,
            base_url=base_url,
            project_name=project_name,
            callbacks=build_cli_callbacks(),
        )
    except KeyboardInterrupt:
        _record_cli_history(
            run_id=run_id,
            task_mode="gen-all",
            run_state="cancelled",
            input_path=excel_path,
            error="cancelled",
            created_at=started_at,
            started_at=started_at,
        )
        print("\n  任务已取消", file=sys.stderr)
        sys.exit(1)
    except CosmicToolError as e:
        _record_cli_history(
            run_id=run_id,
            task_mode="gen-all",
            run_state="error",
            input_path=excel_path,
            error=str(e),
            created_at=started_at,
            started_at=started_at,
        )
        print(f"\n  错误: {e}", file=sys.stderr)
        play_notify_sound()
        sys.exit(1)
    except Exception as e:
        _record_cli_history(
            run_id=run_id,
            task_mode="gen-all",
            run_state="error",
            input_path=excel_path,
            error=str(e),
            created_at=started_at,
            started_at=started_at,
        )
        raise

    _summary_files = _summary_files_from_result(result)
    _record_cli_history(
        run_id=run_id,
        task_mode="gen-all",
        run_state="done",
        input_path=excel_path,
        output_dir=_infer_output_dir(result),
        done_files=_done_files_from_result(result),
        created_at=started_at,
        started_at=started_at,
    )
    print()
    for _label, _path in _summary_files:
        if _path and os.path.exists(_path):
            _size = os.path.getsize(_path)
            _uri = _Path(_path).as_uri()
            print(f"  ✅ {_label}: {_uri} ({_size/1024:.0f} KB)")
        else:
            print(f"  ⏭️  {_label}: 跳过（已存在或未生成）")
    print()

    from ai_gen_reimbursement_docs.cli.logging import write_combined_ai_log
    from ai_gen_reimbursement_docs.cli.notify import play_notify_sound
    write_combined_ai_log('gen-all')
    play_notify_sound()


def _auto_init_config(root: str) -> None:
    """exe 首次运行自动初始化用户配置文件（不覆盖已有配置）。"""
    from ai_gen_reimbursement_docs.config_utils import copy_default_config_files

    home_cfg = _Path(os.path.expanduser('~')) / '.ai-gen-reimbursement-docs'
    cfg_dir = _Path(root) / 'config'
    if not cfg_dir.is_dir():
        return
    for dst in copy_default_config_files(home_cfg, cfg_dir):
        print(f"已自动创建配置文件: {dst}")


def _read_l3_descriptions_from_excel(excel_path: str) -> list[str]:
    """从 Excel 功能清单中读取三级模块整体功能描述（去重，处理合并单元格继承）。"""
    from ai_gen_reimbursement_docs.config_utils import load_sheet_names
    import openpyxl

    _s = load_sheet_names()
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[_s["func_list"]]

    descriptions: list[str] = []
    seen: set[str] = set()
    prev_desc = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        desc = str(row[5]).strip() if len(row) > 5 and row[5] else ""
        if not desc:
            desc = prev_desc  # 合并单元格继承上一行
        else:
            prev_desc = desc
        if desc and desc not in seen:
            seen.add(desc)
            descriptions.append(desc)
    wb.close()
    return descriptions


def _read_meta_field_value(excel_path: str, field_key: str) -> tuple[str, str]:
    """在所有元数据 sheet 中搜索指定字段的值，返回 (sheet名, 值)。"""
    from ai_gen_reimbursement_docs.config_utils import load_sheet_names
    import openpyxl

    _s = load_sheet_names()
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    for sheet_key in ["meta", "fpa_meta", "spec_meta", "cosmic_meta", "list_meta"]:
        sheet_name = _s.get(sheet_key, "")
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            key = str(row[0]).strip() if row[0] else ""
            val = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if key == field_key:
                wb.close()
                return sheet_name, val
    wb.close()
    return "", ""


def _read_meta_all_keys(excel_path: str) -> list[tuple[str, str]]:
    """列出所有元数据 sheet 的字段 key，返回 [(sheet名, key), ...]。"""
    from ai_gen_reimbursement_docs.config_utils import load_sheet_names
    import openpyxl

    _s = load_sheet_names()
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    keys: list[tuple[str, str]] = []
    for sheet_key in ["meta", "fpa_meta", "spec_meta", "cosmic_meta", "list_meta"]:
        sheet_name = _s.get(sheet_key, "")
        if not sheet_name or sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            key = str(row[0]).strip() if row[0] else ""
            if key:
                keys.append((sheet_name, key))
    wb.close()
    return keys


def _read_l3_names_from_excel(excel_path: str) -> list[str]:
    """从 Excel 功能清单中读取三级模块名称（去重，处理合并单元格继承）。"""
    from ai_gen_reimbursement_docs.config_utils import load_sheet_names
    import openpyxl

    _s = load_sheet_names()
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb[_s["func_list"]]

    names: list[str] = []
    seen: set[str] = set()
    prev_name = ""
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = str(row[3]).strip() if len(row) > 3 and row[3] else ""
        if not name:
            name = prev_name
        else:
            prev_name = name
        if name and name not in seen:
            seen.add(name)
            names.append(name)
    wb.close()
    return names


def _build_parser() -> argparse.ArgumentParser:
    """构建 CLI 参数解析器。"""
    parser = argparse.ArgumentParser(
        description="AI生成项目报账文档 — 从功能清单自动生成全套报账交付物",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""  示例:

  ard --init-config                 初始化配置
  ard --from-dir ./某项目/ --gen-all    指定目录全流程
  ard --from-excel 功能清单.xlsx --gen-all      全流程
  ard --from-excel 功能清单.xlsx --gen-fpa      仅 FPA""",
    )

    parser.add_argument('--api-key', '-k', default='',
                        help='API Key（默认从 .env 读取）')
    parser.add_argument('--model', '-m', default='',
                        help='模型名称（默认 deepseek-v4-flash[1m]）')
    parser.add_argument('--from-excel', default='',
                        help='功能清单.xlsx 路径')
    parser.add_argument('--from-dir', default='',
                        help='项目目录（含功能清单.xlsx），自动搜索并以此为输出根目录')
    parser.add_argument('--gen-fpa', action='store_true',
                        help='生成 FPA工作量评估.xlsx')
    parser.add_argument('--gen-spec', action='store_true',
                        help='生成 项目需求说明书.docx')
    parser.add_argument('--gen-cosmic', action='store_true',
                        help='生成 项目功能点拆分表.xlsx')
    parser.add_argument('--gen-list', action='store_true',
                        help='生成 项目需求清单.xlsx')
    parser.add_argument('--gen-basedata', action='store_true',
                        help='生成 基础数据：模块树.md + 元数据.md')
    parser.add_argument('--gen-all', action='store_true',
                        help='全流程自动执行，全套报账文档')
    parser.add_argument('--output-dir', default='',
                        help='交付物输出目录')
    parser.add_argument('--project-name', default='',
                        help='输出文件夹名称')
    parser.add_argument('--fpa-out-template', default='',
                        help='FPA 输出模板路径')
    parser.add_argument('--fpa-profile', default='',
                        help='FPA 规划口径（默认读取 system_config.yaml）')
    parser.add_argument('--fpa-strategy', default='',
                        help='FPA 执行策略：rules_first / ai_first / rules_only / ai_only（默认跟随 profile）')
    parser.add_argument('--fpa-rule-set', default='',
                        help='FPA 规则集名称（默认跟随 profile）')
    parser.add_argument('--fpa-stability-report', nargs='+', default=[],
                        help='读取一个或多个 fpa_audit_trace.json，输出 FPA 稳定性对比报告')
    parser.add_argument('--fpa-stability-output', default='',
                        help='FPA 稳定性对比报告 Markdown 输出路径')
    parser.add_argument('--preview-fpa-module', default='',
                        help='只预览指定三级模块的 FPA 拆分结果，不生成 Excel')
    parser.add_argument('--preview-fpa-module-index', type=int, default=None,
                        help='按序号预览三级模块 FPA 拆分结果，用于同名三级模块')
    parser.add_argument('--preview-fpa-json', action='store_true',
                        help='FPA 预览以 JSON 输出')
    parser.add_argument('--use-preview-cache', action='store_true',
                        help='FPA 预览复用已生成的 fpa-preview-md，不重新解析 Excel')
    parser.add_argument('--keep-preview-files', action='store_true',
                        help='保留 FPA 预览生成的中间 MD 文件，便于调试')
    parser.add_argument('--cosmic-out-template', default='',
                        help='COSMIC 输出模板路径')
    parser.add_argument('--list-out-template', default='',
                        help='需求清单 输出模板路径')
    parser.add_argument('--spec-out-template', default='',
                        help='需求说明书 输出模板路径')
    parser.add_argument('--clean', action='store_true',
                        help='删除旧输出再重新生成')
    parser.add_argument('--init-config', action='store_true',
                        help='初始化配置文件')
    parser.add_argument('--log', nargs='?', const='tail', default=None,
                        help='查看日志')
    parser.add_argument('--history', action='store_true',
                        help='查看 CLI 运行历史')
    parser.add_argument('--history-limit', '--limit', type=int, default=20,
                        help='查看历史时返回的记录数')
    parser.add_argument('--history-json', '--json', action='store_true',
                        help='以 JSON 格式输出运行历史')
    parser.add_argument('--history-open', default='',
                        help='打开指定 run_id 的本机输出目录')
    parser.add_argument('--version', '-v', action='store_true',
                        help='显示版本号')
    parser.add_argument('--test-sound', action='store_true',
                        help='测试提示音')
    parser.add_argument('--log-level', type=str, default='',
                        help='控制台日志级别（DEBUG/INFO/WARNING/ERROR），默认读取配置')
    parser.add_argument('--max-tokens', type=str, default='',
                        help='覆盖 AI max_tokens')
    parser.add_argument('--test-ai-gen-reliability-desc', action='store_true',
                        help='测试"调整因子中的可靠性描述"AI生成（仅控制台+日志输出）')
    parser.add_argument('--test-ai-gen-metadata', type=str, default='',
                        help='测试元数据中指定字段的#AI生成#（仅控制台+日志输出）')
    parser.add_argument('--port', '-p', type=int, default=0,
                        help='Web UI 端口号（默认读取配置文件 web_port 或 8088）')
    parser.add_argument('--web', action='store_true',
                        help='启动 Web UI 界面')
    parser.add_argument('--activate', action='store_true',
                        help='使用 license 文件激活受保护数据')
    parser.add_argument('--license', default='',
                        help='license.ard.json 路径')
    parser.add_argument('--license-secret', default='',
                        help='license secret/token')
    parser.add_argument('--data-enc', default='',
                        help='data.enc 路径（默认使用程序目录下 data.enc）')
    parser.add_argument('--data-output', default='',
                        help='解密后的 data 输出目录（默认使用程序目录下 data）')
    parser.add_argument('--public-key', default='',
                        help='Ed25519 公钥路径（默认使用内置 public_key.pem）')
    parser.add_argument('--activation-path', default='',
                        help='激活元数据路径（默认写入用户配置目录）')

    return parser


def main():
    from ai_gen_reimbursement_docs.cli.logging import (
        init_global_logging, setup_logging, write_combined_ai_log,
    )
    from ai_gen_reimbursement_docs.cli.notify import play_notify_sound
    from ai_gen_reimbursement_docs.excel_source import project_root

    from ai_gen_reimbursement_docs.config_utils import load_log_level

    parser = _build_parser()
    args = parser.parse_args()

    log_level = args.log_level or load_log_level()
    init_global_logging(level=log_level)

    if args.max_tokens:
        os.environ['AI_REIMBURSEMENT_MAX_TOKENS'] = args.max_tokens
    logger.debug(f"CLI args: {args}")

    ver = _get_version()
    run_mode = "exe" if getattr(sys, 'frozen', False) else "源码"
    root = os.path.dirname(os.path.dirname(os.path.dirname(__file__)))
    run_path = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else root
    logger.info(f"[CLI] AI生成项目报账文档 v{ver} ({run_mode}: {run_path})")

    from ai_gen_reimbursement_docs.config_utils import config_dir, migrate_config
    logger.info(f"配置文件目录: {config_dir()}")
    migrate_config()

    try:
        from ai_gen_reimbursement_docs.version_check import check_version
        check_version(ver)
    except Exception:
        pass  # 版本检查失败不影响主流程

    # ── 纯 CLI 功能 ──
    if args.test_sound:
        try:
            import winsound
            _ap = os.path.join(root, 'assets', 'audio', 'ticktick_pop.wav')
            if os.path.isfile(_ap):
                winsound.PlaySound(_ap, winsound.SND_FILENAME | winsound.SND_SYNC)
                print("提示音已播放")
            else:
                print(f"音频文件不存在: {_ap}")
        except Exception as e:
            print(f"提示音播放失败: {e}")
        return

    log_root = os.path.join(os.path.expanduser('~'), '.ai-gen-reimbursement-docs', 'log')
    if args.log:
        if args.log == 'open':
            os.startfile(log_root)
            return
        log_files = sorted([f for f in os.listdir(log_root) if f.endswith('.log')], reverse=True)
        if not log_files:
            logger.error("没有找到日志文件")
            return
        latest = os.path.join(log_root, log_files[0])
        if args.log == 'watch':
            try:
                os.system(f'tail -f "{latest}"')
            except Exception:
                os.system(f'powershell -command "Get-Content \\"{latest}\\" -Wait"')
            return
        with open(latest, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        if args.log == 'tail':
            lines = lines[-30:]
        print(''.join(lines))
        return

    if args.version:
        print(f"AI生成项目报账文档 v{_get_version()}")
        return

    if args.history_open:
        try:
            item = get_run(args.history_open, user_history_path())
        except Exception as exc:
            logger.error("运行历史读取失败: %s", exc)
            return
        if item is None:
            logger.error(f"未找到历史记录: {args.history_open}")
            return
        output_dir = item.get("output_dir") or ""
        if not output_dir or not os.path.isdir(output_dir):
            logger.error(f"输出目录不存在: {output_dir or '-'}")
            return
        os.startfile(output_dir)
        return

    if args.history:
        _print_history(limit=args.history_limit, as_json=args.history_json)
        return

    if args.fpa_stability_report:
        from ai_gen_reimbursement_docs.fpa_stability_report import (
            build_fpa_stability_comparison,
            render_fpa_stability_comparison_markdown,
        )

        comparison = build_fpa_stability_comparison(args.fpa_stability_report)
        markdown = render_fpa_stability_comparison_markdown(comparison)
        report_output = args.fpa_stability_output or args.output_dir
        if report_output:
            output_path = os.path.abspath(report_output)
            if os.path.isdir(output_path):
                output_path = os.path.join(output_path, "fpa-stability-report.md")
            os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
            with open(output_path, "w", encoding="utf-8") as f:
                f.write(markdown)
            print(output_path)
        else:
            print(markdown)
        return

    if args.web:
        _auto_init_config(root)
        _start_web_ui(root, port=args.port)
        return

    if args.activate:
        if not args.license:
            logger.error("缺少 --license 参数")
            return
        if not args.license_secret:
            logger.error("缺少 --license-secret 参数")
            return

        try:
            from ai_gen_reimbursement_docs.licensing import activate, load_public_key
            from ai_gen_reimbursement_docs.licensing.exceptions import LicensingError
        except ModuleNotFoundError as exc:
            if exc.name == "cryptography":
                logger.error("激活功能需要安装依赖: cryptography>=41.0")
                return
            raise

        default_data_enc = os.path.join(run_path, 'data.enc')
        default_data_output = os.path.join(run_path, 'data')
        default_public_key = os.path.join(
            os.path.dirname(os.path.dirname(__file__)),
            'licensing',
            'public_key.pem',
        )

        license_path = _Path(args.license)
        data_enc = _Path(args.data_enc or default_data_enc)
        data_output = _Path(args.data_output or default_data_output)
        public_key_path = _Path(args.public_key or default_public_key)
        activation_path = _Path(args.activation_path) if args.activation_path else None

        if not license_path.exists():
            logger.error(f"license 文件不存在: {license_path}")
            return
        if not data_enc.exists():
            logger.error(f"data.enc 不存在: {data_enc}")
            return
        if not public_key_path.exists():
            logger.error(f"公钥文件不存在: {public_key_path}")
            return

        try:
            public_key = load_public_key(public_key_path)
        except Exception as exc:
            logger.error(f"公钥加载失败: {exc}")
            return

        try:
            result = activate(
                license_path=license_path,
                secret=args.license_secret,
                data_enc=data_enc,
                output_dir=data_output,
                public_key=public_key,
                activation_path=activation_path,
            )
        except LicensingError as exc:
            logger.error(f"激活失败: {exc}")
            return

        logger.info(f"激活成功: {result.license_id} / {result.customer}")
        logger.info(f"数据目录: {data_output}")
        logger.info(f"激活元数据: {result.activation_path}")
        return

    if args.init_config:
        from ai_gen_reimbursement_docs.config_utils import (
            DEFAULT_CONFIG_TEMPLATE_FILES,
            copy_default_config_files,
        )
        cfg_dir = _Path(root) / 'config'
        home_cfg = _Path(os.path.expanduser('~')) / '.ai-gen-reimbursement-docs'
        created = set(copy_default_config_files(home_cfg, cfg_dir))
        for _, target_name in DEFAULT_CONFIG_TEMPLATE_FILES:
            dst = home_cfg / target_name
            if dst in created:
                logger.info(f"已创建: {dst}")
            else:
                logger.info(f"已存在，跳过: {dst}")
        logger.info("请编辑 ~/.ai-gen-reimbursement-docs/.env 填入你的 API Key 后使用")
        return

    # ── 配置加载 ──
    api_key = args.api_key or load_api_key()
    base_url = load_base_url()
    model = args.model or load_model_name()
    logger.debug(f"API Key: {'已设置' if api_key else '未设置'}, 端点: {base_url or '默认'}, 模型: {model}")

    from ai_gen_reimbursement_docs.config_utils import load_max_tokens
    logger.info(f"配置: MAX_TOKENS={load_max_tokens()}")

    if args.preview_fpa_module or args.preview_fpa_module_index is not None:
        excel_path = args.from_excel
        if not excel_path:
            logger.error("FPA 预览需要指定 --from-excel")
            return
        if args.from_dir and not os.path.isabs(excel_path):
            excel_path = os.path.join(os.path.abspath(args.from_dir), excel_path)
        if not os.path.exists(excel_path):
            logger.error(f"文件不存在: {excel_path}")
            return
        template_path = args.fpa_out_template if args.fpa_out_template and os.path.exists(args.fpa_out_template) else ""
        if not template_path:
            from ai_gen_reimbursement_docs.pipeline import _resolve_templates
            template_path = _resolve_templates(excel_path, None).get("fpa", "")
        from ai_gen_reimbursement_docs.config_utils import load_fpa_profile
        from ai_gen_reimbursement_docs.gen_fpa import preview_fpa_module
        preview_work_dir = args.output_dir or ""
        result = preview_fpa_module(
            file_path=excel_path,
            module_name=args.preview_fpa_module,
            module_index=args.preview_fpa_module_index,
            api_key=api_key,
            model=model,
            base_url=base_url,
            template_path=template_path,
            profile_name=args.fpa_profile or load_fpa_profile(),
            strategy=args.fpa_strategy,
            rule_set=args.fpa_rule_set,
            work_dir=preview_work_dir,
            use_preview_cache=args.use_preview_cache,
            keep_preview_files=args.keep_preview_files,
        )
        if args.preview_fpa_json:
            print(json.dumps(result, ensure_ascii=False, indent=2))
            return
        module = result["module"]
        print(f"三级模块：{module['l3']}")
        print(f"功能过程数：{module['process_count']}")
        print()
        print("序号  类型  功能点名称                         归类")
        for i, row in enumerate(result["rows"], 1):
            print(f"{i:<5} {row['type']:<5} {row['name']:<32} {row.get('classification_basis') or '-'}")
        if result.get("warnings"):
            print("\nWarnings:")
            for item in result["warnings"]:
                print(f"- {item}")
        if args.keep_preview_files or args.use_preview_cache:
            print(f"\n预览中间文件: {result.get('preview_md_dir')}")
            print(f"使用缓存: {'是' if result.get('preview_cache_used') else '否'}")
        print("\n说明：")
        for i, row in enumerate(result["rows"], 1):
            print(f"[{i}] {row['name']}")
            print(row.get("explanation", ""))
        return

    # ── 测试：调整因子中的可靠性描述 AI 生成 ──
    if args.test_ai_gen_reliability_desc:
        excel_path = args.from_excel
        if not excel_path:
            import glob
            for name in ["功能清单-录入模板.xlsx", "功能清单.xlsx"]:
                matches = glob.glob(name)
                if matches:
                    excel_path = matches[0]
                    break
            if not excel_path:
                logger.error("未指定 --from-excel，且当前目录未找到功能清单文件")
                return
        if not os.path.exists(excel_path):
            logger.error(f"文件不存在: {excel_path}")
            return

        if not api_key:
            logger.error("未配置 API Key")
            return

        descriptions = _read_l3_descriptions_from_excel(excel_path)
        if not descriptions:
            logger.warning("未找到三级模块整体功能描述")
            return

        logger.info(f"共读取到 {len(descriptions)} 条三级模块整体功能描述")

        user_prompt = (
            f"根据功能清单，提取其中涉及与可靠性方面的模块，生成一句关于可靠性业务描述。不少于50字。\n"
            f"功能清单：\n" + '\n'.join(f'- {d}' for d in descriptions)
        )

        from ai_gen_reimbursement_docs.config_utils import load_ai_system_prompt
        system_prompt = load_ai_system_prompt("reliability_desc")
        if not system_prompt:
            logger.warning("未找到 reliability_desc 系统提示词，将使用默认提示词")

        logger.info("用户提示词:\n%s", user_prompt)
        logger.info("系统提示词:\n%s", system_prompt)

        from ai_gen_reimbursement_docs.llm_client import call_llm
        try:
            result = call_llm(
                prompt=user_prompt,
                system=system_prompt,
                api_key=api_key,
                model=model,
                base_url=base_url,
                tag="test_reliability_desc",
            )
            print(result)
            logger.info("可靠性描述生成结果:\n%s", result)
        except Exception as e:
            logger.error("AI 调用失败: %s", e)
            print(f"AI 调用失败: {e}")
        return

    # ── 测试：元数据 #AI生成# 字段 ──
    if args.test_ai_gen_metadata:
        excel_path = args.from_excel
        if not excel_path:
            import glob
            for name in ["功能清单-录入模板.xlsx", "功能清单.xlsx"]:
                matches = glob.glob(name)
                if matches:
                    excel_path = matches[0]
                    break
            if not excel_path:
                logger.error("未指定 --from-excel，且当前目录未找到功能清单文件")
                return
        if not os.path.exists(excel_path):
            logger.error(f"文件不存在: {excel_path}")
            return

        if not api_key:
            logger.error("未配置 API Key")
            return

        field_key = args.test_ai_gen_metadata
        found_sheet, raw_value = _read_meta_field_value(excel_path, field_key)
        if not raw_value:
            all_keys = _read_meta_all_keys(excel_path)
            logger.warning(f"未找到字段「{field_key}」，所有可用字段: {all_keys}")
            print(f"未找到字段「{field_key}」")
            print("所有元数据字段: ")
            for sn, k in all_keys:
                print(f"  [{sn}] {k}")
            return

        from ai_gen_reimbursement_docs.excel_source import strip_ai_marker
        prompt_template, needs_ai = strip_ai_marker(raw_value)
        if not needs_ai:
            logger.warning(f"字段「{field_key}」不含 #AI生成# 标记，当前值: {raw_value}")
            return

        # 解析占位符
        import openpyxl
        from ai_gen_reimbursement_docs.config_utils import load_sheet_names
        _s = load_sheet_names()
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        # Sheet 1: 工单需求-元数据录入
        project_info: dict[str, str] = {}
        for row in wb[_s["work_order_meta"]].iter_rows(min_row=2, values_only=True):
            k = str(row[0]).strip() if row[0] else ""
            v = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if k:
                project_info[k] = v
        # Sheet 3: FPA工作量评估-元数据录入
        fpa_meta: dict[str, str] = {}
        for row in wb[_s["fpa_meta"]].iter_rows(min_row=2, values_only=True):
            k = str(row[0]).strip() if row[0] else ""
            v = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if k:
                fpa_meta[k] = v
        wb.close()

        user_prompt = prompt_template
        user_prompt = user_prompt.replace('${工单编号}', project_info.get('工单编号', ''))
        user_prompt = user_prompt.replace('${工单名称}', project_info.get('工单标题', ''))
        user_prompt = user_prompt.replace('${工单标题}', project_info.get('工单标题', ''))
        user_prompt = user_prompt.replace('${工单内容}', project_info.get('工单内容', ''))
        user_prompt = user_prompt.replace('${子系统（模块）}', fpa_meta.get('子系统（模块）', ''))

        if '${三级模块}' in user_prompt:
            l3_names = _read_l3_names_from_excel(excel_path)
            user_prompt = user_prompt.replace('${三级模块}', '、'.join(l3_names))
        if '${三级模块整体功能描述}' in user_prompt:
            l3_descs = _read_l3_descriptions_from_excel(excel_path)
            user_prompt = user_prompt.replace('${三级模块整体功能描述}',
                                              '\n'.join(f'- {d}' for d in l3_descs))

        from ai_gen_reimbursement_docs.config_utils import load_ai_system_prompt
        system_prompt = load_ai_system_prompt("metadata_gen")
        if not system_prompt:
            logger.warning("未找到 metadata_gen 系统提示词")

        logger.info("字段来源: [%s] %s", found_sheet, field_key)
        logger.info("用户提示词:\n%s", user_prompt)
        logger.info("系统提示词:\n%s", system_prompt)

        from ai_gen_reimbursement_docs.llm_client import call_llm
        try:
            result = call_llm(
                prompt=user_prompt,
                system=system_prompt,
                api_key=api_key,
                model=model,
                base_url=base_url,
                tag=f"test_meta_{field_key}",
            )
            print(result)
            logger.info("AI 生成结果 [%s]:\n%s", field_key, result)
        except Exception as e:
            logger.error("AI 调用失败: %s", e)
            print(f"AI 调用失败: {e}")
        return

    # ── 零参数模式：自动搜索功能清单并全流程执行 ──
    if not any([args.gen_basedata, args.gen_fpa, args.gen_cosmic, args.gen_list,
                args.gen_spec, args.gen_all]):
        _auto_detect_and_run(api_key, model, base_url,
                             search_dir=args.from_dir)
        # 自动检测失败（无可识别文件）时返回，继续往下走，由 gen-* 块报错

    # ── from-excel / from-dir 管道 ──
    if any([args.gen_basedata, args.gen_fpa, args.gen_cosmic, args.gen_list,
             args.gen_spec, args.gen_all]):

        excel_path = args.from_excel
        if not excel_path:
            import glob
            _search = args.from_dir if args.from_dir else "."
            for name in ["功能清单-录入模板.xlsx", "功能清单.xlsx"]:
                matches = glob.glob(os.path.join(_search, name))
                if matches:
                    excel_path = matches[0]
                    break
            if not excel_path:
                _where = args.from_dir if args.from_dir else "当前目录"
                logger.error(f"未指定 --from-excel，且{_where}未找到 功能清单-录入模板.xlsx")
                return
        elif args.from_dir and not os.path.isabs(excel_path):
            # --from-dir + 相对路径 --from-excel → 拼接
            excel_path = os.path.join(os.path.abspath(args.from_dir), excel_path)
        if not os.path.exists(excel_path):
            logger.error(f"文件不存在: {excel_path}")
            return

        excel_dir = os.path.dirname(os.path.abspath(excel_path))
        # --from-dir 决定输出根目录，--output-dir 可覆盖
        output_root = os.path.abspath(args.from_dir) if args.from_dir else excel_dir
        if args.output_dir:
            out_dir = args.output_dir
        elif args.project_name:
            safe = re.sub(r'[\/:*?"<>|]', '_', args.project_name)
            out_dir = os.path.join(output_root, safe)
        else:
            from ai_gen_reimbursement_docs.pipeline import _try_read_project_name
            auto_name = _try_read_project_name(excel_path)
            if auto_name:
                safe = re.sub(r'[\/:*?"<>|]', '_', auto_name)
                out_dir = os.path.join(output_root, safe)
                args.project_name = auto_name
            else:
                out_dir = output_root

        if args.clean and args.project_name:
            safe = re.sub(r'[\/:*?"<>|]', '_', args.project_name)
            target = os.path.join(output_root, safe)
            if os.path.exists(target):
                shutil.rmtree(target)
                logger.info(f"已删除交付物输出目录: {target}")

        mode_map = {
            'gen_all': 'gen-all', 'gen_basedata': 'gen-basedata',
            'gen_fpa': 'gen-fpa', 'gen_cosmic': 'gen-cosmic',
            'gen_list': 'gen-list', 'gen_spec': 'gen-spec',
        }
        mode = 'gen-all'
        for arg_key, mode_val in mode_map.items():
            if getattr(args, arg_key, False):
                mode = mode_val
                break

        templates = {}
        for key, arg_name in [('fpa', 'fpa_out_template'), ('cosmic', 'cosmic_out_template'),
                              ('list', 'list_out_template'), ('spec', 'spec_out_template')]:
            val = getattr(args, arg_name, '').strip()
            if val and os.path.exists(val):
                templates[key] = val

        # 交互式参数（gen-cosmic/gen-all/gen-list 均由 pipeline 内部处理）
        from ai_gen_reimbursement_docs.cli.logging import build_cli_callbacks
        from ai_gen_reimbursement_docs.pipeline import run_pipeline
        from ai_gen_reimbursement_docs.exceptions import CosmicToolError
        run_id = _new_cli_run_id()
        started_at = now_iso()
        _record_cli_history(
            run_id=run_id,
            task_mode=mode,
            run_state="running",
            input_path=excel_path,
            output_dir=out_dir,
            created_at=started_at,
            started_at=started_at,
        )
        try:
            result = run_pipeline(
                mode=mode,
                file_path=excel_path,
                output_dir=out_dir,
                api_key=api_key,
                model=model,
                base_url=base_url,
                project_name=args.project_name,
                templates=templates or None,
                fpa_profile=args.fpa_profile,
                fpa_strategy=args.fpa_strategy,
                fpa_rule_set=args.fpa_rule_set,
                callbacks=build_cli_callbacks(),
            )
        except KeyboardInterrupt:
            _record_cli_history(
                run_id=run_id,
                task_mode=mode,
                run_state="cancelled",
                input_path=excel_path,
                output_dir=out_dir,
                error="cancelled",
                created_at=started_at,
                started_at=started_at,
            )
            print("\n  任务已取消", file=sys.stderr)
            sys.exit(1)
        except CosmicToolError as e:
            _record_cli_history(
                run_id=run_id,
                task_mode=mode,
                run_state="error",
                input_path=excel_path,
                output_dir=out_dir,
                error=str(e),
                created_at=started_at,
                started_at=started_at,
            )
            print(f"\n  错误: {e}", file=sys.stderr)
            play_notify_sound()
            sys.exit(1)
        except Exception as e:
            _record_cli_history(
                run_id=run_id,
                task_mode=mode,
                run_state="error",
                input_path=excel_path,
                output_dir=out_dir,
                error=str(e),
                created_at=started_at,
                started_at=started_at,
            )
            raise

        _summary_files = _summary_files_from_result(result)
        _record_cli_history(
            run_id=run_id,
            task_mode=mode,
            run_state="done",
            input_path=excel_path,
            output_dir=_infer_output_dir(result, out_dir),
            done_files=_done_files_from_result(result),
            created_at=started_at,
            started_at=started_at,
        )
        print()
        for _label, _path in _summary_files:
            if _path and os.path.exists(_path):
                _size = os.path.getsize(_path)
                print(f"  ✅ {_label}: {_path} ({_size/1024:.0f} KB)")
            else:
                print(f"  ⏭️  {_label}: 跳过（已存在或未生成）")
        print()

        write_combined_ai_log(mode)
        play_notify_sound()
        return


if __name__ == '__main__':
    _exit_code = 0
    try:
        main()
    except KeyboardInterrupt:
        print("\n  任务已取消", file=sys.stderr)
    except Exception as _e:
        _exit_code = 1
        print(f"\n  错误: {_e}", file=sys.stderr)
        try:
            logger.debug("未捕获异常", exc_info=True)
        except Exception:
            pass
        try:
            from ai_gen_reimbursement_docs.cli.notify import play_notify_sound
            play_notify_sound()
        except Exception:
            pass
    if getattr(sys, 'frozen', False):
        input("\n按 Enter 键退出...")
    sys.exit(_exit_code)
